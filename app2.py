# ============================================================
# AZIMUT PORTFOLIO ANALYZER v2.0 — app.py
# Aggiornamento: Schede fondi + Rendimenti da FondiDoc FIDA
# ============================================================

import re
import json
from pathlib import Path
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import load_workbook
import zipfile
from xml.etree import ElementTree as ET
import io
import datetime
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import matplotlib
try:
    matplotlib.use('Agg')
except Exception:
    pass
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, Image as RLImage,
    HRFlowable, PageBreak, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import cm

# ── PAGE CONFIG ─────────────────────────────────────────────
st.set_page_config(
    page_title="Azimut | Analisi Portafoglio — AAS Emilia Romagna Marche Umbria",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CONSTANTS ───────────────────────────────────────────────
GROUP_NAMES = {"ALLOCATION", "AZIONARI (LONG)", "BOND"}
COL_A, COL_B, COL_C, COL_G, COL_K, COL_O, COL_R = 0, 1, 2, 6, 10, 14, 17
PROFILES     = ["CONSERVATIVO", "EQUILIBRATO", "ACCRESCITIVO"]
PROFILE_ICONS = {"CONSERVATIVO":"🛡️","EQUILIBRATO":"⚖️","ACCRESCITIVO":"📈"}
PROFILE_W_COL = {"CONSERVATIVO":"w_cons","EQUILIBRATO":"w_equil","ACCRESCITIVO":"w_accr"}

MACRO_COLORS = {
    "Azionari":"#1B4FBB","Bilanciati/Flessibili":"#C9A84C",
    "Obbligazionari":"#2D9D78","Alternativi":"#8B5CF6","Altro":"#94A3B8",
}
SHADES = {
    "Azionari":             ["#0D3080","#1B4FBB","#2563EB","#3B82F6","#60A5FA","#93C5FD","#BFDBFE"],
    "Bilanciati/Flessibili":["#92650A","#B8860B","#C9A84C","#D4B572","#DFC298","#E9CEB4","#F3DACD"],
    "Obbligazionari":       ["#065F46","#14855F","#2D9D78","#34B98A","#6DE5BC","#9AEFD2","#C5F7E7"],
    "Alternativi":          ["#5B21B6","#7C3AED","#8B5CF6","#A78BFA","#C4B5FD","#DDD6FE"],
    "Altro":                ["#475569","#64748B","#94A3B8","#CBD5E1"],
}
DEFAULT_AZ = {"Azionari":0.92,"Bilanciati/Flessibili":0.50,"Obbligazionari":0.06,"Altro":0.50}
FONDIDOC_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "en-US,en;q=0.9,it;q=0.8",
}
# Override for one fund whose FIDA sheet hyperlink points to class B
MANUAL_URL_OVERRIDES = {
    "AZ F.1 All. Balanced FoF A Cap EUR":
        "https://www.fondidoc.it/d/Index/AZPOA/LU0346933400_az-f1-allocation-balanced-fof-a-az-fund-cap-eur",
}

# ── FUND DATA CACHE ──────────────────────────────────────────────────────────
# fund_cache.json is bundled in the repo and updated by the user after a fresh
# FondiDoc fetch (download button → commit to git).
CACHE_FILE = Path("data/fund_cache.json")

def load_fund_cache() -> tuple:
    """Load cached FondiDoc data. Returns (fund_data_dict, last_updated_str)."""
    try:
        if CACHE_FILE.exists():
            # utf-8-sig strips BOM se presente (file creato su Windows)
            payload = json.loads(CACHE_FILE.read_text(encoding="utf-8-sig"))
            return payload.get("fund_data", {}), payload.get("last_updated", "")
    except Exception:
        pass
    return {}, ""

def save_fund_cache(fund_data: dict):
    """Persist fund data to data/fund_cache.json (overwrites)."""
    try:
        CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        # Preserve existing keys (es. ms_data) while updating fund_data
        payload = {}
        if CACHE_FILE.exists():
            try:
                payload = json.loads(CACHE_FILE.read_text(encoding="utf-8-sig"))
            except Exception:
                payload = {}
        payload["last_updated"] = datetime.date.today().isoformat()
        payload["fund_data"] = fund_data
        CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                              encoding="utf-8")
    except Exception:
        pass


def load_ms_cache() -> dict:
    """Load cached Morningstar ratings. Returns {fund_name: {ms_rating, fo_url}}."""
    try:
        if CACHE_FILE.exists():
            payload = json.loads(CACHE_FILE.read_text(encoding="utf-8-sig"))
            return payload.get("ms_data", {})
    except Exception:
        pass
    return {}


def save_ms_cache(ms_data: dict):
    """Persist Morningstar ratings to data/fund_cache.json alongside fund_data."""
    try:
        CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload = {}
        if CACHE_FILE.exists():
            try:
                payload = json.loads(CACHE_FILE.read_text(encoding="utf-8-sig"))
            except Exception:
                payload = {}
        payload["ms_data"] = ms_data
        CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                              encoding="utf-8")
    except Exception:
        pass

# ── UNP/IUNP CATALOG (Catalogo Prodotti&Servizi Azimut, settembre 2025) ──────
# Fonte: DETTAGLIO AZ FUND — valori %UNP e %IUNP36
# Estratto da pdftotext su tutte le pagine del PDF (93 voci).
# Nomi fondi esattamente come nel PDF; _normalize_for_unp() gestisce la
# corrispondenza con i nomi abbreviati dell'Excel (es. "AZ F.1 Bd. ...").
UNP_CATALOG = {
    # ── AZ Allocation ─────────────────────────────────────────────────────────
    "AZ Allocation - Asset Timing 2026":                               (2.29, 1.14),
    "AZ Allocation - Asset Timing 2028":                               (2.29, 1.14),
    "AZ Allocation - Balanced Brave":                                  (1.89, 0.94),
    "AZ Allocation - Balanced FoF":                                    (1.89, 0.94),
    "AZ Allocation - Balanced Plus":                                   (1.89, 0.94),
    "AZ Allocation - Escalator 2026":                                  (2.29, 1.14),
    "AZ Allocation - Escalator 2028":                                  (2.29, 1.14),
    "AZ Allocation - Escalator 2030":                                  (2.29, 1.14),
    "AZ Allocation - Flexible Equity":                                 (1.92, 0.96),
    "AZ Allocation - Global Aggressive":                               (1.89, 0.94),
    "AZ Allocation - Global Balanced":                                 (1.80, 0.90),
    "AZ Allocation - Global Conservative":                             (1.61, 0.80),
    "AZ Allocation - Global Conservative (Classe C)":                  (0.95, 0.47),
    "AZ Allocation - Italian Trend":                                   (2.01, 1.00),
    "AZ Allocation - Life Plan 2040":                                  (2.51, 1.25),
    "AZ Allocation - PIR Italian Excellence 70%":                      (1.89, 0.94),
    "AZ Allocation - Potential Income Upside 2030":                    (1.67, 0.83),
    "AZ Allocation - Risk Parity Factors":                             (1.89, 0.94),
    "AZ Allocation - Trend":                                           (2.01, 1.00),
    "AZ Allocation - Turkey":                                          (1.89, 0.94),
    # ── AZ Alternative ────────────────────────────────────────────────────────
    "AZ Alternative - Capital Enhanced":                               (0.51, 0.25),
    "AZ Alternative - Commodity":                                      (1.81, 0.90),
    # ── AZ Bond ───────────────────────────────────────────────────────────────
    "AZ Bond - Aggregate Bond Euro":                                   (1.10, 0.55),
    "AZ Bond - Asian Bond":                                            (1.38, 0.69),
    "AZ Bond - Bond Value":                                            (1.53, 0.76),
    "AZ Bond - COCO Bonds":                                            (1.53, 0.76),
    "AZ Bond - Convertible":                                           (1.62, 0.81),
    "AZ Bond - Enhanced Yield":                                        (0.35, 0.17),
    "AZ Bond - Euro Corporate":                                        (1.23, 0.61),
    "AZ Bond - Frontier Markets Debt":                                 (1.62, 0.81),
    "AZ Bond - Global Macro Bond":                                     (1.29, 0.64),
    "AZ Bond - High Income FoF":                                       (1.62, 0.81),
    "AZ Bond - High Yield":                                            (1.47, 0.73),
    "AZ Bond - High Yield Target 2028 Climate Transition":             (1.53, 0.76),
    "AZ Bond - High Yield Target 2028 Climate Transition (Classe C)":  (0.52, 0.26),
    "AZ Bond - Income Dynamic":                                        (0.99, 0.49),
    "AZ Bond - International FoF":                                     (1.62, 0.81),
    "AZ Bond - Latin America Bonds":                                   (1.53, 0.76),
    "AZ Bond - Patriot":                                               (1.36, 0.68),
    "AZ Bond - Renminbi Opportunities":                                (1.23, 0.61),
    "AZ Bond - Short Term Investment Grade Climate Transition":        (1.53, 0.76),
    "AZ Bond - Short Term Investment Grade Climate Transition (Classe C)": (0.43, 0.21),
    "AZ Bond - Sustainable Hybrid":                                    (1.46, 0.73),
    "AZ Bond - Target 2025":                                           (1.11, 0.55),
    "AZ Bond - Target 2026":                                           (1.11, 0.55),
    "AZ Bond - Target 2028":                                           (1.11, 0.55),
    "AZ Bond - Target 2029":                                           (1.11, 0.55),
    "AZ Bond - Target 2029 USD":                                       (1.23, 0.61),
    "AZ Bond - Target 2031":                                           (1.11, 0.55),
    "AZ Bond - Total Return Bond":                                     (1.55, 0.77),
    "AZ Bond - US Dollar Aggregate":                                   (1.24, 0.62),
    # ── AZ Equity ─────────────────────────────────────────────────────────────
    "AZ Equity - Al Mal Mena":                                         (2.51, 1.25),
    "AZ Equity - American Opportunities":                              (2.19, 1.09),
    "AZ Equity - ASEAN Countries":                                     (2.19, 1.09),
    "AZ Equity - Best Value":                                          (2.09, 1.04),
    "AZ Equity - Biotechnology":                                       (2.51, 1.25),
    "AZ Equity - Borletti Global Lifestyle":                           (2.30, 1.15),
    "AZ Equity - Brazil Trend":                                        (2.19, 1.09),
    "AZ Equity - China":                                               (2.19, 1.09),
    "AZ Equity - Egypt":                                               (2.51, 1.25),
    "AZ Equity - Emerging Asia FoF":                                   (2.51, 1.25),
    "AZ Equity - Emerging Markets Technology":                         (2.51, 1.25),
    "AZ Equity - Escalator":                                           (2.29, 1.14),
    "AZ Equity - Europe":                                              (2.19, 1.09),
    "AZ Equity - Food & Agriculture":                                  (2.40, 1.20),
    "AZ Equity - Global Dividend":                                     (2.51, 1.25),
    "AZ Equity - Global Emerging FoF":                                 (2.51, 1.25),
    "AZ Equity - Global ESG":                                          (2.51, 1.25),
    "AZ Equity - Global FoF":                                          (2.51, 1.25),
    "AZ Equity - Global Growth":                                       (2.40, 1.20),
    "AZ Equity - Global Healthcare":                                   (2.40, 1.20),
    "AZ Equity - Global Infrastructure":                               (2.26, 1.13),
    "AZ Equity - Global Quality":                                      (2.18, 1.09),
    "AZ Equity - Global Value FoF":                                    (2.51, 1.25),
    "AZ Equity - Industrial Revolution 4.0":                           (2.51, 1.25),
    "AZ Equity - Japan":                                               (2.20, 1.10),
    "AZ Equity - Mexico":                                              (2.51, 1.25),
    "AZ Equity - Momentum":                                            (2.19, 1.09),
    "AZ Equity - Small Cap Europe FoF":                                (2.51, 1.25),
    "AZ Equity - Special Needs & Inclusion":                           (2.51, 1.25),
    "AZ Equity - Water & Renewable Resources":                         (2.40, 1.20),
    "AZ Equity - World Minimum Volatility":                            (2.19, 1.09),
    # ── AZ Islamic ────────────────────────────────────────────────────────────
    "AZ Islamic - Global Sukuk":                                       (1.36, 0.68),
    # ── Azimut Thematic Fund ──────────────────────────────────────────────────
    "Azimut Thematic Fund - AZ Allocation - Global Goals":             (2.50, 1.25),
    "Azimut Thematic Fund - AZ Equity - New Generation":               (2.79, 1.39),
    "Azimut Thematic Fund - AZ Equity - Space":                        (2.79, 1.39),
    # ── Fondi Economia Reale (sezione separata nel catalogo) ──────────────────
    "AZ Allocation - Italian Long-Term Opp.":                          (2.73, 1.36),
    "AZ Allocation - Long Term Credit Opp.":                           (1.94, 0.97),
    "AZ Allocation - Long-Term Equity Opp.":                           (2.73, 1.36),
    "AZ Bond - ABS":                                                   (1.23, 0.61),
    "AZ Equity - Future Opportunities":                                (2.51, 1.25),
}

# CSS is injected inside main() to keep it within the error-handler scope.


# ════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════

def get_macro(cat: str) -> str:
    if not cat or cat == "-": return "Altro"
    c = cat.lower()
    if "azionari" in c or "equity" in c: return "Azionari"
    if any(x in c for x in ["obbligazionari","bond","credit","debt","sukuk","reddito"]): return "Obbligazionari"
    if any(x in c for x in ["bilanciati","allocation","flessibili","balanced","flexible","prudenti","moderati"]): return "Bilanciati/Flessibili"
    if any(x in c for x in ["alternativi","alternative","commodity"]): return "Alternativi"
    return "Altro"


def assign_colors(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    counter = {k: 0 for k in SHADES}
    colors = []
    for _, row in df.iterrows():
        mc = row.get("macro_cat", "Altro")
        shades = SHADES.get(mc, SHADES["Altro"])
        colors.append(shades[counter.get(mc, 0) % len(shades)])
        counter[mc] = counter.get(mc, 0) + 1
    df["color"] = colors
    return df


def pct_color(val_str: str) -> str:
    """Returns 'pos', 'neg', or 'neu' based on value sign"""
    try:
        v = float(val_str.replace("%","").replace(",",".").strip())
        return "pos" if v > 0 else ("neg" if v < 0 else "neu")
    except:
        return "neu"


# ── UNP/IUNP lookup helpers ──────────────────────────────────

def _normalize_for_unp(name: str) -> str:
    """Normalise a fund name for UNP/factbook catalog lookup.

    Works across all three name formats:
    - Excel  : "AZ F.1 Bd. Global Macro Bond A Cap EUR"  (abbreviation + share class)
    - Excel  : "AZ F.1 Bd Global Macro Bond A Cap EUR"   (abbreviation without period)
    - PDF    : "AZ Bond - Global Macro Bond"              (full name with dash)
    - Web    : "AZ BOND - GLOBAL MACRO BOND"              (full name, ALL CAPS)

    All four normalise to → "az bond global macro bond"
    """
    n = name.strip()
    # ── 1. Expand AZ F.1 abbreviations (period OPTIONAL after family code) ────
    n = re.sub(r'AZ\s+F\.1\s+All\.?\s*',  'AZ Allocation ',  n, flags=re.IGNORECASE)
    n = re.sub(r'AZ\s+F\.1\s+Eq\.?\s*',   'AZ Equity ',      n, flags=re.IGNORECASE)
    n = re.sub(r'AZ\s+F\.1\s+Bd\.?\s*',   'AZ Bond ',        n, flags=re.IGNORECASE)
    n = re.sub(r'AZ\s+F\.1\s+Alt\.?\s*',  'AZ Alternative ', n, flags=re.IGNORECASE)
    n = re.sub(r'AZ\s+F\.1\s+Isl\.?\s*',  'AZ Islamic ',     n, flags=re.IGNORECASE)
    # ── 2. Strip "(Classe C)" / "(classe c)" variants ─────────────────────────
    n = re.sub(r'\s*\(clas[se]+\s+c\)\s*$', '', n, flags=re.IGNORECASE)
    # ── 3. Strip share-class suffix: " A Cap EUR", " B Acc USD", "A-HU Cap EUR Hdg" etc.
    n = re.sub(r'\s+[A-Z][-\w]*\s+(Cap|Acc|Dis|Inc)\s+\w{3}(\s+Hdg)?\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s+[A-Z]\s+(Cap|Acc|Dis|Inc)\s*$',                          '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s+Cap\s+\w{3}\s*$',                                         '', n, flags=re.IGNORECASE)
    # ── 4. Normalise dashes, ampersands, dots and remaining punctuation ────────
    n = re.sub(r'[-–—]', ' ', n)
    n = re.sub(r'[&]',   ' ', n)
    n = re.sub(r'[^\w\s]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip().lower()
    return n


_UNP_CATALOG_NORMALIZED: dict = {
    _normalize_for_unp(k): v for k, v in UNP_CATALOG.items()
}

# Alias table: normalised Excel name → normalised catalog name
# Needed when the Excel name differs semantically from the catalog (e.g. "Hybrids" vs
# "Sustainable Hybrid") or when slight wording changes prevent automatic substring match.
_FUND_ALIASES: dict = {
    "az bond hybrids":              "az bond sustainable hybrid",
    "az bond sustainable hybrids":  "az bond sustainable hybrid",
    # Long-Term Opp variants (catalog uses abbreviated "Opp.")
    "az allocation long term credit opportunities":  "az allocation long term credit opp",
    "az allocation long term equity opportunities":  "az allocation long term equity opp",
    "az allocation italian long term opportunities": "az allocation italian long term opp",
    "az allocation italian long-term opportunities": "az allocation italian long term opp",
}

# ── Credit rating numeric scale (AAA=1 … D=22) for weighted averages ─────────
RATING_SCALE: dict = {
    'AAA': 1,  'AA+': 2,  'AA': 3,  'AA-': 4,
    'A+':  5,  'A':   6,  'A-': 7,
    'BBB+':8,  'BBB': 9,  'BBB-':10,
    'BB+':11,  'BB': 12,  'BB-': 13,
    'B+': 14,  'B':  15,  'B-':  16,
    'CCC+':17, 'CCC':18,  'CCC-':19,
    'CC': 20,  'C':  21,  'D':   22,
}
RATING_INVERSE: dict = {v: k for k, v in RATING_SCALE.items()}


def lookup_unp(fund_name: str):
    """Return (unp_pct, iunp36_pct) for a fund, or (None, None) if not found."""
    norm = _normalize_for_unp(fund_name)
    norm = _FUND_ALIASES.get(norm, norm)          # apply alias if any
    # 1. exact match
    if norm in _UNP_CATALOG_NORMALIZED:
        return _UNP_CATALOG_NORMALIZED[norm]
    # 2. longest substring match (catalog key inside fund name or vice versa)
    best, best_len = None, 0
    for cat_key, val in _UNP_CATALOG_NORMALIZED.items():
        if cat_key in norm or norm in cat_key:
            if len(cat_key) > best_len:
                best, best_len = val, len(cat_key)
    return best if best is not None else (None, None)


# ════════════════════════════════════════════════════════════
# DATA PARSING
# ════════════════════════════════════════════════════════════

def _parse_excel_impl(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    out = {}
    for sname in ["PTF FULL", "PTF SHORT"]:
        if sname in wb.sheetnames:
            out[sname] = _parse_ptf(wb, sname)
    if "FIDA" in wb.sheetnames:
        out["FIDA"] = _parse_fida(wb)
    wb.close()
    out["fida_urls"] = extract_fida_urls(file_bytes)
    return out

# Apply cache decorator safely at module load time — if it fails (Streamlit
# version mismatch or hasher bug), fall back to uncached version.
try:
    parse_excel = st.cache_data(show_spinner=False)(_parse_excel_impl)
except Exception:
    parse_excel = _parse_excel_impl


def parse_factbook(pdf_bytes: bytes) -> dict:
    """Parse the AZ Investments factbook PDF performance summary tables.

    Returns {normalized_fund_name: {ytd, perf_1y, perf_3y, perf_5y}}

    Summary table columns: Fund | AUM | 1M | 3M | 6M | 12M | 24M | 36M | 60M | YTD
    Risk metrics (vol, Sharpe, Sortino) are NOT in the factbook — those still
    come from FondiDoc.
    """
    try:
        import pdfplumber
    except ImportError:
        return {}

    result: dict = {}
    _metrics: dict = {}   # {norm_fund: {credit_rating, duration, ytm}}

    def _to_pct(s):
        """Convert a factbook value string to a signed '%' string, or None."""
        if not s:
            return None
        s = str(s).strip()
        if s in ('-', 'n.d.', 'N/D', '', 'None', 'n/d', 'nd'):
            return None
        s = s.replace('−', '-').replace('–', '-').replace('—', '-')
        s = s.replace(',', '.').replace('%', '').strip()
        try:
            return f"{float(s):+.2f}%"
        except Exception:
            return None

    def _annualize(pct_str, years: float):
        """Convert a cumulative return string to annualized (CAGR)."""
        if not pct_str:
            return None
        try:
            v = float(pct_str.replace('%','').replace('+','').strip()) / 100
            ann = ((1 + v) ** (1 / years) - 1) * 100
            return f"{ann:+.2f}%"
        except Exception:
            return None

    def _plausible(pct_str, max_abs: float = 80.0):
        """Return True if pct_str is None OR in [-max_abs, +max_abs]."""
        if pct_str is None:
            return True
        try:
            return abs(float(pct_str.replace('%','').replace('+',''))) <= max_abs
        except Exception:
            return True  # keep on parse error

    def _get_fi_from_src(src: str):
        """Extract Credit Rating, YTM, Duration from one pdfplumber text source.
        Defined once here (not inside the page loop) to avoid repeated closure
        creation and potential caching issues.
        """
        _lines = [ln.strip() for ln in src.split('\n') if ln.strip()]
        _txt   = '\n'.join(_lines)
        _r = _y = _d = None
        _RATING_PAT = (
            r'(AAA|AA[+\-]|AA|A[+\-]|A'
            r'|BBB[+\-]|BBB|BB[+\-]|BB|B[+\-]|B'
            r'|CCC[+\-]|CCC|CC|C|D)'
            r'(?=[^A-Za-z]|$)')
        # Credit Rating – inline ("...Credit Rating medio  BB+")
        #              or next-line ("...Credit Rating medio\n  BB+")
        _m = re.search(
            r'Credit\s+Rating[^\n]{0,60}?' + _RATING_PAT,
            _txt, re.IGNORECASE | re.MULTILINE)
        if not _m:
            _m = re.search(
                r'Credit\s+Rating[^\n]*\n\s*' + _RATING_PAT,
                _txt, re.IGNORECASE)
        if _m and _m.group(1).upper() in RATING_SCALE:
            _r = _m.group(1).upper()
        # YTM – inline or next-line
        _m = re.search(
            r'Yield\s+To\s+Maturity[^\n]{0,60}?([\d]+[,\.][\d]+)\s*%',
            _txt, re.IGNORECASE | re.MULTILINE)
        if not _m:
            _m = re.search(
                r'Yield\s+To\s+Maturity[^\n]*\n\s*([\d]+[,\.][\d]+)\s*%',
                _txt, re.IGNORECASE)
        if _m:
            try:
                _y = round(float(_m.group(1).replace(',', '.')), 2)
            except Exception:
                pass
        # Duration – scan up to 10 lines after "Portfolio Duration" label
        _dn = next(
            (i for i, ln in enumerate(_lines)
             if re.search(r'portfolio\s+duration', ln, re.IGNORECASE)), None)
        if _dn is not None:
            for _dl in _lines[_dn:_dn + 10]:
                if re.search(r'portfolio\s+duration', _dl, re.IGNORECASE):
                    _dm = re.search(
                        r'Duration\D{0,20}([\d]+[,\.][\d]+)(?!\s*%)',
                        _dl, re.IGNORECASE)
                else:
                    _dm = re.search(r'^([\d]+[,\.][\d]+)\s*$', _dl)
                    if not _dm:
                        _dm = re.search(
                            r'(?<![%\d,\.])([\d]{1,2}[,\.][\d]{2})\s*$', _dl)
                if _dm:
                    try:
                        _v = float(_dm.group(1).replace(',', '.'))
                        if 0 < _v < 40:
                            _d = round(_v, 2)
                            break
                    except Exception:
                        pass
        return _r, _y, _d

    def _store(name_raw: str, cols_vals: list):
        """Normalise name and store performance data if not already seen.

        cols_vals: [1M, 3M, 6M, 12M, 24M, 36M, 60M, YTD]
        3Y (36M) and 5Y (60M) are stored as ANNUALIZED CAGR to match FondiDoc.
        """
        if not name_raw:
            return
        norm = _normalize_for_unp(name_raw.strip())
        norm = _FUND_ALIASES.get(norm, norm)
        if not norm or norm in result:
            return
        # ── extract raw cumulative values ─────────────────────────────
        n = len(cols_vals)
        ytd_raw = _to_pct(cols_vals[7]) if n > 7 else _to_pct(cols_vals[-1] if cols_vals else None)
        p1y_raw = _to_pct(cols_vals[3]) if n > 3 else None
        p3y_cum = _to_pct(cols_vals[5]) if n > 5 else None   # 36M cumulative
        p5y_cum = _to_pct(cols_vals[6]) if n > 6 else None   # 60M cumulative

        # ── sanity check: YTD and 1Y must be in ±80 % range ──────────
        # Values outside this range indicate NAV prices, year numbers, or
        # misaligned columns from individual fund pages (not summary table).
        if not _plausible(ytd_raw, 80) or not _plausible(p1y_raw, 80):
            return

        # ── annualize 3Y and 5Y (factbook stores cumulative) ─────────
        p3y = _annualize(p3y_cum, 3)
        p5y = _annualize(p5y_cum, 5)

        if any(v is not None for v in (ytd_raw, p1y_raw, p3y, p5y)):
            result[norm] = {
                "ytd": ytd_raw, "perf_1y": p1y_raw,
                "perf_3y": p3y, "perf_5y": p5y,
            }

    _ref_date: str = ""   # data di riferimento estratta dal frontespizio

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_idx, page in enumerate(pdf.pages):

                # ── A: structured table extraction (best-case) ────────────
                for tbl in (page.extract_tables() or []):
                    for row in tbl:
                        if not row or len(row) < 6:
                            continue
                        cell0 = str(row[0] or '').replace('\n', ' ').strip()
                        if not re.match(r'^(AZ\b|AZIMUT\b)', cell0, re.IGNORECASE):
                            continue
                        # row layout: [name, AUM, 1M, 3M, 6M, 12M, 24M, 36M, 60M, YTD]
                        data_cols = [str(c or '').strip() for c in row[2:]]
                        _store(cell0, data_cols)

                # ── B: text-line fallback ──────────────────────────────────
                text = page.extract_text() or ""

                # Estrai data di riferimento dal frontespizio (prime 5 pagine)
                if page_idx < 5 and not _ref_date:
                    # Cerca GG/MM/AAAA oppure GG.MM.AAAA
                    _dm = re.search(r'\b(\d{1,2})[/.](\d{2})[/.](\d{4})\b', text)
                    if _dm:
                        _ref_date = f"{_dm.group(1).zfill(2)}/{_dm.group(2)}/{_dm.group(3)}"
                    else:
                        # Cerca "31 marzo 2026" / "31 March 2026"
                        _months = {
                            "gennaio":"01","february":"02","febbraio":"02","march":"03",
                            "marzo":"03","april":"04","aprile":"04","may":"05","maggio":"05",
                            "june":"06","giugno":"06","july":"07","luglio":"07",
                            "august":"08","agosto":"08","september":"09","settembre":"09",
                            "october":"10","ottobre":"10","november":"11","novembre":"11",
                            "december":"12","dicembre":"12","january":"01","gennaio":"01",
                        }
                        _mp = r'\b(\d{1,2})\s+(' + '|'.join(_months) + r')\s+(20\d{2})\b'
                        _dm2 = re.search(_mp, text, re.IGNORECASE)
                        if _dm2:
                            _mn = _months.get(_dm2.group(2).lower(), "??")
                            _ref_date = f"{_dm2.group(1).zfill(2)}/{_mn}/{_dm2.group(3)}"

                for line in text.split('\n'):
                    line = line.strip()
                    if not re.match(r'^(AZ\b|AZIMUT\b)', line, re.IGNORECASE):
                        continue
                    tokens = line.split()
                    if len(tokens) < 6:
                        continue
                    # Find first all-digit token (= AUM in M€, e.g. "81", "1050")
                    # Skip year-like tokens (2010-2030) which appear in annual-return
                    # tables on individual fund pages and would misalign columns.
                    num_idx = next(
                        (i for i, t in enumerate(tokens)
                         if re.match(r'^\d{1,6}$', t)
                         and i > 0
                         and not re.match(r'^20[12]\d$', t)),  # skip 2010-2029
                        None
                    )
                    if num_idx is None:
                        continue
                    name_raw = ' '.join(tokens[:num_idx])
                    # tokens after AUM: [1M, 3M, 6M, 12M, 24M, 36M, 60M, YTD]
                    data_cols = tokens[num_idx + 1:]
                    _store(name_raw, data_cols)

                # ── C: METRICHE FIXED INCOME ─────────────────────────────
                # Strategy: try BOTH pdfplumber text formats.
                #  • default extract_text() merges same-Y chars into one line
                #  • layout=True  preserves spatial layout (like pdftotext -l)
                # Trigger on section header OR individual field names so we
                # never miss a page where only part of the right column is
                # captured.
                _FI_FIELDS = ('METRICHE FIXED INCOME', 'CREDIT RATING MEDIO',
                              'PORTFOLIO DURATION', 'YIELD TO MATURITY')
                _text_layout = ""
                try:
                    _text_layout = page.extract_text(layout=True) or ""
                except Exception:
                    pass
                _text_combined = text + "\n" + _text_layout

                if any(f in _text_combined.upper() for f in _FI_FIELDS):
                    # Use whichever text source is richer (more chars)
                    _txt_c = _text_layout if len(_text_layout) > len(text) \
                        else text
                    _lines_p = _txt_c.split('\n')
                    _fn = None
                    for _i, _ln in enumerate(_lines_p[:40]):
                        _ls = _ln.strip()
                        # Two-line: "AZ BOND" then "HIGH YIELD" on next line
                        if re.match(
                                r'^AZ\s+(BOND|ALLOCATION|EQUITY|ALTERNATIVE|ISLAMIC)\s*$',
                                _ls, re.IGNORECASE):
                            for _j in range(_i + 1, min(_i + 6, len(_lines_p))):
                                _nl = _lines_p[_j].strip()
                                if _nl and not re.match(
                                        r'^([A-Z]{2}\d{10}|ISIN|\d+\s*$)',
                                        _nl, re.IGNORECASE):
                                    _nrm = _normalize_for_unp(_ls + ' ' + _nl)
                                    _fn  = _FUND_ALIASES.get(_nrm, _nrm)
                                    break
                            if _fn:
                                break
                        # Single-line: "AZ BOND - ENHANCED YIELD"
                        _m1 = re.match(
                            r'^(AZ\s+(?:BOND|ALLOCATION|EQUITY|ALTERNATIVE|ISLAMIC))'
                            r'\s*[-–]?\s+(.+)$', _ls, re.IGNORECASE)
                        if _m1 and len(_m1.group(2).strip()) > 2:
                            _sub = _m1.group(2).strip()
                            if not re.match(r'^(\d+|[A-Z]{2}\d{10})', _sub):
                                _nrm = _normalize_for_unp(
                                    _m1.group(1).strip() + ' ' + _sub)
                                _fn  = _FUND_ALIASES.get(_nrm, _nrm)
                                break
                    if _fn:
                        # _get_fi_from_src is defined once above the page loop.
                        # Try layout text first (richer for two-column pages),
                        # then default text; merge – prefer first non-None.
                        _cr1, _yt1, _dur1 = (
                            _get_fi_from_src(_text_layout)
                            if _text_layout else (None, None, None))
                        _cr2, _yt2, _dur2 = _get_fi_from_src(text)

                        _cr  = _cr1  or _cr2
                        _yt  = _yt1  if _yt1  is not None else _yt2
                        _dur = _dur1 if _dur1 is not None else _dur2

                        if _fn not in _metrics:
                            _metrics[_fn] = {}
                        if _cr:              _metrics[_fn]['credit_rating'] = _cr
                        if _yt  is not None: _metrics[_fn]['ytm']           = _yt
                        if _dur is not None: _metrics[_fn]['duration']      = _dur

                # ── D: SCOMPOSIZIONE PORTAFOGLIO - ASSET CLASS ────────────
                # Only AZ ALLOCATION / BALANCED funds have this section.
                # AZ EQUITY and AZ BOND funds do NOT → Excel binary (0/1) is
                # the correct fallback for them; we don't overwrite here.
                #
                # pdfplumber layout quirks on this section:
                #  • "Equity" label appears on its own line WITHOUT a value.
                #    Its value (e.g. "54%") appears AFTER the x-axis scale
                #    labels (-10%, 0%, 10%, …40%) because the Equity bar
                #    physically extends to the right of the shorter bars.
                #    Marker: "Cash Offset" ends the data rows; the first
                #    non-round (not multiple of 10) positive % after that
                #    is the Equity value.
                #  • Bond components ("Sovereign", "Corporate", …) appear
                #    either inline ("Sovereign  28%") or in two lines
                #    ("Sovereign\n28%").  We sum them for fb_obb_pct.
                #  • Two-column merge: pdfplumber may merge rows from both
                #    columns, so we use re.search (CONTAINS) not re.match.
                if re.search(
                        r'SCOMPOSIZIONE\s+PORTAFOGLIO.*?ASSET\s+CLASS',
                        text, re.IGNORECASE):
                    _lp_d = text.split('\n')
                    _fn_d = None
                    for _i_d, _ln_d in enumerate(_lp_d[:25]):
                        _ls_d = _ln_d.strip()
                        if re.match(
                                r'^AZ\s+(BOND|ALLOCATION|EQUITY|ALTERNATIVE|ISLAMIC)\s*$',
                                _ls_d, re.IGNORECASE):
                            for _j_d in range(_i_d + 1, min(_i_d + 5, len(_lp_d))):
                                _nl_d = _lp_d[_j_d].strip()
                                if _nl_d and not re.match(
                                        r'^([A-Z]{2}\d{10}|ISIN|\d+\s*$)',
                                        _nl_d, re.IGNORECASE):
                                    _nrm_d = _normalize_for_unp(_ls_d + ' ' + _nl_d)
                                    _fn_d  = _FUND_ALIASES.get(_nrm_d, _nrm_d)
                                    break
                            if _fn_d:
                                break
                        _m1_d = re.match(
                            r'^(AZ\s+(?:BOND|ALLOCATION|EQUITY|ALTERNATIVE|ISLAMIC))'
                            r'\s*[-–]?\s+(.+)$', _ls_d, re.IGNORECASE)
                        if _m1_d and len(_m1_d.group(2).strip()) > 2:
                            _sub_d = _m1_d.group(2).strip()
                            if not re.match(r'^(\d+|[A-Z]{2}\d{10})', _sub_d):
                                _nrm_d = _normalize_for_unp(
                                    _m1_d.group(1).strip() + ' ' + _sub_d)
                                _fn_d  = _FUND_ALIASES.get(_nrm_d, _nrm_d)
                                break
                    if _fn_d:
                        # Locate the section header
                        _ALLOC_SEC_RE = re.compile(
                            r'SCOMPOSIZIONE\s+PORTAFOGLIO.*?ASSET\s+CLASS',
                            re.IGNORECASE)
                        _sec_start = next(
                            (i for i, l in enumerate(_lp_d)
                             if _ALLOC_SEC_RE.search(l)), None)
                        if _sec_start is not None:
                            # Bond-component keywords (English, as used in factbook)
                            _BD_KW = re.compile(
                                r'\b(sovereign|government|corporate|aggregate|'
                                r'dm\s+corporate|em\s+sovereign|'
                                r'eu\s+sovereign|us\s+sovereign|'
                                r'high\s+yield|convertible|'
                                r'fixed\s+income)\b', re.IGNORECASE)
                            _EQ_KW   = re.compile(r'\bequity\b', re.IGNORECASE)
                            _SKIP_KW = re.compile(
                                r'\b(cash|mixed|commodity|mmkt|offset|'
                                r'allocation)\b', re.IGNORECASE)
                            _STOP_RE = re.compile(
                                r'SCOMPOSIZIONE\s+(OBBLIGAZIONARIA|AZIONARIA)'
                                r'|SHARE\s+CLASS',
                                re.IGNORECASE)
                            # Inline: label + % on same line (handles merged cols)
                            # Decimal separator is optional (e.g. "28%" and "28,5%")
                            _INLINE = re.compile(
                                r'\b(equity|sovereign|government|corporate|aggregate|'
                                r'dm\s+corporate|em\s+sovereign|'
                                r'eu\s+sovereign|us\s+sovereign|'
                                r'high\s+yield|convertible|fixed\s+income)\b'
                                r'.*?(-?[\d]+(?:[,\.]\d*)?)\s*%',
                                re.IGNORECASE)
                            # Standalone %: "-10%", "0%", "54%", "28%" …
                            _PCT_SA = re.compile(
                                r'^(-?[\d]+(?:[,\.][\d]+)?)\s*%\s*$')

                            _equity_v   = None
                            _bond_sum   = 0.0
                            _last_label = None   # 'equity' | 'bond' | None
                            _saw_offset = False

                            for _ll2 in _lp_d[_sec_start: _sec_start + 80]:
                                _ll2s = _ll2.strip()
                                if not _ll2s:
                                    continue
                                if _STOP_RE.search(_ll2s):
                                    break

                                # Cash Offset marks end of data rows
                                if re.search(r'cash\s+offset', _ll2s,
                                             re.IGNORECASE):
                                    _saw_offset = True
                                    _last_label = None
                                    continue

                                # Skip non-data rows (Cash, Mixed, MMkt…)
                                if _SKIP_KW.search(_ll2s):
                                    _last_label = None
                                    continue

                                # ── Try inline: keyword + % on same line ──
                                _mi = _INLINE.search(_ll2s)
                                if _mi:
                                    _lk = _mi.group(1).lower()
                                    try:
                                        _pv = float(
                                            _mi.group(2).replace(',', '.'))
                                    except Exception:
                                        _last_label = None
                                        continue
                                    if _pv <= 0:
                                        _last_label = None
                                        continue
                                    if _EQ_KW.search(_lk) \
                                            and _equity_v is None:
                                        _equity_v = _pv / 100
                                    elif _BD_KW.search(_lk):
                                        _bond_sum += _pv / 100
                                    _last_label = None
                                    continue

                                # ── Keyword-only line (no % found) ──
                                _pm2 = _PCT_SA.match(_ll2s)
                                if not _pm2:
                                    if _EQ_KW.search(_ll2s):
                                        _last_label = 'equity'
                                    elif _BD_KW.search(_ll2s) \
                                            and not _SKIP_KW.search(_ll2s):
                                        _last_label = 'bond'
                                    else:
                                        _last_label = None
                                    continue

                                # ── Standalone % line ──
                                try:
                                    _pv = float(
                                        _pm2.group(1).replace(',', '.'))
                                except Exception:
                                    _last_label = None
                                    continue

                                if _last_label == 'equity' \
                                        and _equity_v is None \
                                        and _pv > 0:
                                    _equity_v = _pv / 100
                                    _last_label = None
                                elif _last_label == 'bond' and _pv > 0:
                                    _bond_sum += _pv / 100
                                    _last_label = None
                                elif _saw_offset \
                                        and _equity_v is None \
                                        and _pv > 0:
                                    # After Cash Offset the Equity bar value
                                    # appears here; axis labels are multiples
                                    # of 10 (or negative), so skip those
                                    if _pv > 0 and round(_pv) % 10 != 0:
                                        _equity_v = _pv / 100
                                else:
                                    _last_label = None

                            # Persist valid values
                            _ok_eq = (_equity_v is not None
                                      and 0.0 < _equity_v <= 1.0)
                            _ok_bd = _bond_sum > 0.005
                            if _ok_eq or _ok_bd:
                                if _fn_d not in _metrics:
                                    _metrics[_fn_d] = {}
                                if _ok_eq \
                                        and 'fb_az_pct' not in _metrics[_fn_d]:
                                    _metrics[_fn_d]['fb_az_pct'] = \
                                        round(_equity_v, 4)
                                if _ok_bd \
                                        and 'fb_obb_pct' not in _metrics[_fn_d]:
                                    _metrics[_fn_d]['fb_obb_pct'] = \
                                        round(_bond_sum, 4)

    except Exception:
        pass

    # Unisci metriche fixed income nelle voci del result
    for _fn, _md in _metrics.items():
        if _fn in result:
            result[_fn].update(_md)
        else:
            result[_fn] = dict(_md)

    # Aggiungi metadato data di riferimento (chiave speciale)
    if _ref_date:
        result["_ref_date"] = _ref_date

    return result

try:
    parse_factbook = st.cache_data(show_spinner=False)(parse_factbook)
except Exception:
    pass  # run uncached if decorator fails


# ── Factbook Excel export / import ──────────────────────────────────────────

_FB_COLS = [
    ("fondo",         "Fondo (chiave normalizzata)"),
    ("credit_rating", "Credit Rating Medio"),
    ("duration",      "Duration (anni)"),
    ("ytm",           "YTM (%)"),
    ("az_pct",        "% Azionario (0-1)"),
    ("obb_pct",       "% Obbligazionario (0-1)"),
    ("ytd",           "YTD"),
    ("perf_1y",       "Rend. 1A"),
    ("perf_3y",       "Rend. 3A ann."),
    ("perf_5y",       "Rend. 5A ann."),
]
_FB_INTERNAL_KEYS = {
    "az_pct":  "fb_az_pct",
    "obb_pct": "fb_obb_pct",
}


def factbook_to_excel_bytes(fb: dict) -> bytes:
    """Serialise the factbook data dict to a downloadable Excel file.

    The Excel has one row per fund with human-readable column headers.
    Numeric fields (duration, ytm, az_pct, obb_pct) are stored as floats.
    The 'fondo' column is the normalised key used internally for matching.
    """
    rows = []
    for norm, data in fb.items():
        if norm == "_ref_date" or not isinstance(data, dict):
            continue
        row = {"fondo": norm}
        row["credit_rating"] = data.get("credit_rating", "")
        row["duration"]      = data.get("duration", "")
        row["ytm"]           = data.get("ytm", "")
        row["az_pct"]        = data.get("fb_az_pct", "")
        row["obb_pct"]       = data.get("fb_obb_pct", "")
        row["ytd"]           = data.get("ytd", "")
        row["perf_1y"]       = data.get("perf_1y", "")
        row["perf_3y"]       = data.get("perf_3y", "")
        row["perf_5y"]       = data.get("perf_5y", "")
        rows.append(row)

    int_keys = [k for k, _ in _FB_COLS]
    hdr_map  = {k: h for k, h in _FB_COLS}
    df = pd.DataFrame(rows, columns=int_keys) if rows else pd.DataFrame(columns=int_keys)
    df.rename(columns=hdr_map, inplace=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Factbook")
        ws = writer.sheets["Factbook"]
        ws.column_dimensions["A"].width = 38
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws.column_dimensions[col].width = 18
    return buf.getvalue()


def factbook_from_excel(excel_bytes: bytes) -> dict:
    """Load a factbook Excel cache (produced by factbook_to_excel_bytes).

    Returns the same dict structure as parse_factbook().
    """
    try:
        df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Factbook",
                           dtype=str)
    except Exception:
        return {}

    # Map header labels back to internal keys
    hdr_to_int = {h: k for k, h in _FB_COLS}
    df.rename(columns=hdr_to_int, inplace=True)

    result: dict = {}
    _float_keys = {"duration", "ytm", "az_pct", "obb_pct"}

    for _, row in df.iterrows():
        norm = str(row.get("fondo", "")).strip()
        if not norm or norm == "nan":
            continue
        entry: dict = {}
        for int_key in [k for k, _ in _FB_COLS if k != "fondo"]:
            raw = str(row.get(int_key, "")).strip()
            if not raw or raw in ("nan", "None", ""):
                continue
            if int_key in _float_keys:
                try:
                    entry[_FB_INTERNAL_KEYS.get(int_key, int_key)] = float(raw)
                except ValueError:
                    pass
            else:
                entry[_FB_INTERNAL_KEYS.get(int_key, int_key)] = raw
        if entry:
            result[norm] = entry

    return result


# ── Factbook JSON persistence (auto-load / GitHub API save) ─────────────────

_FB_REPO      = "albertobeneventi/azimut_portfolio_analyzer"
_FB_REPO_PATH = "data/factbook_dati.json"
_FB_BRANCH    = "master"


def load_factbook_auto() -> dict:
    """Load factbook data from data/factbook_dati.json (committed in the repo).
    Returns {} when the file is absent or empty.
    """
    import json
    try:
        fp = Path(__file__).parent / "data" / "factbook_dati.json"
        if fp.exists() and fp.stat().st_size > 5:
            with open(fp, encoding="utf-8") as fh:
                data = json.load(fh)
                if isinstance(data, dict) and data:
                    return data
    except Exception:
        pass
    return {}


def save_factbook_to_repo(fb_data: dict) -> bool:
    """Commit data/factbook_dati.json to GitHub via the Contents API.

    Requires a Streamlit secret  GITHUB_TOKEN  with 'contents: write'
    permission (fine-grained PAT) or  repo  scope (classic PAT).

    Returns True on success, False on any error.
    """
    import json, base64
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        if not token:
            return False

        content_str = json.dumps(fb_data, ensure_ascii=False, indent=2,
                                 default=str)
        content_b64 = base64.b64encode(
            content_str.encode("utf-8")).decode()

        headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json",
            "X-GitHub-Api-Version": "2022-11-28",
        }
        url = (f"https://api.github.com/repos/{_FB_REPO}"
               f"/contents/{_FB_REPO_PATH}")

        # Need existing SHA to update (not create) the file
        r_get = requests.get(url, headers=headers,
                             params={"ref": _FB_BRANCH}, timeout=10)
        sha = (r_get.json().get("sha")
               if r_get.status_code == 200 else None)

        payload: dict = {
            "message": (f"auto: aggiorna dati factbook "
                        f"{datetime.date.today().isoformat()}"),
            "content": content_b64,
            "branch": _FB_BRANCH,
        }
        if sha:
            payload["sha"] = sha

        r_put = requests.put(url, json=payload, headers=headers, timeout=15)
        return r_put.status_code in (200, 201)
    except Exception:
        return False


def _parse_ptf(wb, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    funds, cur_group = [], {"name":None,"mc":1.0,"me":1.0,"ma":1.0}
    for row in rows:
        name = row[COL_A]
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name in GROUP_NAMES:
            cur_group = {
                "name": name,
                "mc": float(row[COL_C]) if isinstance(row[COL_C],(int,float)) else 1.0,
                "me": float(row[COL_G]) if isinstance(row[COL_G],(int,float)) else 1.0,
                "ma": float(row[COL_K]) if isinstance(row[COL_K],(int,float)) else 1.0,
            }
            continue
        if name.startswith("AZ") and cur_group["name"]:
            rw = row[COL_R]
            if not isinstance(rw,(int,float)) or rw <= 0: continue
            funds.append({
                "nome":     name,
                "categoria": row[COL_B] if isinstance(row[COL_B],str) else "",
                "gruppo":   cur_group["name"],
                "az_pct":   min(1.0,max(0.0,float(row[COL_O]) if isinstance(row[COL_O],(int,float)) else 0.5)),
                "obb_pct":  min(1.0,max(0.0,1.0-(float(row[COL_O]) if isinstance(row[COL_O],(int,float)) else 0.5))),
                "r_weight": float(rw),
                "mc":cur_group["mc"],"me":cur_group["me"],"ma":cur_group["ma"],
            })
    if not funds: return pd.DataFrame()
    df = pd.DataFrame(funds)
    for wcol, mcol in [("w_cons","mc"),("w_equil","me"),("w_accr","ma")]:
        raw = df["r_weight"]*df[mcol]
        df[wcol] = raw/raw.sum() if raw.sum()>0 else raw
    df["macro_cat"] = df["categoria"].apply(get_macro)
    df = assign_colors(df)
    return df


def _parse_fida(wb) -> pd.DataFrame:
    ws = wb["FIDA"]
    rows = list(ws.iter_rows(values_only=True))
    funds = []
    for row in rows[1:]:
        nome = row[0]
        if not nome or not isinstance(nome,str): continue
        nome = nome.strip().replace("\xa0","")
        cat  = (row[2] or "").strip()
        funds.append({"nome":nome,"isin":row[1] or "","categoria":cat,"macro_cat":get_macro(cat)})
    return pd.DataFrame(funds).drop_duplicates(subset=["nome"])


def extract_fida_urls(file_bytes: bytes) -> dict:
    """Extract fondidoc.it URLs from FIDA sheet hyperlinks in the Excel XML."""
    fund_urls = dict(MANUAL_URL_OVERRIDES)
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            ss_root  = ET.fromstring(z.read("xl/sharedStrings.xml"))
            ns_ss    = {"s":"http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            strings  = ["".join(t.text or "" for t in si.findall(".//s:t",ns_ss))
                        for si in ss_root.findall("s:si",ns_ss)]
            fida_root= ET.fromstring(z.read("xl/worksheets/sheet5.xml"))
            ws_ns    = {"ws":"http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            rels_root= ET.fromstring(z.read("xl/worksheets/_rels/sheet5.xml.rels"))
            rels_ns  = {"rel":"http://schemas.openxmlformats.org/package/2006/relationships"}
        rid_to_url = {r.get("Id"):r.get("Target") for r in rels_root.findall("rel:Relationship",rels_ns)}
        hyp_map    = {}
        r_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        for hl in fida_root.findall(".//ws:hyperlink",ws_ns):
            rid = hl.get(r_attr)
            if rid: hyp_map[hl.get("ref")] = rid
        for row in fida_root.findall(".//ws:row",ws_ns):
            for cell in row.findall("ws:c",ws_ns):
                ref = cell.get("r")
                if ref and ref.startswith("A") and ref in hyp_map:
                    v = cell.find("ws:v",ws_ns)
                    if v is not None and cell.get("t") == "s":
                        name = strings[int(v.text)].strip().replace("\xa0","")
                        url  = rid_to_url.get(hyp_map[ref],"")
                        if "fondidoc.it" in url:
                            fund_urls[name] = url
    except Exception:
        pass
    return fund_urls


# ════════════════════════════════════════════════════════════
# FONDIDOC SCRAPING
# ════════════════════════════════════════════════════════════

def _to_en_url(url: str) -> str:
    """Ensure URL uses /en/ locale."""
    if "/en/" in url: return url
    return url.replace("fondidoc.it/d/","fondidoc.it/en/d/")


def _to_ana_url(index_url: str) -> str:
    return index_url.replace("/d/Index/","/d/Ana/").replace("/en/d/Index/","/en/d/Ana/")


def _fetch_html(url: str, timeout: int = 8) -> str | None:
    try:
        r = requests.get(_to_en_url(url), headers=FONDIDOC_HEADERS, timeout=timeout)
        return r.text if r.status_code == 200 else None
    except Exception:
        return None


def _parse_overview(html: str) -> dict:
    soup = BeautifulSoup(html, "lxml")
    lines = [l.strip() for l in soup.get_text(separator="\n").split("\n") if l.strip()]
    d = {}
    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1<len(lines) else ""
        if line == "SRRI (risk value)":           d["srri"]         = nxt
        elif line == "Start date":                d["start_date"]   = nxt
        elif line == "Assogestioni category":     d["cat_assog"]    = nxt
        elif line == "Income distribution":       d["income"]       = nxt
        elif line == "Management Fee":            d["mgmt_fee"]     = nxt
        elif line == "Performance Fee":           d["perf_fee"]     = nxt
        elif line == "Subscription fee":          d["sub_fee"]      = nxt
        elif line == "Rating" and "fida_rating" not in d: d["fida_rating"] = nxt
        elif line == "Score":                     d["fida_score"]   = nxt
        elif line == "Category" and "fida_cat" not in d: d["fida_cat"] = nxt
    return d


def _parse_analysis(html: str) -> dict:
    soup = BeautifulSoup(html, "lxml")
    d = {}

    # NAV section
    lines = [l.strip() for l in soup.get_text(separator="\n").split("\n") if l.strip()]
    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1<len(lines) else ""
        if line == "Last update": d["last_update"] = nxt
        elif line == "NAV":       d["nav"]          = nxt
        elif line == "Daily change (%)": d["daily_change"] = nxt

    # Tables
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows: continue
        header = [td.get_text(strip=True) for td in rows[0].find_all(["th","td"])]

        # Performance table (has YTD column)
        if "YTD" in header and "1 year" in header:
            def sg(cells, key, hdr):
                try: return cells[hdr.index(key)] if hdr.index(key)<len(cells) else "—"
                except ValueError: return "—"
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if not cells: continue
                lbl = cells[0].lower()
                if "performance" in lbl:
                    d["ytd"]     = sg(cells,"YTD",header)
                    d["perf_1y"] = sg(cells,"1 year",header)
                    d["perf_3y"] = sg(cells,"3 years",header)
                    d["perf_5y"] = sg(cells,"5 years",header)

        # Risk table (1 year / 3 years / 5 years, NO YTD)
        elif "1 year" in header and "YTD" not in header and len(header) >= 4:
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if not cells: continue
                lbl = cells[0].lower()
                def gv(idx): return cells[idx] if idx<len(cells) else "—"
                if "volatility" in lbl and "negative" not in lbl:
                    d["vol_1y"],d["vol_3y"],d["vol_5y"] = gv(1),gv(2),gv(3)
                elif "negative" in lbl:
                    d["neg_vol_1y"] = gv(1)
                    if len(header) > 2: d["neg_vol_3y"] = gv(2)
                    if len(header) > 3: d["neg_vol_5y"] = gv(3)
                elif "sharpe" in lbl:
                    d["sharpe_1y"],d["sharpe_3y"],d["sharpe_5y"] = gv(1),gv(2),gv(3)
                elif "sortino" in lbl:
                    d["sortino_1y"] = gv(1)
                elif "var" in lbl or "value at risk" in lbl:
                    d["var_1y"] = gv(1)
                    if len(header) > 2: d["var_3y"] = gv(2)

        # Annual performance (header contains year digits)
        elif any(h.isdigit() and len(h)==4 for h in header):
            years = [h for h in header if h.isdigit() and len(h)==4]
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if len(cells)<2: continue
                # Take first data row that has % values (fund row, not benchmark)
                if any("%" in c for c in cells) and "annual_perf" not in d:
                    annual = {}
                    for yr in years:
                        try:
                            idx = header.index(yr)
                            annual[yr] = cells[idx] if idx<len(cells) else "—"
                        except ValueError: pass
                    if annual: d["annual_perf"] = annual
                    break
    return d


def _extract_isin(url: str) -> str:
    """Estrae ISIN da URL FondiDoc (formato: /CATCODE/ISIN_slug-nome-fondo)."""
    m = re.search(r'/([A-Z]{2}[A-Z0-9]{10})[_/]', url)
    return m.group(1) if m else ""


def fetch_fund_data(index_url: str) -> dict:
    """Fetch overview + analysis for one fund. Cached 1h."""
    result = {"url": index_url}
    isin = _extract_isin(index_url)
    if isin:
        result["isin"] = isin
    html_idx = _fetch_html(index_url)
    if html_idx: result["overview"] = _parse_overview(html_idx)
    html_ana = _fetch_html(_to_ana_url(index_url))
    if html_ana: result["analysis"] = _parse_analysis(html_ana)
    return result

try:
    fetch_fund_data = st.cache_data(ttl=3600, show_spinner=False)(fetch_fund_data)
except Exception:
    pass  # run uncached if decorator fails


def fetch_all_fund_data(df: pd.DataFrame, fida_urls: dict,
                         progress_cb=None) -> dict:
    """Parallel fetch for all portfolio funds."""
    tasks = {}
    for nome in df["nome"].unique():
        url = fida_urls.get(nome)
        if url: tasks[nome] = url

    results = {}
    total = len(tasks)
    done  = 0

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {pool.submit(fetch_fund_data, url): nome for nome, url in tasks.items()}
        for future in as_completed(futures):
            nome = futures[future]
            try: results[nome] = future.result()
            except Exception: results[nome] = {}
            done += 1
            if progress_cb: progress_cb(done/total)

    return results


# ════════════════════════════════════════════════════════════
# FONDIONLINE API — Morningstar rating
# ════════════════════════════════════════════════════════════
# FondiOnline exposes a JSON API used by its fund screener page.
# One HTTP request returns all funds for a company with Rating field.

FONDIONLINE_BASE    = "https://www.fondionline.it"
FO_API_URL          = "https://www.fondionline.it/offers-list"
FO_AZ_COMPANY_ID    = "0C00001L0E"   # Azimut Investments S.A. (Morningstar ID)
FONDIONLINE_HDR     = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0",
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
    "Referer":         "https://www.fondionline.it/fondi/elenco_prodotti.html",
}


def _fo_fetch_company_ratings(company_id: str) -> dict:
    """Fetch all Morningstar ratings for one company via FondiOnline JSON API.

    Single HTTP request — returns {ISIN: {"ms_rating": int|None, "fo_url": str|None}}.
    The API paginates; we request pageSize=1000 to get everything in one shot
    (Azimut has ~310 funds total).
    """
    result: dict = {}
    try:
        r = requests.get(
            FO_API_URL,
            params={
                "productType":      "OICR",
                "sortOrder":        "asc",
                "pageNumber":       1,
                "pageSize":         1000,
                "tab":              0,
                "fundId":           "",
                "orderBy":          "Name",
                "brandingCompanyId": company_id,
                "distribution":     -1,
            },
            headers=FONDIONLINE_HDR,
            timeout=15,
        )
        if r.status_code == 200:
            data = r.json()
            for fund in data.get("funds", []):
                isin   = (fund.get("ISIN") or "").strip()
                rating = fund.get("Rating")          # "1"…"5" or absent
                url    = (f"{FONDIONLINE_BASE}/elenco-fondi/{fund['detailsUrl']}"
                          if fund.get("detailsUrl") else None)
                if isin:
                    result[isin] = {
                        "ms_rating": int(rating) if rating else None,
                        "fo_url":    url,
                    }
    except Exception:
        pass
    return result


def fetch_all_ms_ratings(df: pd.DataFrame, fida_df: pd.DataFrame,
                          progress_cb=None) -> dict:
    """Fetch Morningstar ratings for all portfolio funds via FondiOnline API.

    Replaces the old per-page scraping approach with a single JSON API call.
    Returns {fund_name: {"ms_rating": int_or_None, "fo_url": str_or_None}}.
    """
    # 1. Build nome → ISIN map from FIDA sheet
    nome_to_isin: dict = {}
    if not fida_df.empty and "isin" in fida_df.columns:
        for _, fr in fida_df.iterrows():
            isin = str(fr.get("isin") or "").strip()
            if isin:
                nome_to_isin[fr["nome"]] = isin

    portfolio_names = list(df["nome"].unique()) if not df.empty else []

    # 2. One API call → ISIN → {ms_rating, fo_url}
    isin_to_ms = _fo_fetch_company_ratings(FO_AZ_COMPANY_ID)
    if progress_cb:
        progress_cb(1.0)

    # 3. Match portfolio funds by ISIN
    results: dict = {}
    for nome in portfolio_names:
        isin = nome_to_isin.get(nome, "")
        results[nome] = isin_to_ms.get(isin, {"ms_rating": None, "fo_url": None})

    return results


# ════════════════════════════════════════════════════════════
# PLOTLY CHARTS (unchanged)
# ════════════════════════════════════════════════════════════

def make_fund_pie(df, wcol, profile):
    d = df[df[wcol]>0.005].copy()
    d["pct"] = d[wcol]*100
    labels = d["nome"].apply(lambda x: (x[:38]+"…") if len(x)>38 else x)
    fig = go.Figure(go.Pie(
        labels=labels, values=d["pct"],
        marker=dict(colors=d["color"].tolist(), line=dict(color="#fff",width=2.5)),
        hovertemplate="<b>%{label}</b><br>Peso: <b>%{value:.1f}%</b><extra></extra>",
        textinfo="percent", textfont=dict(size=10,family="DM Sans"),
        hole=0.40, pull=[0.04 if v==d["pct"].max() else 0 for v in d["pct"]],
        sort=False, direction="clockwise",
    ))
    fig.update_layout(
        margin=dict(t=10,b=10,l=10,r=180), showlegend=True,
        legend=dict(x=1.01,y=0.5,orientation="v",font=dict(size=9.5,family="DM Sans"),bgcolor="rgba(0,0,0,0)"),
        paper_bgcolor="rgba(0,0,0,0)", height=430,
        annotations=[dict(text=f"<b>{profile[:4]}</b>",x=0.5,y=0.5,showarrow=False,
                          font=dict(size=17,color="#0d1b2a",family="Cormorant Garamond"))],
    )
    return fig


def make_macro_bar(df, wcol):
    agg = df[df[wcol]>0.001].groupby("macro_cat")[wcol].sum().reset_index().sort_values(wcol)
    agg["pct"] = agg[wcol]*100
    agg["color"] = agg["macro_cat"].map(MACRO_COLORS)
    fig = go.Figure(go.Bar(
        x=agg["pct"], y=agg["macro_cat"], orientation="h",
        marker=dict(color=agg["color"].tolist(), line=dict(color="#fff",width=1)),
        hovertemplate="<b>%{y}</b>: %{x:.1f}%<extra></extra>",
        text=agg["pct"].apply(lambda v:f"{v:.1f}%"),
        textposition="inside", insidetextanchor="middle",
        textfont=dict(color="#fff",size=11,family="DM Sans"),
    ))
    fig.update_layout(
        margin=dict(t=10,b=10,l=10,r=10), paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(showgrid=False,showticklabels=False,range=[0,105]),
        yaxis=dict(showgrid=False,tickfont=dict(size=11,family="DM Sans")),
        height=max(160,len(agg)*48), bargap=0.28,
    )
    return fig


# ════════════════════════════════════════════════════════════
# PDF — MATPLOTLIB HELPERS
# ════════════════════════════════════════════════════════════

def _mpl_portfolio_pie(df, wcol, profile) -> io.BytesIO:
    """Donut only — la leggenda viene resa in ReportLab per permettere hyperlink."""
    d = df[df[wcol] > 0.005].sort_values(wcol, ascending=False)
    fig, ax = plt.subplots(figsize=(5.5, 5.5))
    _, _, autotexts = ax.pie(
        d[wcol], colors=d["color"].tolist(),
        autopct=lambda p: f"{p:.1f}%" if p > 3.5 else "",
        pctdistance=0.72,
        wedgeprops=dict(width=0.58, edgecolor="white", linewidth=2),
        startangle=90,
    )
    for at in autotexts:
        at.set_fontsize(9); at.set_color("white"); at.set_fontweight("bold")
    ax.text(0, 0, profile[:4], ha="center", va="center",
            fontsize=16, fontweight="bold", color="#0D1B2A")
    fig.patch.set_facecolor("#FFFFFF")
    plt.tight_layout(pad=0.3)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig); buf.seek(0)
    return buf


def _mpl_macro_pie(df, wcol) -> io.BytesIO | None:
    """Asset allocation donut Azionario/Obbligazionario — leggenda in ReportLab."""
    w_az  = (df[wcol] * df["az_pct"]).sum()
    w_obb = (df[wcol] * df["obb_pct"]).sum()
    if w_az + w_obb < 0.001:
        return None
    fig, ax = plt.subplots(figsize=(5.5, 5.5))
    _, _, autotexts = ax.pie(
        [w_az, w_obb], colors=["#1B4FBB", "#2D9D78"],
        autopct=lambda p: f"{p:.1f}%" if p >= 5 else "",
        pctdistance=0.70,
        wedgeprops=dict(width=0.58, edgecolor="white", linewidth=2.5),
        startangle=90,
    )
    for at in autotexts:
        at.set_fontsize(10); at.set_color("white"); at.set_fontweight("bold")
    ax.text(0, 0, "Asset\nAlloc.", ha="center", va="center",
            fontsize=11, fontweight="bold", color="#0D1B2A")
    fig.patch.set_facecolor("#FFFFFF")
    plt.tight_layout(pad=0.3)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig); buf.seek(0)
    return buf


def _mpl_annual_bar(annual_perf: dict, fund_name: str) -> io.BytesIO | None:
    if not annual_perf: return None
    years, vals = [], []
    for y,v in sorted(annual_perf.items()):
        try:
            num = float(v.replace("%","").replace(",",".").strip())
            years.append(y); vals.append(num)
        except: pass
    if not years: return None
    fig, ax = plt.subplots(figsize=(6,2.2))
    colors = ["#2D9D78" if v>=0 else "#E05252" for v in vals]
    bars = ax.bar(years, vals, color=colors, edgecolor="white", linewidth=0.8, width=0.6)
    ax.axhline(0, color="#94A3B8", linewidth=0.8, linestyle="-")
    for bar,val in zip(bars,vals):
        ax.text(bar.get_x()+bar.get_width()/2,
                bar.get_height()+(0.3 if val>=0 else -0.9),
                f"{val:+.1f}%", ha="center", va="bottom", fontsize=7, fontweight="bold",
                color="#2D9D78" if val>=0 else "#E05252")
    ax.set_ylabel("%", fontsize=8, color="#64748B")
    ax.tick_params(axis="both",labelsize=8,colors="#475569")
    ax.spines[["top","right","left"]].set_visible(False)
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.0f%%"))
    ax.grid(axis="y",alpha=0.3,linestyle="--")
    fig.patch.set_facecolor("#FFFFFF"); plt.tight_layout(pad=0.5)
    buf=io.BytesIO(); fig.savefig(buf,format="png",dpi=130,bbox_inches="tight",facecolor="white")
    plt.close(fig); buf.seek(0); return buf


# ════════════════════════════════════════════════════════════
# PDF GENERATION
# ════════════════════════════════════════════════════════════

def generate_pdf(df: pd.DataFrame, wcol: str, profile: str,
                 ptf_name: str, fund_data: dict = None,
                 fida_df: pd.DataFrame = None,
                 factbook_data: dict = None,
                 cache_date: str = "") -> bytes:

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2.2*cm, bottomMargin=2.2*cm)

    ss = getSampleStyleSheet()
    def S(name,**kw): return ParagraphStyle(name,parent=ss["Normal"],**kw)

    T  = S("T",  fontName="Helvetica-Bold",  fontSize=22, textColor=rl_colors.HexColor("#0D1B2A"), spaceAfter=4,leading=28)
    EY = S("EY", fontName="Helvetica",       fontSize=8,  textColor=rl_colors.HexColor("#94A3B8"), spaceAfter=4,letterSpacing=1.5)
    SU = S("SU", fontName="Helvetica",       fontSize=10, textColor=rl_colors.HexColor("#64748B"), spaceAfter=4)
    SC = S("SC", fontName="Helvetica-Bold",  fontSize=11, textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=14,spaceAfter=8)
    BD = S("BD", fontName="Helvetica",       fontSize=8.5,textColor=rl_colors.HexColor("#1E293B"), leading=13)
    SM = S("SM", fontName="Helvetica",       fontSize=7.5,textColor=rl_colors.HexColor("#1E293B"), leading=11)
    FT = S("FT", fontName="Helvetica-Oblique",fontSize=7, textColor=rl_colors.HexColor("#94A3B8"), leading=10)
    FS = S("FS", fontName="Helvetica-Bold",  fontSize=13, textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=4,spaceAfter=2)
    FK = S("FK", fontName="Helvetica",       fontSize=7.5,textColor=rl_colors.HexColor("#64748B"), spaceAfter=2)
    LK = S("LK", fontName="Helvetica",       fontSize=7.5,textColor=rl_colors.HexColor("#1B4FBB"), spaceAfter=2)
    # HDR: sempre bianco+grassetto — per celle intestazione su sfondo scuro
    # (TEXTCOLOR di TableStyle NON sovrascrive il colore dei Paragraph — serve lo stile dedicato)
    HDR= S("HDR",fontName="Helvetica-Bold",  fontSize=7.5,textColor=rl_colors.white, leading=11)

    story = []
    d_act = df[df[wcol]>0.001].copy()
    n_fondi = len(d_act)

    # ISIN da foglio FIDA (fallback per fondi senza URL FondiDoc)
    isin_map = {}
    if fida_df is not None and not fida_df.empty and "isin" in fida_df.columns:
        isin_map = {r["nome"]: str(r["isin"]).strip() for _, r in fida_df.iterrows()
                    if r.get("isin") and str(r.get("isin","")).strip()}
    w_az  = (d_act[wcol]*d_act["az_pct"]).sum()*100
    w_obb = (d_act[wcol]*d_act["obb_pct"]).sum()*100

    # ── ACCENT BAR ──────────────────────────────────────────
    story.append(Table([[""]], colWidths=[17*cm], rowHeights=[10],
        style=TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),rl_colors.HexColor("#0D1B2A")),
            ("LINEBELOW",(0,0),(-1,-1),3,rl_colors.HexColor("#C9A84C")),
        ])))
    story.append(Spacer(1,14))

    # ── TITLE BLOCK ─────────────────────────────────────────
    story.append(Paragraph("AZIMUT INVESTMENTS  ·  AAS EMILIA ROMAGNA MARCHE UMBRIA", EY))
    story.append(Spacer(1,4))
    story.append(Paragraph(f"Portafoglio {ptf_name}", T))
    story.append(Paragraph(
        f"{PROFILE_ICONS.get(profile,'●')} Profilo {profile.title()}  ·  "
        f"Dati al {datetime.date.today().strftime('%d %B %Y')}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=14))

    # ── KPI ─────────────────────────────────────────────────
    KC = S("KC", fontName="Helvetica", fontSize=8.5,
           textColor=rl_colors.HexColor("#1E293B"), leading=13, alignment=1)
    def kpi_cell(v,l):
        return Paragraph(f'<font size="18"><b>{v}</b></font><br/>'
                         f'<font size="8" color="#64748B">{l}</font>', KC)
    kpi = Table(
        [[kpi_cell(str(n_fondi),"Fondi"),kpi_cell(f"{w_az:.1f}%","Quota Azionaria"),
          kpi_cell(f"{w_obb:.1f}%","Quota Obbligazionaria"),
          kpi_cell(datetime.date.today().strftime("%m/%Y"),"Data Report")]],
        colWidths=[4.25*cm]*4,
        rowHeights=[2.2*cm],
    )
    kpi.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.8,rl_colors.HexColor("#E2E8F0")),
        ("INNERGRID",(0,0),(-1,-1),0.8,rl_colors.HexColor("#E2E8F0")),
        ("BACKGROUND",(0,0),(-1,-1),rl_colors.HexColor("#F8FAFC")),
        ("PADDING",(0,0),(-1,-1),12),("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(kpi)

    # ── PIE CHARTS con legende ReportLab ────────────────────
    # Titolo sezione con meno spazio sopra per avvicinare il grafico al KPI
    SC_PIE = S("SCPIE", fontName="Helvetica-Bold", fontSize=11,
               textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=6, spaceAfter=5)
    story.append(Paragraph("Allocazione del Portafoglio", SC_PIE))

    PIE_W = 7.5 * cm
    LEG_W = 17 * cm - PIE_W   # 9.5 cm

    LG = S("LG", fontName="Helvetica", fontSize=10,
           textColor=rl_colors.HexColor("#1E293B"), leading=15)

    def _dot(hex_color):
        t = Table([[""]], colWidths=[0.28*cm], rowHeights=[0.28*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), rl_colors.HexColor(hex_color)),
        ]))
        return t

    # — Grafico 1: fondi con hyperlink (torta + didascalia interattiva affiancate) —
    pie_buf = _mpl_portfolio_pie(d_act, wcol, profile)
    pie_img = RLImage(pie_buf, width=PIE_W, height=PIE_W)
    d_leg   = d_act[d_act[wcol] > 0.005].sort_values(wcol, ascending=False)
    leg_rows = []
    for _, r in d_leg.iterrows():
        url    = (fund_data or {}).get(r["nome"], {}).get("url", "")
        name_s = (r["nome"][:38] + "…") if len(r["nome"]) > 38 else r["nome"]
        pct_s  = f"{r[wcol]*100:.1f}%"
        if url:
            lbl = Paragraph(
                f'<link href="{url}"><font color="#1B4FBB"><u>{name_s}</u></font></link>'
                f'  <b>{pct_s}</b>', LG)
        else:
            lbl = Paragraph(f'{name_s}  <b>{pct_s}</b>', LG)
        leg_rows.append([_dot(r["color"]), lbl])
    leg_tbl = Table(leg_rows, colWidths=[0.45*cm, LEG_W - 0.45*cm])
    leg_tbl.setStyle(TableStyle([
        ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",     (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",  (0,0), (-1,-1), 3),
        ("LEFTPADDING",    (1,0), (1,-1),  6),
        ("LEFTPADDING",    (0,0), (0,-1),  0),
        ("RIGHTPADDING",   (0,0), (-1,-1), 4),
    ]))
    combo1 = Table([[pie_img, leg_tbl]], colWidths=[PIE_W, LEG_W])
    combo1.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE")]))
    story.append(combo1)

    # — Grafico 2: asset allocation — torta centrata sotto la prima,
    #   righe illustrative Azionario/Obbligazionario sotto la torta —
    macro_buf = _mpl_macro_pie(d_act, wcol)
    if macro_buf:
        story.append(Spacer(1, 8))
        macro_img = RLImage(macro_buf, width=PIE_W, height=PIE_W)
        w_az_v  = (d_act[wcol] * d_act["az_pct"]).sum()
        w_obb_v = (d_act[wcol] * d_act["obb_pct"]).sum()

        # Torta centrata orizzontalmente sulla pagina
        pie2_tbl = Table([[macro_img]], colWidths=[17 * cm])
        pie2_tbl.setStyle(TableStyle([
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
            ("BOTTOMPADDING", (0,0), (-1,-1), 0),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
        ]))
        story.append(pie2_tbl)

        # Righe illustrative sotto la torta, centrate
        story.append(Spacer(1, 6))
        macro_leg_rows = [
            [_dot("#1B4FBB"), Paragraph(f'Azionario  <b>{w_az_v*100:.1f}%</b>', LG)],
            [_dot("#2D9D78"), Paragraph(f'Obbligazionario  <b>{w_obb_v*100:.1f}%</b>', LG)],
        ]
        macro_leg_inner = Table(macro_leg_rows, colWidths=[0.45*cm, 5.5*cm])
        macro_leg_inner.setStyle(TableStyle([
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (1,0), (1,-1),  6),
            ("LEFTPADDING",   (0,0), (0,-1),  0),
        ]))
        macro_leg_wrapper = Table([[macro_leg_inner]], colWidths=[17 * cm])
        macro_leg_wrapper.setStyle(TableStyle([
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
            ("BOTTOMPADDING", (0,0), (-1,-1), 0),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
        ]))
        story.append(macro_leg_wrapper)

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════
    # PAGE 2: RENDIMENTI 1-3-5 ANNI
    # ════════════════════════════════════════════════════════
    story.append(Paragraph("AZIMUT INVESTMENTS  ·  AAS EMILIA ROMAGNA MARCHE UMBRIA", EY))
    story.append(Spacer(1,4))
    story.append(Paragraph("Tavola dei Rendimenti", T))
    _fb_loaded = bool(factbook_data)
    _fb_ref    = (factbook_data or {}).get("_ref_date", "")      # data dal frontespizio
    _fd_ref    = cache_date or datetime.date.today().strftime("%d/%m/%Y")  # data FondiDoc
    if _fb_loaded:
        _rend_src = (f"Rendimenti: Factbook AZ Investments"
                     + (f" al {_fb_ref}" if _fb_ref else "")
                     + f"  ·  Rischio: FondiDoc aggiornata al {_fd_ref}")
    else:
        _rend_src = f"Fonte: FIDA FondiDoc aggiornata al {_fd_ref}"
    story.append(Paragraph(
        f"Performance per fondo  ·  Profilo {profile.title()}  ·  {_rend_src}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=12))

    # ── Helper: look up a performance metric from the factbook ────────────────
    _fb = factbook_data or {}
    _PERF_KEYS = {"ytd", "perf_1y", "perf_3y", "perf_5y"}

    def get_fb(nome: str, key: str) -> str:
        """Return factbook value for fund `nome` and metric `key`, or ''."""
        if key not in _PERF_KEYS or not _fb:
            return ""
        norm = _normalize_for_unp(nome)
        norm = _FUND_ALIASES.get(norm, norm)
        entry = _fb.get(norm)
        if not entry:
            # Substring fallback (same logic as lookup_unp)
            best, best_len = None, 0
            for fb_key, fb_val in _fb.items():
                if (fb_key in norm or norm in fb_key) and len(fb_key) > best_len:
                    best, best_len = fb_val, len(fb_key)
            entry = best
        if not entry:
            return ""
        return entry.get(key) or ""

    # ── Helper: weighted average of a metric across all funds ─────────────────
    def ptf_wavg(keys_list):
        """Weighted average per metric. Prefers factbook for return keys."""
        totals = {k: 0.0 for k in keys_list}
        cov_w  = {k: 0.0 for k in keys_list}
        for _, row in d_sorted.iterrows():
            fd  = (fund_data or {}).get(row["nome"], {})
            ana = fd.get("analysis", {})
            w   = row[wcol]
            for k in keys_list:
                raw = get_fb(row["nome"], k) or ana.get(k, "")
                try:
                    num = float(raw.replace("%","").replace(",",".").strip())
                    totals[k] += num * w
                    cov_w[k]  += w
                except Exception:
                    pass
        out = {}
        for k in keys_list:
            out[k] = f"{totals[k]/cov_w[k]:+.2f}%" if cov_w[k] > 0.01 else "N/D"
        return out

    # Paragraph style for portfolio summary row
    WH = S("WH", fontName="Helvetica-Bold", fontSize=8,
           textColor=rl_colors.white, leading=11)

    def pstyle_w(val):
        """White bold text coloured green/red for portfolio row."""
        try:
            v = float(val.replace("%","").replace(",",".").strip())
            c = "#7EFFC0" if v > 0 else ("#FFB3B3" if v < 0 else "#E2E8F0")
        except Exception:
            c = "#E2E8F0"
        return Paragraph(f'<font color="{c}"><b>{val}</b></font>', WH)

    def pstyle(val):
        try:
            v = float(val.replace("%","").replace(",","."))
            color = "#1A7A4A" if v>0 else ("#C0392B" if v<0 else "#475569")
            return f'<font color="{color}"><b>{val}</b></font>'
        except: return val

    # ── d_sorted deve essere definito PRIMA di chiamare ptf_wavg ──
    d_sorted = d_act.sort_values(wcol, ascending=False)

    # ── PERFORMANCE TABLE ────────────────────────────────────
    perf_keys = ["ytd","perf_1y","perf_3y","perf_5y","vol_1y","sharpe_1y"]
    ptf_p = ptf_wavg(perf_keys)

    perf_hdr = [Paragraph(f"<b>{t}</b>", HDR) for t in
                ["Fondo","Peso","YTD","1 Anno","3 Anni","5 Anni","Vol. 1A","Sharpe 1A"]]

    # Portfolio summary row (row index 1 — gold background)
    ptf_perf_row = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph(f"<b>100%</b>", WH),
        pstyle_w(ptf_p.get("ytd","N/D")),
        pstyle_w(ptf_p.get("perf_1y","N/D")),
        pstyle_w(ptf_p.get("perf_3y","N/D")),
        pstyle_w(ptf_p.get("perf_5y","N/D")),
        Paragraph(ptf_p.get("vol_1y","N/D"), WH),
        Paragraph(ptf_p.get("sharpe_1y","N/D"), WH),
    ]

    perf_rows = [perf_hdr, ptf_perf_row]

    for _, row in d_sorted.iterrows():
        fd  = (fund_data or {}).get(row["nome"], {})
        ana = fd.get("analysis", {})
        def gv(key, nome=row["nome"]):
            # Return factbook value (for return metrics) or FondiDoc value
            return get_fb(nome, key) or ana.get(key, "N/D")
        perf_rows.append([
            Paragraph(row["nome"][:48], SM),
            Paragraph(f"<b>{row[wcol]*100:.1f}%</b>", SM),
            Paragraph(pstyle(gv("ytd")),     SM),
            Paragraph(pstyle(gv("perf_1y")), SM),
            Paragraph(pstyle(gv("perf_3y")), SM),
            Paragraph(pstyle(gv("perf_5y")), SM),
            Paragraph(gv("vol_1y"),           SM),
            Paragraph(gv("sharpe_1y"),        SM),
        ])

    perf_tbl = Table(perf_rows,
        colWidths=[5.2*cm,1.4*cm,1.4*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm],
        repeatRows=1)
    ts_perf = [
        ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),  # header
        ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
        # Portfolio summary row — deep forest green + gold accent
        ("BACKGROUND",(0,1),(-1,1), rl_colors.HexColor("#1B4332")),
        ("LINEBELOW",(0,1),(-1,1),  2, rl_colors.HexColor("#C9A84C")),
        # Funds
        ("FONTSIZE",(0,0),(-1,-1),  8),
        ("PADDING",(0,0),(-1,-1),   5),
        ("ROWBACKGROUNDS",(0,2),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",(1,0),(-1,-1),     "CENTER"),
        ("VALIGN",(0,0),(-1,-1),    "MIDDLE"),
    ]
    perf_tbl.setStyle(TableStyle(ts_perf))
    # KeepTogether: la tabella rendimenti non viene spezzata su due pagine
    NOTE_P = S("NTP", fontName="Helvetica-Oblique", fontSize=6.5,
               textColor=rl_colors.HexColor("#94A3B8"), leading=9)
    story.append(KeepTogether([
        perf_tbl,
        Spacer(1, 5),
        Paragraph(
            "◆ La riga <b>Portafoglio</b> riporta la media ponderata dei rendimenti dei singoli fondi, "
            "usando i pesi del profilo selezionato. Il calcolo include solo i fondi per cui il dato è "
            "disponibile su FondiDoc, rinormalizzando i pesi su di essi. "
            "I rendimenti a 3 e 5 anni sono tassi annualizzati: la media ponderata è un'approssimazione "
            "(il rendimento composito effettivo dipende dalla sequenza temporale e dalla correlazione tra fondi). "
            "Dati a titolo indicativo — non costituiscono consulenza di investimento.", NOTE_P),
    ]))
    story.append(Spacer(1,12))

    # ── RISK TABLE ───────────────────────────────────────────
    risk_keys = ["vol_1y","vol_3y","vol_5y","neg_vol_1y","sharpe_3y","sortino_1y"]
    ptf_r = ptf_wavg(risk_keys)

    risk_hdr = [Paragraph(f"<b>{t}</b>", HDR) for t in
                ["Fondo","Peso","Vol. 1A","Vol. 3A","Vol. 5A","Vol. Neg. 1A","Sharpe 3A","Sortino 1A"]]

    ptf_risk_row = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph("<b>100%</b>", WH),
        Paragraph(ptf_r.get("vol_1y","N/D"),     WH),
        Paragraph(ptf_r.get("vol_3y","N/D"),     WH),
        Paragraph(ptf_r.get("vol_5y","N/D"),     WH),
        Paragraph(ptf_r.get("neg_vol_1y","N/D"), WH),
        Paragraph(ptf_r.get("sharpe_3y","N/D"),  WH),
        Paragraph(ptf_r.get("sortino_1y","N/D"), WH),
    ]

    risk_rows = [risk_hdr, ptf_risk_row]
    for _, row in d_sorted.iterrows():
        fd  = (fund_data or {}).get(row["nome"], {})
        ana = fd.get("analysis", {})
        def gv_r(k): return ana.get(k,"N/D")
        risk_rows.append([
            Paragraph(row["nome"][:48], SM),
            Paragraph(f"{row[wcol]*100:.1f}%", SM),
            Paragraph(gv_r("vol_1y"),     SM), Paragraph(gv_r("vol_3y"),     SM), Paragraph(gv_r("vol_5y"),    SM),
            Paragraph(gv_r("neg_vol_1y"), SM), Paragraph(gv_r("sharpe_3y"),  SM), Paragraph(gv_r("sortino_1y"),SM),
        ])

    risk_tbl = Table(risk_rows,
        colWidths=[5.2*cm,1.4*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm],
        repeatRows=1)
    ts_risk = [
        ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),
        ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
        ("BACKGROUND",(0,1),(-1,1), rl_colors.HexColor("#1B4332")),
        ("LINEBELOW",(0,1),(-1,1),  2, rl_colors.HexColor("#C9A84C")),
        ("FONTSIZE",(0,0),(-1,-1),  8),
        ("PADDING",(0,0),(-1,-1),   5),
        ("ROWBACKGROUNDS",(0,2),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",(1,0),(-1,-1),     "CENTER"),
        ("VALIGN",(0,0),(-1,-1),    "MIDDLE"),
    ]
    risk_tbl.setStyle(TableStyle(ts_risk))

    NOTE = S("NT", fontName="Helvetica-Oblique", fontSize=6.5,
             textColor=rl_colors.HexColor("#94A3B8"), leading=9)
    # KeepTogether: titolo + tabella rischio + nota nella stessa pagina
    story.append(KeepTogether([
        Paragraph("Metriche di Rischio", SC),
        risk_tbl,
        Spacer(1,6),
        Paragraph(
            "◆ I valori del Portafoglio sono medie ponderate per peso dei singoli fondi. "
            "La volatilità di portafoglio effettiva dipende anche dalle correlazioni tra i fondi. "
            "Dati forniti a titolo indicativo.", NOTE),
    ]))

    story.append(Spacer(1, 14))

    # ── ALLOCATION TABLE ─────────────────────────────────────
    def get_fi_metric(nome: str, key: str):
        """Return fixed-income metric (duration/credit_rating/ytm) from factbook or None."""
        if not _fb:
            return None
        norm = _normalize_for_unp(nome)
        norm = _FUND_ALIASES.get(norm, norm)
        entry = _fb.get(norm)
        if not entry or not isinstance(entry, dict):
            best, best_len = None, 0
            for fb_key, fb_val in _fb.items():
                if (isinstance(fb_val, dict)
                        and (fb_key in norm or norm in fb_key)
                        and len(fb_key) > best_len):
                    best, best_len = fb_val, len(fb_key)
            entry = best
        if not entry or not isinstance(entry, dict):
            return None
        return entry.get(key)

    # Per-fund helpers: prefer factbook composition over binary Excel values
    def _az_eff(row):
        v = get_fi_metric(row["nome"], "fb_az_pct")
        return v if v is not None else row["az_pct"]

    def _obb_eff(row):
        v = get_fi_metric(row["nome"], "fb_obb_pct")
        return v if v is not None else row["obb_pct"]

    # Portfolio-level composition (factbook-aware weighted averages)
    _ptf_az_wtd  = sum(_row[wcol] * _az_eff(_row)  for _, _row in d_sorted.iterrows())
    _ptf_obb_wtd = sum(_row[wcol] * _obb_eff(_row) for _, _row in d_sorted.iterrows())

    # Duration and rating weighted by effective obb% × w
    _ptf_dur_num = _ptf_dur_den = 0.0
    _ptf_rat_num = _ptf_rat_den = 0.0
    for _, _row in d_sorted.iterrows():
        _w   = _row[wcol]
        _obb = _obb_eff(_row)            # factbook obb% if available, else Excel
        _dur = get_fi_metric(_row["nome"], "duration")
        _rat = get_fi_metric(_row["nome"], "credit_rating")
        if isinstance(_dur, (int, float)) and _obb > 0:
            _ptf_dur_num += _w * _obb * _dur
            _ptf_dur_den += _w * _obb
        if isinstance(_rat, str) and _rat in RATING_SCALE and _obb > 0:
            _ptf_rat_num += _w * _obb * RATING_SCALE[_rat]
            _ptf_rat_den += _w * _obb

    _ptf_dur_str = (f"{_ptf_dur_num / _ptf_dur_den:.2f}"
                    if _ptf_dur_den > 0.001 else "N/D")
    if _ptf_rat_den > 0.001:
        _ri = max(1, min(22, round(_ptf_rat_num / _ptf_rat_den)))
        _ptf_rat_str = RATING_INVERSE.get(_ri, "N/D")
    else:
        _ptf_rat_str = "N/D"

    alloc_hdr = [Paragraph(f"<b>{t}</b>", HDR) for t in
                 ["Fondo", "Peso", "% Azionario", "% Obbligazionario",
                  "Duration", "Rating Medio", "Cat. FIDA", "FIDArating",
                  "Morningstar"]]
    alloc_ptf = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph("<b>100%</b>",                       WH),
        Paragraph(f"<b>{_ptf_az_wtd*100:.1f}%</b>",   WH),
        Paragraph(f"<b>{_ptf_obb_wtd*100:.1f}%</b>",  WH),
        Paragraph(f"<b>{_ptf_dur_str}</b>",            WH),
        Paragraph(f"<b>{_ptf_rat_str}</b>",            WH),
        Paragraph("",                                  WH),
        Paragraph("",                                  WH),
        Paragraph("",                                  WH),
    ]
    # FIDArating visual badges in PDF:
    # BACKGROUND per-cell works in ReportLab TableStyle and overrides
    # ROWBACKGROUNDS when placed after it.  Paragraph textColor must be
    # set via ParagraphStyle (not TEXTCOLOR in TableStyle — that is ignored
    # when the cell contains a Paragraph object).
    # 5=dark green bg / 4=medium green / 3=light green — all white text.
    # 1-2: no background, bold black (no red for print).
    _FIDA_BG_HEX = {"5": "#166534", "4": "#15803d", "3": "#16a34a"}

    def _fida_para(val: str):
        _v = str(val).strip()
        if _v in _FIDA_BG_HEX:
            return Paragraph(
                _v,
                S(f"SMF{_v}", fontName="Helvetica-Bold", fontSize=7.5,
                  textColor=rl_colors.white, leading=11))
        if _v in ("1", "2"):
            return Paragraph(
                _v,
                S(f"SMF{_v}", fontName="Helvetica-Bold", fontSize=7.5,
                  textColor=rl_colors.HexColor("#1E293B"), leading=11))
        return Paragraph(val, SM)   # "—" or unknown

    # Morningstar data for PDF (loaded from cache / session state)
    _ms_pdf = st.session_state.get("_ms_data") or load_ms_cache()

    # Morningstar amber/gold palette for PDF
    _MS_BG_HEX = {"5": "#78350F", "4": "#92400E", "3": "#B45309"}

    def _ms_para(val) -> Paragraph:
        """ReportLab Paragraph for a Morningstar rating value.
        Shows numeric value with star count in ASCII to stay within Helvetica charset.
        """
        try:
            v = int(val)
        except (TypeError, ValueError):
            return Paragraph("—", SM)
        label = f"{v} {'*'*v}"   # e.g. "4 ****" — ASCII-safe, no Unicode stars
        if str(v) in _MS_BG_HEX:
            return Paragraph(
                label,
                S(f"SMMSP{v}", fontName="Helvetica-Bold", fontSize=7,
                  textColor=rl_colors.white, leading=11))
        return Paragraph(
            label,
            S(f"SMMSd{v}", fontName="Helvetica", fontSize=7,
              textColor=rl_colors.HexColor("#475569"), leading=11))

    alloc_fund_rows = []
    _fida_vals = []   # keep to build BACKGROUND commands after the loop
    _ms_vals   = []
    for _, _row in d_sorted.iterrows():
        _dur2  = get_fi_metric(_row["nome"], "duration")
        _rat2  = get_fi_metric(_row["nome"], "credit_rating")
        _az_s  = _az_eff(_row)  * 100
        _obb_s = _obb_eff(_row) * 100
        _fd_ov2 = (fund_data or {}).get(_row["nome"], {}).get("overview", {})
        _cat2   = _fd_ov2.get("cat_assog") or "—"
        _fida2  = _fd_ov2.get("fida_rating") or "—"
        _ms2    = _ms_pdf.get(_row["nome"], {}).get("ms_rating")
        _fida_vals.append(str(_fida2).strip())
        _ms_vals.append(str(_ms2).strip() if _ms2 is not None else "—")
        alloc_fund_rows.append([
            Paragraph(_row["nome"][:48], SM),
            Paragraph(f"{_row[wcol]*100:.1f}%",                          SM),
            Paragraph(f"{_az_s:.1f}%",                                   SM),
            Paragraph(f"{_obb_s:.1f}%",                                  SM),
            Paragraph(f"{_dur2:.2f}" if isinstance(_dur2, (int, float)) else "—", SM),
            Paragraph(_rat2 if isinstance(_rat2, str) else "—",           SM),
            Paragraph(_cat2,                                               SM),
            _fida_para(_fida2),
            _ms_para(_ms2),
        ])

    # Build per-row BACKGROUND commands for FIDArating (col 7) and Morningstar (col 8).
    _fida_bg_cmds = []
    for _fi, _fv in enumerate(_fida_vals):
        _bg_hex = _FIDA_BG_HEX.get(_fv)
        if _bg_hex:
            _tr = _fi + 2   # row 0=hdr, 1=ptf summary, 2+=fund rows
            _fida_bg_cmds.append(
                ("BACKGROUND", (7, _tr), (7, _tr),
                 rl_colors.HexColor(_bg_hex)))
    for _mi, _mv in enumerate(_ms_vals):
        _bg_hex_ms = _MS_BG_HEX.get(_mv)
        if _bg_hex_ms:
            _tr = _mi + 2
            _fida_bg_cmds.append(
                ("BACKGROUND", (8, _tr), (8, _tr),
                 rl_colors.HexColor(_bg_hex_ms)))

    alloc_tbl = Table(
        [alloc_hdr, alloc_ptf] + alloc_fund_rows,
        # Slightly reduced cols to fit new Morningstar column (total ~17.1 cm)
        colWidths=[3.8*cm, 1.1*cm, 1.3*cm, 1.5*cm, 1.4*cm, 1.8*cm, 2.6*cm, 1.5*cm, 2.1*cm],
        repeatRows=1,
    )
    alloc_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0,0), (-1,0),  rl_colors.HexColor("#0D1B2A")),
        ("TEXTCOLOR",      (0,0), (-1,0),  rl_colors.white),
        ("FONTNAME",       (0,0), (-1,0),  "Helvetica-Bold"),
        ("BACKGROUND",     (0,1), (-1,1),  rl_colors.HexColor("#1B4332")),
        ("LINEBELOW",      (0,1), (-1,1),  2, rl_colors.HexColor("#C9A84C")),
        ("FONTSIZE",       (0,0), (-1,-1), 8),
        ("PADDING",        (0,0), (-1,-1), 5),
        ("ROWBACKGROUNDS", (0,2), (-1,-1),
         [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",      (0,0), (-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",          (1,0), (-1,-1), "CENTER"),
        ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
        *_fida_bg_cmds,   # coloured cell backgrounds — placed last to override
    ]))

    NOTE_A = S("NTA", fontName="Helvetica-Oblique", fontSize=6.5,
               textColor=rl_colors.HexColor("#94A3B8"), leading=9)
    story.append(KeepTogether([
        Paragraph("Scomposizione Azionario / Obbligazionario", SC),
        alloc_tbl,
        Spacer(1, 6),
        Paragraph(
            "◆ Le quote azionaria e obbligazionaria derivano dal Factbook AZ Investments"
            + (f" al {_fb_ref}" if _fb_ref else "")
            + " (pagine singoli fondi); in assenza del dato factbook si usa la "
              "classificazione binaria del foglio Excel. "
              "Duration e Rating Medio riguardano la sola componente obbligazionaria. "
              "Il Rating Medio di Portafoglio è la media ponderata degli score numerici "
              "(AAA=1 … D=22) × peso × quota obbligazionaria. "
              "Il simbolo — indica dato non disponibile.",
            NOTE_A),
    ]))

    story.append(Spacer(1, 14))

    # ── UNP / IUNP TABLE ─────────────────────────────────────
    # Pre-compute per-fund UNP/IUNP and portfolio weighted average
    _fund_unp: dict = {}
    _wtd_unp = _wtd_iunp = _cov_w = 0.0
    for _, _row in d_sorted.iterrows():
        _u, _iu = lookup_unp(_row["nome"])
        _fund_unp[_row["nome"]] = (_u, _iu)
        if _u is not None:
            _w = _row[wcol]
            _wtd_unp  += _u  * _w
            _wtd_iunp += _iu * _w
            _cov_w    += _w

    if _cov_w > 0.01:
        _ptf_unp_str  = f"{_wtd_unp  / _cov_w:.2f}%"
        _ptf_iunp_str = f"{_wtd_iunp / _cov_w:.2f}%"
    else:
        _ptf_unp_str = _ptf_iunp_str = "N/D"

    unp_hdr_row = [Paragraph(f"<b>{t}</b>", HDR) for t in
                   ["Fondo", "Peso", "%UNP", "%IUNP36", "FIDArating", "Morningstar"]]
    unp_ptf_row = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph("<b>100%</b>", WH),
        Paragraph(f"<b>{_ptf_unp_str}</b>",  WH),
        Paragraph(f"<b>{_ptf_iunp_str}</b>", WH),
        Paragraph("", WH),
        Paragraph("", WH),
    ]
    unp_fund_rows = []
    _unp_fida_vals: list = []
    _unp_ms_vals:   list = []
    for _, _row in d_sorted.iterrows():
        _u, _iu   = _fund_unp[_row["nome"]]
        _fd_ov_u  = (fund_data or {}).get(_row["nome"], {}).get("overview", {})
        _fida_u   = str(_fd_ov_u.get("fida_rating") or "—").strip()
        _ms_u     = _ms_pdf.get(_row["nome"], {}).get("ms_rating")
        _unp_fida_vals.append(_fida_u)
        _unp_ms_vals.append(str(_ms_u).strip() if _ms_u is not None else "—")
        unp_fund_rows.append([
            Paragraph(_row["nome"][:50], SM),
            Paragraph(f"{_row[wcol]*100:.1f}%", SM),
            Paragraph(f"{_u:.2f}%"  if _u  is not None else "—", SM),
            Paragraph(f"{_iu:.2f}%" if _iu is not None else "—", SM),
            _fida_para(_fida_u),
            _ms_para(_ms_u),
        ])

    # Per-cell background for FIDArating (col 4) and Morningstar (col 5)
    _unp_bg_cmds: list = []
    for _fi, _fv in enumerate(_unp_fida_vals):
        _bh = _FIDA_BG_HEX.get(_fv)
        if _bh:
            _unp_bg_cmds.append(("BACKGROUND", (4, _fi+2), (4, _fi+2), rl_colors.HexColor(_bh)))
    for _mi, _mv in enumerate(_unp_ms_vals):
        _bh = _MS_BG_HEX.get(_mv)
        if _bh:
            _unp_bg_cmds.append(("BACKGROUND", (5, _mi+2), (5, _mi+2), rl_colors.HexColor(_bh)))

    unp_tbl = Table(
        [unp_hdr_row, unp_ptf_row] + unp_fund_rows,
        colWidths=[5.0*cm, 1.5*cm, 2.0*cm, 2.0*cm, 2.0*cm, 4.5*cm],
        repeatRows=1,
    )
    unp_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0,0), (-1,0),  rl_colors.HexColor("#0D1B2A")),
        ("TEXTCOLOR",      (0,0), (-1,0),  rl_colors.white),
        ("FONTNAME",       (0,0), (-1,0),  "Helvetica-Bold"),
        ("BACKGROUND",     (0,1), (-1,1),  rl_colors.HexColor("#1B4332")),
        ("LINEBELOW",      (0,1), (-1,1),  2, rl_colors.HexColor("#C9A84C")),
        ("FONTSIZE",       (0,0), (-1,-1), 8),
        ("PADDING",        (0,0), (-1,-1), 5),
        ("ROWBACKGROUNDS", (0,2), (-1,-1),
         [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",      (0,0), (-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",          (1,0), (-1,-1), "CENTER"),
        ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
        *_unp_bg_cmds,
    ]))

    NOTE_U = S("NTU", fontName="Helvetica-Oblique", fontSize=6.5,
               textColor=rl_colors.HexColor("#94A3B8"), leading=9)
    story.append(KeepTogether([
        Paragraph("UNP e IUNP dei Fondi in Portafoglio", SC),
        unp_tbl,
        Spacer(1, 6),
        Paragraph(
            "◆ UNP (Utile Netto di Portafoglio): commissione annua netta percepita dal consulente. "
            "IUNP36: indice UNP calcolato su orizzonte triennale. "
            "Fonte: Catalogo Prodotti & Servizi Azimut, settembre 2025. "
            "La riga Portafoglio è la media ponderata per peso dei fondi per cui il dato è disponibile. "
            "Il simbolo — indica che il fondo non è presente nel catalogo.",
            NOTE_U),
    ]))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════
    # PAGES 3+: SCHEDE SINGOLI FONDI
    # ════════════════════════════════════════════════════════
    story.append(Paragraph("AZIMUT INVESTMENTS  ·  AAS EMILIA ROMAGNA MARCHE UMBRIA", EY))
    story.append(Spacer(1,4))
    story.append(Paragraph("Schede Analitiche dei Fondi", T))
    story.append(Paragraph(
        f"Profilo {profile.title()}  ·  Fonte: FIDA FondiDoc  ·  {datetime.date.today().strftime('%d %B %Y')}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=6))
    story.append(Paragraph(
        '🔍 <link href="https://www.morningstar.it/it/funds/SecuritySearchResults.aspx">'
        '<u>Motore di ricerca Morningstar</u></link>', LK))
    story.append(Spacer(1, 10))

    for idx, (_, row) in enumerate(d_sorted.iterrows()):
        fd  = (fund_data or {}).get(row["nome"], {})
        ov  = fd.get("overview",  {})
        ana = fd.get("analysis",  {})

        def gv(k,src=ana,fallback="N/D"): return src.get(k,fallback)

        # Fund header block
        srri_str = f"SRRI {gv('srri',ov,'—')}/7" if gv('srri',ov) != "N/D" else ""
        nav_str  = f"NAV {gv('nav')} € ({gv('last_update')})" if gv('nav') != "N/D" else ""
        rating_s = f"FIDArating {gv('fida_rating',ov)}" if gv('fida_rating',ov) not in ("N/D","—") else ""

        # ── Intestazione fondo (3 righe × 1 colonna) ─────────
        meta_extra = "  ·  ".join(x for x in [srri_str, rating_s, nav_str] if x)

        # ISIN: estratto dall'URL FondiDoc oppure dal foglio FIDA
        isin = fd.get("isin", "") or isin_map.get(row["nome"], "")
        isin_str = f"  ·  ISIN: <b>{isin}</b>" if isin else ""

        hdr_rows = [
            [Paragraph(f"<b>{row['nome']}</b>", FS)],
            [Paragraph(f"Peso: <b>{row[wcol]*100:.1f}%</b>  ·  {row['categoria']}{isin_str}", FK)],
            [Paragraph(meta_extra or "—", FK)],
        ]

        hdr_tbl = Table(hdr_rows, colWidths=[17*cm])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), rl_colors.HexColor("#F0F4F9")),
            ("LEFTPADDING",(0,0),(-1,-1), 10),
            ("RIGHTPADDING",(0,0),(-1,-1), 10),
            ("TOPPADDING",(0,0),(-1,0), 10),
            ("BOTTOMPADDING",(0,-1),(-1,-1), 10),
            ("TOPPADDING",(0,1),(-1,-1), 2),
            ("BOTTOMPADDING",(0,0),(-1,-2), 2),
            ("LINEBELOW",(0,-1),(-1,-1), 2, rl_colors.HexColor("#C9A84C")),
        ]))

        # ── Tabella rendimenti fondo ──────────────────────────
        def pval(v):
            try:
                num = float(v.replace("%","").replace(",","."))
                c = "#1A7A4A" if num>0 else ("#C0392B" if num<0 else "#475569")
                return Paragraph(f'<font color="{c}"><b>{v}</b></font>', BD)
            except: return Paragraph(v, BD)

        perf_data = [
            [Paragraph("<b>Metrica</b>",HDR), Paragraph("<b>YTD</b>",HDR),
             Paragraph("<b>1 Anno</b>",HDR), Paragraph("<b>3 Anni</b>",HDR), Paragraph("<b>5 Anni</b>",HDR)],
            [Paragraph("Performance",SM),
             pval(gv("ytd")), pval(gv("perf_1y")), pval(gv("perf_3y")), pval(gv("perf_5y"))],
            [Paragraph("Volatilità",SM),
             Paragraph("—",SM), Paragraph(gv("vol_1y"),SM), Paragraph(gv("vol_3y"),SM), Paragraph(gv("vol_5y"),SM)],
            [Paragraph("Vol. Neg.",SM),
             Paragraph("—",SM), Paragraph(gv("neg_vol_1y"),SM), Paragraph(gv("neg_vol_3y"),SM), Paragraph(gv("neg_vol_5y"),SM)],
            [Paragraph("Sharpe",SM),
             Paragraph("—",SM), Paragraph("—",SM), Paragraph(gv("sharpe_3y"),SM), Paragraph(gv("sharpe_5y"),SM)],
            [Paragraph("Sortino",SM),
             Paragraph("—",SM), Paragraph(gv("sortino_1y"),SM), Paragraph("—",SM), Paragraph("—",SM)],
        ]
        perf_tbl2 = Table(perf_data, colWidths=[2.4*cm,1.5*cm,1.8*cm,1.8*cm,1.8*cm])
        perf_tbl2.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),
            ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
            ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),  7.5),
            ("PADDING",(0,0),(-1,-1),   4),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
            ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
            ("ALIGN",(1,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))

        # ── Dettagli fondo ───────────────────────────────────
        det_data = [
            [Paragraph("<b>Dettagli Fondo</b>", BD)],
            [Paragraph(f"Data avvio: {gv('start_date',ov,'—')}", SM)],
            [Paragraph(f"Distribuzione: {gv('income',ov,'—')}", SM)],
            [Paragraph(f"Categoria: {gv('cat_assog',ov,'—')}", SM)],
            [Paragraph(f"Gestione: {gv('mgmt_fee',ov,'—')}  |  Perf.: {gv('perf_fee',ov,'—')}", SM)],
            [Paragraph(f"Sottoscrizione: {gv('sub_fee',ov,'—')}", SM)],
            [Paragraph(f"<b>FIDArating:</b> {gv('fida_rating',ov,'—')}  |  Score: {gv('fida_score',ov,'—')}", SM)],
        ]
        det_tbl = Table([[d[0]] for d in det_data], colWidths=[7.3*cm])
        det_tbl.setStyle(TableStyle([
            ("PADDING",(0,0),(-1,-1), 3),
            ("TOPPADDING",(0,0),(-1,0), 6),
            ("LINEBELOW",(0,0),(0,0), 0.8, rl_colors.HexColor("#C9A84C")),
            ("BACKGROUND",(0,0),(0,-1), rl_colors.HexColor("#F8FAFC")),
        ]))

        mid_row = Table([[perf_tbl2, det_tbl]], colWidths=[9.7*cm, 7.3*cm])
        mid_row.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1), "TOP"),
            ("PADDING",(0,0),(-1,-1), 0),
            ("LEFTPADDING",(1,0),(1,-1), 10),
        ]))

        # ── Grafico rendimenti annuali ───────────────────────
        annual  = ana.get("annual_perf")
        bar_buf = _mpl_annual_bar(annual, row["nome"]) if annual else None

        # ── KeepTogether: tutta la scheda su stessa pagina ───
        card = [Spacer(1,6), hdr_tbl, Spacer(1,6), mid_row]
        if bar_buf:
            card += [Spacer(1,4),
                     Paragraph("<b>Performance Annuale (%)</b>", SM),
                     RLImage(bar_buf, width=14*cm, height=3.2*cm)]
        story.append(KeepTogether(card))

        # Separatore tra fondi
        if idx < len(d_sorted)-1:
            story.append(HRFlowable(width="100%", thickness=0.5,
                                    color=rl_colors.HexColor("#CBD5E1"),
                                    spaceBefore=8, spaceAfter=8))

    # ── FOOTER ─────────────────────────────────────────────
    story.append(PageBreak())
    story.append(HRFlowable(width="100%",thickness=0.5,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=8))
    story.append(Paragraph(
        "Documento generato automaticamente a scopo illustrativo. I dati di performance provengono da FIDA FondiDoc "
        "(fondidoc.it). I pesi indicati sono riferiti al portafoglio modello e non costituiscono offerta o consulenza "
        "di investimento. Rendimenti passati non garantiscono risultati futuri. © Azimut Group — uso interno.", FT))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════
# FREE PORTFOLIO BUILDER
# ════════════════════════════════════════════════════════════

def free_portfolio_ui(data):
    fida = data.get("FIDA", pd.DataFrame())
    if fida.empty:
        st.error("❌ Foglio FIDA non trovato."); return None

    az_lookup = {}
    for sname in ["PTF FULL","PTF SHORT"]:
        if sname in data and not data[sname].empty:
            for _,r in data[sname].iterrows(): az_lookup[r["nome"]] = r["az_pct"]

    if "free_ptf" not in st.session_state: st.session_state.free_ptf = []
    st.markdown('<p class="sec-title">Costruttore Portafoglio Libero</p>',unsafe_allow_html=True)

    # ── FIDArating filter ─────────────────────────────────────────────────────
    _fd_live_free = st.session_state.get("_scomp_fd", {})
    # Build rating map: fund_name → "5"/"4"/"3"/"2"/"1"/"—"
    _fr_map = {
        r["nome"]: (str(_fd_live_free.get(r["nome"], {})
                        .get("overview", {}).get("fida_rating", "") or "").strip()
                    or "—")
        for _, r in fida.iterrows()
    }
    _has_ratings = any(v != "—" for v in _fr_map.values())

    # ── Morningstar filter ────────────────────────────────────────────────────
    _ms_live_free = st.session_state.get("_ms_data") or load_ms_cache()
    _ms_fr_map = {
        r["nome"]: (str(_ms_live_free.get(r["nome"], {}).get("ms_rating", "") or "").strip()
                    or "—")
        for _, r in fida.iterrows()
    }
    _has_ms_ratings = any(v != "—" for v in _ms_fr_map.values())

    _RATING_OPTS  = ["5", "4", "3", "2", "1", "—"]
    _FIDA_LABEL = {
        "5": "⭐⭐⭐⭐⭐  FIDArating 5",
        "4": "⭐⭐⭐⭐  FIDArating 4",
        "3": "⭐⭐⭐  FIDArating 3",
        "2": "⭐⭐  FIDArating 2",
        "1": "⭐  FIDArating 1",
        "—": "Nessun rating",
    }
    _MS_LABEL = {
        "5": "★★★★★  Morningstar 5",
        "4": "★★★★  Morningstar 4",
        "3": "★★★  Morningstar 3",
        "2": "★★  Morningstar 2",
        "1": "★  Morningstar 1",
        "—": "Nessun rating",
    }

    # Layout: two filter columns side by side
    _fcol1, _fcol2 = st.columns(2)
    with _fcol1:
        if _has_ratings:
            _sel_fida = st.multiselect(
                "🔵  Filtra per FIDArating",
                options=_RATING_OPTS,
                default=_RATING_OPTS,
                format_func=lambda x: _FIDA_LABEL[x],
                key="free_fida_filter",
            )
            _active_fida = set(_sel_fida) if _sel_fida else set(_RATING_OPTS)
        else:
            _active_fida = set(_RATING_OPTS)
            st.caption("ℹ️ FIDArating — scarica dati FondiDoc")
    with _fcol2:
        if _has_ms_ratings:
            _sel_ms = st.multiselect(
                "⭐  Filtra per Morningstar",
                options=_RATING_OPTS,
                default=_RATING_OPTS,
                format_func=lambda x: _MS_LABEL[x],
                key="free_ms_filter",
            )
            _active_ms = set(_sel_ms) if _sel_ms else set(_RATING_OPTS)
        else:
            _active_ms = set(_RATING_OPTS)
            st.caption("ℹ️ Morningstar — scarica rating FondiOnline")

    fida_filtered = fida[fida["nome"].apply(
        lambda n: (
            _fr_map.get(n, "—") in _active_fida
            and _ms_fr_map.get(n, "—") in _active_ms
        )
    )]
    if fida_filtered.empty:
        st.warning("⚠️ Nessun fondo corrisponde ai filtri selezionati.")
        fida_filtered = fida  # fallback: show all

    # Build option labels: include FIDArating and Morningstar badges
    def _fund_option(r):
        fr   = _fr_map.get(r["nome"], "—")
        ms_r = _ms_fr_map.get(r["nome"], "—")
        ftag = f" · F{fr}"   if fr   != "—" else ""
        mtag = f" · M{ms_r}" if ms_r != "—" else ""
        if r["macro_cat"] != "Altro":
            return f"{r['nome']}{ftag}{mtag}  [{r['macro_cat']}]"
        return f"{r['nome']}{ftag}{mtag}"

    options = fida_filtered.apply(_fund_option, axis=1).tolist()

    # st.multiselect has native live search built into Streamlit (no Enter needed).
    # max_selections=1 limits it to a single fund, giving us a searchable picker.
    c1,c2,c3 = st.columns([3.5,1,0.8])
    with c1:
        _sel_list = st.multiselect(
            "🔍  Seleziona / cerca fondo:",
            options=options,
            max_selections=1,
            placeholder="Digita per cercare, es. «comm», «glob», «targ»…",
            key="sel_fund_ms",
        )
        sel = _sel_list[0] if _sel_list else (options[0] if options else "")
    with c2: w = st.number_input("Peso %",0.1,100.0,10.0,0.5,key="sel_w")
    with c3:
        st.markdown("<br>",unsafe_allow_html=True)
        if st.button("➕ Aggiungi",use_container_width=True):
            # Strip FIDArating tag "· FN", Morningstar tag "· MN" and macro-cat "  [...]"
            fname = re.split(r'\s+·\s+[FM]\d|\s{2}\[', sel)[0].strip()
            if any(f["nome"]==fname for f in st.session_state.free_ptf):
                st.toast("⚠️ Fondo già presente!",icon="⚠️")
            else:
                fd = fida[fida["nome"]==fname].iloc[0] if not fida[fida["nome"]==fname].empty else None
                mc = fd["macro_cat"] if fd is not None else "Altro"
                az = az_lookup.get(fname,DEFAULT_AZ.get(mc,0.5))
                st.session_state.free_ptf.append({"nome":fname,"categoria":fd["categoria"] if fd is not None else "","macro_cat":mc,"az_pct":az,"w_input":w})
                st.rerun()

    if not st.session_state.free_ptf: st.info("☝️ Aggiungi fondi."); return None
    st.markdown("**Fondi nel portafoglio:**")
    total_w = 0.0
    for i, fund in enumerate(st.session_state.free_ptf):
        r1,r2,r3 = st.columns([4,1.5,0.6])
        with r1: st.markdown(f"**{fund['nome']}** <span style='color:#64748b;font-size:.8rem;'>— {fund['macro_cat']}</span>",unsafe_allow_html=True)
        with r2:
            nw = st.number_input("Peso",0.0,100.0,float(fund["w_input"]),0.5,key=f"fw_{i}",label_visibility="collapsed")
            st.session_state.free_ptf[i]["w_input"] = nw
        with r3:
            if st.button("🗑️",key=f"del_{i}",use_container_width=True):
                st.session_state.free_ptf.pop(i); st.rerun()
        total_w += st.session_state.free_ptf[i]["w_input"]

    diff = abs(total_w-100.0)
    if diff<0.05: st.markdown(f'<div class="w-ok">✅ Somma pesi: <b>{total_w:.1f}%</b> — OK!</div>',unsafe_allow_html=True)
    else:         st.markdown(f'<div class="w-warn">⚠️ Somma pesi: <b>{total_w:.1f}%</b> (mancano {100-total_w:+.1f}%)</div>',unsafe_allow_html=True)
    if diff>0.5: return None

    records = [{"nome":f["nome"],"categoria":f["categoria"],"gruppo":f["macro_cat"],"macro_cat":f["macro_cat"],"az_pct":f["az_pct"],"obb_pct":1-f["az_pct"],"r_weight":f["w_input"]/100,"w_cons":f["w_input"]/100,"w_equil":f["w_input"]/100,"w_accr":f["w_input"]/100} for f in st.session_state.free_ptf]
    df = pd.DataFrame(records)
    df["macro_cat"] = df["categoria"].apply(get_macro)
    df = assign_colors(df)
    for wc in ["w_cons","w_equil","w_accr"]:
        t = df[wc].sum(); df[wc] = df[wc]/t if t>0 else df[wc]
    return df


# ════════════════════════════════════════════════════════════
# GLOBAL PERSPECTIVES — PDF parsing & SUGGERITO portfolio
# ════════════════════════════════════════════════════════════

def _resolve_nome_for_fd(nome_pdf: str, fund_data: dict) -> str:
    """Map a PDF-format name (e.g. "AZ Allocation - Balanced Plus") to the
    actual key present in fund_data (usually an Excel-abbreviated form like
    "AZ F.1 All. Balanced Plus A Cap EUR").  Uses the same normalisation logic
    as UNP lookup.  Falls back to nome_pdf if no match is found."""
    if not fund_data or nome_pdf in fund_data:
        return nome_pdf
    norm = _normalize_for_unp(nome_pdf)
    norm = _FUND_ALIASES.get(norm, norm)
    best_key, best_len = None, 0
    for key in fund_data:
        k_norm = _normalize_for_unp(key)
        k_norm = _FUND_ALIASES.get(k_norm, k_norm)
        if k_norm == norm:
            return key                               # exact normalised match
        if (k_norm in norm or norm in k_norm) and len(k_norm) > best_len:
            best_key, best_len = key, len(k_norm)
    return best_key if best_key else nome_pdf


def parse_global_perspectives(pdf_bytes: bytes):
    """Parse a *Global Perspectives* quarterly PDF and return the three
    Azimut View scenario portfolios (Base, Bear, Bull), excluding private-
    market funds (ELTIF, RAIF, Demos, …).

    Returns
    -------
    dict | None
        ``{
            "Base": {
                "info": "Equity 32% · Bond 38% · Private Markets 30%",
                "funds": [
                    {"nome": "AZ Allocation - Balanced Plus",
                     "gruppo": "ALLOCATION",
                     "categoria": "Bilanciati/Flessibili",
                     "az_pct": 0.50, "obb_pct": 0.50, "weight": 0.045},
                    ...
                ],
                "subcat_weights": {"alloc_balanced": 25, ...},
            },
            "Bear": {...},
            "Bull": {...},
        }``
    or ``None`` if the PDF could not be recognised.
    """
    try:
        import pdfplumber
    except ImportError:
        return None
    import io as _io

    # ── 1. Extract text page by page ──────────────────────────────────────────
    try:
        pages = []
        with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
            for pg in pdf.pages:
                t = pg.extract_text()
                if t:
                    pages.append(t)
    except Exception:
        return None
    if not pages:
        return None
    full = "\n".join(pages)

    # ── 2. Locate scenario section boundaries ─────────────────────────────────
    _SC_PATS: dict = {
        "Base": [r"AZIMUT\s+VIEW\s+SCENARIO\s+BASE", r"Scenario\s+Base"],
        "Bear": [r"AZIMUT\s+VIEW\s+SCENARIO\s+BEAR", r"Scenario\s+Bear"],
        "Bull": [r"AZIMUT\s+VIEW\s+SCENARIO\s+BULL", r"Scenario\s+Bull"],
    }
    positions: dict = {}
    for sc, pats in _SC_PATS.items():
        for pat in pats:
            m = re.search(pat, full, re.IGNORECASE)
            if m:
                positions[sc] = m.start()
                break
    if len(positions) < 3:
        return None

    sorted_sc = sorted(positions, key=lambda k: positions[k])
    sections: dict = {}
    for i, sc in enumerate(sorted_sc):
        start = positions[sc]
        end   = positions[sorted_sc[i + 1]] if i + 1 < len(sorted_sc) else len(full)
        sections[sc] = full[start:end]

    # ── 3. Sub-category meta ──────────────────────────────────────────────────
    _SUBCAT_LABELS: list = [
        ("alloc_balanced",  r"Allocation\s*[-–]\s*Balanced"),
        ("alloc_flexible",  r"Allocation\s*[-–]\s*Flex"),
        ("bond_aggregate",  r"Bond\s*[-–]\s*Aggregate"),
        ("bond_thematic",   r"Bond\s*[-–]\s*Thematic"),
        ("bond_em",         r"Bond\s*[-–]\s*Paesi\s+emergenti"),
        ("bond_target",     r"Bond\s*[-–]\s*Target"),
        ("equity_thematic", r"Equity\s*[-–]\s*Thematic"),
        ("equity_dev",      r"Equity\s*[-–]\s*Paesi\s+sviluppati"),
        ("equity_em",       r"Equity\s*[-–]\s*Paesi\s+emergenti"),
    ]
    _SUBCAT_GROUP = {
        "alloc_balanced": "ALLOCATION",   "alloc_flexible": "ALLOCATION",
        "bond_aggregate": "BOND",         "bond_thematic":  "BOND",
        "bond_em":        "BOND",         "bond_target":    "BOND",
        "equity_thematic":"AZIONARI (LONG)",
        "equity_dev":     "AZIONARI (LONG)", "equity_em": "AZIONARI (LONG)",
    }
    _SUBCAT_CAT = {
        "alloc_balanced":  "Bilanciati/Flessibili",
        "alloc_flexible":  "Bilanciati/Flessibili",
        "bond_aggregate":  "Obbligazionari", "bond_thematic": "Obbligazionari",
        "bond_em":         "Obbligazionari", "bond_target":   "Obbligazionari",
        "equity_thematic": "Azionari",
        "equity_dev":      "Azionari",       "equity_em":     "Azionari",
    }
    # Private-market keywords → skip these fund lines
    _PRIV_KW = frozenset([
        "raif", "eltif", "demos ", "yhox", "direct investments",
        "hybrid growth", "automobile", "infrastrutture", "real assets",
        "digitech fund", "young group", "alicrowd", "hipstr", " p103",
        "italia 500", "globALinvest", "borletti", "broadlight", "highpost",
        "roundshield", "pensinsula", "ophelia", "gp stakes", "kennedy lewis",
        "digital assets", "bcp asia", "valsabbina", "d-orbit",
        "escalator 1", "escalator 2",
    ])

    def _is_priv(name: str) -> bool:
        n = name.lower()
        return any(k in n for k in _PRIV_KW)

    # ── 4. Parse each scenario ────────────────────────────────────────────────
    result: dict = {}

    for sc_name, sect in sections.items():
        # ── 4a. Sub-category weights from pie-chart text ──────────────────────
        fc_m  = re.search(r"Fondi\s+consigliati", sect, re.IGNORECASE)
        pie   = sect[:fc_m.start()] if fc_m else sect
        sw: dict = {}
        for key, lbl_pat in _SUBCAT_LABELS:
            for m in re.finditer(lbl_pat, pie, re.IGNORECASE):
                # Search for "N%" within ±200 chars of the label match
                win  = pie[max(0, m.start() - 200): m.end() + 50]
                nums = re.findall(r'(\d{1,2})\s*%', win)
                for n in nums:
                    v = int(n)
                    if 1 <= v <= 50:
                        sw[key] = v
                        break
                if key in sw:
                    break

        # ── 4b. Parse "Fondi consigliati" section ─────────────────────────────
        if not fc_m:
            continue
        fc_txt = sect[fc_m.start():]

        cur_group:  str | None = None
        cur_subcat: str | None = None
        fund_subcat: dict = {}      # fund_name → subcat_key

        for line in fc_txt.split("\n"):
            l = line.strip()
            if not l:
                continue
            # — Group headers —
            if   re.match(r'^ALLOCATION$',    l, re.I): cur_group = "allocation";  cur_subcat = None
            elif re.match(r'^BOND$',          l, re.I): cur_group = "bond";         cur_subcat = None
            elif re.match(r'^EQUITY$',        l, re.I): cur_group = "equity";       cur_subcat = None
            elif re.match(r'^PRIVATE\s',      l, re.I): cur_group = "private";      cur_subcat = None
            # — Sub-category headers —
            elif re.match(r'^BALANCED$',      l, re.I): cur_subcat = "alloc_balanced"
            elif re.match(r'^FLEXIBLE$',      l, re.I): cur_subcat = "alloc_flexible"
            elif re.match(r'^AGGREGATE',      l, re.I): cur_subcat = "bond_aggregate"
            elif re.match(r'^THEMATIC$',      l, re.I):
                cur_subcat = "bond_thematic" if cur_group == "bond" else "equity_thematic"
            elif re.match(r'^TARGET',         l, re.I): cur_subcat = "bond_target"
            elif re.match(r'^PAESI\s+EMERGENTI$', l, re.I):
                cur_subcat = "bond_em" if cur_group == "bond" else "equity_em"
            elif re.match(r'^PAESI\s+SVILUPPATI$', l, re.I): cur_subcat = "equity_dev"
            elif re.match(r'^EMERGENTI$',     l, re.I):
                cur_subcat = "bond_em" if cur_group == "bond" else "equity_em"
            elif re.match(r'^SVILUPPATI$',    l, re.I): cur_subcat = "equity_dev"
            # — Fund lines (split on each "AZ Fund 1 -" to handle 2-column layout) —
            elif re.search(r'AZ\s+Fund\s+1\s*[-–]', l, re.I):
                for _part in re.split(r'(?=AZ\s+Fund\s+1\s*[-–])', l, flags=re.I):
                    _part = _part.strip()
                    _m = re.match(r'AZ\s+Fund\s+1\s*[-–]\s*(AZ\s+\S.+)', _part, re.I)
                    if _m and cur_subcat and cur_group != "private":
                        # Truncate at any additional "AZ Fund 1" still present
                        raw_nm = re.split(
                            r'\s+AZ\s+Fund\s+1\s*[-–]', _m.group(1), flags=re.I
                        )[0].strip().rstrip("*").strip()
                        if raw_nm and not _is_priv(raw_nm):
                            fund_subcat[raw_nm] = cur_subcat

        # ── 4c. Compute equal-weight per sub-category ─────────────────────────
        subcat_funds: dict = {}
        for fname, sc_key in fund_subcat.items():
            subcat_funds.setdefault(sc_key, []).append(fname)

        total_liq = sum(sw.get(k, 0) for k in subcat_funds)
        if total_liq == 0:
            total_liq = sum(sw.values()) or 70

        records: list = []
        for sc_key, funds in subcat_funds.items():
            w_sc = sw.get(sc_key, 0)
            if not funds:
                continue
            w_per = ((w_sc / len(funds)) / total_liq) if w_sc else (1.0 / max(len(fund_subcat), 1))
            grp = _SUBCAT_GROUP.get(sc_key, "ALLOCATION")
            cat = _SUBCAT_CAT.get(sc_key, "Altro")
            az  = DEFAULT_AZ.get(get_macro(cat), 0.5)
            for fname in funds:
                records.append({
                    "nome":     fname,
                    "gruppo":   grp,
                    "categoria": cat,
                    "az_pct":   az,
                    "obb_pct":  1.0 - az,
                    "weight":   w_per,
                    "subcat":   sc_key,
                })

        # ── 4d. Info string from summary paragraph ────────────────────────────
        info_m = re.search(
            r'azioni\s+(\d+)%.*?obbligazioni\s+(\d+)%.*?private\s+markets\s+(\d+)%',
            sect, re.IGNORECASE | re.DOTALL)
        info = (f"Equity {info_m.group(1)}% · Bond {info_m.group(2)}%"
                f" · Private Markets {info_m.group(3)}%") if info_m else ""

        result[sc_name] = {
            "info":           info,
            "funds":          records,
            "subcat_weights": sw,
        }

    return result if result else None


# ── Module-level badge helpers (used in suggerito_portfolio_ui) ──────────────
_FIDA_BG_GP  = {5:"#166534", 4:"#15803d", 3:"#16a34a", 2:"#64748B", 1:"#94A3B8"}
_MS_BG_GP    = {5:"#78350F", 4:"#92400E", 3:"#B45309"}
_MS_COL_GP   = {5:"#78350F", 4:"#92400E", 3:"#B45309", 2:"#475569", 1:"#94A3B8"}

def _fida_badge_gp(r) -> str:
    try:    v = int(r)
    except (TypeError, ValueError): return "<span style='color:#94A3B8;'>—</span>"
    bg = _FIDA_BG_GP.get(v)
    return (f"<span style='background:{bg};color:#fff;padding:2px 8px;"
            f"border-radius:4px;font-weight:700;font-size:.8rem;'>{v}</span>"
            if bg else f"<span style='color:#64748B;font-weight:700;font-size:.8rem;'>{v}</span>")

def _ms_badge_gp(ms_r) -> str:
    try:    v = int(ms_r)
    except (TypeError, ValueError): return "<span style='color:#94A3B8;'>—</span>"
    filled = "★" * v
    bg = _MS_BG_GP.get(v)
    if bg:
        return (f"<span style='background:{bg};color:#fff;padding:2px 8px;"
                f"border-radius:4px;font-weight:700;font-size:.8rem;'>{filled}</span>")
    _col = _MS_COL_GP.get(v, "#64748B")
    return (f"<span style='color:{_col};font-weight:700;"
            f"font-size:.8rem;'>{filled}</span>")

# Sub-category display names (Italian labels)
_SUBCAT_DISPLAY = {
    "alloc_balanced":  "Allocation – Balanced",
    "alloc_flexible":  "Allocation – Flexible",
    "bond_aggregate":  "Bond – Aggregate / Gov",
    "bond_thematic":   "Bond – Thematic",
    "bond_em":         "Bond – Paesi Emergenti",
    "bond_target":     "Bond – Target Maturity",
    "equity_thematic": "Equity – Thematic",
    "equity_dev":      "Equity – Paesi Sviluppati",
    "equity_em":       "Equity – Paesi Emergenti",
}


def suggerito_portfolio_ui(sc_name: str, gp_scenario: dict,
                           fund_data: dict, ms_data: dict):
    """Interactive portfolio builder for a SUGGERITO scenario.

    Shows macro-category headers with the scenario-suggested weight, then
    lists the recommended funds with FIDArating + Morningstar badges and a
    free peso-% input for each.  Returns a ready DataFrame when weights sum
    to 100 %, or None while the user is still editing.
    """
    funds = gp_scenario.get("funds", [])
    if not funds:
        return None

    sw = gp_scenario.get("subcat_weights", {})

    # Group funds by subcategory, preserving parse order
    subcat_funds: dict = {}
    for f in funds:
        subcat_funds.setdefault(f["subcat"], []).append(f)

    # Per-scenario session-state key so weights reset when switching scenarios
    ss_key = f"_sg_w_{sc_name}"
    if ss_key not in st.session_state:
        # Initialise with equal-weight defaults from the scenario
        st.session_state[ss_key] = {
            f["nome"]: round(f["weight"] * 100, 1) for f in funds
        }
    ww: dict = st.session_state[ss_key]

    # ── Page header ───────────────────────────────────────────────────────────
    st.markdown('<p class="sec-title">Costruisci il Portafoglio Suggerito</p>',
                unsafe_allow_html=True)
    st.caption(
        "I pesi mostrati sono distribuiti equamente all'interno di ogni "
        "sottocategoria.  Modifica liberamente i valori e l'analisi si "
        "aggiorna automaticamente quando la somma raggiunge 100 %."
    )

    # ── Column headers (only once, above all subcategories) ───────────────────
    _h1, _h2, _h3, _h4 = st.columns([4.5, 1.2, 1.2, 1.4])
    _h1.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>Fondo</span>",
                 unsafe_allow_html=True)
    _h2.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>FIDArating</span>",
                 unsafe_allow_html=True)
    _h3.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>Morningstar</span>",
                 unsafe_allow_html=True)
    _h4.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>Peso %</span>",
                 unsafe_allow_html=True)
    st.markdown("<hr style='margin:.15rem 0 .3rem 0;border-color:#e2e8f0;'>",
                unsafe_allow_html=True)

    # ── Per-subcategory sections ──────────────────────────────────────────────
    for sc_key, sc_funds in subcat_funds.items():
        w_sc   = sw.get(sc_key, 0)
        sc_lbl = _SUBCAT_DISPLAY.get(sc_key, sc_key)

        # — Subcategory header bar —
        st.markdown(
            f"<div style='background:linear-gradient(90deg,#0D1B2A,#162e52);"
            f"color:#fff;padding:.45rem 1rem;border-radius:6px;margin-top:.7rem;"
            f"display:flex;align-items:center;gap:.8rem;'>"
            f"<span style='font-weight:700;font-size:.88rem;flex:1;'>{sc_lbl}</span>"
            f"<span style='background:#C9A84C;color:#0D1B2A;padding:2px 9px;"
            f"border-radius:4px;font-size:.73rem;font-weight:700;white-space:nowrap;'>"
            f"Peso suggerito: {w_sc}%</span></div>",
            unsafe_allow_html=True)

        # — Fund rows —
        for f in sc_funds:
            fname    = f["nome"]
            resolved = _resolve_nome_for_fd(fname, fund_data)

            # Ratings from cache
            fd_ov  = (fund_data or {}).get(resolved, {}).get("overview", {})
            fida_r = str(fd_ov.get("fida_rating") or "").strip() or "—"
            ms_r   = (ms_data or {}).get(resolved, {}).get("ms_rating")

            # Display name: strip "AZ [Family] - " prefix
            short = re.sub(r'^AZ\s+(?:Allocation|Bond|Equity)\s*[-–]\s*',
                           '', fname, flags=re.I).strip()

            c1, c2, c3, c4 = st.columns([4.5, 1.2, 1.2, 1.4])
            with c1:
                st.markdown(
                    f"<div style='font-size:.84rem;font-weight:500;color:#1e293b;"
                    f"padding:.55rem 0 .3rem 0;'>{short}</div>",
                    unsafe_allow_html=True)
            with c2:
                st.markdown(
                    f"<div style='padding:.5rem 0 .25rem 0;'>"
                    f"{_fida_badge_gp(fida_r)}</div>",
                    unsafe_allow_html=True)
            with c3:
                st.markdown(
                    f"<div style='padding:.5rem 0 .25rem 0;'>"
                    f"{_ms_badge_gp(ms_r)}</div>",
                    unsafe_allow_html=True)
            with c4:
                default_w = float(ww.get(fname, round(f["weight"] * 100, 1)))
                new_w = st.number_input(
                    "w", min_value=0.0, max_value=100.0,
                    value=default_w, step=0.5,
                    key=f"sg_{sc_name}_{fname[:35]}",
                    label_visibility="collapsed",
                )
                ww[fname] = new_w

        st.markdown("<hr style='margin:.25rem 0 0 0;border-color:#f1f5f9;'>",
                    unsafe_allow_html=True)

    # ── Total weight indicator ────────────────────────────────────────────────
    total_w = sum(ww.get(f["nome"], 0.0) for f in funds)
    diff    = abs(total_w - 100.0)
    st.markdown("<br>", unsafe_allow_html=True)
    if diff < 0.15:
        st.markdown(
            f'<div class="w-ok">✅ Somma pesi: <b>{total_w:.1f}%</b>'
            f' — Portafoglio pronto!</div>', unsafe_allow_html=True)
    else:
        left = 100.0 - total_w
        st.markdown(
            f'<div class="w-warn">⚠️ Somma pesi: <b>{total_w:.1f}%</b>'
            f' ({"mancano" if left>0 else "eccedono"} {abs(left):.1f}%)</div>',
            unsafe_allow_html=True)

    if diff > 1.0:
        return None   # analysis only when weights are balanced

    # ── Build DataFrame ───────────────────────────────────────────────────────
    records: list = []
    for f in funds:
        peso = ww.get(f["nome"], 0.0)
        if peso <= 0:
            continue
        nome = _resolve_nome_for_fd(f["nome"], fund_data)
        records.append({
            "nome":      nome,
            "categoria": f["categoria"],
            "gruppo":    f["gruppo"],
            "macro_cat": get_macro(f["categoria"]),
            "az_pct":    f["az_pct"],
            "obb_pct":   f["obb_pct"],
            "r_weight":  peso / 100.0,
            "w_cons":    peso / 100.0,
            "w_equil":   peso / 100.0,
            "w_accr":    peso / 100.0,
        })
    if not records:
        return None
    df = pd.DataFrame(records)
    for wc in ("w_cons", "w_equil", "w_accr"):
        t = df[wc].sum()
        if t > 0:
            df[wc] /= t
    df["macro_cat"] = df["categoria"].apply(get_macro)
    df = assign_colors(df)
    st.markdown("<br>", unsafe_allow_html=True)
    return df


# ════════════════════════════════════════════════════════════
# MAIN APP
# ════════════════════════════════════════════════════════════

_APP_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@600;700&family=DM+Sans:opsz,wght@9..40,300;9..40,400;9..40,500;9..40,600&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
h1,h2,h3{font-family:'Cormorant Garamond',serif !important;}
[data-testid="stSidebar"]{background:linear-gradient(170deg,#06101e 0%,#0d1f3c 55%,#0a1628 100%);border-right:1px solid #1a3050;}
[data-testid="stSidebar"] .stFileUploader label,[data-testid="stSidebar"] .stRadio > label,[data-testid="stSidebar"] .stSelectbox > label{color:#4a6582 !important;font-size:.68rem !important;letter-spacing:.12em !important;text-transform:uppercase !important;font-weight:600 !important;}
[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p{color:#c0cfe0 !important;font-size:.9rem !important;}
[data-testid="stSidebar"] .stSelectbox>div>div{background:#132035 !important;border:1px solid #243d5a !important;color:#dde6f0 !important;border-radius:6px !important;}
[data-testid="stSidebar"] .stSelectbox svg{fill:#C9A84C !important;width:22px !important;height:22px !important;opacity:1 !important;}
[data-testid="stSidebar"] .stFileUploader>div{background:#132035 !important;border:1px dashed #2a4a6a !important;border-radius:8px !important;}
[data-testid="stSidebar"] .stFileUploader p,[data-testid="stSidebar"] .stFileUploader span{color:#8aa5c0 !important;font-size:.8rem !important;}
[data-testid="stSidebar"] ::-webkit-scrollbar{width:6px;}
[data-testid="stSidebar"] ::-webkit-scrollbar-track{background:#06101e;}
[data-testid="stSidebar"] ::-webkit-scrollbar-thumb{background:#C9A84C;border-radius:3px;}
[data-testid="stSidebar"] ::-webkit-scrollbar-thumb:hover{background:#d4b87a;}
.main{background:#f6f8fb !important;}.block-container{padding-top:1.8rem !important;max-width:1300px;}
.az-header{background:linear-gradient(130deg,#081420 0%,#0f2644 50%,#162e52 100%);border-radius:16px;padding:2rem 2.5rem;margin-bottom:1.8rem;position:relative;overflow:hidden;box-shadow:0 8px 32px rgba(0,0,0,.15);}
.az-header::after{content:'';position:absolute;bottom:-60px;right:-40px;width:220px;height:220px;border-radius:50%;background:radial-gradient(circle,rgba(201,168,76,.18) 0%,transparent 70%);}
.az-eyebrow{font-size:.65rem;letter-spacing:.2em;color:#4a7098;text-transform:uppercase;font-weight:600;}
.az-title{font-family:'Cormorant Garamond',serif;font-size:2.1rem;font-weight:700;color:#f0f6ff;margin:.2rem 0 .4rem;line-height:1.1;}
.az-rule{width:38px;height:3px;background:#C9A84C;border-radius:2px;margin:.6rem 0;}
.az-meta{font-size:.88rem;color:#6b8fb0;}
.kpi{background:#fff;border:1px solid #e4eaf3;border-radius:12px;padding:1.2rem 1.4rem;box-shadow:0 1px 4px rgba(0,0,0,.05);}
.kpi-label{font-size:.65rem;text-transform:uppercase;letter-spacing:.1em;color:#94a3b8;font-weight:500;margin-bottom:.3rem;}
.kpi-value{font-size:1.9rem;font-weight:700;color:#0d1b2a;font-family:'Cormorant Garamond',serif;line-height:1;}
.kpi-sub{font-size:.75rem;color:#64748b;margin-top:.3rem;}
.sec-title{font-family:'Cormorant Garamond',serif;font-size:1.25rem;font-weight:600;color:#0d1b2a;border-bottom:2px solid #c9a84c;display:inline-block;padding-bottom:.4rem;margin-bottom:.9rem;}
.fund-group-hdr{background:#f0f4f9;padding:.45rem 1rem;font-size:.65rem;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:#64748b;border-bottom:1px solid #e2e8f0;}
.fund-row{display:flex;align-items:center;gap:10px;padding:.65rem 1rem;border-bottom:1px solid #f1f5f9;}
.fund-row:last-child{border-bottom:none;}
.fund-dot{width:8px;height:34px;border-radius:3px;flex-shrink:0;}
.fund-name{font-size:.83rem;color:#1e293b;font-weight:500;flex:1;min-width:0;}
.fund-cat{font-size:.68rem;color:#64748b;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.fund-pct{font-size:1rem;font-weight:700;color:#0d1b2a;min-width:2.8rem;text-align:right;}
[data-testid="stDownloadButton"]>button{background:linear-gradient(135deg,#0f2d6b 0%,#1b4fbb 100%) !important;color:#fff !important;font-size:1.05rem !important;font-weight:600 !important;padding:.9rem 2rem !important;border-radius:10px !important;border:none !important;width:100% !important;letter-spacing:.02em !important;box-shadow:0 4px 18px rgba(27,79,187,.35) !important;}
[data-testid="stDownloadButton"]>button:hover{box-shadow:0 6px 24px rgba(27,79,187,.55) !important;transform:translateY(-2px) !important;}
.w-ok{background:#d1fae5;border:1px solid #6ee7b7;border-radius:8px;padding:.7rem 1rem;font-size:.84rem;color:#065f46;}
.w-warn{background:#fef3c7;border:1px solid #fcd34d;border-radius:8px;padding:.7rem 1rem;font-size:.84rem;color:#92400e;}
</style>
"""


def main():
    st.markdown(_APP_CSS, unsafe_allow_html=True)
    _ms_with_rating = 0   # default; updated inside sidebar block below
    with st.sidebar:
        st.markdown("""<div style='padding:1.2rem 0 .4rem 0;'><div style='font-size:.6rem;letter-spacing:.22em;color:#3a5a78;text-transform:uppercase;font-weight:700;'>Analisi Portafoglio</div><div style='font-family:"Cormorant Garamond",serif;font-size:1.3rem;color:#dde8f5;font-weight:700;margin-top:4px;line-height:1.3;'>AAS Emilia<br>Romagna<br>Marche Umbria</div><div style='width:32px;height:3px;background:#C9A84C;border-radius:2px;margin-top:8px;'></div><div style='font-size:.6rem;color:#2a4a6a;margin-top:5px;'>v2.2 — factbook Excel cache</div></div>""", unsafe_allow_html=True)
        st.markdown("<hr style='margin:.4rem 0 .5rem 0;border-color:#1a3050;'>", unsafe_allow_html=True)
        uploaded   = st.file_uploader("FILE EXCEL (PTF FULL + PTF SHORT + FIDA)", type=["xlsx","xls"])
        uploaded_fb = st.file_uploader(
            "FACTBOOK PDF (prima estrazione)",
            type=["pdf"],
            help="Carica il Factbook PDF per estrarre Duration, Rating e Asset "
                 "Allocation. Dopo la prima estrazione scarica il file Excel "
                 "e ricaricalo la prossima volta: è più veloce.",
        )
        uploaded_fb_xl = st.file_uploader(
            "DATI FACTBOOK (Excel, dopo prima estrazione)",
            type=["xlsx","xls"],
            help="Carica il file Excel scaricato dopo la prima estrazione del "
                 "Factbook PDF. Evita di ricaricare il PDF ogni volta.",
        )
        # ── FondiDoc + Morningstar — unico tasto ─────────────────────────────
        st.markdown("---")
        _fd_now = st.session_state.get("_scomp_fd") or load_fund_cache()[0]
        _ms_now = st.session_state.get("_ms_data") or load_ms_cache()
        _ms_with_rating = sum(1 for v in _ms_now.values() if v.get("ms_rating"))

        # Card stato dati
        _fd_line = (f"✅ <b>FondiDoc</b> — {len(_fd_now)} fondi"
                    if _fd_now else "⚠️ <b>FondiDoc</b> — non ancora scaricato")
        _ms_line = (f"⭐ <b>Morningstar</b> — {_ms_with_rating} rating"
                    if _ms_with_rating else "⚠️ <b>Morningstar</b> — non ancora scaricato")
        _card_bg  = "#0d2b1a" if (_fd_now and _ms_with_rating) else "#1a1a08"
        _card_brd = "#166534" if (_fd_now and _ms_with_rating) else "#854d0e"
        _card_clr = "#86efac" if (_fd_now and _ms_with_rating) else "#fde68a"
        st.markdown(
            f"<div style='background:{_card_bg};border:1px solid {_card_brd};"
            f"border-radius:8px;padding:.5rem .85rem;font-size:.73rem;"
            f"color:{_card_clr};margin-bottom:.5rem;line-height:1.8;'>"
            f"{_fd_line}<br>{_ms_line}</div>",
            unsafe_allow_html=True)

        if uploaded:
            if st.button("📥  Aggiorna Dati (FondiDoc + Morningstar)",
                         use_container_width=True,
                         help="Scarica in sequenza: Cat. FIDA, FIDArating, "
                              "rendimenti e metriche di rischio da FondiDoc, "
                              "poi i rating Morningstar da FondiOnline."):
                st.session_state["_fetch_fd_requested"] = True
                st.session_state["_fetch_ms_requested"] = True
        else:
            st.caption("⬆️ Carica prima il file Excel")

        # ── Global Perspectives PDF ──────────────────────────────────────────────
        st.markdown("---")
        uploaded_gp = st.file_uploader(
            "GLOBAL PERSPECTIVES PDF",
            type=["pdf"],
            help="Carica il PDF Global Perspectives trimestrale per attivare "
                 "la modalità SUGGERITO con i 3 scenari (Base / Bear / Bull). "
                 "Il file viene re-parsato solo quando cambia.",
        )
        if uploaded_gp is not None:
            # Re-parse only when a new file is uploaded (name change = new edition)
            if st.session_state.get("_gp_filename") != uploaded_gp.name:
                with st.spinner("📄 Parsing Global Perspectives…"):
                    _gp_parsed = parse_global_perspectives(uploaded_gp.read())
                if _gp_parsed:
                    st.session_state["_gp_data"]     = _gp_parsed
                    st.session_state["_gp_filename"]  = uploaded_gp.name
                    _n_gp = sum(len(v["funds"]) for v in _gp_parsed.values())
                    st.success(
                        f"✅ GP — {_n_gp} fondi · "
                        f"{', '.join(_gp_parsed.keys())}")
                else:
                    st.session_state.pop("_gp_data", None)
                    st.warning("⚠️ PDF non riconosciuto — verifica che sia un "
                               "Global Perspectives Azimut.")
            elif st.session_state.get("_gp_data"):
                _gp_ok = st.session_state["_gp_data"]
                _n_gp  = sum(len(v["funds"]) for v in _gp_ok.values())
                st.markdown(
                    f"<div style='background:#0d2b1a;border:1px solid #166534;"
                    f"border-radius:8px;padding:.5rem .85rem;font-size:.73rem;"
                    f"color:#86efac;margin-bottom:.5rem;line-height:1.5;'>"
                    f"✅ <b>Global Perspectives</b> — {_n_gp} fondi "
                    f"({', '.join(_gp_ok.keys())})</div>",
                    unsafe_allow_html=True)
        else:
            # If file removed, clear cached data
            if st.session_state.get("_gp_filename"):
                st.session_state.pop("_gp_data",     None)
                st.session_state.pop("_gp_filename",  None)

        st.markdown("---")
        _gp_loaded    = bool(st.session_state.get("_gp_data"))
        _ptf_options  = ["📋  PTF FULL", "⚡  PTF SHORT", "🎨  LIBERO"]
        if _gp_loaded:
            _ptf_options.append("🌐  SUGGERITO")
        ptf_choice = st.radio("TIPO PORTAFOGLIO", _ptf_options)
        st.markdown("---")
        profile    = st.selectbox("PROFILO DI RISCHIO", PROFILES, index=0)
        if "LIBERO" not in ptf_choice and "free_ptf" in st.session_state:
            del st.session_state["free_ptf"]
        # ── Scenario sub-selector (SUGGERITO only) ────────────────────────────
        if "SUGGERITO" in ptf_choice:
            _gp_keys = list(st.session_state.get("_gp_data", {}).keys())
            _SC_LABELS = {
                "Base": "⚖️  Scenario Base",
                "Bear": "🐻  Scenario Bear",
                "Bull": "🐂  Scenario Bull",
            }
            _sc_opts = [_SC_LABELS.get(k, k) for k in _gp_keys]
            if _sc_opts:
                _sc_sel = st.radio("SCENARIO", _sc_opts, key="_gp_sc_radio")
                st.session_state["_gp_sc_key"] = next(
                    (k for k, v in _SC_LABELS.items() if v == _sc_sel),
                    _gp_keys[0])
                # Show scenario info
                _sc_info = st.session_state["_gp_data"].get(
                    st.session_state["_gp_sc_key"], {}).get("info", "")
                if _sc_info:
                    st.caption(f"📊 {_sc_info}")

    ptf_label = ptf_choice.split("  ",1)[1] if "  " in ptf_choice else ptf_choice
    _is_suggerito = "SUGGERITO" in ptf_choice
    if _is_suggerito:
        _sc_key_hdr = st.session_state.get("_gp_sc_key", "Base")
        ptf_label   = f"SUGGERITO — Scenario {_sc_key_hdr}"

    # ── Invalidate cached PDF when portfolio type or profile changes ──────────
    _ptf_key = f"{ptf_choice}|{profile}"
    if st.session_state.get("_last_ptf_key") != _ptf_key:
        for _k in ("_pdf_bytes_ready", "_pdf_fname_ready", "_pdf_lbl"):
            st.session_state.pop(_k, None)
        st.session_state["_last_ptf_key"] = _ptf_key

    st.markdown(f"""<div class="az-header"><div class="az-eyebrow">AZIMUT INVESTMENTS · AAS EMILIA ROMAGNA MARCHE UMBRIA</div><div class="az-rule"></div><div class="az-title">{ptf_label}</div><div class="az-meta">{PROFILE_ICONS.get(profile,'●')} Profilo {profile.title()} &nbsp;·&nbsp; {datetime.date.today().strftime('%d %B %Y')}</div></div>""",unsafe_allow_html=True)

    if uploaded is None and not _is_suggerito:
        st.info("⬅️ **Carica il file Excel** nella barra laterale per iniziare.")
        return

    if uploaded is not None:
        with st.spinner("⏳ Caricamento dati…"):
            file_bytes = uploaded.read()
            raw = parse_excel(file_bytes)
    else:
        raw = {}

    # ── Sidebar-triggered FondiDoc / MS fetch (only when Excel is loaded) ───────
    if uploaded is not None:
        if st.session_state.pop("_fetch_fd_requested", False):
            _fida_urls_all = raw.get("fida_urls", {})
            _sheets = [raw[s] for s in ("PTF FULL", "PTF SHORT")
                       if s in raw and not raw[s].empty]
            _df_all = (pd.concat(_sheets, ignore_index=True)
                       .drop_duplicates(subset=["nome"]) if _sheets else pd.DataFrame())
            if not _df_all.empty:
                _pb_fd = st.progress(0, text="Scarico dati FondiDoc…")
                def _upd_fd(v): _pb_fd.progress(v, text=f"FondiDoc: {int(v*100)}%…")
                _fd_new = fetch_all_fund_data(_df_all, _fida_urls_all, _upd_fd)
                _pb_fd.empty()
                save_fund_cache(_fd_new)
                st.session_state["_scomp_fd"] = _fd_new
                st.rerun()
            else:
                st.warning("⚠️ Nessun fondo trovato — verifica il file Excel.")

        if st.session_state.pop("_fetch_ms_requested", False):
            _fida_df   = raw.get("FIDA", pd.DataFrame())
            _sheets_ms = [raw[s] for s in ("PTF FULL", "PTF SHORT")
                          if s in raw and not raw[s].empty]
            _df_ms = (pd.concat(_sheets_ms, ignore_index=True)
                      .drop_duplicates(subset=["nome"]) if _sheets_ms else pd.DataFrame())
            if not _df_ms.empty:
                with st.spinner("⭐ Scarico rating Morningstar da FondiOnline…"):
                    _ms_new = fetch_all_ms_ratings(_df_ms, _fida_df)
                save_ms_cache(_ms_new)
                st.session_state["_ms_data"] = _ms_new
                _n_found = sum(1 for v in _ms_new.values() if v.get("ms_rating"))
                st.success(f"⭐ Morningstar: {_n_found}/{len(_ms_new)} rating trovati")
                st.rerun()
            else:
                st.warning("⚠️ Nessun fondo trovato — verifica il file Excel.")
    else:
        # Drain any stale fetch flags so they don't fire when Excel is later loaded
        st.session_state.pop("_fetch_fd_requested", None)
        st.session_state.pop("_fetch_ms_requested", None)

    # ── Factbook data ──────────────────────────────────────────────────────────
    # Priority:
    #   1. Excel uploaded manually (override / fix)
    #   2. PDF uploaded (first-time extraction → auto-save to repo)
    #   3. Auto-load from data/factbook_dati.json committed in the repo
    factbook_data: dict = load_factbook_auto()
    _fb_source = f"repository ({len(factbook_data)} fondi)" if factbook_data else ""

    if uploaded_fb is not None:
        # First-time (or refresh): parse PDF, auto-save, offer Excel
        with st.spinner("📖 Estraggo dati dal Factbook PDF…"):
            _new = parse_factbook(uploaded_fb.read())
        if _new:
            factbook_data = _new
            _fb_source = f"PDF ({len(_new)} fondi)"
            st.success(f"✅ Factbook estratto — {len(_new)} fondi trovati")
            # Auto-save to GitHub repo (needs GITHUB_TOKEN secret)
            with st.spinner("💾 Salvo dati nel repository…"):
                _saved = save_factbook_to_repo(_new)
            if _saved:
                st.info(
                    "🔄 Dati salvati nel repository. "
                    "Al prossimo accesso non servirà ricaricare il PDF "
                    "(l'app si riavvierà automaticamente entro ~1 min).")
            else:
                # Fallback: offer Excel download so user can re-use data
                xl_bytes = factbook_to_excel_bytes(_new)
                st.download_button(
                    label="💾  Scarica dati Factbook (Excel)",
                    data=xl_bytes,
                    file_name="factbook_dati.xlsx",
                    mime="application/vnd.openxmlformats-officedocument"
                         ".spreadsheetml.sheet",
                    help="Il salvataggio automatico non è riuscito (token "
                         "mancante?). Scarica questo Excel e caricalo "
                         "nella casella 'DATI FACTBOOK (Excel)' oppure "
                         "aggiorna il secret GITHUB_TOKEN.",
                )
        else:
            st.warning("⚠️ Factbook PDF caricato ma nessun dato estratto — "
                       "verrà usato FondiDoc")

    if uploaded_fb_xl is not None:
        # Manual override: user uploaded a corrected Excel
        _xl = factbook_from_excel(uploaded_fb_xl.read())
        if _xl:
            factbook_data = _xl
            _fb_source = f"Excel ({len(_xl)} fondi)"
            st.success(f"✅ Dati Factbook caricati da Excel — {len(_xl)} fondi")
        else:
            st.warning("⚠️ Excel Factbook vuoto — uso dati precedenti")

    if _is_suggerito:
        _gp_data_main = st.session_state.get("_gp_data", {})
        _sc_key_main  = st.session_state.get("_gp_sc_key", "Base")
        _sc_data_main = _gp_data_main.get(_sc_key_main)
        if not _sc_data_main:
            st.warning("📄 Carica il PDF **Global Perspectives** nella barra "
                       "laterale per vedere i portafogli suggeriti.")
            return
        _fd_for_gp = st.session_state.get("_scomp_fd") or load_fund_cache()[0]
        _ms_for_gp = st.session_state.get("_ms_data") or load_ms_cache()
        df = suggerito_portfolio_ui(_sc_key_main, _sc_data_main,
                                    _fd_for_gp, _ms_for_gp)
        if df is None or df.empty:
            return  # weights not balanced yet — builder is shown, analysis waits
    elif "LIBERO" in ptf_choice:
        df = free_portfolio_ui(raw)
    else:
        key = "PTF FULL" if "FULL" in ptf_choice else "PTF SHORT"
        if key not in raw or raw[key].empty:
            st.error(f"❌ Foglio '{key}' non trovato o vuoto."); return
        df = raw[key]

    if df is None or df.empty: return

    wcol   = PROFILE_W_COL[profile]
    df_act = df[df[wcol]>0.001].copy()

    # KPI row
    n_fondi = len(df_act)
    w_az    = (df_act[wcol]*df_act["az_pct"]).sum()*100
    w_obb   = (df_act[wcol]*df_act["obb_pct"]).sum()*100
    srri    = max(1,min(7,round(w_az/100*6+1)))

    _SRRI_LABELS = {
        1: "Rischio Molto Basso",
        2: "Rischio Basso",
        3: "Rischio Medio-Basso",
        4: "Rischio Medio",
        5: "Rischio Medio-Alto",
        6: "Rischio Alto",
        7: "Rischio Molto Alto",
    }
    _srri_sub = (
        f"<span style='font-weight:600;color:#0d1b2a;'>"
        f"{_SRRI_LABELS.get(srri,'—')}</span><br>"
        f"<span style='font-size:.68rem;line-height:1.5;'>"
        f"Indicatore sintetico europeo di rischio/rendimento<br>"
        f"Scala 1 (min) → 7 (max) · stima da quota azionaria</span>"
    )

    c1,c2,c3,c4 = st.columns(4)
    for col,val,lbl,sub in [
        (c1,str(n_fondi),"Fondi in Portafoglio",f"{df_act['gruppo'].nunique()} gruppi"),
        (c2,f"{w_az:.1f}%","Quota Azionaria","ponderata per peso"),
        (c3,f"{w_obb:.1f}%","Quota Obbligazionaria","ponderata per peso"),
        (c4,f"{srri} / 7","Risk Score (SRRI proxy)", _srri_sub),
    ]:
        col.markdown(f'<div class="kpi"><div class="kpi-label">{lbl}</div><div class="kpi-value">{val}</div><div class="kpi-sub">{sub}</div></div>',unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)

    # Charts + fund list
    col_l,col_r = st.columns([1.15,0.85],gap="large")
    with col_l:
        st.markdown('<p class="sec-title">Allocazione per Fondo</p>',unsafe_allow_html=True)
        st.plotly_chart(make_fund_pie(df_act,wcol,profile),use_container_width=True,config={"displayModeBar":False})
        st.markdown('<p class="sec-title">Allocazione per Macro-Categoria</p>',unsafe_allow_html=True)
        st.plotly_chart(make_macro_bar(df_act,wcol),use_container_width=True,config={"displayModeBar":False})
    with col_r:
        st.markdown('<p class="sec-title">Composizione del Portafoglio</p>',unsafe_allow_html=True)
        _gruppi = list(df_act["gruppo"].unique())
        # Icon map for known group names
        _GRP_ICON = {
            "ALLOCATION":     "🔀",
            "AZIONARI (LONG)":"📈",
            "BOND":           "🏛️",
        }
        _grp_tabs = st.tabs([
            f"{_GRP_ICON.get(g, '📁')}  {g}" for g in _gruppi
        ])
        for _gtab, gruppo in zip(_grp_tabs, _gruppi):
            with _gtab:
                sub = df_act[df_act["gruppo"]==gruppo].sort_values(wcol,ascending=False)
                rows_html = "".join([
                    f'<div class="fund-row">'
                    f'<div class="fund-dot" style="background:{r["color"]};"></div>'
                    f'<div style="flex:1;min-width:0;">'
                    f'<div class="fund-name">{r["nome"]}</div>'
                    f'<div class="fund-cat">'
                    f'{r["categoria"][:48]+"…" if r["categoria"] and len(r["categoria"])>48 else (r["categoria"] or "—")}'
                    f'</div></div>'
                    f'<div class="fund-pct">{r[wcol]*100:.1f}%</div>'
                    f'</div>'
                    for _, r in sub.iterrows()
                ])
                st.markdown(
                    f'<div style="background:#fff;border:1px solid #e2e8f0;'
                    f'border-radius:10px;overflow:hidden;">{rows_html}</div>',
                    unsafe_allow_html=True)

    # ── Load cached FondiDoc data (bundled in repo) ──────────────────────────
    cached_fd, cache_date = load_fund_cache()

    # ── TABBED ANALYTICS TABLES ──────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<p class="sec-title">Analisi del Portafoglio</p>',
                unsafe_allow_html=True)

    # ── Shared helpers for all four tabs ──────────────────────────────────────
    def _fb_metric(nome: str, key: str):
        """Return metric from factbook_data (duration / credit_rating / az_pct …)."""
        if not factbook_data:
            return None
        norm = _normalize_for_unp(nome)
        norm = _FUND_ALIASES.get(norm, norm)
        entry = factbook_data.get(norm)
        if not entry or not isinstance(entry, dict):
            best, best_len = None, 0
            for _fk, _fv in factbook_data.items():
                if (isinstance(_fv, dict)
                        and (_fk in norm or norm in _fk)
                        and len(_fk) > best_len):
                    best, best_len = _fv, len(_fk)
            entry = best
        if not entry or not isinstance(entry, dict):
            return None
        return entry.get(key)

    # Prefer live FondiDoc data fetched in this session over on-disk cache
    _fd_live = st.session_state.get("_scomp_fd") or cached_fd

    # Morningstar ratings — from this session or from disk cache
    _ms_live = st.session_state.get("_ms_data") or load_ms_cache()

    # Morningstar color scale (amber/gold palette)
    _MS_COL = {5: "#78350F", 4: "#92400E", 3: "#B45309", 2: "#475569", 1: "#94A3B8"}
    _MS_BG  = {5: "#78350F", 4: "#92400E", 3: "#B45309"}   # bg only for top-3

    def _ms_badge_html(ms_r) -> str:
        """Return HTML span for a Morningstar rating integer.

        Only filled stars are shown (no empty stars): ☆ on a coloured background
        is visually indistinguishable from ★ in white text, which caused ratings
        like 3★ to appear as 5★.  Showing only the earned stars is unambiguous.
        """
        try:
            v = int(ms_r)
        except (TypeError, ValueError):
            return "<span style='color:#94A3B8;'>—</span>"
        filled = "★" * v          # e.g. "★★★" for 3 — no empty stars
        bg = _MS_BG.get(v)
        if bg:
            return (f"<span style='background:{bg};color:#fff;padding:2px 8px;"
                    f"border-radius:4px;font-weight:700;font-size:.8rem;'>"
                    f"{filled}</span>")
        col = _MS_COL.get(v, "#64748B")
        return (f"<span style='color:{col};font-weight:700;font-size:.8rem;'>"
                f"{filled}</span>")

    # Shared HTML style tokens
    _TH  = ("background:#0D1B2A;color:#fff;font-size:.74rem;"
            "padding:8px 10px;white-space:nowrap;")
    _TC  = "font-size:.77rem;padding:6px 10px;border-bottom:1px solid #f1f5f9;"
    _TP  = ("background:#1B4332;color:#fff;font-size:.77rem;font-weight:700;"
            "padding:6px 10px;border-bottom:2px solid #C9A84C;")

    # FIDArating color scale: 5=best (dark green) … 1=worst (dark red)
    _FIDA_COL = {5: "#166534", 4: "#15803d", 3: "#22c55e",
                 2: "#f87171", 1: "#b91c1c"}
    _FIDA_BG  = {5: "#166534", 4: "#15803d", 3: "#22c55e"}

    # Source note labels (reused in all tab footers)
    _note_fb  = ("Factbook AZ Investments" if factbook_data
                 else "n.d. — carica il Factbook PDF nella barra laterale")
    if st.session_state.get("_scomp_fd"):
        _note_fd = "FondiDoc live (questa sessione)"
    elif cached_fd:
        _note_fd = f"FondiDoc (cache {cache_date})"
    else:
        _note_fd = "n.d. — clicca «Genera PDF» per popolare i dati FondiDoc"
    _note_style = "font-size:.71rem;color:#94A3B8;margin-top:5px;"

    # Pre-sort funds by weight (descending) — shared across all tabs
    _df_sorted = df_act.sort_values(wcol, ascending=False)

    def _perf_val_col(raw) -> str:
        """Wrap a performance string in green/red HTML span."""
        s = str(raw) if raw is not None else "N/D"
        try:
            v   = float(s.replace("%", "").replace(",", ".").strip())
            col = "#1A7A4A" if v > 0 else ("#C0392B" if v < 0 else "#475569")
            return f"<span style='color:{col};font-weight:700;'>{s}</span>"
        except Exception:
            return f"<span style='color:#94A3B8;'>{s}</span>"

    def _perf_wavg(keys: list) -> dict:
        """Weighted average of performance/risk metrics across active funds."""
        totals = {k: 0.0 for k in keys}
        cov_w  = {k: 0.0 for k in keys}
        for _, _row in df_act.iterrows():
            _w   = _row[wcol]
            _ana = _fd_live.get(_row["nome"], {}).get("analysis", {})
            for k in keys:
                raw = _fb_metric(_row["nome"], k) or _ana.get(k, "")
                try:
                    num = float(str(raw).replace("%", "").replace(",", ".").strip())
                    totals[k] += num * _w
                    cov_w[k]  += _w
                except Exception:
                    pass
        return {k: (f"{totals[k]/cov_w[k]:+.2f}%" if cov_w[k] > 0.01 else "N/D")
                for k in keys}

    def _html_table(hdr_cols: list, ptf_row: list, fund_rows: list) -> str:
        """Render a styled HTML table with header, portfolio summary row, fund rows."""
        def _align(i):
            return "left" if i == 0 else "center"
        hdr_html = "".join(
            f"<th style='{_TH}text-align:{_align(i)};'>{h}</th>"
            for i, h in enumerate(hdr_cols)
        )
        ptf_html = "".join(
            f"<td style='{_TP}text-align:{_align(i)};'>{v}</td>"
            for i, v in enumerate(ptf_row)
        )
        body_html = f"<tr>{ptf_html}</tr>"
        for fr in fund_rows:
            row_html = "".join(
                f"<td style='{_TC}text-align:{_align(i)};'>{v}</td>"
                for i, v in enumerate(fr)
            )
            body_html += f"<tr>{row_html}</tr>"
        return (
            f"<div style='overflow-x:auto;border-radius:10px;"
            f"border:1px solid #e2e8f0;background:#fff;'>"
            f"<table style='width:100%;border-collapse:collapse;'>"
            f"<thead><tr>{hdr_html}</tr></thead>"
            f"<tbody>{body_html}</tbody>"
            f"</table></div>"
        )

    _ptf_row_label = f"◆ PORTAFOGLIO {ptf_label.upper()}"

    # URL lookup: Excel hyperlinks first, FondiDoc cache as enriched fallback
    _fida_urls_raw = raw.get("fida_urls", {})

    def _fund_url(nome: str) -> str:
        """Return the FondiDoc URL for a fund, or '' if not available."""
        return (_fd_live.get(nome, {}).get("url", "")
                or _fida_urls_raw.get(nome, ""))

    def _fund_link(nome: str) -> str:
        """Return fund name as HTML — hyperlinked if URL is available."""
        url = _fund_url(nome)
        if url:
            return (f'<a href="{url}" target="_blank" rel="noopener noreferrer" '
                    f'style="color:#1B4FBB;text-decoration:underline;'
                    f'text-underline-offset:2px;">{nome}</a>')
        return nome

    tab1, tab2, tab3, tab4 = st.tabs([
        "📊  Scomposizione Az/Obb",
        "📈  Rendimenti",
        "⚠️  Rischio",
        "💰  UNP / IUNP",
    ])

    # ── TAB 1 — SCOMPOSIZIONE ────────────────────────────────────────────────
    with tab1:
        _scomp_hdr = (
            f"<tr>"
            f"<th style='{_TH}text-align:left;'>Fondo</th>"
            f"<th style='{_TH}text-align:center;'>Peso</th>"
            f"<th style='{_TH}text-align:center;'>% Az.</th>"
            f"<th style='{_TH}text-align:center;'>% Obb.</th>"
            f"<th style='{_TH}text-align:center;'>Duration</th>"
            f"<th style='{_TH}text-align:center;'>Rating Medio</th>"
            f"<th style='{_TH}text-align:left;'>Cat. FIDA</th>"
            f"<th style='{_TH}text-align:center;'>FIDArating</th>"
            f"<th style='{_TH}text-align:center;'>Morningstar</th>"
            f"</tr>"
        )
        _tbl_body = ""
        for _, _tr in _df_sorted.iterrows():
            _dur   = _fb_metric(_tr["nome"], "duration")
            _rat   = _fb_metric(_tr["nome"], "credit_rating")
            _azfb  = _fb_metric(_tr["nome"], "fb_az_pct")
            _obfb  = _fb_metric(_tr["nome"], "fb_obb_pct")
            _az_d  = (_azfb if _azfb is not None else _tr["az_pct"]) * 100
            _ob_d  = (_obfb if _obfb is not None else _tr["obb_pct"]) * 100
            _fd_ov = _fd_live.get(_tr["nome"], {}).get("overview", {})
            _cat   = _fd_ov.get("cat_assog") or "—"
            _fida  = _fd_ov.get("fida_rating") or "—"
            _dur_s = f"{_dur:.2f} y" if isinstance(_dur, (int, float)) else "—"
            _rat_s = _rat if isinstance(_rat, str) else "—"
            _rat_w = "600" if _rat_s != "—" else "400"
            try:
                _fr_int = int(_fida)
                _fr_col = _FIDA_COL.get(_fr_int, "#64748B")
                _fr_bg  = _FIDA_BG.get(_fr_int)
            except (ValueError, TypeError):
                _fr_col, _fr_bg = "#64748B", None
            _fida_cell = (
                f"<span style='background:{_fr_bg};color:#fff;padding:2px 8px;"
                f"border-radius:4px;font-weight:700;'>{_fida}</span>"
                if _fr_bg else
                f"<span style='color:{_fr_col};font-weight:700;'>{_fida}</span>"
            )
            _ms_r = _ms_live.get(_tr["nome"], {}).get("ms_rating")
            _ms_cell = _ms_badge_html(_ms_r)
            _tbl_body += (
                f"<tr>"
                f"<td style='{_TC}font-weight:500;'>{_fund_link(_tr['nome'])}</td>"
                f"<td style='{_TC}text-align:center;color:#1B4FBB;font-weight:600;'>"
                f"{_tr[wcol]*100:.1f}%</td>"
                f"<td style='{_TC}text-align:center;'>{_az_d:.1f}%</td>"
                f"<td style='{_TC}text-align:center;'>{_ob_d:.1f}%</td>"
                f"<td style='{_TC}text-align:center;'>{_dur_s}</td>"
                f"<td style='{_TC}text-align:center;font-weight:{_rat_w};'>{_rat_s}</td>"
                f"<td style='{_TC}color:#64748B;'>{_cat}</td>"
                f"<td style='{_TC}text-align:center;'>{_fida_cell}</td>"
                f"<td style='{_TC}text-align:center;'>{_ms_cell}</td>"
                f"</tr>"
            )
        if _tbl_body:
            st.markdown(
                f"<div style='overflow-x:auto;border-radius:10px;"
                f"border:1px solid #e2e8f0;background:#fff;'>"
                f"<table style='width:100%;border-collapse:collapse;'>"
                f"<thead>{_scomp_hdr}</thead><tbody>{_tbl_body}</tbody>"
                f"</table></div>",
                unsafe_allow_html=True)
            _ms_note = (f" &nbsp;·&nbsp; Morningstar: FondiOnline"
                        if _ms_with_rating else
                        " &nbsp;·&nbsp; Morningstar: clicca «Scarica Rating Morningstar»")
            st.markdown(
                f"<p style='{_note_style}'>"
                f"Duration &amp; Rating Medio: {_note_fb}"
                f" &nbsp;·&nbsp; Cat. FIDA &amp; FIDArating: {_note_fd}"
                f"{_ms_note}</p>",
                unsafe_allow_html=True)

    # ── TAB 2 — RENDIMENTI ───────────────────────────────────────────────────
    with tab2:
        _pk      = ["ytd", "perf_1y", "perf_3y", "perf_5y", "vol_1y", "sharpe_1y"]
        _ptf_p   = _perf_wavg(_pk)
        _p_hdr   = ["Fondo", "Peso", "YTD", "1 Anno", "3 Anni", "5 Anni",
                    "Vol. 1A", "Sharpe 1A"]
        _p_ptf   = [_ptf_row_label, "100%",
                    _ptf_p.get("ytd",      "N/D"),
                    _ptf_p.get("perf_1y",  "N/D"),
                    _ptf_p.get("perf_3y",  "N/D"),
                    _ptf_p.get("perf_5y",  "N/D"),
                    _ptf_p.get("vol_1y",   "N/D"),
                    _ptf_p.get("sharpe_1y","N/D")]
        _p_funds = []
        for _, _pr in _df_sorted.iterrows():
            _np  = _pr["nome"]
            _ana = _fd_live.get(_np, {}).get("analysis", {})
            def _gp(k, _n=_np, _a=_ana):
                v = _fb_metric(_n, k) or _a.get(k, "") or ""
                return str(v) if v else "N/D"
            _p_funds.append([
                _fund_link(_np),
                f"{_pr[wcol]*100:.1f}%",
                _perf_val_col(_gp("ytd")),
                _perf_val_col(_gp("perf_1y")),
                _perf_val_col(_gp("perf_3y")),
                _perf_val_col(_gp("perf_5y")),
                _gp("vol_1y"),
                _gp("sharpe_1y"),
            ])
        st.markdown(_html_table(_p_hdr, _p_ptf, _p_funds), unsafe_allow_html=True)
        st.markdown(
            f"<p style='{_note_style}'>"
            f"YTD, 1A, 3A, 5A: {_note_fb} &nbsp;·&nbsp; Vol. e Sharpe: {_note_fd}</p>",
            unsafe_allow_html=True)

    # ── TAB 3 — RISCHIO ──────────────────────────────────────────────────────
    with tab3:
        _rk      = ["vol_1y", "vol_3y", "vol_5y", "neg_vol_1y", "sharpe_3y", "sortino_1y"]
        _ptf_r   = _perf_wavg(_rk)
        _r_hdr   = ["Fondo", "Peso", "Vol. 1A", "Vol. 3A", "Vol. 5A",
                    "Vol. Neg. 1A", "Sharpe 3A", "Sortino 1A"]
        _r_ptf   = [_ptf_row_label, "100%",
                    _ptf_r.get("vol_1y",     "N/D"),
                    _ptf_r.get("vol_3y",     "N/D"),
                    _ptf_r.get("vol_5y",     "N/D"),
                    _ptf_r.get("neg_vol_1y", "N/D"),
                    _ptf_r.get("sharpe_3y",  "N/D"),
                    _ptf_r.get("sortino_1y", "N/D")]
        _r_funds = []
        for _, _rr in _df_sorted.iterrows():
            _nr  = _rr["nome"]
            _ana = _fd_live.get(_nr, {}).get("analysis", {})
            def _gr(k, _a=_ana):
                return str(_a.get(k, "") or "") or "N/D"
            _r_funds.append([
                _fund_link(_nr),
                f"{_rr[wcol]*100:.1f}%",
                _gr("vol_1y"),
                _gr("vol_3y"),
                _gr("vol_5y"),
                _gr("neg_vol_1y"),
                _gr("sharpe_3y"),
                _gr("sortino_1y"),
            ])
        st.markdown(_html_table(_r_hdr, _r_ptf, _r_funds), unsafe_allow_html=True)
        st.markdown(
            f"<p style='{_note_style}'>Metriche di rischio: {_note_fd}</p>",
            unsafe_allow_html=True)

    # ── TAB 4 — UNP / IUNP ───────────────────────────────────────────────────
    with tab4:
        _u_wtd = _iu_wtd = _u_covw = 0.0
        _u_funds = []
        for _, _ur in _df_sorted.iterrows():
            _nu       = _ur["nome"]
            _uu, _iuu = lookup_unp(_nu)
            _wu       = _ur[wcol]
            if _uu is not None:
                _u_wtd  += _uu  * _wu
                _iu_wtd += _iuu * _wu
                _u_covw += _wu
            # FIDArating badge
            _fd_ov_u  = _fd_live.get(_nu, {}).get("overview", {})
            _fida_u   = _fd_ov_u.get("fida_rating") or "—"
            try:
                _fri_u  = int(_fida_u)
                _fcol_u = _FIDA_COL.get(_fri_u, "#64748B")
                _fbg_u  = _FIDA_BG.get(_fri_u)
            except (ValueError, TypeError):
                _fcol_u, _fbg_u = "#64748B", None
            _fida_cell_u = (
                f"<span style='background:{_fbg_u};color:#fff;padding:2px 8px;"
                f"border-radius:4px;font-weight:700;'>{_fida_u}</span>"
                if _fbg_u else
                f"<span style='color:{_fcol_u};font-weight:700;'>{_fida_u}</span>"
            )
            # Morningstar badge
            _ms_r_u    = _ms_live.get(_nu, {}).get("ms_rating")
            _ms_cell_u = _ms_badge_html(_ms_r_u)
            _u_funds.append([
                _fund_link(_nu),
                f"{_wu*100:.1f}%",
                f"{_uu:.2f}%"  if _uu  is not None else "—",
                f"{_iuu:.2f}%" if _iuu is not None else "—",
                _fida_cell_u,
                _ms_cell_u,
            ])
        _ptf_unp  = f"{_u_wtd/_u_covw:.2f}%"  if _u_covw > 0.01 else "N/D"
        _ptf_iunp = f"{_iu_wtd/_u_covw:.2f}%"  if _u_covw > 0.01 else "N/D"
        st.markdown(
            _html_table(
                ["Fondo", "Peso", "%UNP", "%IUNP36", "FIDArating", "Morningstar"],
                [_ptf_row_label, "100%", _ptf_unp, _ptf_iunp, "", ""],
                _u_funds,
            ),
            unsafe_allow_html=True)
        st.markdown(
            f"<p style='{_note_style}'>"
            f"UNP = Utile Netto di Portafoglio · IUNP36 = indice su orizzonte triennale. "
            f"Fonte: Catalogo Prodotti &amp; Servizi Azimut, settembre 2025.</p>",
            unsafe_allow_html=True)

    # ── DOWNLOAD SECTION ─────────────────────────────────────
    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<p class="sec-title">Esporta Report PDF Completo</p>',unsafe_allow_html=True)

    fida_urls = raw.get("fida_urls", {})
    n_urls = sum(1 for nome in df_act["nome"].unique() if nome in fida_urls)

    col_btn,col_inf = st.columns([1,2])
    with col_inf:
        if factbook_data:
            _src_note = f"📖 Factbook: rendimenti per <b>{len(factbook_data)}</b> fondi"
        elif cached_fd:
            _src_note = f"💾 Cache FondiDoc aggiornata al: <b>{cache_date}</b>"
        else:
            _src_note = f"🌐 Dati live da FondiDoc per {n_urls}/{n_fondi} fondi"
        st.markdown(
            f"<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;"
            f"padding:1rem 1.25rem;'>"
            f"<div style='font-size:.8rem;color:#1d4ed8;font-weight:600;margin-bottom:.4rem;'>"
            f"Il report PDF (3 sezioni) contiene:</div>"
            f"<div style='font-size:.82rem;color:#1e40af;line-height:1.9;'>"
            f"✓ <b>Pag. 1</b> — Grafico a torta + KPI di portafoglio<br>"
            f"✓ <b>Pag. 2</b> — Tavola rendimenti YTD / 1A / 3A / 5A + Rischio<br>"
            f"✓ <b>Pag. 3+</b> — Schede analitiche per {n_fondi} fondi<br>"
            f"<span style='color:#3b82f6;'>{_src_note}</span></div></div>",
            unsafe_allow_html=True)

    fida_df = raw.get("FIDA", pd.DataFrame())

    with col_btn:
        if st.session_state.get("_pdf_bytes_ready"):
            # PDF già generato per questo portafoglio/profilo: mostra solo download
            st.download_button(
                "📥   Scarica Report PDF",
                data=st.session_state["_pdf_bytes_ready"],
                file_name=st.session_state.get("_pdf_fname_ready", "report.pdf"),
                mime="application/pdf",
                use_container_width=True,
            )
            st.success(f"✅ PDF pronto — {st.session_state.get('_pdf_lbl','')}")
            st.caption("Cambia portafoglio, profilo o fondi per rigenerare.")
        if not st.session_state.get("_pdf_bytes_ready") and \
                st.button("⚡  Genera PDF", use_container_width=True, type="primary"):
            # Clear any stale PDF from a previous run
            for _k in ("_pdf_bytes_ready", "_pdf_fname_ready", "_pdf_lbl"):
                st.session_state.pop(_k, None)

            pb = st.progress(0, text="Scarico dati FondiDoc…")
            def upd(v): pb.progress(v, text=f"FondiDoc: {int(v*100)}%…")
            fund_data = fetch_all_fund_data(df_act, fida_urls, upd)
            pb.progress(1.0, text="✅ Genero PDF…")
            save_fund_cache(fund_data)
            # Store fund data so the Scomposizione table gets populated
            # on the immediate rerun triggered below
            st.session_state["_scomp_fd"] = fund_data
            try:
                pdf_bytes = generate_pdf(
                    df_act, wcol, profile, ptf_label, fund_data,
                    fida_df=fida_df, factbook_data=factbook_data,
                    cache_date=cache_date)
                fname = (f"Azimut_{ptf_label.replace(' ','_')}_{profile}_"
                         f"{datetime.date.today().strftime('%Y%m%d')}.pdf")
                st.session_state["_pdf_bytes_ready"] = pdf_bytes
                st.session_state["_pdf_fname_ready"] = fname
                st.session_state["_pdf_lbl"] = f"{len(fund_data)} schede da FondiDoc"
            except Exception as _pe:
                st.error(f"Errore PDF: {_pe}")
            pb.empty()
            # Force immediate rerun so the Scomposizione table and download
            # button both reflect the freshly fetched FondiDoc data
            st.rerun()

    st.markdown("<br><br>",unsafe_allow_html=True)


# ── Entry point ─────────────────────────────────────────────────────────────
# Call main() directly (not behind __name__ guard) so it always runs on
# Streamlit Cloud regardless of how the script is executed.
import traceback as _tb
try:
    main()
except BaseException as _e:
    # Re-raise Streamlit-internal control-flow exceptions so the framework
    # can handle them (RerunException → st.rerun(), StopException → st.stop(),
    # SystemExit, KeyboardInterrupt, etc.).
    _ename = type(_e).__name__
    if _ename in ("RerunException", "StopException") or \
       isinstance(_e, (SystemExit, KeyboardInterrupt)):
        raise
    # For everything else, show a readable error in the UI.
    try:
        st.error(f"**Errore imprevisto:** {_ename}: {_e}")
        with st.expander("🔍 Dettaglio tecnico (per il debug)"):
            st.code(_tb.format_exc())
    except Exception:
        pass
