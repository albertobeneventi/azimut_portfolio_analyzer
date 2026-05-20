# ============================================================
# AZIMUT PORTFOLIO ANALYZER v2.0 — app.py
# Aggiornamento: Schede fondi + Rendimenti da FondiDoc FIDA
# ============================================================

import re
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import load_workbook
import zipfile
from xml.etree import ElementTree as ET
import io
import datetime
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, Image as RLImage,
    HRFlowable, PageBreak, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import cm

# ── PAGE CONFIG ─────────────────────────────────────────────
st.set_page_config(
    page_title="Azimut | Portfolio Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CONSTANTS ───────────────────────────────────────────────
GROUP_NAMES = {"ALLOCATION", "AZIONARI (LONG)", "BOND"}
COL_A, COL_B, COL_C, COL_G, COL_K, COL_O, COL_R = 0, 1, 2, 6, 10, 14, 17
PROFILES     = ["CONSERVATIVO", "EQUILIBRATO", "ACCRESCITIVO", "CENTRALE"]
PROFILE_ICONS = {"CONSERVATIVO":"🛡️","EQUILIBRATO":"⚖️","ACCRESCITIVO":"📈","CENTRALE":"🎯"}
PROFILE_W_COL = {"CONSERVATIVO":"w_cons","EQUILIBRATO":"w_equil","ACCRESCITIVO":"w_accr","CENTRALE":"w_centrale"}

MACRO_COLORS = {
    "Azionari":"#1B4FBB","Bilanciati/Flessibili":"#C9A84C",
    "Obbligazionari":"#2D9D78","Alternativi":"#8B5CF6","Altro":"#94A3B8",
}
SHADES = {
    "Azionari":             ["#0D3080","#1B4FBB","#2563EB","#3B82F6","#60A5FA","#93C5FD","#BFDBFE"],
    "Bilanciati/Flessibili":["#92650A","#B8860B","#C9A84C","#D4B572","#DFC298","#E9CEB4","#F3DACD"],
    "Obbligazionari":       ["#065F46","#14855F","#2D9D78","#34B98A","#6DE5BC","#9AEFD2","#C5F7E7"],
    "Alternativi":          ["#5B21B6","#7C3AED","#8B5CF6","#A78BFA","#C4B5FD","#DDD6FE"],
    "Altro":                ["#475569","#64748B","#94A3B8","#CBD5E1"],
}
DEFAULT_AZ = {"Azionari":0.92,"Bilanciati/Flessibili":0.50,"Obbligazionari":0.06,"Altro":0.50}
FONDIDOC_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "en-US,en;q=0.9,it;q=0.8",
}
# Override for one fund whose FIDA sheet hyperlink points to class B
MANUAL_URL_OVERRIDES = {
    "AZ F.1 All. Balanced FoF A Cap EUR":
        "https://www.fondidoc.it/d/Index/AZPOA/LU0346933400_az-f1-allocation-balanced-fof-a-az-fund-cap-eur",
}

# ── CSS ──────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@600;700&family=DM+Sans:opsz,wght@9..40,300;9..40,400;9..40,500;9..40,600&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
h1,h2,h3{font-family:'Cormorant Garamond',serif !important;}
[data-testid="stSidebar"]{background:linear-gradient(170deg,#06101e 0%,#0d1f3c 55%,#0a1628 100%);border-right:1px solid #1a3050;}
[data-testid="stSidebar"] .stFileUploader label,[data-testid="stSidebar"] .stRadio > label,[data-testid="stSidebar"] .stSelectbox > label{color:#4a6582 !important;font-size:.68rem !important;letter-spacing:.12em !important;text-transform:uppercase !important;font-weight:600 !important;}
[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p{color:#c0cfe0 !important;font-size:.9rem !important;}
[data-testid="stSidebar"] .stSelectbox>div>div{background:#132035 !important;border:1px solid #243d5a !important;color:#dde6f0 !important;border-radius:6px !important;}
[data-testid="stSidebar"] .stFileUploader>div{background:#132035 !important;border:1px dashed #2a4a6a !important;border-radius:8px !important;}
[data-testid="stSidebar"] .stFileUploader p,[data-testid="stSidebar"] .stFileUploader span{color:#8aa5c0 !important;font-size:.8rem !important;}
.main{background:#f6f8fb !important;}.block-container{padding-top:1.8rem !important;max-width:1300px;}
.az-header{background:linear-gradient(130deg,#081420 0%,#0f2644 50%,#162e52 100%);border-radius:16px;padding:2rem 2.5rem;margin-bottom:1.8rem;position:relative;overflow:hidden;box-shadow:0 8px 32px rgba(0,0,0,.15);}
.az-header::after{content:'';position:absolute;bottom:-60px;right:-40px;width:220px;height:220px;border-radius:50%;background:radial-gradient(circle,rgba(201,168,76,.18) 0%,transparent 70%);}
.az-eyebrow{font-size:.65rem;letter-spacing:.2em;color:#4a7098;text-transform:uppercase;font-weight:600;}
.az-title{font-family:'Cormorant Garamond',serif;font-size:2.1rem;font-weight:700;color:#f0f6ff;margin:.2rem 0 .4rem;line-height:1.1;}
.az-rule{width:38px;height:3px;background:#C9A84C;border-radius:2px;margin:.6rem 0;}
.az-meta{font-size:.88rem;color:#6b8fb0;}
.kpi{background:#fff;border:1px solid #e4eaf3;border-radius:12px;padding:1.2rem 1.4rem;box-shadow:0 1px 4px rgba(0,0,0,.05);}
.kpi-label{font-size:.65rem;text-transform:uppercase;letter-spacing:.1em;color:#94a3b8;font-weight:500;margin-bottom:.3rem;}
.kpi-value{font-size:1.9rem;font-weight:700;color:#0d1b2a;font-family:'Cormorant Garamond',serif;line-height:1;}
.kpi-sub{font-size:.75rem;color:#64748b;margin-top:.3rem;}
.sec-title{font-family:'Cormorant Garamond',serif;font-size:1.25rem;font-weight:600;color:#0d1b2a;border-bottom:2px solid #c9a84c;display:inline-block;padding-bottom:.4rem;margin-bottom:.9rem;}
.fund-group-hdr{background:#f0f4f9;padding:.45rem 1rem;font-size:.65rem;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:#64748b;border-bottom:1px solid #e2e8f0;}
.fund-row{display:flex;align-items:center;gap:10px;padding:.65rem 1rem;border-bottom:1px solid #f1f5f9;}
.fund-row:last-child{border-bottom:none;}
.fund-dot{width:8px;height:34px;border-radius:3px;flex-shrink:0;}
.fund-name{font-size:.83rem;color:#1e293b;font-weight:500;flex:1;min-width:0;}
.fund-cat{font-size:.68rem;color:#64748b;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.fund-pct{font-size:1rem;font-weight:700;color:#0d1b2a;min-width:2.8rem;text-align:right;}
[data-testid="stDownloadButton"]>button{background:linear-gradient(135deg,#0f2d6b 0%,#1b4fbb 100%) !important;color:#fff !important;font-size:1.05rem !important;font-weight:600 !important;padding:.9rem 2rem !important;border-radius:10px !important;border:none !important;width:100% !important;letter-spacing:.02em !important;box-shadow:0 4px 18px rgba(27,79,187,.35) !important;}
[data-testid="stDownloadButton"]>button:hover{box-shadow:0 6px 24px rgba(27,79,187,.55) !important;transform:translateY(-2px) !important;}
.w-ok{background:#d1fae5;border:1px solid #6ee7b7;border-radius:8px;padding:.7rem 1rem;font-size:.84rem;color:#065f46;}
.w-warn{background:#fef3c7;border:1px solid #fcd34d;border-radius:8px;padding:.7rem 1rem;font-size:.84rem;color:#92400e;}
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════

def get_macro(cat: str) -> str:
    if not cat or cat == "-": return "Altro"
    c = cat.lower()
    if "azionari" in c or "equity" in c: return "Azionari"
    if any(x in c for x in ["obbligazionari","bond","credit","debt","sukuk","reddito"]): return "Obbligazionari"
    if any(x in c for x in ["bilanciati","allocation","flessibili","balanced","flexible","prudenti","moderati"]): return "Bilanciati/Flessibili"
    if any(x in c for x in ["alternativi","alternative","commodity"]): return "Alternativi"
    return "Altro"


def assign_colors(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    counter = {k: 0 for k in SHADES}
    colors = []
    for _, row in df.iterrows():
        mc = row.get("macro_cat", "Altro")
        shades = SHADES.get(mc, SHADES["Altro"])
        colors.append(shades[counter.get(mc, 0) % len(shades)])
        counter[mc] = counter.get(mc, 0) + 1
    df["color"] = colors
    return df


def pct_color(val_str: str) -> str:
    """Returns 'pos', 'neg', or 'neu' based on value sign"""
    try:
        v = float(val_str.replace("%","").replace(",",".").strip())
        return "pos" if v > 0 else ("neg" if v < 0 else "neu")
    except:
        return "neu"


# ════════════════════════════════════════════════════════════
# DATA PARSING
# ════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def parse_excel(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    out = {}
    for sname in ["PTF FULL", "PTF SHORT"]:
        if sname in wb.sheetnames:
            out[sname] = _parse_ptf(wb, sname)
    if "FIDA" in wb.sheetnames:
        out["FIDA"] = _parse_fida(wb)
    wb.close()
    out["fida_urls"] = extract_fida_urls(file_bytes)
    return out


def _parse_ptf(wb, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    funds, cur_group = [], {"name":None,"mc":1.0,"me":1.0,"ma":1.0}
    for row in rows:
        name = row[COL_A]
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name in GROUP_NAMES:
            cur_group = {
                "name": name,
                "mc": float(row[COL_C]) if isinstance(row[COL_C],(int,float)) else 1.0,
                "me": float(row[COL_G]) if isinstance(row[COL_G],(int,float)) else 1.0,
                "ma": float(row[COL_K]) if isinstance(row[COL_K],(int,float)) else 1.0,
            }
            continue
        if name.startswith("AZ") and cur_group["name"]:
            rw = row[COL_R]
            if not isinstance(rw,(int,float)) or rw <= 0: continue
            funds.append({
                "nome":     name,
                "categoria": row[COL_B] if isinstance(row[COL_B],str) else "",
                "gruppo":   cur_group["name"],
                "az_pct":   min(1.0,max(0.0,float(row[COL_O]) if isinstance(row[COL_O],(int,float)) else 0.5)),
                "obb_pct":  min(1.0,max(0.0,1.0-(float(row[COL_O]) if isinstance(row[COL_O],(int,float)) else 0.5))),
                "r_weight": float(rw),
                "mc":cur_group["mc"],"me":cur_group["me"],"ma":cur_group["ma"],
            })
    if not funds: return pd.DataFrame()
    df = pd.DataFrame(funds)
    for wcol, mcol in [("w_cons","mc"),("w_equil","me"),("w_accr","ma")]:
        raw = df["r_weight"]*df[mcol]
        df[wcol] = raw/raw.sum() if raw.sum()>0 else raw
    tot = df["r_weight"].sum()
    df["w_centrale"] = df["r_weight"]/tot if tot>0 else df["r_weight"]
    df["macro_cat"] = df["categoria"].apply(get_macro)
    df = assign_colors(df)
    return df


def _parse_fida(wb) -> pd.DataFrame:
    ws = wb["FIDA"]
    rows = list(ws.iter_rows(values_only=True))
    funds = []
    for row in rows[1:]:
        nome = row[0]
        if not nome or not isinstance(nome,str): continue
        nome = nome.strip().replace("\xa0","")
        cat  = (row[2] or "").strip()
        funds.append({"nome":nome,"isin":row[1] or "","categoria":cat,"macro_cat":get_macro(cat)})
    return pd.DataFrame(funds).drop_duplicates(subset=["nome"])


def extract_fida_urls(file_bytes: bytes) -> dict:
    """Extract fondidoc.it URLs from FIDA sheet hyperlinks in the Excel XML."""
    fund_urls = dict(MANUAL_URL_OVERRIDES)
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            ss_root  = ET.fromstring(z.read("xl/sharedStrings.xml"))
            ns_ss    = {"s":"http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            strings  = ["".join(t.text or "" for t in si.findall(".//s:t",ns_ss))
                        for si in ss_root.findall("s:si",ns_ss)]
            fida_root= ET.fromstring(z.read("xl/worksheets/sheet5.xml"))
            ws_ns    = {"ws":"http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            rels_root= ET.fromstring(z.read("xl/worksheets/_rels/sheet5.xml.rels"))
            rels_ns  = {"rel":"http://schemas.openxmlformats.org/package/2006/relationships"}
        rid_to_url = {r.get("Id"):r.get("Target") for r in rels_root.findall("rel:Relationship",rels_ns)}
        hyp_map    = {}
        r_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        for hl in fida_root.findall(".//ws:hyperlink",ws_ns):
            rid = hl.get(r_attr)
            if rid: hyp_map[hl.get("ref")] = rid
        for row in fida_root.findall(".//ws:row",ws_ns):
            for cell in row.findall("ws:c",ws_ns):
                ref = cell.get("r")
                if ref and ref.startswith("A") and ref in hyp_map:
                    v = cell.find("ws:v",ws_ns)
                    if v is not None and cell.get("t") == "s":
                        name = strings[int(v.text)].strip().replace("\xa0","")
                        url  = rid_to_url.get(hyp_map[ref],"")
                        if "fondidoc.it" in url:
                            fund_urls[name] = url
    except Exception:
        pass
    return fund_urls


# ════════════════════════════════════════════════════════════
# FONDIDOC SCRAPING
# ════════════════════════════════════════════════════════════

def _to_en_url(url: str) -> str:
    """Ensure URL uses /en/ locale."""
    if "/en/" in url: return url
    return url.replace("fondidoc.it/d/","fondidoc.it/en/d/")


def _to_ana_url(index_url: str) -> str:
    return index_url.replace("/d/Index/","/d/Ana/").replace("/en/d/Index/","/en/d/Ana/")


def _fetch_html(url: str, timeout: int = 8) -> str | None:
    try:
        r = requests.get(_to_en_url(url), headers=FONDIDOC_HEADERS, timeout=timeout)
        return r.text if r.status_code == 200 else None
    except Exception:
        return None


def _parse_overview(html: str) -> dict:
    soup = BeautifulSoup(html, "lxml")
    lines = [l.strip() for l in soup.get_text(separator="\n").split("\n") if l.strip()]
    d = {}
    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1<len(lines) else ""
        if line == "SRRI (risk value)":           d["srri"]         = nxt
        elif line == "Start date":                d["start_date"]   = nxt
        elif line == "Assogestioni category":     d["cat_assog"]    = nxt
        elif line == "Income distribution":       d["income"]       = nxt
        elif line == "Management Fee":            d["mgmt_fee"]     = nxt
        elif line == "Performance Fee":           d["perf_fee"]     = nxt
        elif line == "Subscription fee":          d["sub_fee"]      = nxt
        elif line == "Rating" and "fida_rating" not in d: d["fida_rating"] = nxt
        elif line == "Score":                     d["fida_score"]   = nxt
        elif line == "Category" and "fida_cat" not in d: d["fida_cat"] = nxt
    return d


def _parse_analysis(html: str) -> dict:
    soup = BeautifulSoup(html, "lxml")
    d = {}

    # NAV section
    lines = [l.strip() for l in soup.get_text(separator="\n").split("\n") if l.strip()]
    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1<len(lines) else ""
        if line == "Last update": d["last_update"] = nxt
        elif line == "NAV":       d["nav"]          = nxt
        elif line == "Daily change (%)": d["daily_change"] = nxt

    # Tables
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows: continue
        header = [td.get_text(strip=True) for td in rows[0].find_all(["th","td"])]

        # Performance table (has YTD column)
        if "YTD" in header and "1 year" in header:
            def sg(cells, key, hdr):
                try: return cells[hdr.index(key)] if hdr.index(key)<len(cells) else "—"
                except ValueError: return "—"
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if not cells: continue
                lbl = cells[0].lower()
                if "performance" in lbl:
                    d["ytd"]     = sg(cells,"YTD",header)
                    d["perf_1y"] = sg(cells,"1 year",header)
                    d["perf_3y"] = sg(cells,"3 years",header)
                    d["perf_5y"] = sg(cells,"5 years",header)

        # Risk table (1 year / 3 years / 5 years, NO YTD)
        elif "1 year" in header and "YTD" not in header and len(header) >= 4:
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if not cells: continue
                lbl = cells[0].lower()
                def gv(idx): return cells[idx] if idx<len(cells) else "—"
                if "volatility" in lbl and "negative" not in lbl:
                    d["vol_1y"],d["vol_3y"],d["vol_5y"] = gv(1),gv(2),gv(3)
                elif "negative" in lbl:
                    d["neg_vol_1y"] = gv(1)
                elif "sharpe" in lbl:
                    d["sharpe_1y"],d["sharpe_3y"],d["sharpe_5y"] = gv(1),gv(2),gv(3)
                elif "sortino" in lbl:
                    d["sortino_1y"] = gv(1)
                elif "var" in lbl or "value at risk" in lbl:
                    d["var_1y"] = gv(1)
                    if len(header) > 2: d["var_3y"] = gv(2)

        # Annual performance (header contains year digits)
        elif any(h.isdigit() and len(h)==4 for h in header):
            years = [h for h in header if h.isdigit() and len(h)==4]
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if len(cells)<2: continue
                # Take first data row that has % values (fund row, not benchmark)
                if any("%" in c for c in cells) and "annual_perf" not in d:
                    annual = {}
                    for yr in years:
                        try:
                            idx = header.index(yr)
                            annual[yr] = cells[idx] if idx<len(cells) else "—"
                        except ValueError: pass
                    if annual: d["annual_perf"] = annual
                    break
    return d


def _extract_isin(url: str) -> str:
    """Estrae ISIN da URL FondiDoc (formato: /CATCODE/ISIN_slug-nome-fondo)."""
    m = re.search(r'/([A-Z]{2}[A-Z0-9]{10})[_/]', url)
    return m.group(1) if m else ""


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_fund_data(index_url: str) -> dict:
    """Fetch overview + analysis for one fund. Cached 1h."""
    result = {"url": index_url}
    isin = _extract_isin(index_url)
    if isin:
        result["isin"] = isin
    html_idx = _fetch_html(index_url)
    if html_idx: result["overview"] = _parse_overview(html_idx)
    html_ana = _fetch_html(_to_ana_url(index_url))
    if html_ana: result["analysis"] = _parse_analysis(html_ana)
    return result


def fetch_all_fund_data(df: pd.DataFrame, fida_urls: dict,
                         progress_cb=None) -> dict:
    """Parallel fetch for all portfolio funds."""
    tasks = {}
    for nome in df["nome"].unique():
        url = fida_urls.get(nome)
        if url: tasks[nome] = url

    results = {}
    total = len(tasks)
    done  = 0

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {pool.submit(fetch_fund_data, url): nome for nome, url in tasks.items()}
        for future in as_completed(futures):
            nome = futures[future]
            try: results[nome] = future.result()
            except Exception: results[nome] = {}
            done += 1
            if progress_cb: progress_cb(done/total)

    return results


# ════════════════════════════════════════════════════════════
# PLOTLY CHARTS (unchanged)
# ════════════════════════════════════════════════════════════

def make_fund_pie(df, wcol, profile):
    d = df[df[wcol]>0.005].copy()
    d["pct"] = d[wcol]*100
    labels = d["nome"].apply(lambda x: (x[:38]+"…") if len(x)>38 else x)
    fig = go.Figure(go.Pie(
        labels=labels, values=d["pct"],
        marker=dict(colors=d["color"].tolist(), line=dict(color="#fff",width=2.5)),
        hovertemplate="<b>%{label}</b><br>Peso: <b>%{value:.1f}%</b><extra></extra>",
        textinfo="percent", textfont=dict(size=10,family="DM Sans"),
        hole=0.40, pull=[0.04 if v==d["pct"].max() else 0 for v in d["pct"]],
        sort=False, direction="clockwise",
    ))
    fig.update_layout(
        margin=dict(t=10,b=10,l=10,r=180), showlegend=True,
        legend=dict(x=1.01,y=0.5,orientation="v",font=dict(size=9.5,family="DM Sans"),bgcolor="rgba(0,0,0,0)"),
        paper_bgcolor="rgba(0,0,0,0)", height=430,
        annotations=[dict(text=f"<b>{profile[:4]}</b>",x=0.5,y=0.5,showarrow=False,
                          font=dict(size=17,color="#0d1b2a",family="Cormorant Garamond"))],
    )
    return fig


def make_macro_bar(df, wcol):
    agg = df[df[wcol]>0.001].groupby("macro_cat")[wcol].sum().reset_index().sort_values(wcol)
    agg["pct"] = agg[wcol]*100
    agg["color"] = agg["macro_cat"].map(MACRO_COLORS)
    fig = go.Figure(go.Bar(
        x=agg["pct"], y=agg["macro_cat"], orientation="h",
        marker=dict(color=agg["color"].tolist(), line=dict(color="#fff",width=1)),
        hovertemplate="<b>%{y}</b>: %{x:.1f}%<extra></extra>",
        text=agg["pct"].apply(lambda v:f"{v:.1f}%"),
        textposition="inside", insidetextanchor="middle",
        textfont=dict(color="#fff",size=11,family="DM Sans"),
    ))
    fig.update_layout(
        margin=dict(t=10,b=10,l=10,r=10), paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(showgrid=False,showticklabels=False,range=[0,105]),
        yaxis=dict(showgrid=False,tickfont=dict(size=11,family="DM Sans")),
        height=max(160,len(agg)*48), bargap=0.28,
    )
    return fig


# ════════════════════════════════════════════════════════════
# PDF — MATPLOTLIB HELPERS
# ════════════════════════════════════════════════════════════

def _mpl_portfolio_pie(df, wcol, profile) -> io.BytesIO:
    d = df[df[wcol]>0.005].sort_values(wcol, ascending=False)
    fig, (ax_pie,ax_leg) = plt.subplots(1,2,figsize=(11,5.5),gridspec_kw={"width_ratios":[1.3,1]})
    wedges,_,autotexts = ax_pie.pie(
        d[wcol], colors=d["color"].tolist(),
        autopct=lambda p:f"{p:.1f}%" if p>3.5 else "",
        pctdistance=0.72,
        wedgeprops=dict(width=0.58,edgecolor="white",linewidth=2), startangle=90,
    )
    for at in autotexts: at.set_fontsize(8); at.set_color("white"); at.set_fontweight("bold")
    ax_pie.text(0,0,profile[:4],ha="center",va="center",fontsize=15,fontweight="bold",color="#0D1B2A")
    ax_leg.axis("off")
    handles = [mpatches.Patch(color=r["color"],
               label=f"{r['nome'][:32]}{'…' if len(r['nome'])>32 else ''}  {r[wcol]*100:.1f}%")
               for _,r in d.iterrows()]
    ax_leg.legend(handles=handles,loc="center left",frameon=False,fontsize=7.8,labelspacing=0.85,handlelength=1.2)
    fig.patch.set_facecolor("#FFFFFF"); plt.tight_layout(pad=1.5)
    buf=io.BytesIO(); fig.savefig(buf,format="png",dpi=150,bbox_inches="tight",facecolor="white")
    plt.close(fig); buf.seek(0); return buf


def _mpl_macro_pie(df, wcol) -> io.BytesIO | None:
    """Asset-allocation donut per macro-categoria — stesso stile di _mpl_portfolio_pie."""
    agg = df[df[wcol]>0.001].groupby("macro_cat")[wcol].sum().sort_values(ascending=False)
    if agg.empty: return None
    colors = [MACRO_COLORS.get(k, "#94A3B8") for k in agg.index]
    fig, (ax_pie, ax_leg) = plt.subplots(1, 2, figsize=(10, 4),
                                          gridspec_kw={"width_ratios": [1.2, 1]})
    wedges, _, autotexts = ax_pie.pie(
        agg.values, colors=colors,
        autopct=lambda p: f"{p:.1f}%" if p >= 5 else "",
        pctdistance=0.70,
        wedgeprops=dict(width=0.58, edgecolor="white", linewidth=2.5),
        startangle=90,
    )
    for at in autotexts:
        at.set_fontsize(10); at.set_color("white"); at.set_fontweight("bold")
    ax_pie.text(0, 0, "Asset\nAlloc.", ha="center", va="center",
                fontsize=11, fontweight="bold", color="#0D1B2A")
    ax_leg.axis("off")
    handles = [mpatches.Patch(color=colors[i],
               label=f"{k}  {v*100:.1f}%")
               for i, (k, v) in enumerate(agg.items())]
    ax_leg.legend(handles=handles, loc="center left", frameon=False,
                  fontsize=9.5, labelspacing=1.2, handlelength=1.4)
    fig.patch.set_facecolor("#FFFFFF")
    plt.tight_layout(pad=1.2)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig); buf.seek(0); return buf


def _mpl_annual_bar(annual_perf: dict, fund_name: str) -> io.BytesIO | None:
    if not annual_perf: return None
    years, vals = [], []
    for y,v in sorted(annual_perf.items()):
        try:
            num = float(v.replace("%","").replace(",",".").strip())
            years.append(y); vals.append(num)
        except: pass
    if not years: return None
    fig, ax = plt.subplots(figsize=(6,2.2))
    colors = ["#2D9D78" if v>=0 else "#E05252" for v in vals]
    bars = ax.bar(years, vals, color=colors, edgecolor="white", linewidth=0.8, width=0.6)
    ax.axhline(0, color="#94A3B8", linewidth=0.8, linestyle="-")
    for bar,val in zip(bars,vals):
        ax.text(bar.get_x()+bar.get_width()/2,
                bar.get_height()+(0.3 if val>=0 else -0.9),
                f"{val:+.1f}%", ha="center", va="bottom", fontsize=7, fontweight="bold",
                color="#2D9D78" if val>=0 else "#E05252")
    ax.set_ylabel("%", fontsize=8, color="#64748B")
    ax.tick_params(axis="both",labelsize=8,colors="#475569")
    ax.spines[["top","right","left"]].set_visible(False)
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.0f%%"))
    ax.grid(axis="y",alpha=0.3,linestyle="--")
    fig.patch.set_facecolor("#FFFFFF"); plt.tight_layout(pad=0.5)
    buf=io.BytesIO(); fig.savefig(buf,format="png",dpi=130,bbox_inches="tight",facecolor="white")
    plt.close(fig); buf.seek(0); return buf


# ════════════════════════════════════════════════════════════
# PDF GENERATION
# ════════════════════════════════════════════════════════════

def generate_pdf(df: pd.DataFrame, wcol: str, profile: str,
                 ptf_name: str, fund_data: dict = None,
                 fida_df: pd.DataFrame = None) -> bytes:

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2.2*cm, bottomMargin=2.2*cm)

    ss = getSampleStyleSheet()
    def S(name,**kw): return ParagraphStyle(name,parent=ss["Normal"],**kw)

    T  = S("T",  fontName="Helvetica-Bold",  fontSize=22, textColor=rl_colors.HexColor("#0D1B2A"), spaceAfter=4,leading=28)
    EY = S("EY", fontName="Helvetica",       fontSize=8,  textColor=rl_colors.HexColor("#94A3B8"), spaceAfter=4,letterSpacing=1.5)
    SU = S("SU", fontName="Helvetica",       fontSize=10, textColor=rl_colors.HexColor("#64748B"), spaceAfter=4)
    SC = S("SC", fontName="Helvetica-Bold",  fontSize=11, textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=14,spaceAfter=8)
    BD = S("BD", fontName="Helvetica",       fontSize=8.5,textColor=rl_colors.HexColor("#1E293B"), leading=13)
    SM = S("SM", fontName="Helvetica",       fontSize=7.5,textColor=rl_colors.HexColor("#1E293B"), leading=11)
    FT = S("FT", fontName="Helvetica-Oblique",fontSize=7, textColor=rl_colors.HexColor("#94A3B8"), leading=10)
    FS = S("FS", fontName="Helvetica-Bold",  fontSize=13, textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=4,spaceAfter=2)
    FK = S("FK", fontName="Helvetica",       fontSize=7.5,textColor=rl_colors.HexColor("#64748B"), spaceAfter=2)
    # HDR: sempre bianco+grassetto — per celle intestazione su sfondo scuro
    # (TEXTCOLOR di TableStyle NON sovrascrive il colore dei Paragraph — serve lo stile dedicato)
    HDR= S("HDR",fontName="Helvetica-Bold",  fontSize=7.5,textColor=rl_colors.white, leading=11)

    story = []
    d_act = df[df[wcol]>0.001].copy()
    n_fondi = len(d_act)

    # ISIN da foglio FIDA (fallback per fondi senza URL FondiDoc)
    isin_map = {}
    if fida_df is not None and not fida_df.empty and "isin" in fida_df.columns:
        isin_map = {r["nome"]: str(r["isin"]).strip() for _, r in fida_df.iterrows()
                    if r.get("isin") and str(r.get("isin","")).strip()}
    w_az  = (d_act[wcol]*d_act["az_pct"]).sum()*100
    w_obb = (d_act[wcol]*d_act["obb_pct"]).sum()*100

    # ── ACCENT BAR ──────────────────────────────────────────
    story.append(Table([[""]], colWidths=[17*cm], rowHeights=[10],
        style=TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),rl_colors.HexColor("#0D1B2A")),
            ("LINEBELOW",(0,0),(-1,-1),3,rl_colors.HexColor("#C9A84C")),
        ])))
    story.append(Spacer(1,14))

    # ── TITLE BLOCK ─────────────────────────────────────────
    story.append(Paragraph("AZIMUT INVESTMENTS  ·  PORTAFOGLI MODELLO", EY))
    story.append(Spacer(1,4))
    story.append(Paragraph(f"Portafoglio {ptf_name}", T))
    story.append(Paragraph(
        f"{PROFILE_ICONS.get(profile,'●')} Profilo {profile.title()}  ·  "
        f"Dati al {datetime.date.today().strftime('%d %B %Y')}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=14))

    # ── KPI ─────────────────────────────────────────────────
    def kpi_cell(v,l):
        return Paragraph(f'<font size="18"><b>{v}</b></font><br/>'
                         f'<font size="8" color="#64748B">{l}</font>', BD)
    kpi = Table(
        [[kpi_cell(str(n_fondi),"Fondi"),kpi_cell(f"{w_az:.1f}%","Quota Azionaria"),
          kpi_cell(f"{w_obb:.1f}%","Quota Obbligazionaria"),
          kpi_cell(datetime.date.today().strftime("%m/%Y"),"Data Report")]],
        colWidths=[4.25*cm]*4
    )
    kpi.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.8,rl_colors.HexColor("#E2E8F0")),
        ("INNERGRID",(0,0),(-1,-1),0.8,rl_colors.HexColor("#E2E8F0")),
        ("BACKGROUND",(0,0),(-1,-1),rl_colors.HexColor("#F8FAFC")),
        ("PADDING",(0,0),(-1,-1),12),("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(kpi)

    # ── PIE CHARTS: allocazione fondi + asset allocation ────
    story.append(Paragraph("Allocazione del Portafoglio", SC))
    pie_buf   = _mpl_portfolio_pie(d_act, wcol, profile)
    macro_buf = _mpl_macro_pie(d_act, wcol)
    # Pie fondi — larghezza piena
    story.append(RLImage(pie_buf, width=15*cm, height=6.5*cm))
    # Pie asset allocation — stessa larghezza, stesso stile, sotto
    if macro_buf:
        story.append(Spacer(1, 8))
        story.append(RLImage(macro_buf, width=15*cm, height=5.5*cm))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════
    # PAGE 2: RENDIMENTI 1-3-5 ANNI
    # ════════════════════════════════════════════════════════
    story.append(Paragraph("AZIMUT INVESTMENTS  ·  PORTAFOGLI MODELLO", EY))
    story.append(Spacer(1,4))
    story.append(Paragraph("Tavola dei Rendimenti", T))
    story.append(Paragraph(
        f"Performance per fondo  ·  Profilo {profile.title()}  ·  "
        f"Fonte: FIDA FondiDoc  ·  Dati al {datetime.date.today().strftime('%d %B %Y')}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=12))

    # ── Helper: weighted average of a metric across all funds ──
    def ptf_wavg(keys_list):
        """
        Returns dict key→weighted-avg-string for each key in keys_list.
        Uses only funds that have a valid numeric value for ALL keys in the row.
        """
        totals  = {k: 0.0 for k in keys_list}
        cov_w   = {k: 0.0 for k in keys_list}   # weight covered for each metric

        for _, row in d_sorted.iterrows():
            fd  = (fund_data or {}).get(row["nome"], {})
            ana = fd.get("analysis", {})
            w   = row[wcol]
            for k in keys_list:
                raw = ana.get(k, "")
                try:
                    num = float(raw.replace("%","").replace(",",".").strip())
                    totals[k]  += num * w
                    cov_w[k]   += w
                except Exception:
                    pass

        result = {}
        for k in keys_list:
            if cov_w[k] > 0.01:
                avg = totals[k] / cov_w[k]
                result[k] = f"{avg:+.2f}%"
            else:
                result[k] = "N/D"
        return result

    # Paragraph style for portfolio summary row
    WH = S("WH", fontName="Helvetica-Bold", fontSize=8,
           textColor=rl_colors.white, leading=11)

    def pstyle_w(val):
        """White bold text coloured green/red for portfolio row."""
        try:
            v = float(val.replace("%","").replace(",",".").strip())
            c = "#7EFFC0" if v > 0 else ("#FFB3B3" if v < 0 else "#E2E8F0")
        except Exception:
            c = "#E2E8F0"
        return Paragraph(f'<font color="{c}"><b>{val}</b></font>', WH)

    def pstyle(val):
        try:
            v = float(val.replace("%","").replace(",","."))
            color = "#1A7A4A" if v>0 else ("#C0392B" if v<0 else "#475569")
            return f'<font color="{color}"><b>{val}</b></font>'
        except: return val

    # ── d_sorted deve essere definito PRIMA di chiamare ptf_wavg ──
    d_sorted = d_act.sort_values(wcol, ascending=False)

    # ── PERFORMANCE TABLE ────────────────────────────────────
    perf_keys = ["ytd","perf_1y","perf_3y","perf_5y","vol_1y","sharpe_1y"]
    ptf_p = ptf_wavg(perf_keys)

    perf_hdr = [Paragraph(f"<b>{t}</b>", HDR) for t in
                ["Fondo","Peso","YTD","1 Anno","3 Anni","5 Anni","Vol. 1A","Sharpe 1A"]]

    # Portfolio summary row (row index 1 — gold background)
    ptf_perf_row = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph(f"<b>100%</b>", WH),
        pstyle_w(ptf_p.get("ytd","N/D")),
        pstyle_w(ptf_p.get("perf_1y","N/D")),
        pstyle_w(ptf_p.get("perf_3y","N/D")),
        pstyle_w(ptf_p.get("perf_5y","N/D")),
        Paragraph(ptf_p.get("vol_1y","N/D"), WH),
        Paragraph(ptf_p.get("sharpe_1y","N/D"), WH),
    ]

    perf_rows = [perf_hdr, ptf_perf_row]

    for _, row in d_sorted.iterrows():
        fd  = (fund_data or {}).get(row["nome"], {})
        ana = fd.get("analysis", {})
        def gv(key): return ana.get(key,"N/D")
        perf_rows.append([
            Paragraph(row["nome"][:48], SM),
            Paragraph(f"<b>{row[wcol]*100:.1f}%</b>", SM),
            Paragraph(pstyle(gv("ytd")),     SM),
            Paragraph(pstyle(gv("perf_1y")), SM),
            Paragraph(pstyle(gv("perf_3y")), SM),
            Paragraph(pstyle(gv("perf_5y")), SM),
            Paragraph(gv("vol_1y"),           SM),
            Paragraph(gv("sharpe_1y"),        SM),
        ])

    perf_tbl = Table(perf_rows,
        colWidths=[5.2*cm,1.4*cm,1.4*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm],
        repeatRows=1)
    ts_perf = [
        ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),  # header
        ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
        # Portfolio summary row — gold/navy
        ("BACKGROUND",(0,1),(-1,1), rl_colors.HexColor("#1B4FBB")),
        ("LINEBELOW",(0,1),(-1,1),  2, rl_colors.HexColor("#C9A84C")),
        # Funds
        ("FONTSIZE",(0,0),(-1,-1),  8),
        ("PADDING",(0,0),(-1,-1),   5),
        ("ROWBACKGROUNDS",(0,2),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",(1,0),(-1,-1),     "CENTER"),
        ("VALIGN",(0,0),(-1,-1),    "MIDDLE"),
    ]
    perf_tbl.setStyle(TableStyle(ts_perf))
    # KeepTogether: la tabella rendimenti non viene spezzata su due pagine
    story.append(KeepTogether([perf_tbl]))
    story.append(Spacer(1,12))

    # ── RISK TABLE ───────────────────────────────────────────
    risk_keys = ["vol_1y","vol_3y","vol_5y","var_1y","sharpe_3y","sortino_1y"]
    ptf_r = ptf_wavg(risk_keys)

    risk_hdr = [Paragraph(f"<b>{t}</b>", HDR) for t in
                ["Fondo","Peso","Vol. 1A","Vol. 3A","Vol. 5A","VaR 1A","Sharpe 3A","Sortino 1A"]]

    ptf_risk_row = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph("<b>100%</b>", WH),
        Paragraph(ptf_r.get("vol_1y","N/D"),    WH),
        Paragraph(ptf_r.get("vol_3y","N/D"),    WH),
        Paragraph(ptf_r.get("vol_5y","N/D"),    WH),
        Paragraph(ptf_r.get("var_1y","N/D"),    WH),
        Paragraph(ptf_r.get("sharpe_3y","N/D"), WH),
        Paragraph(ptf_r.get("sortino_1y","N/D"),WH),
    ]

    risk_rows = [risk_hdr, ptf_risk_row]
    for _, row in d_sorted.iterrows():
        fd  = (fund_data or {}).get(row["nome"], {})
        ana = fd.get("analysis", {})
        def gv_r(k): return ana.get(k,"N/D")
        risk_rows.append([
            Paragraph(row["nome"][:48], SM),
            Paragraph(f"{row[wcol]*100:.1f}%", SM),
            Paragraph(gv_r("vol_1y"),    SM), Paragraph(gv_r("vol_3y"),    SM), Paragraph(gv_r("vol_5y"),    SM),
            Paragraph(gv_r("var_1y"),    SM), Paragraph(gv_r("sharpe_3y"), SM), Paragraph(gv_r("sortino_1y"),SM),
        ])

    risk_tbl = Table(risk_rows,
        colWidths=[5.2*cm,1.4*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm],
        repeatRows=1)
    ts_risk = [
        ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),
        ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
        ("BACKGROUND",(0,1),(-1,1), rl_colors.HexColor("#1B4FBB")),
        ("LINEBELOW",(0,1),(-1,1),  2, rl_colors.HexColor("#C9A84C")),
        ("FONTSIZE",(0,0),(-1,-1),  8),
        ("PADDING",(0,0),(-1,-1),   5),
        ("ROWBACKGROUNDS",(0,2),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",(1,0),(-1,-1),     "CENTER"),
        ("VALIGN",(0,0),(-1,-1),    "MIDDLE"),
    ]
    risk_tbl.setStyle(TableStyle(ts_risk))

    NOTE = S("NT", fontName="Helvetica-Oblique", fontSize=6.5,
             textColor=rl_colors.HexColor("#94A3B8"), leading=9)
    # KeepTogether: titolo + tabella rischio + nota nella stessa pagina
    story.append(KeepTogether([
        Paragraph("Metriche di Rischio", SC),
        risk_tbl,
        Spacer(1,6),
        Paragraph(
            "◆ I valori del Portafoglio sono medie ponderate per peso dei singoli fondi. "
            "La volatilità di portafoglio effettiva dipende anche dalle correlazioni tra i fondi. "
            "Dati forniti a titolo indicativo.", NOTE),
    ]))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════
    # PAGES 3+: SCHEDE SINGOLI FONDI
    # ════════════════════════════════════════════════════════
    story.append(Paragraph("AZIMUT INVESTMENTS  ·  PORTAFOGLI MODELLO", EY))
    story.append(Spacer(1,4))
    story.append(Paragraph("Schede Analitiche dei Fondi", T))
    story.append(Paragraph(
        f"Profilo {profile.title()}  ·  Fonte: FIDA FondiDoc  ·  {datetime.date.today().strftime('%d %B %Y')}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=10))

    for idx, (_, row) in enumerate(d_sorted.iterrows()):
        fd  = (fund_data or {}).get(row["nome"], {})
        ov  = fd.get("overview",  {})
        ana = fd.get("analysis",  {})

        def gv(k,src=ana,fallback="N/D"): return src.get(k,fallback)

        # Fund header block
        srri_str = f"SRRI {gv('srri',ov,'—')}/7" if gv('srri',ov) != "N/D" else ""
        nav_str  = f"NAV {gv('nav')} € ({gv('last_update')})" if gv('nav') != "N/D" else ""
        rating_s = f"FIDArating {gv('fida_rating',ov)}" if gv('fida_rating',ov) not in ("N/D","—") else ""

        # ── Intestazione fondo (3 righe × 1 colonna) ─────────
        meta_extra = "  ·  ".join(x for x in [srri_str, rating_s, nav_str] if x)

        # ISIN: estratto dall'URL FondiDoc oppure dal foglio FIDA
        isin = fd.get("isin", "") or isin_map.get(row["nome"], "")
        isin_str = f"  ·  ISIN: <b>{isin}</b>" if isin else ""

        hdr_rows = [
            [Paragraph(f"<b>{row['nome']}</b>", FS)],
            [Paragraph(f"Peso: <b>{row[wcol]*100:.1f}%</b>  ·  {row['categoria']}{isin_str}", FK)],
            [Paragraph(meta_extra or "—", FK)],
        ]

        hdr_tbl = Table(hdr_rows, colWidths=[17*cm])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), rl_colors.HexColor("#F0F4F9")),
            ("LEFTPADDING",(0,0),(-1,-1), 10),
            ("RIGHTPADDING",(0,0),(-1,-1), 10),
            ("TOPPADDING",(0,0),(-1,0), 10),
            ("BOTTOMPADDING",(0,-1),(-1,-1), 10),
            ("TOPPADDING",(0,1),(-1,-1), 2),
            ("BOTTOMPADDING",(0,0),(-1,-2), 2),
            ("LINEBELOW",(0,-1),(-1,-1), 2, rl_colors.HexColor("#C9A84C")),
        ]))

        # ── Tabella rendimenti fondo ──────────────────────────
        def pval(v):
            try:
                num = float(v.replace("%","").replace(",","."))
                c = "#1A7A4A" if num>0 else ("#C0392B" if num<0 else "#475569")
                return Paragraph(f'<font color="{c}"><b>{v}</b></font>', BD)
            except: return Paragraph(v, BD)

        perf_data = [
            [Paragraph("<b>Metrica</b>",HDR), Paragraph("<b>YTD</b>",HDR),
             Paragraph("<b>1 Anno</b>",HDR), Paragraph("<b>3 Anni</b>",HDR), Paragraph("<b>5 Anni</b>",HDR)],
            [Paragraph("Performance",SM),
             pval(gv("ytd")), pval(gv("perf_1y")), pval(gv("perf_3y")), pval(gv("perf_5y"))],
            [Paragraph("Volatilità",SM),
             Paragraph("—",SM), Paragraph(gv("vol_1y"),SM), Paragraph(gv("vol_3y"),SM), Paragraph(gv("vol_5y"),SM)],
            [Paragraph("VaR",SM),
             Paragraph("—",SM), Paragraph(gv("var_1y"),SM), Paragraph(gv("var_3y"),SM), Paragraph("—",SM)],
            [Paragraph("Sharpe",SM),
             Paragraph("—",SM), Paragraph("—",SM), Paragraph(gv("sharpe_3y"),SM), Paragraph(gv("sharpe_5y"),SM)],
            [Paragraph("Sortino",SM),
             Paragraph("—",SM), Paragraph(gv("sortino_1y"),SM), Paragraph("—",SM), Paragraph("—",SM)],
        ]
        perf_tbl2 = Table(perf_data, colWidths=[2.4*cm,1.5*cm,1.8*cm,1.8*cm,1.8*cm])
        perf_tbl2.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),
            ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
            ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),  7.5),
            ("PADDING",(0,0),(-1,-1),   4),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
            ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
            ("ALIGN",(1,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))

        # ── Dettagli fondo ───────────────────────────────────
        det_data = [
            [Paragraph("<b>Dettagli Fondo</b>", BD)],
            [Paragraph(f"Data avvio: {gv('start_date',ov,'—')}", SM)],
            [Paragraph(f"Distribuzione: {gv('income',ov,'—')}", SM)],
            [Paragraph(f"Categoria: {gv('cat_assog',ov,'—')}", SM)],
            [Paragraph(f"Gestione: {gv('mgmt_fee',ov,'—')}  |  Perf.: {gv('perf_fee',ov,'—')}", SM)],
            [Paragraph(f"Sottoscrizione: {gv('sub_fee',ov,'—')}", SM)],
            [Paragraph(f"<b>FIDArating:</b> {gv('fida_rating',ov,'—')}  |  Score: {gv('fida_score',ov,'—')}", SM)],
        ]
        det_tbl = Table([[d[0]] for d in det_data], colWidths=[7.3*cm])
        det_tbl.setStyle(TableStyle([
            ("PADDING",(0,0),(-1,-1), 3),
            ("TOPPADDING",(0,0),(-1,0), 6),
            ("LINEBELOW",(0,0),(0,0), 0.8, rl_colors.HexColor("#C9A84C")),
            ("BACKGROUND",(0,0),(0,-1), rl_colors.HexColor("#F8FAFC")),
        ]))

        mid_row = Table([[perf_tbl2, det_tbl]], colWidths=[9.7*cm, 7.3*cm])
        mid_row.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1), "TOP"),
            ("PADDING",(0,0),(-1,-1), 0),
            ("LEFTPADDING",(1,0),(1,-1), 10),
        ]))

        # ── Grafico rendimenti annuali ───────────────────────
        annual  = ana.get("annual_perf")
        bar_buf = _mpl_annual_bar(annual, row["nome"]) if annual else None

        # ── KeepTogether: tutta la scheda su stessa pagina ───
        card = [Spacer(1,6), hdr_tbl, Spacer(1,6), mid_row]
        if bar_buf:
            card += [Spacer(1,4),
                     Paragraph("<b>Performance Annuale (%)</b>", SM),
                     RLImage(bar_buf, width=14*cm, height=3.2*cm)]
        story.append(KeepTogether(card))

        # Separatore tra fondi
        if idx < len(d_sorted)-1:
            story.append(HRFlowable(width="100%", thickness=0.5,
                                    color=rl_colors.HexColor("#CBD5E1"),
                                    spaceBefore=8, spaceAfter=8))

    # ── FOOTER ─────────────────────────────────────────────
    story.append(PageBreak())
    story.append(HRFlowable(width="100%",thickness=0.5,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=8))
    story.append(Paragraph(
        "Documento generato automaticamente a scopo illustrativo. I dati di performance provengono da FIDA FondiDoc "
        "(fondidoc.it). I pesi indicati sono riferiti al portafoglio modello e non costituiscono offerta o consulenza "
        "di investimento. Rendimenti passati non garantiscono risultati futuri. © Azimut Group — uso interno.", FT))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════
# FREE PORTFOLIO BUILDER
# ════════════════════════════════════════════════════════════

def free_portfolio_ui(data):
    fida = data.get("FIDA", pd.DataFrame())
    if fida.empty:
        st.error("❌ Foglio FIDA non trovato."); return None

    az_lookup = {}
    for sname in ["PTF FULL","PTF SHORT"]:
        if sname in data and not data[sname].empty:
            for _,r in data[sname].iterrows(): az_lookup[r["nome"]] = r["az_pct"]

    if "free_ptf" not in st.session_state: st.session_state.free_ptf = []
    st.markdown('<p class="sec-title">Costruttore Portafoglio Libero</p>',unsafe_allow_html=True)
    options = fida.apply(lambda r: f"{r['nome']}  [{r['macro_cat']}]" if r['macro_cat']!='Altro' else r['nome'],axis=1).tolist()

    c1,c2,c3 = st.columns([3.5,1,0.8])
    with c1: sel = st.selectbox("Seleziona fondo:",options,key="sel_fund")
    with c2: w   = st.number_input("Peso %",0.1,100.0,10.0,0.5,key="sel_w")
    with c3:
        st.markdown("<br>",unsafe_allow_html=True)
        if st.button("➕ Aggiungi",use_container_width=True):
            fname = sel.split("  [")[0].strip()
            if any(f["nome"]==fname for f in st.session_state.free_ptf):
                st.toast("⚠️ Fondo già presente!",icon="⚠️")
            else:
                fd = fida[fida["nome"]==fname].iloc[0] if not fida[fida["nome"]==fname].empty else None
                mc = fd["macro_cat"] if fd is not None else "Altro"
                az = az_lookup.get(fname,DEFAULT_AZ.get(mc,0.5))
                st.session_state.free_ptf.append({"nome":fname,"categoria":fd["categoria"] if fd is not None else "","macro_cat":mc,"az_pct":az,"w_input":w})
                st.rerun()

    if not st.session_state.free_ptf: st.info("☝️ Aggiungi fondi."); return None
    st.markdown("**Fondi nel portafoglio:**")
    total_w = 0.0
    for i, fund in enumerate(st.session_state.free_ptf):
        r1,r2,r3 = st.columns([4,1.5,0.6])
        with r1: st.markdown(f"**{fund['nome']}** <span style='color:#64748b;font-size:.8rem;'>— {fund['macro_cat']}</span>",unsafe_allow_html=True)
        with r2:
            nw = st.number_input("Peso",0.0,100.0,float(fund["w_input"]),0.5,key=f"fw_{i}",label_visibility="collapsed")
            st.session_state.free_ptf[i]["w_input"] = nw
        with r3:
            if st.button("🗑️",key=f"del_{i}",use_container_width=True):
                st.session_state.free_ptf.pop(i); st.rerun()
        total_w += st.session_state.free_ptf[i]["w_input"]

    diff = abs(total_w-100.0)
    if diff<0.05: st.markdown(f'<div class="w-ok">✅ Somma pesi: <b>{total_w:.1f}%</b> — OK!</div>',unsafe_allow_html=True)
    else:         st.markdown(f'<div class="w-warn">⚠️ Somma pesi: <b>{total_w:.1f}%</b> (mancano {100-total_w:+.1f}%)</div>',unsafe_allow_html=True)
    if diff>0.5: return None

    records = [{"nome":f["nome"],"categoria":f["categoria"],"gruppo":f["macro_cat"],"macro_cat":f["macro_cat"],"az_pct":f["az_pct"],"obb_pct":1-f["az_pct"],"r_weight":f["w_input"]/100,"w_cons":f["w_input"]/100,"w_equil":f["w_input"]/100,"w_accr":f["w_input"]/100,"w_centrale":f["w_input"]/100} for f in st.session_state.free_ptf]
    df = pd.DataFrame(records)
    df["macro_cat"] = df["categoria"].apply(get_macro)
    df = assign_colors(df)
    for wc in ["w_cons","w_equil","w_accr","w_centrale"]:
        t = df[wc].sum(); df[wc] = df[wc]/t if t>0 else df[wc]
    return df


# ════════════════════════════════════════════════════════════
# MAIN APP
# ════════════════════════════════════════════════════════════

def main():
    with st.sidebar:
        st.markdown("""<div style='padding:1.4rem 0 .8rem 0;'><div style='font-size:.6rem;letter-spacing:.22em;color:#3a5a78;text-transform:uppercase;font-weight:700;'>Strumento di Analisi</div><div style='font-family:"Cormorant Garamond",serif;font-size:1.6rem;color:#dde8f5;font-weight:700;margin-top:4px;line-height:1.2;'>Portfolio<br>Analyzer</div><div style='width:32px;height:3px;background:#C9A84C;border-radius:2px;margin-top:10px;'></div></div>""", unsafe_allow_html=True)
        st.markdown("---")
        uploaded   = st.file_uploader("FILE EXCEL (PTF FULL + PTF SHORT + FIDA)", type=["xlsx","xls"])
        st.markdown("---")
        ptf_choice = st.radio("TIPO PORTAFOGLIO", ["📋  PTF FULL","⚡  PTF SHORT","🎨  LIBERO"])
        st.markdown("---")
        profile    = st.selectbox("PROFILO DI RISCHIO", PROFILES, index=3)
        if "LIBERO" not in ptf_choice and "free_ptf" in st.session_state:
            del st.session_state["free_ptf"]

    ptf_label = ptf_choice.split("  ",1)[1] if "  " in ptf_choice else ptf_choice
    st.markdown(f"""<div class="az-header"><div class="az-eyebrow">AZIMUT INVESTMENTS · PORTAFOGLI MODELLO</div><div class="az-rule"></div><div class="az-title">{ptf_label}</div><div class="az-meta">{PROFILE_ICONS.get(profile,'●')} Profilo {profile.title()} &nbsp;·&nbsp; {datetime.date.today().strftime('%d %B %Y')}</div></div>""",unsafe_allow_html=True)

    if uploaded is None:
        st.info("⬅️ **Carica il file Excel** nella barra laterale per iniziare.")
        return

    with st.spinner("⏳ Caricamento dati…"):
        file_bytes = uploaded.read()
        raw = parse_excel(file_bytes)

    if "LIBERO" in ptf_choice:
        df = free_portfolio_ui(raw)
    else:
        key = "PTF FULL" if "FULL" in ptf_choice else "PTF SHORT"
        if key not in raw or raw[key].empty:
            st.error(f"❌ Foglio '{key}' non trovato o vuoto."); return
        df = raw[key]

    if df is None or df.empty: return

    wcol   = PROFILE_W_COL[profile]
    df_act = df[df[wcol]>0.001].copy()

    # KPI row
    n_fondi = len(df_act)
    w_az    = (df_act[wcol]*df_act["az_pct"]).sum()*100
    w_obb   = (df_act[wcol]*df_act["obb_pct"]).sum()*100
    srri    = max(1,min(7,round(w_az/100*6+1)))

    c1,c2,c3,c4 = st.columns(4)
    for col,val,lbl,sub in [
        (c1,str(n_fondi),"Fondi in Portafoglio",f"{df_act['gruppo'].nunique()} gruppi"),
        (c2,f"{w_az:.1f}%","Quota Azionaria","ponderata per peso"),
        (c3,f"{w_obb:.1f}%","Quota Obbligazionaria","ponderata per peso"),
        (c4,f"{srri} / 7","Risk Score (SRRI proxy)","basato su quota azionaria"),
    ]:
        col.markdown(f'<div class="kpi"><div class="kpi-label">{lbl}</div><div class="kpi-value">{val}</div><div class="kpi-sub">{sub}</div></div>',unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)

    # Charts + fund list
    col_l,col_r = st.columns([1.15,0.85],gap="large")
    with col_l:
        st.markdown('<p class="sec-title">Allocazione per Fondo</p>',unsafe_allow_html=True)
        st.plotly_chart(make_fund_pie(df_act,wcol,profile),use_container_width=True,config={"displayModeBar":False})
        st.markdown('<p class="sec-title">Allocazione per Macro-Categoria</p>',unsafe_allow_html=True)
        st.plotly_chart(make_macro_bar(df_act,wcol),use_container_width=True,config={"displayModeBar":False})
    with col_r:
        st.markdown('<p class="sec-title">Composizione del Portafoglio</p>',unsafe_allow_html=True)
        for gruppo in df_act["gruppo"].unique():
            sub = df_act[df_act["gruppo"]==gruppo].sort_values(wcol,ascending=False)
            rows_html = "".join([f"""<div class="fund-row"><div class="fund-dot" style="background:{r['color']};"></div><div style="flex:1;min-width:0;"><div class="fund-name">{r['nome']}</div><div class="fund-cat">{r['categoria'][:48]+'…' if r['categoria'] and len(r['categoria'])>48 else (r['categoria'] or '—')}</div></div><div class="fund-pct">{r[wcol]*100:.1f}%</div></div>""" for _,r in sub.iterrows()])
            st.markdown(f'<div style="background:#fff;border:1px solid #e2e8f0;border-radius:10px;margin-bottom:12px;overflow:hidden;"><div class="fund-group-hdr">{gruppo}</div>{rows_html}</div>',unsafe_allow_html=True)

    # ── DOWNLOAD SECTION ─────────────────────────────────────
    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<p class="sec-title">Esporta Report PDF Completo</p>',unsafe_allow_html=True)

    fida_urls = raw.get("fida_urls", {})
    n_urls = sum(1 for nome in df_act["nome"].unique() if nome in fida_urls)

    col_btn,col_inf = st.columns([1,2])
    with col_inf:
        st.markdown(f"""<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:1rem 1.25rem;'><div style='font-size:.8rem;color:#1d4ed8;font-weight:600;margin-bottom:.4rem;'>Il report PDF (3 sezioni) contiene:</div><div style='font-size:.82rem;color:#1e40af;line-height:1.9;'>✓ <b>Pag. 1</b> — Grafico a torta + KPI di portafoglio<br>✓ <b>Pag. 2</b> — Tavola rendimenti YTD / 1A / 3A / 5A + Rischio<br>✓ <b>Pag. 3+</b> — Scheda analitica per ciascuno degli {n_fondi} fondi<br><span style='color:#3b82f6;'>🌐 Dati live da FondiDoc per {n_urls}/{n_fondi} fondi</span></div></div>""",unsafe_allow_html=True)

    with col_btn:
        if st.button("🔄  Carica Dati da FondiDoc + Genera PDF",use_container_width=True,
                     type="primary"):
            progress_bar = st.progress(0, text="Scarico dati FondiDoc…")
            def upd(v): progress_bar.progress(v, text=f"FondiDoc: {int(v*100)}% completato…")
            fund_data = fetch_all_fund_data(df_act, fida_urls, upd)
            progress_bar.progress(1.0, text="✅ Dati ricevuti — Genero PDF…")

            try:
                fida_df = raw.get("FIDA", pd.DataFrame())
                pdf_bytes = generate_pdf(df_act, wcol, profile, ptf_label, fund_data,
                                         fida_df=fida_df)
                fname = f"Azimut_{ptf_label.replace(' ','_')}_{profile}_{datetime.date.today().strftime('%Y%m%d')}.pdf"
                progress_bar.empty()
                st.download_button(
                    label="📥   Scarica Report PDF",
                    data=pdf_bytes, file_name=fname, mime="application/pdf",
                    use_container_width=True,
                )
                st.success(f"✅ PDF pronto — {n_fondi} fondi, {len(fund_data)} schede caricate da FondiDoc")
            except Exception as e:
                st.error(f"Errore PDF: {e}")

    st.markdown("<br><br>",unsafe_allow_html=True)


if __name__ == "__main__":
    main()
