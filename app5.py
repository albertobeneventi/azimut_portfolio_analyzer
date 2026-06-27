# ============================================================
# AZIMUT PORTFOLIO ANALYZER v2.0 — app.py
# Aggiornamento: Schede fondi + Rendimenti da FondiDoc FIDA
# ============================================================

import re
import json
from pathlib import Path
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import load_workbook
import zipfile
from xml.etree import ElementTree as ET
import io
import datetime
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
import matplotlib
try:
    matplotlib.use('Agg')
except Exception:
    pass
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, Image as RLImage,
    HRFlowable, PageBreak, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import cm

# ── PAGE CONFIG ─────────────────────────────────────────────
st.set_page_config(
    page_title="Azimut | Demo Analisi",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Playwright / Chromium bootstrap (Streamlit Cloud) ───────
@st.cache_resource(show_spinner=False)
def _ensure_playwright_chromium():
    """Installa il browser Chromium per Playwright (solo al primo avvio del server).
    Usa solo 'install chromium' senza --with-deps: le dipendenze di sistema
    sono gestite da packages.txt e --with-deps richiede sudo su Streamlit Cloud."""
    import subprocess, sys
    try:
        result = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            capture_output=True, timeout=180
        )
        print(f"[playwright install] rc={result.returncode}", flush=True)
        if result.stderr:
            print(f"[playwright install stderr] {result.stderr.decode('utf-8','replace')[:300]}", flush=True)
    except Exception as _e:
        print(f"[playwright install] errore: {_e}", flush=True)

_ensure_playwright_chromium()

# ── CONSTANTS ───────────────────────────────────────────────
GROUP_NAMES = {"ALLOCATION", "AZIONARI (LONG)", "BOND"}
COL_A, COL_B, COL_C, COL_G, COL_K, COL_O, COL_R = 0, 1, 2, 6, 10, 14, 17
PROFILES     = ["CONSERVATIVO", "EQUILIBRATO", "ACCRESCITIVO"]
PROFILE_ICONS = {"CONSERVATIVO":"🛡️","EQUILIBRATO":"⚖️","ACCRESCITIVO":"📈"}
PROFILE_W_COL = {"CONSERVATIVO":"w_cons","EQUILIBRATO":"w_equil","ACCRESCITIVO":"w_accr"}

MACRO_COLORS = {
    "Azionari":"#1B4FBB","Bilanciati/Flessibili":"#C9A84C",
    "Obbligazionari":"#2D9D78","Alternativi":"#8B5CF6","Altro":"#94A3B8",
}
SHADES = {
    "Azionari":             ["#0D3080","#1B4FBB","#2563EB","#3B82F6","#60A5FA","#93C5FD","#BFDBFE"],
    "Bilanciati/Flessibili":["#92650A","#B8860B","#C9A84C","#D4B572","#DFC298","#E9CEB4","#F3DACD"],
    "Obbligazionari":       ["#065F46","#14855F","#2D9D78","#34B98A","#6DE5BC","#9AEFD2","#C5F7E7"],
    "Alternativi":          ["#5B21B6","#7C3AED","#8B5CF6","#A78BFA","#C4B5FD","#DDD6FE"],
    "Altro":                ["#475569","#64748B","#94A3B8","#CBD5E1"],
}
DEFAULT_AZ = {"Azionari":0.92,"Bilanciati/Flessibili":0.50,"Obbligazionari":0.06,"Altro":0.50}
FONDIDOC_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "en-US,en;q=0.9,it;q=0.8",
}
# ISIN mancanti nel foglio FIDA dell'Excel.
# Aggiungere qui i fondi per cui l'Excel non riporta l'ISIN ma è noto.
MANUAL_ISIN_OVERRIDES: dict[str, str] = {
    "AZ Equity - Global Infrastructure":    "LU1621767737",
    "AZ Bond - CoCo Bonds (EUR-hedged)":    "LU2622195936",
    "AZ Bond - Convertible Bond":           "LU1422848470",
    "AZ F.1 All. Balanced FoF A Cap EUR":   "LU0346933400",  # ISIN classe A (FondiDoc URL)
    # Fondi GP (nome breve) senza ISIN nel foglio FIDA — necessari per Quantalys SUGGERITO
    "AZ Allocation - Global Balanced":      "LU0262757841",
    "AZ Bond - Aggregate Bond Euro":        "LU0194809330",
    "AZ Bond - Income Dynamic":             "LU0108019232",
    "AZ Equity - Global FoF":              "LU0262760399",
    "AZ Equity - Global Healthcare":        "LU2384058314",
}

# Override for one fund whose FIDA sheet hyperlink points to class B
# Aggiungere qui i fondi assenti da FondiDoc o con URL errati.
MANUAL_URL_OVERRIDES = {
    "AZ F.1 All. Balanced FoF A Cap EUR":
        "https://www.fondidoc.it/d/Index/AZPOA/LU0346933400_az-f1-allocation-balanced-fof-a-az-fund-cap-eur",
    # AZ Bond - Convertible Bond: pagina FondiDoc (classe A HU Cap EUR Hdg)
    "AZ Bond - Convertible Bond":
        "https://www.fondidoc.it/d/Index/AZF11671/LU1422848470_az-f1-bd-convertible-a-hu-cap-eur-hdg",
    # AZ Equity - Global Infrastructure: classe A Cap EUR
    "AZ Equity - Global Infrastructure":
        "https://www.fondidoc.it/d/Index/GIU_5366/LU1621767737_az-f1-eq-gl-infrastructure-a-az-fund-cap-eur",
    # AZ Bond - CoCo Bonds (EUR-hedged): classe A-HU Cap EUR Hdg
    "AZ Bond - CoCo Bonds (EUR-hedged)":
        "https://www.fondidoc.it/d/Index/FDFM4346/LU2622195936_az-f1-bd-coco-bonds-a-az-fund-cap-eur-hdg",
    # AZ Bond - CoCo Bonds (senza hedging — nome alternativo nei portafogli)
    "AZ Bond - CoCo Bonds":
        "https://www.fondidoc.it/d/Index/FIDFM857/LU2622195423_az-f1-bd-coco-bonds-a-az-fund-cap-eur",
    # ── Fondi GP presenti nel portafoglio suggerito ma assenti dall'Excel ────
    # Aggiunti per garantire link nel PDF e dati analisi nel tab Rischio
    "AZ F.1 All. Balanced Plus A Cap EUR":
        "https://www.fondidoc.it/d/Index/FDM03292/LU3081370317_az-f1-all-balanced-plus-a-cap-eur",
    "AZ F.1 Bd Total Return Bond A Cap EUR":
        "https://www.fondidoc.it/d/Index/AZF1BNDC/LU2168561392_az-f1-bd-total-return-bond-a-cap-eur",
    "AZ F.1 Eq. Global Emerging FoF A Cap EUR":
        "https://www.fondidoc.it/d/Index/AZ1278/LU1225037040_az-f1-eq-global-emerging-fof-a-cap-eur",
    "AZ F.1 Eq. Future Opportunities A Cap EUR":
        "https://www.fondidoc.it/d/Index/G1U16575/LU2332973481_az-f1-eq-future-opportunities-a-cap-eur",
    "AZ F.1 Eq. Global Growth A Cap EUR":
        "https://www.fondidoc.it/d/Index/AZFGGSA/LU0804221488_az-f1-eq-global-growth-a-cap-eur",
    "AZ F.1 Eq. World Minimum Volatility A Cap EUR":
        "https://www.fondidoc.it/d/Index/AZFABAAZ/LU0262757098_az-f1-eq-world-minimum-volatility-a-cap-eur",
    "AZ F.1 Eq. Global Value FoF A Cap EUR":
        "https://www.fondidoc.it/d/Index/FDFM4094/LU2622203623_az-f1-eq-global-value-fof-a-cap-eur",
    "AZ F.1 Eq. Small Cap Europe FoF A Cap EUR":
        "https://www.fondidoc.it/d/Index/AZSEUAAZ/LU0262753857_az-f1-eq-small-cap-europe-fof-a-cap-eur",
}

# ── GITHUB REPO PERSISTENCE ──────────────────────────────────────────────────
# Tutti i cache JSON vengono committati su GitHub via Contents API subito dopo
# il salvataggio su disco. Richiede il secret GITHUB_TOKEN (contents:write).
# In questo modo i dati sopravvivono ai riavvii di Streamlit Cloud e sono
# immediatamente disponibili a tutti gli utenti dell'app.
_REPO   = "albertobeneventi/azimut_portfolio_analyzer"
_BRANCH = "master"


def _push_json_to_repo(payload: dict, repo_path: str, commit_msg: str) -> bool:
    """Commita un JSON su GitHub via Contents API (crea o aggiorna il file).

    Silenzioso: non mostra UI. Restituisce True se ok, False altrimenti.
    """
    import base64
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        if not token:
            return False
        headers = {
            "Authorization":    f"token {token}",
            "Accept":           "application/vnd.github.v3+json",
            "X-GitHub-Api-Version": "2022-11-28",
        }
        url = f"https://api.github.com/repos/{_REPO}/contents/{repo_path}"
        r_get = requests.get(url, headers=headers,
                             params={"ref": _BRANCH}, timeout=10)
        sha = r_get.json().get("sha") if r_get.status_code == 200 else None
        body: dict = {
            "message": commit_msg,
            "content": base64.b64encode(
                json.dumps(payload, ensure_ascii=False, indent=2,
                           default=str).encode("utf-8")
            ).decode(),
            "branch": _BRANCH,
        }
        if sha:
            body["sha"] = sha
        r_put = requests.put(url, json=body, headers=headers, timeout=20)
        return r_put.status_code in (200, 201)
    except Exception:
        return False


# ── FUND DATA CACHE ──────────────────────────────────────────────────────────
CACHE_FILE = Path("data/fund_cache.json")

def load_fund_cache() -> tuple:
    """Load cached FondiDoc data. Returns (fund_data_dict, last_updated_str)."""
    try:
        if CACHE_FILE.exists():
            # utf-8-sig strips BOM se presente (file creato su Windows)
            payload = json.loads(CACHE_FILE.read_text(encoding="utf-8-sig"))
            return payload.get("fund_data", {}), payload.get("last_updated", "")
    except Exception:
        pass
    return {}, ""

def save_fund_cache(fund_data: dict):
    """Persist fund data to fund_cache.json e lo pusha su GitHub."""
    try:
        CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload = {}
        if CACHE_FILE.exists():
            try:
                payload = json.loads(CACHE_FILE.read_text(encoding="utf-8-sig"))
            except Exception:
                payload = {}
        payload["last_updated"] = datetime.date.today().isoformat()
        payload["fund_data"] = fund_data
        CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                              encoding="utf-8")
        _push_json_to_repo(payload, "data/fund_cache.json",
                           f"auto: aggiorna fund_cache {datetime.date.today().isoformat()}")
    except Exception:
        pass


def load_ms_cache() -> dict:
    """Load cached Morningstar ratings. Returns {fund_name: {ms_rating, fo_url}}."""
    try:
        if CACHE_FILE.exists():
            payload = json.loads(CACHE_FILE.read_text(encoding="utf-8-sig"))
            return payload.get("ms_data", {})
    except Exception:
        pass
    return {}


def save_ms_cache(ms_data: dict):
    """Persist Morningstar ratings in fund_cache.json e lo pusha su GitHub."""
    try:
        CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload = {}
        if CACHE_FILE.exists():
            try:
                payload = json.loads(CACHE_FILE.read_text(encoding="utf-8-sig"))
            except Exception:
                payload = {}
        payload["ms_data"] = ms_data
        CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                              encoding="utf-8")
        _push_json_to_repo(payload, "data/fund_cache.json",
                           f"auto: aggiorna ms_cache {datetime.date.today().isoformat()}")
    except Exception:
        pass


# ── EXCEL / GP PERSISTENT CACHE ──────────────────────────────────────────────
# Persiste i dati parsed tra sessioni diverse: zero upload nel normale utilizzo,
# ricaricamento solo quando ci sono aggiornamenti (mensile Excel, trimestrale GP).
EXCEL_CACHE_FILE = Path("data/excel_cache.json")
GP_CACHE_FILE    = Path("data/gp_cache.json")
FP_CACHE_FILE    = Path("data/fp_factbook_cache.json")
SAVED_PTF_FILE   = Path("data/saved_portfolios.json")


def _df_to_records(df: pd.DataFrame) -> list:
    """Serializza un DataFrame in una lista di dizionari (JSON-safe)."""
    if df is None or df.empty:
        return []
    return df.where(pd.notna(df), None).to_dict(orient="records")


def _records_to_df(records: list) -> pd.DataFrame:
    """Ricostruisce un DataFrame da una lista di dizionari."""
    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def load_excel_cache() -> tuple:
    """Carica i dati Excel salvati su disco.

    Returns (raw_dict, last_updated_str) oppure (None, "") se assente/corrotto.
    raw_dict ha la stessa struttura di parse_excel():
      {"PTF FULL": DataFrame, "PTF SHORT": DataFrame,
       "FIDA": DataFrame, "fida_urls": dict}
    """
    try:
        if EXCEL_CACHE_FILE.exists() and EXCEL_CACHE_FILE.stat().st_size > 10:
            payload = json.loads(EXCEL_CACHE_FILE.read_text(encoding="utf-8-sig"))
            raw: dict = {}
            for sname in ("PTF FULL", "PTF SHORT"):
                recs = payload.get(sname, [])
                if recs:
                    raw[sname] = _records_to_df(recs)
            fida_recs = payload.get("FIDA", [])
            if fida_recs:
                raw["FIDA"] = _records_to_df(fida_recs)
            raw["fida_urls"] = payload.get("fida_urls") or {}
            # Valida: almeno uno dei due sheet deve essere non-vuoto
            if raw.get("PTF FULL") is not None and not raw["PTF FULL"].empty:
                return raw, payload.get("last_updated", "")
    except Exception:
        pass
    return None, ""


def save_excel_cache(raw: dict):
    """Salva i dati Excel su disco e li pusha su GitHub."""
    try:
        EXCEL_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload: dict = {"last_updated": datetime.date.today().isoformat()}
        for sname in ("PTF FULL", "PTF SHORT"):
            df = raw.get(sname)
            payload[sname] = _df_to_records(df) if df is not None else []
        fida = raw.get("FIDA")
        payload["FIDA"] = _df_to_records(fida) if fida is not None else []
        payload["fida_urls"] = raw.get("fida_urls") or {}
        EXCEL_CACHE_FILE.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        _push_json_to_repo(payload, "data/excel_cache.json",
                           f"auto: aggiorna excel_cache {datetime.date.today().isoformat()}")
    except Exception:
        pass


def load_gp_cache() -> tuple:
    """Carica i dati Global Perspectives salvati su disco.

    Returns (gp_data, filename_str, last_updated_str) oppure (None, "", "").
    """
    try:
        if GP_CACHE_FILE.exists() and GP_CACHE_FILE.stat().st_size > 10:
            payload = json.loads(GP_CACHE_FILE.read_text(encoding="utf-8-sig"))
            gp_data = payload.get("gp_data")
            if gp_data and isinstance(gp_data, dict) and len(gp_data) >= 3:
                return (gp_data,
                        payload.get("filename", ""),
                        payload.get("last_updated", ""))
    except Exception:
        pass
    return None, "", ""


def save_gp_cache(gp_data: dict, filename: str = ""):
    """Salva i dati Global Perspectives su disco e li pusha su GitHub."""
    try:
        GP_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "last_updated": datetime.date.today().isoformat(),
            "filename":     filename,
            "gp_data":      gp_data,
        }
        GP_CACHE_FILE.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        _push_json_to_repo(payload, "data/gp_cache.json",
                           f"auto: aggiorna gp_cache {datetime.date.today().isoformat()}")
    except Exception:
        pass


# ── FONDI PENSIONE FACTBOOK CACHE ────────────────────────────────────────────

def load_fp_cache() -> dict:
    """Carica i dati Factbook Fondi Pensione.
    Returns dict {fund_name: {ytd, perf_1y, perf_3y, perf_5y, ...}, "_ref_date": "..."}
    oppure {} se assente."""
    try:
        if FP_CACHE_FILE.exists() and FP_CACHE_FILE.stat().st_size > 10:
            return json.loads(FP_CACHE_FILE.read_text(encoding="utf-8-sig"))
    except Exception:
        pass
    return {}


def save_fp_cache(fp_data: dict):
    """Salva i dati Factbook Fondi Pensione su disco e li pusha su GitHub."""
    try:
        FP_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        FP_CACHE_FILE.write_text(
            json.dumps(fp_data, ensure_ascii=False, indent=2), encoding="utf-8")
        _push_json_to_repo(fp_data, "data/fp_factbook_cache.json",
                           f"auto: aggiorna fp_cache {datetime.date.today().isoformat()}")
    except Exception:
        pass


# ── PORTAFOGLI LIBERI SALVATI ─────────────────────────────────────────────────

def load_saved_portfolios() -> dict:
    """Carica i portafogli liberi salvati.
    Returns {nome: {"date": "YYYY-MM-DD", "fondi": [{"nome": ..., "peso": ...}, ...]}}"""
    try:
        if SAVED_PTF_FILE.exists() and SAVED_PTF_FILE.stat().st_size > 10:
            return json.loads(SAVED_PTF_FILE.read_text(encoding="utf-8-sig"))
    except Exception:
        pass
    return {}


def save_portfolio(name: str, fondi: list):
    """Salva un portafoglio libero con nome. fondi = [{"nome": ..., "peso": float}]"""
    try:
        saved = load_saved_portfolios()
        saved[name] = {
            "date":  datetime.date.today().isoformat(),
            "fondi": fondi,
        }
        SAVED_PTF_FILE.parent.mkdir(parents=True, exist_ok=True)
        SAVED_PTF_FILE.write_text(
            json.dumps(saved, ensure_ascii=False, indent=2), encoding="utf-8")
        _push_json_to_repo(saved, "data/saved_portfolios.json",
                           f"auto: salva portafoglio '{name}' {datetime.date.today().isoformat()}")
    except Exception:
        pass


def delete_portfolio(name: str):
    """Elimina un portafoglio salvato."""
    try:
        saved = load_saved_portfolios()
        if name in saved:
            del saved[name]
            SAVED_PTF_FILE.write_text(
                json.dumps(saved, ensure_ascii=False, indent=2), encoding="utf-8")
            _push_json_to_repo(saved, "data/saved_portfolios.json",
                               f"auto: elimina portafoglio '{name}' {datetime.date.today().isoformat()}")
    except Exception:
        pass


# ── UNP/IUNP CATALOG (Catalogo Prodotti&Servizi Azimut, settembre 2025) ──────
# Fonte: DETTAGLIO AZ FUND — valori %UNP e %IUNP36
# Estratto da pdftotext su tutte le pagine del PDF (93 voci).
# Nomi fondi esattamente come nel PDF; _normalize_for_unp() gestisce la
# corrispondenza con i nomi abbreviati dell'Excel (es. "AZ F.1 Bd. ...").
UNP_CATALOG = {
    # ── AZ Allocation ─────────────────────────────────────────────────────────
    "AZ Allocation - Asset Timing 2026":                               (2.29, 1.14),
    "AZ Allocation - Asset Timing 2028":                               (2.29, 1.14),
    "AZ Allocation - Balanced Brave":                                  (1.89, 0.94),
    "AZ Allocation - Balanced FoF":                                    (1.89, 0.94),
    "AZ Allocation - Balanced Plus":                                   (1.89, 0.94),
    "AZ Allocation - Escalator 2026":                                  (2.29, 1.14),
    "AZ Allocation - Escalator 2028":                                  (2.29, 1.14),
    "AZ Allocation - Escalator 2030":                                  (2.29, 1.14),
    "AZ Allocation - Flexible Equity":                                 (1.92, 0.96),
    "AZ Allocation - Global Aggressive":                               (1.89, 0.94),
    "AZ Allocation - Global Balanced":                                 (1.80, 0.90),
    "AZ Allocation - Global Conservative":                             (1.61, 0.80),
    "AZ Allocation - Global Conservative (Classe C)":                  (0.95, 0.47),
    "AZ Allocation - Italian Trend":                                   (2.01, 1.00),
    "AZ Allocation - Life Plan 2040":                                  (2.51, 1.25),
    "AZ Allocation - PIR Italian Excellence 70%":                      (1.89, 0.94),
    "AZ Allocation - Potential Income Upside 2030":                    (1.67, 0.83),
    "AZ Allocation - Risk Parity Factors":                             (1.89, 0.94),
    "AZ Allocation - Trend":                                           (2.01, 1.00),
    "AZ Allocation - Turkey":                                          (1.89, 0.94),
    # ── AZ Alternative ────────────────────────────────────────────────────────
    "AZ Alternative - Capital Enhanced":                               (0.51, 0.25),
    "AZ Alternative - Commodity":                                      (1.81, 0.90),
    # ── AZ Bond ───────────────────────────────────────────────────────────────
    "AZ Bond - Aggregate Bond Euro":                                   (1.10, 0.55),
    "AZ Bond - Asian Bond":                                            (1.38, 0.69),
    "AZ Bond - Bond Value":                                            (1.53, 0.76),
    "AZ Bond - COCO Bonds":                                            (1.53, 0.76),
    "AZ Bond - Convertible":                                           (1.62, 0.81),
    "AZ Bond - Enhanced Yield":                                        (0.35, 0.17),
    "AZ Bond - Euro Corporate":                                        (1.23, 0.61),
    "AZ Bond - Frontier Markets Debt":                                 (1.62, 0.81),
    "AZ Bond - Global Macro Bond":                                     (1.29, 0.64),
    "AZ Bond - High Income FoF":                                       (1.62, 0.81),
    "AZ Bond - High Yield":                                            (1.47, 0.73),
    "AZ Bond - High Yield Target 2028 Climate Transition":             (1.53, 0.76),
    "AZ Bond - High Yield Target 2028 Climate Transition (Classe C)":  (0.52, 0.26),
    "AZ Bond - Income Dynamic":                                        (0.99, 0.49),
    "AZ Bond - International FoF":                                     (1.62, 0.81),
    "AZ Bond - Latin America Bonds":                                   (1.53, 0.76),
    "AZ Bond - Patriot":                                               (1.36, 0.68),
    "AZ Bond - Renminbi Opportunities":                                (1.23, 0.61),
    "AZ Bond - Short Term Investment Grade Climate Transition":        (1.53, 0.76),
    "AZ Bond - Short Term Investment Grade Climate Transition (Classe C)": (0.43, 0.21),
    "AZ Bond - Sustainable Hybrid":                                    (1.46, 0.73),
    "AZ Bond - Target 2025":                                           (1.11, 0.55),
    "AZ Bond - Target 2026":                                           (1.11, 0.55),
    "AZ Bond - Target 2028":                                           (1.11, 0.55),
    "AZ Bond - Target 2029":                                           (1.11, 0.55),
    "AZ Bond - Target 2029 USD":                                       (1.23, 0.61),
    "AZ Bond - Target 2031":                                           (1.11, 0.55),
    "AZ Bond - Total Return Bond":                                     (1.55, 0.77),
    "AZ Bond - US Dollar Aggregate":                                   (1.24, 0.62),
    # ── AZ Equity ─────────────────────────────────────────────────────────────
    "AZ Equity - Al Mal Mena":                                         (2.51, 1.25),
    "AZ Equity - American Opportunities":                              (2.19, 1.09),
    "AZ Equity - ASEAN Countries":                                     (2.19, 1.09),
    "AZ Equity - Best Value":                                          (2.09, 1.04),
    "AZ Equity - Biotechnology":                                       (2.51, 1.25),
    "AZ Equity - Borletti Global Lifestyle":                           (2.30, 1.15),
    "AZ Equity - Brazil Trend":                                        (2.19, 1.09),
    "AZ Equity - China":                                               (2.19, 1.09),
    "AZ Equity - Egypt":                                               (2.51, 1.25),
    "AZ Equity - Emerging Asia FoF":                                   (2.51, 1.25),
    "AZ Equity - Emerging Markets Technology":                         (2.51, 1.25),
    "AZ Equity - Escalator":                                           (2.29, 1.14),
    "AZ Equity - Europe":                                              (2.19, 1.09),
    "AZ Equity - Food & Agriculture":                                  (2.40, 1.20),
    "AZ Equity - Global Dividend":                                     (2.51, 1.25),
    "AZ Equity - Global Emerging FoF":                                 (2.51, 1.25),
    "AZ Equity - Global ESG":                                          (2.51, 1.25),
    "AZ Equity - Global FoF":                                          (2.51, 1.25),
    "AZ Equity - Global Growth":                                       (2.40, 1.20),
    "AZ Equity - Global Healthcare":                                   (2.40, 1.20),
    "AZ Equity - Global Infrastructure":                               (2.26, 1.13),
    "AZ Equity - Global Quality":                                      (2.18, 1.09),
    "AZ Equity - Global Value FoF":                                    (2.51, 1.25),
    "AZ Equity - Industrial Revolution 4.0":                           (2.51, 1.25),
    "AZ Equity - Japan":                                               (2.20, 1.10),
    "AZ Equity - Mexico":                                              (2.51, 1.25),
    "AZ Equity - Momentum":                                            (2.19, 1.09),
    "AZ Equity - Small Cap Europe FoF":                                (2.51, 1.25),
    "AZ Equity - Special Needs & Inclusion":                           (2.51, 1.25),
    "AZ Equity - Water & Renewable Resources":                         (2.40, 1.20),
    "AZ Equity - World Minimum Volatility":                            (2.19, 1.09),
    # ── AZ Islamic ────────────────────────────────────────────────────────────
    "AZ Islamic - Global Sukuk":                                       (1.36, 0.68),
    # ── Azimut Thematic Fund ──────────────────────────────────────────────────
    "Azimut Thematic Fund - AZ Allocation - Global Goals":             (2.50, 1.25),
    "Azimut Thematic Fund - AZ Equity - New Generation":               (2.79, 1.39),
    "Azimut Thematic Fund - AZ Equity - Space":                        (2.79, 1.39),
    # ── Fondi Economia Reale (sezione separata nel catalogo) ──────────────────
    "AZ Allocation - Italian Long-Term Opp.":                          (2.73, 1.36),
    "AZ Allocation - Long Term Credit Opp.":                           (1.94, 0.97),
    "AZ Allocation - Long-Term Equity Opp.":                           (2.73, 1.36),
    "AZ Bond - ABS":                                                   (1.23, 0.61),
    "AZ Equity - Future Opportunities":                                (2.51, 1.25),
}

# CSS is injected inside main() to keep it within the error-handler scope.


# ════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════

def get_macro(cat: str) -> str:
    if not cat or cat == "-": return "Altro"
    c = cat.lower()
    if "azionari" in c or "equity" in c: return "Azionari"
    if any(x in c for x in ["obbligazionari","bond","credit","debt","sukuk","reddito"]): return "Obbligazionari"
    if any(x in c for x in ["bilanciati","allocation","flessibili","balanced","flexible","prudenti","moderati"]): return "Bilanciati/Flessibili"
    if any(x in c for x in ["alternativi","alternative","commodity"]): return "Alternativi"
    return "Altro"


def assign_colors(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    counter = {k: 0 for k in SHADES}
    colors = []
    for _, row in df.iterrows():
        mc = row.get("macro_cat", "Altro")
        shades = SHADES.get(mc, SHADES["Altro"])
        colors.append(shades[counter.get(mc, 0) % len(shades)])
        counter[mc] = counter.get(mc, 0) + 1
    df["color"] = colors
    return df


def pct_color(val_str: str) -> str:
    """Returns 'pos', 'neg', or 'neu' based on value sign"""
    try:
        v = float(val_str.replace("%","").replace(",",".").strip())
        return "pos" if v > 0 else ("neg" if v < 0 else "neu")
    except:
        return "neu"


# ── UNP/IUNP lookup helpers ──────────────────────────────────

def _normalize_for_unp(name: str) -> str:
    """Normalise a fund name for UNP/factbook catalog lookup.

    Works across all three name formats:
    - Excel  : "AZ F.1 Bd. Global Macro Bond A Cap EUR"  (abbreviation + share class)
    - Excel  : "AZ F.1 Bd Global Macro Bond A Cap EUR"   (abbreviation without period)
    - PDF    : "AZ Bond - Global Macro Bond"              (full name with dash)
    - Web    : "AZ BOND - GLOBAL MACRO BOND"              (full name, ALL CAPS)

    All four normalise to → "az bond global macro bond"
    """
    n = name.strip()
    # ── 1. Expand AZ F.1 abbreviations (period OPTIONAL after family code) ────
    n = re.sub(r'AZ\s+F\.1\s+All\.?\s*',  'AZ Allocation ',  n, flags=re.IGNORECASE)
    n = re.sub(r'AZ\s+F\.1\s+Eq\.?\s*',   'AZ Equity ',      n, flags=re.IGNORECASE)
    n = re.sub(r'AZ\s+F\.1\s+Bd\.?\s*',   'AZ Bond ',        n, flags=re.IGNORECASE)
    n = re.sub(r'AZ\s+F\.1\s+Alt\.?\s*',  'AZ Alternative ', n, flags=re.IGNORECASE)
    n = re.sub(r'AZ\s+F\.1\s+Isl\.?\s*',  'AZ Islamic ',     n, flags=re.IGNORECASE)
    # ── 2. Strip "(Classe C)" / "(classe c)" variants ─────────────────────────
    n = re.sub(r'\s*\(clas[se]+\s+c\)\s*$', '', n, flags=re.IGNORECASE)
    # ── 3. Strip share-class suffix: " A Cap EUR", " B Acc USD", "A-HU Cap EUR Hdg" etc.
    n = re.sub(r'\s+[A-Z][-\w]*\s+(Cap|Acc|Dis|Inc)\s+\w{3}(\s+Hdg)?\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s+[A-Z]\s+(Cap|Acc|Dis|Inc)\s*$',                          '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s+Cap\s+\w{3}\s*$',                                         '', n, flags=re.IGNORECASE)
    # ── 4. Normalise dashes, ampersands, dots and remaining punctuation ────────
    n = re.sub(r'[-–—]', ' ', n)
    n = re.sub(r'[&]',   ' ', n)
    n = re.sub(r'[^\w\s]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip().lower()
    return n


_UNP_CATALOG_NORMALIZED: dict = {
    _normalize_for_unp(k): v for k, v in UNP_CATALOG.items()
}

# Alias table: normalised Excel name → normalised catalog name
# Needed when the Excel name differs semantically from the catalog (e.g. "Hybrids" vs
# "Sustainable Hybrid") or when slight wording changes prevent automatic substring match.
_FUND_ALIASES: dict = {
    "az bond hybrids":              "az bond sustainable hybrid",
    "az bond sustainable hybrids":  "az bond sustainable hybrid",
    # Long-Term Opp variants (catalog uses abbreviated "Opp.")
    "az allocation long term credit opportunities":  "az allocation long term credit opp",
    "az allocation long term equity opportunities":  "az allocation long term equity opp",
    "az allocation italian long term opportunities": "az allocation italian long term opp",
    "az allocation italian long-term opportunities": "az allocation italian long term opp",
}

# ── Credit rating numeric scale (AAA=1 … D=22) for weighted averages ─────────
RATING_SCALE: dict = {
    'AAA': 1,  'AA+': 2,  'AA': 3,  'AA-': 4,
    'A+':  5,  'A':   6,  'A-': 7,
    'BBB+':8,  'BBB': 9,  'BBB-':10,
    'BB+':11,  'BB': 12,  'BB-': 13,
    'B+': 14,  'B':  15,  'B-':  16,
    'CCC+':17, 'CCC':18,  'CCC-':19,
    'CC': 20,  'C':  21,  'D':   22,
}
RATING_INVERSE: dict = {v: k for k, v in RATING_SCALE.items()}


def lookup_unp(fund_name: str):
    """Return (unp_pct, iunp36_pct) for a fund, or (None, None) if not found."""
    norm = _normalize_for_unp(fund_name)
    norm = _FUND_ALIASES.get(norm, norm)          # apply alias if any
    # 1. exact match
    if norm in _UNP_CATALOG_NORMALIZED:
        return _UNP_CATALOG_NORMALIZED[norm]
    # 2. longest substring match (catalog key inside fund name or vice versa)
    best, best_len = None, 0
    for cat_key, val in _UNP_CATALOG_NORMALIZED.items():
        if cat_key in norm or norm in cat_key:
            if len(cat_key) > best_len:
                best, best_len = val, len(cat_key)
    return best if best is not None else (None, None)


# ════════════════════════════════════════════════════════════
# DATA PARSING
# ════════════════════════════════════════════════════════════

def _parse_excel_impl(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    out = {}
    for sname in ["PTF FULL", "PTF SHORT"]:
        if sname in wb.sheetnames:
            out[sname] = _parse_ptf(wb, sname)
    if "FIDA" in wb.sheetnames:
        out["FIDA"] = _parse_fida(wb)
    wb.close()
    out["fida_urls"] = extract_fida_urls(file_bytes)
    return out

# Apply cache decorator safely at module load time — if it fails (Streamlit
# version mismatch or hasher bug), fall back to uncached version.
try:
    parse_excel = st.cache_data(show_spinner=False)(_parse_excel_impl)
except Exception:
    parse_excel = _parse_excel_impl


def parse_factbook(pdf_bytes: bytes) -> dict:
    """Parse the AZ Investments factbook PDF performance summary tables.

    Returns {normalized_fund_name: {ytd, perf_1y, perf_3y, perf_5y}}

    Summary table columns: Fund | AUM | 1M | 3M | 6M | 12M | 24M | 36M | 60M | YTD
    Risk metrics (vol, Sharpe, Sortino) are NOT in the factbook — those still
    come from FondiDoc.
    """
    try:
        import pdfplumber
    except ImportError:
        return {}

    result: dict = {}
    _metrics: dict = {}   # {norm_fund: {credit_rating, duration, ytm}}

    def _to_pct(s):
        """Convert a factbook value string to a signed '%' string, or None."""
        if not s:
            return None
        s = str(s).strip()
        if s in ('-', 'n.d.', 'N/D', '', 'None', 'n/d', 'nd'):
            return None
        s = s.replace('−', '-').replace('–', '-').replace('—', '-')
        s = s.replace(',', '.').replace('%', '').strip()
        try:
            return f"{float(s):+.2f}%"
        except Exception:
            return None

    def _annualize(pct_str, years: float):
        """Convert a cumulative return string to annualized (CAGR)."""
        if not pct_str:
            return None
        try:
            v = float(pct_str.replace('%','').replace('+','').strip()) / 100
            ann = ((1 + v) ** (1 / years) - 1) * 100
            return f"{ann:+.2f}%"
        except Exception:
            return None

    def _plausible(pct_str, max_abs: float = 80.0):
        """Return True if pct_str is None OR in [-max_abs, +max_abs]."""
        if pct_str is None:
            return True
        try:
            return abs(float(pct_str.replace('%','').replace('+',''))) <= max_abs
        except Exception:
            return True  # keep on parse error

    def _get_fi_from_src(src: str):
        """Extract Credit Rating, YTM, Duration from one pdfplumber text source.
        Defined once here (not inside the page loop) to avoid repeated closure
        creation and potential caching issues.
        """
        _lines = [ln.strip() for ln in src.split('\n') if ln.strip()]
        _txt   = '\n'.join(_lines)
        _r = _y = _d = None
        _RATING_PAT = (
            r'(AAA|AA[+\-]|AA|A[+\-]|A'
            r'|BBB[+\-]|BBB|BB[+\-]|BB|B[+\-]|B'
            r'|CCC[+\-]|CCC|CC|C|D)'
            r'(?=[^A-Za-z]|$)')
        # Credit Rating – inline ("...Credit Rating medio  BB+")
        #              or next-line ("...Credit Rating medio\n  BB+")
        _m = re.search(
            r'Credit\s+Rating[^\n]{0,60}?' + _RATING_PAT,
            _txt, re.IGNORECASE | re.MULTILINE)
        if not _m:
            _m = re.search(
                r'Credit\s+Rating[^\n]*\n\s*' + _RATING_PAT,
                _txt, re.IGNORECASE)
        if _m and _m.group(1).upper() in RATING_SCALE:
            _r = _m.group(1).upper()
        # YTM – inline or next-line
        _m = re.search(
            r'Yield\s+To\s+Maturity[^\n]{0,60}?([\d]+[,\.][\d]+)\s*%',
            _txt, re.IGNORECASE | re.MULTILINE)
        if not _m:
            _m = re.search(
                r'Yield\s+To\s+Maturity[^\n]*\n\s*([\d]+[,\.][\d]+)\s*%',
                _txt, re.IGNORECASE)
        if _m:
            try:
                _y = round(float(_m.group(1).replace(',', '.')), 2)
            except Exception:
                pass
        # Duration – scan up to 10 lines after "Portfolio Duration" label
        _dn = next(
            (i for i, ln in enumerate(_lines)
             if re.search(r'portfolio\s+duration', ln, re.IGNORECASE)), None)
        if _dn is not None:
            for _dl in _lines[_dn:_dn + 10]:
                if re.search(r'portfolio\s+duration', _dl, re.IGNORECASE):
                    _dm = re.search(
                        r'Duration\D{0,20}([\d]+[,\.][\d]+)(?!\s*%)',
                        _dl, re.IGNORECASE)
                else:
                    _dm = re.search(r'^([\d]+[,\.][\d]+)\s*$', _dl)
                    if not _dm:
                        _dm = re.search(
                            r'(?<![%\d,\.])([\d]{1,2}[,\.][\d]{2})\s*$', _dl)
                if _dm:
                    try:
                        _v = float(_dm.group(1).replace(',', '.'))
                        if 0 < _v < 40:
                            _d = round(_v, 2)
                            break
                    except Exception:
                        pass
        return _r, _y, _d

    def _store(name_raw: str, cols_vals: list):
        """Normalise name and store performance data if not already seen.

        cols_vals: [1M, 3M, 6M, 12M, 24M, 36M, 60M, YTD]
        3Y (36M) and 5Y (60M) are stored as ANNUALIZED CAGR to match FondiDoc.
        """
        if not name_raw:
            return
        norm = _normalize_for_unp(name_raw.strip())
        norm = _FUND_ALIASES.get(norm, norm)
        if not norm or norm in result:
            return
        # ── extract raw cumulative values ─────────────────────────────
        n = len(cols_vals)
        ytd_raw = _to_pct(cols_vals[7]) if n > 7 else _to_pct(cols_vals[-1] if cols_vals else None)
        p1y_raw = _to_pct(cols_vals[3]) if n > 3 else None
        p3y_cum = _to_pct(cols_vals[5]) if n > 5 else None   # 36M cumulative
        p5y_cum = _to_pct(cols_vals[6]) if n > 6 else None   # 60M cumulative

        # ── sanity check: YTD and 1Y must be in ±80 % range ──────────
        # Values outside this range indicate NAV prices, year numbers, or
        # misaligned columns from individual fund pages (not summary table).
        if not _plausible(ytd_raw, 80) or not _plausible(p1y_raw, 80):
            return

        # ── annualize 3Y and 5Y (factbook stores cumulative) ─────────
        p3y = _annualize(p3y_cum, 3)
        p5y = _annualize(p5y_cum, 5)

        if any(v is not None for v in (ytd_raw, p1y_raw, p3y, p5y)):
            result[norm] = {
                "ytd": ytd_raw, "perf_1y": p1y_raw,
                "perf_3y": p3y, "perf_5y": p5y,
            }

    _ref_date: str = ""   # data di riferimento estratta dal frontespizio

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_idx, page in enumerate(pdf.pages):

                # ── A0: Tabella Bond & Performance (pagine riepilogative) ─────
                # Struttura: [Nome fondo | YTM% | Rating | Duration | Freq. cedola]
                # Trigger: almeno una riga ha "YTM" e "Duration" come intestazioni
                for tbl in (page.extract_tables() or []):
                    _cells_flat = [str(c or '').strip().upper()
                                   for row in tbl for c in row]
                    _has_ytm = any('YTM' in c or 'YIELD' in c for c in _cells_flat)
                    _has_dur = any('DURATION' in c for c in _cells_flat)
                    if not (_has_ytm and _has_dur):
                        continue
                    # Trova indici colonne dall'header
                    _hdr_row = next(
                        (r for r in tbl
                         if any('YTM' in str(c or '').upper() or 'YIELD' in str(c or '').upper()
                                for c in r)),
                        None)
                    if _hdr_row is None:
                        continue
                    _hdr_u = [str(c or '').upper() for c in _hdr_row]
                    _ci_ytm  = next((i for i, h in enumerate(_hdr_u)
                                     if 'YTM' in h or 'YIELD' in h), None)
                    _ci_rat  = next((i for i, h in enumerate(_hdr_u)
                                     if 'RATING' in h), None)
                    _ci_dur  = next((i for i, h in enumerate(_hdr_u)
                                     if 'DURATION' in h), None)
                    for _brow in tbl:
                        _b0 = str(_brow[0] or '').replace('\n', ' ').strip()
                        if not re.match(r'^AZ\s', _b0, re.IGNORECASE):
                            continue
                        _bnorm = _normalize_for_unp(_b0)
                        _bnorm = _FUND_ALIASES.get(_bnorm, _bnorm)
                        if not _bnorm:
                            continue
                        if _bnorm not in _metrics:
                            _metrics[_bnorm] = {}
                        # YTM
                        if _ci_ytm is not None and _ci_ytm < len(_brow):
                            _yt_s = str(_brow[_ci_ytm] or '').replace('%','').replace(',','.').strip()
                            try:
                                _yt_v = round(float(_yt_s), 2)
                                if 'ytm' not in _metrics[_bnorm]:
                                    _metrics[_bnorm]['ytm'] = _yt_v
                            except ValueError:
                                pass
                        # Rating
                        if _ci_rat is not None and _ci_rat < len(_brow):
                            _rat_s = str(_brow[_ci_rat] or '').strip().upper()
                            if _rat_s and _rat_s in RATING_SCALE and 'credit_rating' not in _metrics[_bnorm]:
                                _metrics[_bnorm]['credit_rating'] = _rat_s
                        # Duration
                        if _ci_dur is not None and _ci_dur < len(_brow):
                            _dur_s = str(_brow[_ci_dur] or '').replace(',','.').strip()
                            try:
                                _dur_v = round(float(_dur_s), 2)
                                if 0 < _dur_v < 40 and 'duration' not in _metrics[_bnorm]:
                                    _metrics[_bnorm]['duration'] = _dur_v
                            except ValueError:
                                pass

                # ── A: structured table extraction (best-case) ────────────
                for tbl in (page.extract_tables() or []):
                    for row in tbl:
                        if not row or len(row) < 6:
                            continue
                        cell0 = str(row[0] or '').replace('\n', ' ').strip()
                        if not re.match(r'^(AZ\b|AZIMUT\b)', cell0, re.IGNORECASE):
                            continue
                        # row layout: [name, AUM, 1M, 3M, 6M, 12M, 24M, 36M, 60M, YTD]
                        data_cols = [str(c or '').strip() for c in row[2:]]
                        _store(cell0, data_cols)

                # ── B: text-line fallback ──────────────────────────────────
                text = page.extract_text() or ""

                # Estrai data di riferimento dal frontespizio (prime 5 pagine)
                if page_idx < 5 and not _ref_date:
                    # Cerca GG/MM/AAAA oppure GG.MM.AAAA
                    _dm = re.search(r'\b(\d{1,2})[/.](\d{2})[/.](\d{4})\b', text)
                    if _dm:
                        _ref_date = f"{_dm.group(1).zfill(2)}/{_dm.group(2)}/{_dm.group(3)}"
                    else:
                        # Cerca "31 marzo 2026" / "31 March 2026"
                        _months = {
                            "gennaio":"01","february":"02","febbraio":"02","march":"03",
                            "marzo":"03","april":"04","aprile":"04","may":"05","maggio":"05",
                            "june":"06","giugno":"06","july":"07","luglio":"07",
                            "august":"08","agosto":"08","september":"09","settembre":"09",
                            "october":"10","ottobre":"10","november":"11","novembre":"11",
                            "december":"12","dicembre":"12","january":"01","gennaio":"01",
                        }
                        _mp = r'\b(\d{1,2})\s+(' + '|'.join(_months) + r')\s+(20\d{2})\b'
                        _dm2 = re.search(_mp, text, re.IGNORECASE)
                        if _dm2:
                            _mn = _months.get(_dm2.group(2).lower(), "??")
                            _ref_date = f"{_dm2.group(1).zfill(2)}/{_mn}/{_dm2.group(3)}"

                for line in text.split('\n'):
                    line = line.strip()
                    if not re.match(r'^(AZ\b|AZIMUT\b)', line, re.IGNORECASE):
                        continue
                    tokens = line.split()
                    if len(tokens) < 6:
                        continue
                    # Find first all-digit token (= AUM in M€, e.g. "81", "1050")
                    # Skip year-like tokens (2010-2030) which appear in annual-return
                    # tables on individual fund pages and would misalign columns.
                    num_idx = next(
                        (i for i, t in enumerate(tokens)
                         if re.match(r'^\d{1,6}$', t)
                         and i > 0
                         and not re.match(r'^20[12]\d$', t)),  # skip 2010-2029
                        None
                    )
                    if num_idx is None:
                        continue
                    name_raw = ' '.join(tokens[:num_idx])
                    # tokens after AUM: [1M, 3M, 6M, 12M, 24M, 36M, 60M, YTD]
                    data_cols = tokens[num_idx + 1:]
                    _store(name_raw, data_cols)

                # ── C: METRICHE FIXED INCOME ─────────────────────────────
                # Strategy: try BOTH pdfplumber text formats.
                #  • default extract_text() merges same-Y chars into one line
                #  • layout=True  preserves spatial layout (like pdftotext -l)
                # Trigger on section header OR individual field names so we
                # never miss a page where only part of the right column is
                # captured.
                _FI_FIELDS = ('METRICHE FIXED INCOME', 'CREDIT RATING MEDIO',
                              'PORTFOLIO DURATION', 'YIELD TO MATURITY')
                _text_layout = ""
                try:
                    _text_layout = page.extract_text(layout=True) or ""
                except Exception:
                    pass
                _text_combined = text + "\n" + _text_layout

                if any(f in _text_combined.upper() for f in _FI_FIELDS):
                    # Use whichever text source is richer (more chars)
                    _txt_c = _text_layout if len(_text_layout) > len(text) \
                        else text
                    _lines_p = _txt_c.split('\n')
                    _fn = None
                    for _i, _ln in enumerate(_lines_p[:40]):
                        _ls = _ln.strip()
                        # Two-line: "AZ BOND" then "HIGH YIELD" on next line
                        if re.match(
                                r'^AZ\s+(BOND|ALLOCATION|EQUITY|ALTERNATIVE|ISLAMIC)\s*$',
                                _ls, re.IGNORECASE):
                            for _j in range(_i + 1, min(_i + 6, len(_lines_p))):
                                _nl = _lines_p[_j].strip()
                                if _nl and not re.match(
                                        r'^([A-Z]{2}\d{10}|ISIN|\d+\s*$)',
                                        _nl, re.IGNORECASE):
                                    _nrm = _normalize_for_unp(_ls + ' ' + _nl)
                                    _fn  = _FUND_ALIASES.get(_nrm, _nrm)
                                    break
                            if _fn:
                                break
                        # Single-line: "AZ BOND - ENHANCED YIELD"
                        _m1 = re.match(
                            r'^(AZ\s+(?:BOND|ALLOCATION|EQUITY|ALTERNATIVE|ISLAMIC))'
                            r'\s*[-–]?\s+(.+)$', _ls, re.IGNORECASE)
                        if _m1 and len(_m1.group(2).strip()) > 2:
                            _sub = _m1.group(2).strip()
                            if not re.match(r'^(\d+|[A-Z]{2}\d{10})', _sub):
                                _nrm = _normalize_for_unp(
                                    _m1.group(1).strip() + ' ' + _sub)
                                _fn  = _FUND_ALIASES.get(_nrm, _nrm)
                                break
                    if _fn:
                        # _get_fi_from_src is defined once above the page loop.
                        # Try layout text first (richer for two-column pages),
                        # then default text; merge – prefer first non-None.
                        _cr1, _yt1, _dur1 = (
                            _get_fi_from_src(_text_layout)
                            if _text_layout else (None, None, None))
                        _cr2, _yt2, _dur2 = _get_fi_from_src(text)

                        _cr  = _cr1  or _cr2
                        _yt  = _yt1  if _yt1  is not None else _yt2
                        _dur = _dur1 if _dur1 is not None else _dur2

                        if _fn not in _metrics:
                            _metrics[_fn] = {}
                        if _cr:              _metrics[_fn]['credit_rating'] = _cr
                        if _yt  is not None: _metrics[_fn]['ytm']           = _yt
                        if _dur is not None: _metrics[_fn]['duration']      = _dur

                # ── D: SCOMPOSIZIONE PORTAFOGLIO - ASSET CLASS ────────────
                # Only AZ ALLOCATION / BALANCED funds have this section.
                # AZ EQUITY and AZ BOND funds do NOT → Excel binary (0/1) is
                # the correct fallback for them; we don't overwrite here.
                #
                # pdfplumber layout quirks on this section:
                #  • "Equity" label appears on its own line WITHOUT a value.
                #    Its value (e.g. "54%") appears AFTER the x-axis scale
                #    labels (-10%, 0%, 10%, …40%) because the Equity bar
                #    physically extends to the right of the shorter bars.
                #    Marker: "Cash Offset" ends the data rows; the first
                #    non-round (not multiple of 10) positive % after that
                #    is the Equity value.
                #  • Bond components ("Sovereign", "Corporate", …) appear
                #    either inline ("Sovereign  28%") or in two lines
                #    ("Sovereign\n28%").  We sum them for fb_obb_pct.
                #  • Two-column merge: pdfplumber may merge rows from both
                #    columns, so we use re.search (CONTAINS) not re.match.
                if re.search(
                        r'SCOMPOSIZIONE\s+PORTAFOGLIO.*?ASSET\s+CLASS',
                        text, re.IGNORECASE):
                    _lp_d = text.split('\n')
                    _fn_d = None
                    for _i_d, _ln_d in enumerate(_lp_d[:25]):
                        _ls_d = _ln_d.strip()
                        if re.match(
                                r'^AZ\s+(BOND|ALLOCATION|EQUITY|ALTERNATIVE|ISLAMIC)\s*$',
                                _ls_d, re.IGNORECASE):
                            for _j_d in range(_i_d + 1, min(_i_d + 5, len(_lp_d))):
                                _nl_d = _lp_d[_j_d].strip()
                                if _nl_d and not re.match(
                                        r'^([A-Z]{2}\d{10}|ISIN|\d+\s*$)',
                                        _nl_d, re.IGNORECASE):
                                    _nrm_d = _normalize_for_unp(_ls_d + ' ' + _nl_d)
                                    _fn_d  = _FUND_ALIASES.get(_nrm_d, _nrm_d)
                                    break
                            if _fn_d:
                                break
                        _m1_d = re.match(
                            r'^(AZ\s+(?:BOND|ALLOCATION|EQUITY|ALTERNATIVE|ISLAMIC))'
                            r'\s*[-–]?\s+(.+)$', _ls_d, re.IGNORECASE)
                        if _m1_d and len(_m1_d.group(2).strip()) > 2:
                            _sub_d = _m1_d.group(2).strip()
                            if not re.match(r'^(\d+|[A-Z]{2}\d{10})', _sub_d):
                                _nrm_d = _normalize_for_unp(
                                    _m1_d.group(1).strip() + ' ' + _sub_d)
                                _fn_d  = _FUND_ALIASES.get(_nrm_d, _nrm_d)
                                break
                    if _fn_d:
                        # Locate the section header
                        _ALLOC_SEC_RE = re.compile(
                            r'SCOMPOSIZIONE\s+PORTAFOGLIO.*?ASSET\s+CLASS',
                            re.IGNORECASE)
                        _sec_start = next(
                            (i for i, l in enumerate(_lp_d)
                             if _ALLOC_SEC_RE.search(l)), None)
                        if _sec_start is not None:
                            # Bond-component keywords (English, as used in factbook)
                            _BD_KW = re.compile(
                                r'\b(sovereign|government|corporate|aggregate|'
                                r'dm\s+corporate|em\s+sovereign|'
                                r'eu\s+sovereign|us\s+sovereign|'
                                r'high\s+yield|convertible|'
                                r'fixed\s+income)\b', re.IGNORECASE)
                            _EQ_KW   = re.compile(r'\bequity\b', re.IGNORECASE)
                            _SKIP_KW = re.compile(
                                r'\b(cash|mixed|commodity|mmkt|offset|'
                                r'allocation)\b', re.IGNORECASE)
                            _STOP_RE = re.compile(
                                r'SCOMPOSIZIONE\s+(OBBLIGAZIONARIA|AZIONARIA)'
                                r'|SHARE\s+CLASS',
                                re.IGNORECASE)
                            # Inline: label + % on same line (handles merged cols)
                            # Decimal separator is optional (e.g. "28%" and "28,5%")
                            _INLINE = re.compile(
                                r'\b(equity|sovereign|government|corporate|aggregate|'
                                r'dm\s+corporate|em\s+sovereign|'
                                r'eu\s+sovereign|us\s+sovereign|'
                                r'high\s+yield|convertible|fixed\s+income)\b'
                                r'.*?(-?[\d]+(?:[,\.]\d*)?)\s*%',
                                re.IGNORECASE)
                            # Standalone %: "-10%", "0%", "54%", "28%" …
                            _PCT_SA = re.compile(
                                r'^(-?[\d]+(?:[,\.][\d]+)?)\s*%\s*$')

                            _equity_v   = None
                            _bond_sum   = 0.0
                            _last_label = None   # 'equity' | 'bond' | None
                            _saw_offset = False

                            for _ll2 in _lp_d[_sec_start: _sec_start + 80]:
                                _ll2s = _ll2.strip()
                                if not _ll2s:
                                    continue
                                if _STOP_RE.search(_ll2s):
                                    break

                                # Cash Offset marks end of data rows
                                if re.search(r'cash\s+offset', _ll2s,
                                             re.IGNORECASE):
                                    _saw_offset = True
                                    _last_label = None
                                    continue

                                # Skip non-data rows (Cash, Mixed, MMkt…)
                                if _SKIP_KW.search(_ll2s):
                                    _last_label = None
                                    continue

                                # ── Try inline: keyword + % on same line ──
                                _mi = _INLINE.search(_ll2s)
                                if _mi:
                                    _lk = _mi.group(1).lower()
                                    try:
                                        _pv = float(
                                            _mi.group(2).replace(',', '.'))
                                    except Exception:
                                        _last_label = None
                                        continue
                                    if _pv <= 0:
                                        _last_label = None
                                        continue
                                    if _EQ_KW.search(_lk) \
                                            and _equity_v is None:
                                        _equity_v = _pv / 100
                                    elif _BD_KW.search(_lk):
                                        _bond_sum += _pv / 100
                                    _last_label = None
                                    continue

                                # ── Keyword-only line (no % found) ──
                                _pm2 = _PCT_SA.match(_ll2s)
                                if not _pm2:
                                    if _EQ_KW.search(_ll2s):
                                        _last_label = 'equity'
                                    elif _BD_KW.search(_ll2s) \
                                            and not _SKIP_KW.search(_ll2s):
                                        _last_label = 'bond'
                                    else:
                                        _last_label = None
                                    continue

                                # ── Standalone % line ──
                                try:
                                    _pv = float(
                                        _pm2.group(1).replace(',', '.'))
                                except Exception:
                                    _last_label = None
                                    continue

                                if _last_label == 'equity' \
                                        and _equity_v is None \
                                        and _pv > 0:
                                    _equity_v = _pv / 100
                                    _last_label = None
                                elif _last_label == 'bond' and _pv > 0:
                                    _bond_sum += _pv / 100
                                    _last_label = None
                                elif _saw_offset \
                                        and _equity_v is None \
                                        and _pv > 0:
                                    # After Cash Offset the Equity bar value
                                    # appears here; axis labels are multiples
                                    # of 10 (or negative), so skip those
                                    if _pv > 0 and round(_pv) % 10 != 0:
                                        _equity_v = _pv / 100
                                else:
                                    _last_label = None

                            # Persist valid values
                            _ok_eq = (_equity_v is not None
                                      and 0.0 < _equity_v <= 1.0)
                            _ok_bd = _bond_sum > 0.005
                            if _ok_eq or _ok_bd:
                                if _fn_d not in _metrics:
                                    _metrics[_fn_d] = {}
                                if _ok_eq \
                                        and 'fb_az_pct' not in _metrics[_fn_d]:
                                    _metrics[_fn_d]['fb_az_pct'] = \
                                        round(_equity_v, 4)
                                if _ok_bd \
                                        and 'fb_obb_pct' not in _metrics[_fn_d]:
                                    _metrics[_fn_d]['fb_obb_pct'] = \
                                        round(_bond_sum, 4)

    except Exception:
        pass

    # Unisci metriche fixed income nelle voci del result
    for _fn, _md in _metrics.items():
        if _fn in result:
            result[_fn].update(_md)
        else:
            result[_fn] = dict(_md)

    # Aggiungi metadato data di riferimento (chiave speciale)
    if _ref_date:
        result["_ref_date"] = _ref_date

    return result

try:
    parse_factbook = st.cache_data(show_spinner=False)(parse_factbook)
except Exception:
    pass  # run uncached if decorator fails


# ── Factbook Excel export / import ──────────────────────────────────────────

_FB_COLS = [
    ("fondo",         "Fondo (chiave normalizzata)"),
    ("credit_rating", "Credit Rating Medio"),
    ("duration",      "Duration (anni)"),
    ("ytm",           "YTM (%)"),
    ("az_pct",        "% Azionario (0-1)"),
    ("obb_pct",       "% Obbligazionario (0-1)"),
    ("ytd",           "YTD"),
    ("perf_1y",       "Rend. 1A"),
    ("perf_3y",       "Rend. 3A ann."),
    ("perf_5y",       "Rend. 5A ann."),
]
_FB_INTERNAL_KEYS = {
    "az_pct":  "fb_az_pct",
    "obb_pct": "fb_obb_pct",
}


def factbook_to_excel_bytes(fb: dict) -> bytes:
    """Serialise the factbook data dict to a downloadable Excel file.

    The Excel has one row per fund with human-readable column headers.
    Numeric fields (duration, ytm, az_pct, obb_pct) are stored as floats.
    The 'fondo' column is the normalised key used internally for matching.
    """
    rows = []
    for norm, data in fb.items():
        if norm == "_ref_date" or not isinstance(data, dict):
            continue
        row = {"fondo": norm}
        row["credit_rating"] = data.get("credit_rating", "")
        row["duration"]      = data.get("duration", "")
        row["ytm"]           = data.get("ytm", "")
        row["az_pct"]        = data.get("fb_az_pct", "")
        row["obb_pct"]       = data.get("fb_obb_pct", "")
        row["ytd"]           = data.get("ytd", "")
        row["perf_1y"]       = data.get("perf_1y", "")
        row["perf_3y"]       = data.get("perf_3y", "")
        row["perf_5y"]       = data.get("perf_5y", "")
        rows.append(row)

    int_keys = [k for k, _ in _FB_COLS]
    hdr_map  = {k: h for k, h in _FB_COLS}
    df = pd.DataFrame(rows, columns=int_keys) if rows else pd.DataFrame(columns=int_keys)
    df.rename(columns=hdr_map, inplace=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Factbook")
        ws = writer.sheets["Factbook"]
        ws.column_dimensions["A"].width = 38
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws.column_dimensions[col].width = 18
    return buf.getvalue()


def factbook_from_excel(excel_bytes: bytes) -> dict:
    """Load a factbook Excel cache (produced by factbook_to_excel_bytes).

    Returns the same dict structure as parse_factbook().
    """
    try:
        df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Factbook",
                           dtype=str)
    except Exception:
        return {}

    # Map header labels back to internal keys
    hdr_to_int = {h: k for k, h in _FB_COLS}
    df.rename(columns=hdr_to_int, inplace=True)

    result: dict = {}
    _float_keys = {"duration", "ytm", "az_pct", "obb_pct"}

    for _, row in df.iterrows():
        norm = str(row.get("fondo", "")).strip()
        if not norm or norm == "nan":
            continue
        entry: dict = {}
        for int_key in [k for k, _ in _FB_COLS if k != "fondo"]:
            raw = str(row.get(int_key, "")).strip()
            if not raw or raw in ("nan", "None", ""):
                continue
            if int_key in _float_keys:
                try:
                    entry[_FB_INTERNAL_KEYS.get(int_key, int_key)] = float(raw)
                except ValueError:
                    pass
            else:
                entry[_FB_INTERNAL_KEYS.get(int_key, int_key)] = raw
        if entry:
            result[norm] = entry

    return result


# ── PARSER FACTBOOK FONDI PENSIONE ──────────────────────────────────────────

def parse_fp_factbook(pdf_bytes: bytes) -> dict:
    """Estrae rendimenti dal Factbook Fondi Pensione AZ Previdenza / AZ Sustainable Future.

    Formato atteso (tabella riepilogativa):
      colonne: nome | AUM | 1 mese | 3 mesi | 6 mesi | 12 mesi | 24 mesi | 36 mesi | 60 mesi | YTD
      header riga: ['', 'AUM', '1 mese', '3 mesi', '6 mesi', '12 mesi', '24 mesi', '36 mesi', '60 mesi', 'YTD']

    Restituisce {nome_comparto: {ytd, perf_1y, perf_3y, perf_5y}, "_ref_date": "DD/MM/YYYY"}
    """
    try:
        import pdfplumber
    except ImportError:
        return {}

    result: dict = {}
    ref_date = ""

    def _to_pct(s):
        if not s:
            return None
        s = (str(s).strip()
             .replace('−', '-').replace('–', '-')
             .replace(',', '.').replace('%', '').strip())
        if s in ('-', 'n.d.', 'N/D', '', 'None', 'nd', 'n/d'):
            return None
        try:
            return f"{float(s):+.2f}%"
        except Exception:
            return None

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                page_txt = page.extract_text() or ""

                # Cerca data di riferimento "Performance at DD/MM/YYYY"
                if not ref_date:
                    _dm = re.search(r'Performance\s+at\s+(\d{2}/\d{2}/\d{4})', page_txt, re.IGNORECASE)
                    if _dm:
                        ref_date = _dm.group(1)

                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 6:
                        continue

                    # Cerca riga header con colonna "12 mesi" (formato AZ Previdenza)
                    header_idx = None
                    col_map = {}
                    for i, row in enumerate(table):
                        cells = [str(c or "").strip().lower() for c in row]
                        row_str = " ".join(cells)
                        if "12 mesi" in row_str or ("ytd" in row_str and "mesi" in row_str):
                            header_idx = i
                            # Mappa indici colonne
                            for j, h in enumerate(row):
                                ht = str(h or "").strip().lower()
                                if ht == "ytd":              col_map["ytd"]  = j
                                elif "12 mesi" in ht:        col_map["1y"]   = j
                                elif "36 mesi" in ht:        col_map["3y"]   = j
                                elif "60 mesi" in ht:        col_map["5y"]   = j
                            break

                    if header_idx is None or not col_map:
                        continue

                    # Estrai righe dati
                    for row in table[header_idx + 1:]:
                        if not row:
                            continue
                        nome = str(row[0] or "").replace("\n", " ").strip()
                        if not nome or len(nome) < 5:
                            continue
                        # Salta righe vuote o di spaziatura
                        if not any(str(c or "").strip() for c in row[1:]):
                            continue
                        entry = {}
                        if "ytd" in col_map: entry["ytd"]     = _to_pct(row[col_map["ytd"]])
                        if "1y"  in col_map: entry["perf_1y"] = _to_pct(row[col_map["1y"]])
                        if "3y"  in col_map: entry["perf_3y"] = _to_pct(row[col_map["3y"]])
                        if "5y"  in col_map: entry["perf_5y"] = _to_pct(row[col_map["5y"]])
                        if any(v for v in entry.values()):
                            result[nome] = entry

    except Exception:
        pass

    if ref_date:
        result["_ref_date"] = ref_date
    return result


# ── Factbook JSON persistence (auto-load / GitHub API save) ─────────────────

def load_factbook_auto() -> dict:
    """Load factbook data from data/factbook_dati.json (committed in the repo).
    Returns {} when the file is absent or empty.
    """
    import json
    try:
        fp = Path(__file__).parent / "data" / "factbook_dati.json"
        if fp.exists() and fp.stat().st_size > 5:
            with open(fp, encoding="utf-8") as fh:
                data = json.load(fh)
                if isinstance(data, dict) and data:
                    return data
    except Exception:
        pass
    return {}


def save_factbook_to_repo(fb_data: dict) -> bool:
    """Commit data/factbook_dati.json su GitHub. Richiede secret GITHUB_TOKEN."""
    return _push_json_to_repo(
        fb_data,
        "data/factbook_dati.json",
        f"auto: aggiorna dati factbook {datetime.date.today().isoformat()}",
    )


@st.cache_data(ttl=3600)
def load_quantalys_cache() -> dict:
    """Load ISIN → Quantalys URL mapping from data/quantalys_cache.json.
    Returns {} if the file does not exist yet (run build_quantalys_cache.py first).
    """
    try:
        fp = Path(__file__).parent / "data" / "quantalys_cache.json"
        if fp.exists() and fp.stat().st_size > 5:
            with open(fp, encoding="utf-8") as fh:
                data = json.load(fh)
            # Keep only non-empty URLs
            return {k: v for k, v in data.items() if v}
    except Exception:
        pass
    return {}


@st.cache_data(ttl=3600)
def load_morningstar_cache() -> dict:
    """Load ISIN → Morningstar URL mapping from data/morningstar_cache.json.
    Returns {} if the file does not exist yet (run build_morningstar_cache.py first).
    """
    try:
        fp = Path(__file__).parent / "data" / "morningstar_cache.json"
        if fp.exists() and fp.stat().st_size > 5:
            with open(fp, encoding="utf-8") as fh:
                data = json.load(fh)
            return {k: v for k, v in data.items() if v}
    except Exception:
        pass
    return {}


@st.cache_data(ttl=3600)
def load_quantalys_ratings() -> dict:
    """Load ISIN → {score, globes} from data/quantalys_ratings.json."""
    try:
        fp = Path(__file__).parent / "data" / "quantalys_ratings.json"
        if fp.exists() and fp.stat().st_size > 5:
            with open(fp, encoding="utf-8") as fh:
                return json.load(fh)
    except Exception:
        pass
    return {}


# ── Quantalys chart capture (Playwright) ────────────────────────────────────

def _qtl_to_historique_url(url: str) -> str:
    """Trasforma URL Quantalys fondo → URL pagina Historique con i 6 grafici."""
    m = re.search(r'quantalys\.it/Fonds(?:/[A-Za-z]+)?/(\d+)', url)
    if m:
        return f"https://www.quantalys.it/Fonds/Historique/{m.group(1)}"
    return url


def _capture_qtl_6charts(hist_url: str) -> bytes | None:
    """Screenshot viewport della sezione principale Quantalys Historique.
    Approccio semplice: scroll al pannello → screenshot del viewport 1400×780.
    Cache su file — scrive solo su successo."""
    import hashlib, sys as _sys

    _key       = hashlib.md5(hist_url.encode()).hexdigest()[:14]
    _cache_dir = Path(__file__).parent / "data" / "qtl_chart_cache"
    _cache_dir.mkdir(parents=True, exist_ok=True)
    _cache_fp  = _cache_dir / f"{_key}.png"
    if _cache_fp.exists() and _cache_fp.stat().st_size > 1000:
        return _cache_fp.read_bytes()

    try:
        from playwright.sync_api import sync_playwright  # type: ignore
    except ImportError as _e:
        print(f"[QTL] Import error: {_e}", file=_sys.stderr)
        return None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            # Viewport 1400×780 — mostra il pannello principale completo
            page = browser.new_page(viewport={"width": 1400, "height": 780})
            page.goto(hist_url, wait_until="domcontentloaded", timeout=45_000)

            # Attende rendering amCharts
            try:
                page.wait_for_selector(".qtjs-panel-reloaded-graph svg", timeout=18_000)
            except Exception:
                pass
            page.wait_for_timeout(3_000)

            # Chiude cookie banner se presente
            try:
                _cb = page.locator(
                    "button:has-text('Ok, accetta tutto'),"
                    "button:has-text('Accetta tutto'),"
                    "button:has-text('Accept all')"
                )
                if _cb.count() > 0:
                    _cb.first.click()
                    page.wait_for_timeout(700)
            except Exception:
                pass

            # Scrolla al pannello principale (grafico storico)
            page.evaluate("""() => {
                const el = document.querySelector('.qtjs-panel-reloaded-graph')
                        || document.querySelector('[class*="qtjs-panel"]');
                if (el) el.scrollIntoView({behavior: 'instant', block: 'start'});
            }""")
            page.wait_for_timeout(400)

            # Sposta mouse fuori dai grafici per evitare tooltip hover
            page.mouse.move(10, 10)
            page.wait_for_timeout(600)

            # Calcola clip: bounds pannello (esclude margini grigi) × altezza fino alla toolbar
            clip_info = page.evaluate("""() => {
                const panel = document.querySelector('.qtjs-panel-reloaded-graph')
                           || document.querySelector('[class*="qtjs-panel"]');
                if (!panel) return null;
                const pr = panel.getBoundingClientRect();

                // Trova il top della toolbar blu (contiene <select>)
                let toolbarY = null;
                const ctrl = panel.querySelector('select')
                          || panel.querySelector('input[type="text"]');
                if (ctrl) {
                    let el = ctrl;
                    while (el.parentElement && el.parentElement !== panel) {
                        el = el.parentElement;
                    }
                    toolbarY = Math.round(el.getBoundingClientRect().top);
                }

                const x = Math.max(0, Math.round(pr.left));
                const y = Math.max(0, Math.round(pr.top));
                const w = Math.round(pr.width);
                const h = (toolbarY && toolbarY > 100)
                          ? toolbarY - y - 2
                          : Math.round(pr.height);
                return { x, y, w, h: Math.max(80, h), toolbarY };
            }""")

            svg_n = page.evaluate("() => document.querySelectorAll('svg').length")
            print(f"[QTL] SVG={svg_n}  clip={clip_info}  url={hist_url}", file=_sys.stderr)

            if clip_info and clip_info["w"] > 100 and clip_info["h"] > 80:
                png_bytes = page.screenshot(clip={
                    "x":      clip_info["x"],
                    "y":      clip_info["y"],
                    "width":  clip_info["w"],
                    "height": clip_info["h"],
                })
            else:
                # Fallback: viewport intero
                png_bytes = page.screenshot()
            browser.close()

        _cache_fp.write_bytes(png_bytes)
        return png_bytes

    except Exception as _ex:
        print(f"[QTL] Errore capture {hist_url}: {_ex}", file=_sys.stderr)
        return None


def _parse_ptf(wb, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    funds, cur_group = [], {"name":None,"mc":1.0,"me":1.0,"ma":1.0}
    for row in rows:
        name = row[COL_A]
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name in GROUP_NAMES:
            cur_group = {
                "name": name,
                "mc": float(row[COL_C]) if isinstance(row[COL_C],(int,float)) else 1.0,
                "me": float(row[COL_G]) if isinstance(row[COL_G],(int,float)) else 1.0,
                "ma": float(row[COL_K]) if isinstance(row[COL_K],(int,float)) else 1.0,
            }
            continue
        if name.startswith("AZ") and cur_group["name"]:
            rw = row[COL_R]
            if not isinstance(rw,(int,float)) or rw <= 0: continue
            funds.append({
                "nome":     name,
                "categoria": row[COL_B] if isinstance(row[COL_B],str) else "",
                "gruppo":   cur_group["name"],
                "az_pct":   min(1.0,max(0.0,float(row[COL_O]) if isinstance(row[COL_O],(int,float)) else 0.5)),
                "obb_pct":  min(1.0,max(0.0,1.0-(float(row[COL_O]) if isinstance(row[COL_O],(int,float)) else 0.5))),
                "r_weight": float(rw),
                "mc":cur_group["mc"],"me":cur_group["me"],"ma":cur_group["ma"],
            })
    if not funds: return pd.DataFrame()
    df = pd.DataFrame(funds)
    for wcol, mcol in [("w_cons","mc"),("w_equil","me"),("w_accr","ma")]:
        raw = df["r_weight"]*df[mcol]
        df[wcol] = raw/raw.sum() if raw.sum()>0 else raw
    df["macro_cat"] = df["categoria"].apply(get_macro)
    df = assign_colors(df)
    return df


def _parse_fida(wb) -> pd.DataFrame:
    ws = wb["FIDA"]
    rows = list(ws.iter_rows(values_only=True))
    funds = []
    for row in rows[1:]:
        nome = row[0]
        if not nome or not isinstance(nome,str): continue
        nome = nome.strip().replace("\xa0","")
        cat  = (row[2] or "").strip()
        funds.append({"nome":nome,"isin":row[1] or "","categoria":cat,"macro_cat":get_macro(cat)})
    return pd.DataFrame(funds).drop_duplicates(subset=["nome"])


def extract_fida_urls(file_bytes: bytes) -> dict:
    """Extract fondidoc.it URLs from FIDA sheet hyperlinks in the Excel XML."""
    fund_urls = dict(MANUAL_URL_OVERRIDES)
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            ss_root  = ET.fromstring(z.read("xl/sharedStrings.xml"))
            ns_ss    = {"s":"http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            strings  = ["".join(t.text or "" for t in si.findall(".//s:t",ns_ss))
                        for si in ss_root.findall("s:si",ns_ss)]
            fida_root= ET.fromstring(z.read("xl/worksheets/sheet5.xml"))
            ws_ns    = {"ws":"http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            rels_root= ET.fromstring(z.read("xl/worksheets/_rels/sheet5.xml.rels"))
            rels_ns  = {"rel":"http://schemas.openxmlformats.org/package/2006/relationships"}
        rid_to_url = {r.get("Id"):r.get("Target") for r in rels_root.findall("rel:Relationship",rels_ns)}
        hyp_map    = {}
        r_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        for hl in fida_root.findall(".//ws:hyperlink",ws_ns):
            rid = hl.get(r_attr)
            if rid: hyp_map[hl.get("ref")] = rid
        for row in fida_root.findall(".//ws:row",ws_ns):
            for cell in row.findall("ws:c",ws_ns):
                ref = cell.get("r")
                if ref and ref.startswith("A") and ref in hyp_map:
                    v = cell.find("ws:v",ws_ns)
                    if v is not None and cell.get("t") == "s":
                        name = strings[int(v.text)].strip().replace("\xa0","")
                        url  = rid_to_url.get(hyp_map[ref],"")
                        if "fondidoc.it" in url:
                            fund_urls[name] = url
    except Exception:
        pass
    return fund_urls


# ════════════════════════════════════════════════════════════
# FONDIDOC SCRAPING
# ════════════════════════════════════════════════════════════

def _to_en_url(url: str) -> str:
    """Ensure URL uses /en/ locale."""
    if "/en/" in url: return url
    return url.replace("fondidoc.it/d/","fondidoc.it/en/d/")


def _to_ana_url(index_url: str) -> str:
    return index_url.replace("/d/Index/","/d/Ana/").replace("/en/d/Index/","/en/d/Ana/")


def _fetch_html(url: str, timeout: int = 8) -> str | None:
    try:
        r = requests.get(_to_en_url(url), headers=FONDIDOC_HEADERS, timeout=timeout)
        return r.text if r.status_code == 200 else None
    except Exception:
        return None


def _parse_overview(html: str) -> dict:
    soup = BeautifulSoup(html, "lxml")
    lines = [l.strip() for l in soup.get_text(separator="\n").split("\n") if l.strip()]
    d = {}
    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1<len(lines) else ""
        if line == "SRRI (risk value)":           d["srri"]         = nxt
        elif line == "Start date":                d["start_date"]   = nxt
        elif line == "Assogestioni category":     d["cat_assog"]    = nxt
        elif line == "Income distribution":       d["income"]       = nxt
        elif line == "Management Fee":            d["mgmt_fee"]     = nxt
        elif line == "Performance Fee":           d["perf_fee"]     = nxt
        elif line == "Subscription fee":          d["sub_fee"]      = nxt
        elif line == "Rating" and "fida_rating" not in d: d["fida_rating"] = nxt
        elif line == "Score":                     d["fida_score"]   = nxt
        elif line == "Category" and "fida_cat" not in d: d["fida_cat"] = nxt
    return d


def _parse_analysis(html: str) -> dict:
    soup = BeautifulSoup(html, "lxml")
    d = {}

    # NAV section
    lines = [l.strip() for l in soup.get_text(separator="\n").split("\n") if l.strip()]
    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1<len(lines) else ""
        if line == "Last update": d["last_update"] = nxt
        elif line == "NAV":       d["nav"]          = nxt
        elif line == "Daily change (%)": d["daily_change"] = nxt

    # Tables
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows: continue
        header = [td.get_text(strip=True) for td in rows[0].find_all(["th","td"])]

        # Performance table (has YTD column)
        if "YTD" in header and "1 year" in header:
            def sg(cells, key, hdr):
                try: return cells[hdr.index(key)] if hdr.index(key)<len(cells) else "—"
                except ValueError: return "—"
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if not cells: continue
                lbl = cells[0].lower()
                if "performance" in lbl:
                    d["ytd"]     = sg(cells,"YTD",header)
                    d["perf_1y"] = sg(cells,"1 year",header)
                    d["perf_3y"] = sg(cells,"3 years",header)
                    d["perf_5y"] = sg(cells,"5 years",header)

        # Risk table (1 year / 3 years / 5 years, NO YTD)
        elif "1 year" in header and "YTD" not in header and len(header) >= 4:
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if not cells: continue
                lbl = cells[0].lower()
                def gv(idx): return cells[idx] if idx<len(cells) else "—"
                if "volatility" in lbl and "negative" not in lbl:
                    d["vol_1y"],d["vol_3y"],d["vol_5y"] = gv(1),gv(2),gv(3)
                elif "negative" in lbl:
                    d["neg_vol_1y"] = gv(1)
                    if len(header) > 2: d["neg_vol_3y"] = gv(2)
                    if len(header) > 3: d["neg_vol_5y"] = gv(3)
                elif "sharpe" in lbl:
                    d["sharpe_1y"],d["sharpe_3y"],d["sharpe_5y"] = gv(1),gv(2),gv(3)
                elif "sortino" in lbl:
                    d["sortino_1y"] = gv(1)
                elif "var" in lbl or "value at risk" in lbl:
                    d["var_1y"] = gv(1)
                    if len(header) > 2: d["var_3y"] = gv(2)

        # Annual performance (header contains year digits)
        elif any(h.isdigit() and len(h)==4 for h in header):
            years = [h for h in header if h.isdigit() and len(h)==4]
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if len(cells)<2: continue
                # Take first data row that has % values (fund row, not benchmark)
                if any("%" in c for c in cells) and "annual_perf" not in d:
                    # Rileva shift: la prima colonna-anno contiene il nome del fondo
                    # (label column), non un valore numerico → tutte le chiavi
                    # anno vanno decrement ate di 1.
                    _yr_shift = 0
                    if years:
                        try:
                            _fi  = header.index(years[0])
                            _fv  = (cells[_fi] if _fi < len(cells) else "").replace(
                                "%","").replace("+","").replace("-","").replace(",",".").strip()
                            float(_fv)           # se riesce → no shift
                        except (ValueError, IndexError):
                            _yr_shift = -1       # label column → shift
                    annual = {}
                    for yr in years:
                        try:
                            idx = header.index(yr)
                            val = cells[idx] if idx < len(cells) else "—"
                            # Salta celle non numeriche (es. nome del fondo)
                            _vc = val.replace("%","").replace("+","").replace(
                                "-","").replace(",",".").strip()
                            float(_vc)
                            actual_yr = str(int(yr) + _yr_shift)
                            annual[actual_yr] = val
                        except (ValueError, IndexError):
                            pass
                    if annual: d["annual_perf"] = annual
                    break
    return d


def _extract_isin(url: str) -> str:
    """Estrae ISIN da URL FondiDoc (formato: /CATCODE/ISIN_slug-nome-fondo)."""
    m = re.search(r'/([A-Z]{2}[A-Z0-9]{10})[_/]', url)
    return m.group(1) if m else ""


def fetch_fund_data(index_url: str) -> dict:
    """Fetch overview + analysis for one fund. Cached 1h."""
    result = {"url": index_url}
    isin = _extract_isin(index_url)
    if isin:
        result["isin"] = isin
    html_idx = _fetch_html(index_url)
    if html_idx: result["overview"] = _parse_overview(html_idx)
    html_ana = _fetch_html(_to_ana_url(index_url))
    if html_ana: result["analysis"] = _parse_analysis(html_ana)
    return result

try:
    fetch_fund_data = st.cache_data(ttl=3600, show_spinner=False)(fetch_fund_data)
except Exception:
    pass  # run uncached if decorator fails


def _fondidoc_search_url(query: str) -> str | None:
    """Cerca un fondo su FondiDoc per nome o ISIN.

    Interroga la pagina di ricerca inglese e restituisce il primo URL
    Index trovato, o None se non trovato / errore di rete.
    """
    try:
        import urllib.parse
        q = urllib.parse.quote(query)
        r = requests.get(
            f"https://www.fondidoc.it/en/Search?q={q}",
            headers=FONDIDOC_HEADERS,
            timeout=8,
            allow_redirects=True,
        )
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            # Fund detail pages live under /d/Index/ or /d/Ana/
            if "/d/Index/" in href or "/d/Ana/" in href:
                href = href.replace("/d/Ana/", "/d/Index/")
                if href.startswith("http"):
                    return href
                return "https://www.fondidoc.it" + href
    except Exception:
        pass
    return None


def fetch_gp_urls_missing(gp_data: dict, existing_cache: dict,
                           progress_cb=None, quick_urls: dict | None = None) -> dict:
    """Cerca su FondiDoc i fondi del GP che non sono già in cache.

    Per ogni fondo mancante tenta prima con il nome PDF normalizzato,
    poi con il nome risolto (Excel abbreviato), poi con il nome breve.
    quick_urls (es. _fida_urls_raw dall'Excel) vengono usati direttamente
    senza chiamate di rete per i fondi corrispondenti.
    Restituisce {nome_risolto: fund_data_dict} da aggiungere alla cache.
    """
    quick_urls = quick_urls or {}

    # Raccoglie tutti i nomi GP unici (PDF → risolto)
    # Include fondi assenti dal cache E fondi in cache ma senza URL
    missing: dict = {}   # resolved_name → pdf_name
    for sc_data in gp_data.values():
        if not isinstance(sc_data, dict):
            continue
        for f in sc_data.get("funds", []):
            pdf_name = f["nome"]
            res_name = _resolve_nome_for_fd(pdf_name, existing_cache)
            # Controlla solo la cache reale — non quick_urls.
            # Il pre-pass sotto gestisce i fondi trovati in quick_urls
            # e li salva nel cache (così _gp_miss scende a 0).
            has_url_in_cache = (
                existing_cache.get(res_name, {}).get("url", "")
                or existing_cache.get(pdf_name, {}).get("url", "")
            )
            if not has_url_in_cache:
                missing[res_name] = pdf_name

    if not missing and not quick_urls:
        return {}

    results: dict = {}

    # Pre-pass: URL già disponibili in quick_urls (Excel hyperlinks) — nessuna rete
    still_missing: dict = {}
    for res_name, pdf_name in missing.items():
        url = quick_urls.get(res_name) or quick_urls.get(pdf_name)
        if url:
            results[res_name] = {"url": url}
        else:
            still_missing[res_name] = pdf_name

    total = len(still_missing)
    done  = 0

    def _try_fetch(res_name: str, pdf_name: str):
        # Query 1: nome risolto/abbreviato Excel
        url = _fondidoc_search_url(res_name)
        # Query 2: nome PDF completo (es. "AZ Allocation - Balanced Plus")
        if not url and pdf_name != res_name:
            url = _fondidoc_search_url(pdf_name)
        # Nome breve (strip "AZ [Famiglia] - "), usato in più query
        _short = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', pdf_name, flags=re.I).strip()
        # Famiglia (es. "Bond", "Allocation", "Equity") → abbreviazione FondiDoc
        _fam_m = re.match(r'^AZ\s+(\S+)\s*[-–]', pdf_name, re.I)
        _fam   = _fam_m.group(1) if _fam_m else ""
        _FABBR = {"Bond": "AZ F.1 Bd", "Allocation": "AZ F.1 All.", "Equity": "AZ F.1 Eq."}
        _fabbr = _FABBR.get(_fam, "")
        # Query 3a: abbreviazione famiglia + nome breve (es. "AZ F.1 Bd Convertible Bond")
        if not url and _fabbr and _short:
            _q3a = f"{_fabbr} {_short}"
            if _q3a not in (res_name, pdf_name):
                url = _fondidoc_search_url(_q3a)
        # Query 3b: "AZ Fund 1 - " + nome breve  (es. "AZ Fund 1 - Convertible Bond")
        if not url and _short:
            _fund1_short = "AZ Fund 1 - " + _short
            if _fund1_short not in (res_name, pdf_name):
                url = _fondidoc_search_url(_fund1_short)
        # Query 3c: "AZ Fund 1 - " + nome completo PDF (es. "AZ Fund 1 - AZ Allocation - Balanced Plus")
        if not url:
            _fund1 = "AZ Fund 1 - " + pdf_name
            if _fund1 not in (res_name, pdf_name):
                url = _fondidoc_search_url(_fund1)
        # Query 4 & 5: nome breve da solo e variante con "AZ " prefisso
        if not url and _short and _short not in (res_name, pdf_name):
            url = _fondidoc_search_url(_short)
        if not url and _short:
            _short_az = "AZ " + _short
            if _short_az not in (res_name, pdf_name, _short):
                url = _fondidoc_search_url(_short_az)
        if url:
            try:
                data = fetch_fund_data(url)
                return res_name, data if data else {"url": url}
            except Exception:
                return res_name, {"url": url}
        return res_name, {}

    if still_missing:
        with ThreadPoolExecutor(max_workers=4) as pool:
            futures = {
                pool.submit(_try_fetch, rn, pn): rn
                for rn, pn in still_missing.items()
            }
            for future in as_completed(futures):
                rn = futures[future]
                try:
                    key, data = future.result()
                    if data:
                        results[key] = data
                except Exception:
                    pass
                done += 1
                if progress_cb:
                    progress_cb(done / total)

    return results


def fetch_all_fund_data(df: pd.DataFrame, fida_urls: dict,
                         progress_cb=None) -> dict:
    """Parallel fetch for all portfolio funds."""
    tasks = {}
    for nome in df["nome"].unique():
        url = fida_urls.get(nome)
        if url: tasks[nome] = url

    results = {}
    total = len(tasks)
    done  = 0

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {pool.submit(fetch_fund_data, url): nome for nome, url in tasks.items()}
        for future in as_completed(futures):
            nome = futures[future]
            try: results[nome] = future.result()
            except Exception: results[nome] = {}
            done += 1
            if progress_cb: progress_cb(done/total)

    return results


# ════════════════════════════════════════════════════════════
# MORNINGSTAR RATINGS — FondiOnline (primario) + lt.morningstar.com (fallback)
# ════════════════════════════════════════════════════════════
# Sorgente primaria: FondiOnline.it — un'unica chiamata JSON restituisce
# tutti i fondi Azimut con Rating Morningstar.
# Fallback: lt.morningstar.com — screener/snapshot per ISIN (più lento).

FONDIONLINE_BASE = "https://www.fondionline.it"
FO_API_URL       = "https://www.fondionline.it/offers-list"
FO_AZ_COMPANY_ID = "0C00001L0E"   # Azimut Investments S.A. (Morningstar ID)
FONDIONLINE_HDR  = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0",
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
    "Referer":         "https://www.fondionline.it/fondi/elenco_prodotti.html",
}

_MS_LT_BASE = "https://lt.morningstar.com/api/rest.svc/klr5zyak8x"
_MS_HDR = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/124.0.0.0 Safari/537.36",
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "it-IT,it;q=0.9",
    "Referer":         "https://lt.morningstar.com/",
}


def _fo_fetch_company_ratings(company_id: str) -> dict:
    """Fetch tutti i rating Azimut via FondiOnline JSON API (una sola chiamata).

    Restituisce {ISIN: {"ms_rating": int|None, "fo_url": str|None}}.
    """
    result: dict = {}
    try:
        r = requests.get(
            FO_API_URL,
            params={
                "productType":       "OICR",
                "sortOrder":         "asc",
                "pageNumber":        1,
                "pageSize":          1000,
                "tab":               0,
                "fundId":            "",
                "orderBy":           "Name",
                "brandingCompanyId": company_id,
                "distribution":      -1,
            },
            headers=FONDIONLINE_HDR,
            timeout=15,
        )
        if r.status_code == 200:
            data = r.json()
            for fund in data.get("funds", []):
                isin   = (fund.get("ISIN") or "").strip()
                rating = fund.get("Rating")
                url    = (f"{FONDIONLINE_BASE}/elenco-fondi/{fund['detailsUrl']}"
                          if fund.get("detailsUrl") else None)
                if isin:
                    result[isin] = {
                        "ms_rating": int(rating) if rating else None,
                        "fo_url":    url,
                    }
    except Exception:
        pass
    return result


def _ms_rating_screener(isin: str, sess: requests.Session) -> int | None:
    """Fallback screener lt.morningstar.com: term=ISIN → starRating."""
    try:
        r = sess.get(
            f"{_MS_LT_BASE}/screener",
            params={
                "output":             "json",
                "currencyId":         "EUR",
                "languageId":         "it-IT",
                "limit":              5,
                "securityDataPoints": "SecId,isin,starRating",
                "filters":            "",
                "term":               isin,
                "resultPage":         1,
            },
            headers=_MS_HDR,
            timeout=10,
        )
        if r.status_code == 200:
            for row in r.json().get("rows", []):
                if str(row.get("isin", "")).strip() == isin:
                    v = row.get("starRating")
                    if v is not None:
                        try:
                            return int(v)
                        except (ValueError, TypeError):
                            pass
    except Exception:
        pass
    return None


def _ms_rating_snapshot(isin: str, sess: requests.Session) -> int | None:
    """Fallback security_details lt.morningstar.com: itype=isin → XML → starRating."""
    import xml.etree.ElementTree as ET
    try:
        r = sess.get(
            f"{_MS_LT_BASE}/security_details/{isin}",
            params={"viewId": "MFsnapshot", "currencyId": "EUR",
                    "itype": "isin", "languageId": "it"},
            headers=_MS_HDR,
            timeout=10,
        )
        if r.status_code == 200 and r.text.strip():
            root = ET.fromstring(r.text)
            for tag in ("starRating", "Rating_MStarOverall",
                        "StarRatingM255", "mStarRating"):
                el = root.find(f".//{tag}")
                if el is not None and el.text:
                    try:
                        v = int(el.text.strip())
                        if 1 <= v <= 5:
                            return v
                    except (ValueError, TypeError):
                        pass
    except Exception:
        pass
    return None


def _ms_rating_for_isin(isin: str, sess: requests.Session) -> int | None:
    """Screener lt.morningstar.com, poi snapshot come secondo fallback."""
    r = _ms_rating_screener(isin, sess)
    if r is None:
        r = _ms_rating_snapshot(isin, sess)
    return r


def fetch_all_ms_ratings(df: pd.DataFrame, fida_df: pd.DataFrame,
                          progress_cb=None) -> dict:
    """Fetch rating Morningstar: FondiOnline (primario) → lt.morningstar.com (fallback).

    Restituisce {fund_name: {"ms_rating": int|None, "fo_url": str|None}}.
    """
    # 1. Build nome → ISIN map from FIDA sheet
    nome_to_isin: dict = {}
    if not fida_df.empty and "isin" in fida_df.columns:
        for _, fr in fida_df.iterrows():
            isin = str(fr.get("isin") or "").strip()
            if isin:
                nome_to_isin[fr["nome"]] = isin

    portfolio_names = list(df["nome"].unique()) if not df.empty else []

    # 2. Tentativo primario: FondiOnline (1 chiamata per tutti i fondi)
    isin_to_ms = _fo_fetch_company_ratings(FO_AZ_COMPANY_ID)

    # 3. Fallback lt.morningstar.com per gli ISIN non trovati su FondiOnline
    isins_needed = {
        nome: nome_to_isin[nome]
        for nome in portfolio_names
        if nome in nome_to_isin
    }
    missing_isins = [
        isin for isin in set(isins_needed.values())
        if isin not in isin_to_ms
    ]

    if missing_isins:
        sess = requests.Session()
        total = max(len(missing_isins), 1)
        done  = 0

        def _fetch_one(isin: str):
            return isin, _ms_rating_for_isin(isin, sess)

        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = {pool.submit(_fetch_one, isin): isin for isin in missing_isins}
            for future in as_completed(futures):
                try:
                    isin, rating = future.result()
                    if rating is not None:
                        isin_to_ms[isin] = {"ms_rating": rating, "fo_url": None}
                except Exception:
                    pass
                done += 1
                if progress_cb:
                    progress_cb(done / total)
    else:
        if progress_cb:
            progress_cb(1.0)

    # 4. Map back to fund names
    results: dict = {}
    for nome in portfolio_names:
        isin = isins_needed.get(nome, "")
        results[nome] = isin_to_ms.get(isin, {"ms_rating": None, "fo_url": None})

    return results


# ════════════════════════════════════════════════════════════
# PLOTLY CHARTS (unchanged)
# ════════════════════════════════════════════════════════════

def make_fund_pie(df, wcol, profile):
    d = df[df[wcol]>0.005].copy()
    d["pct"] = d[wcol]*100
    # Use nome_orig (GP display name) when available, else nome
    _disp = d["nome_orig"] if "nome_orig" in d.columns else d["nome"]
    labels = _disp.apply(lambda x: (x[:38]+"…") if len(x)>38 else x)
    fig = go.Figure(go.Pie(
        labels=labels, values=d["pct"],
        marker=dict(colors=d["color"].tolist(), line=dict(color="#fff",width=2.5)),
        hovertemplate="<b>%{label}</b><br>Peso: <b>%{value:.1f}%</b><extra></extra>",
        textinfo="percent", textfont=dict(size=10,family="DM Sans"),
        hole=0.40, pull=[0.04 if v==d["pct"].max() else 0 for v in d["pct"]],
        sort=False, direction="clockwise",
    ))
    fig.update_layout(
        margin=dict(t=10,b=10,l=10,r=180), showlegend=True,
        legend=dict(x=1.01,y=0.5,orientation="v",font=dict(size=9.5,family="DM Sans"),bgcolor="rgba(0,0,0,0)"),
        paper_bgcolor="rgba(0,0,0,0)", height=430,
        annotations=[dict(text=f"<b>{profile[:4]}</b>",x=0.5,y=0.5,showarrow=False,
                          font=dict(size=17,color="#0d1b2a",family="Cormorant Garamond"))],
    )
    return fig


def make_macro_bar(df, wcol):
    agg = df[df[wcol]>0.001].groupby("macro_cat")[wcol].sum().reset_index().sort_values(wcol)
    agg["pct"] = agg[wcol]*100
    agg["color"] = agg["macro_cat"].map(MACRO_COLORS)
    fig = go.Figure(go.Bar(
        x=agg["pct"], y=agg["macro_cat"], orientation="h",
        marker=dict(color=agg["color"].tolist(), line=dict(color="#fff",width=1)),
        hovertemplate="<b>%{y}</b>: %{x:.1f}%<extra></extra>",
        text=agg["pct"].apply(lambda v:f"{v:.1f}%"),
        textposition="inside", insidetextanchor="middle",
        textfont=dict(color="#fff",size=11,family="DM Sans"),
    ))
    fig.update_layout(
        margin=dict(t=10,b=10,l=10,r=10), paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(showgrid=False,showticklabels=False,range=[0,105]),
        yaxis=dict(showgrid=False,tickfont=dict(size=11,family="DM Sans")),
        height=max(160,len(agg)*48), bargap=0.28,
    )
    return fig


# ════════════════════════════════════════════════════════════
# PDF — MATPLOTLIB HELPERS
# ════════════════════════════════════════════════════════════

def _mpl_portfolio_pie(df, wcol, profile) -> io.BytesIO:
    """Donut only — la leggenda viene resa in ReportLab per permettere hyperlink."""
    d = df[df[wcol] > 0.005].sort_values(wcol, ascending=False)
    fig, ax = plt.subplots(figsize=(5.5, 5.5))
    _, _, autotexts = ax.pie(
        d[wcol], colors=d["color"].tolist(),
        autopct=lambda p: f"{p:.1f}%" if p > 3.5 else "",
        pctdistance=0.72,
        wedgeprops=dict(width=0.58, edgecolor="white", linewidth=2),
        startangle=90,
    )
    for at in autotexts:
        at.set_fontsize(9); at.set_color("white"); at.set_fontweight("bold")
    ax.text(0, 0, profile[:4], ha="center", va="center",
            fontsize=16, fontweight="bold", color="#0D1B2A")
    fig.patch.set_facecolor("#FFFFFF")
    plt.tight_layout(pad=0.3)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig); buf.seek(0)
    return buf


def _mpl_macro_pie(df, wcol) -> io.BytesIO | None:
    """Asset allocation donut Azionario/Obbligazionario — leggenda in ReportLab."""
    w_az  = (df[wcol] * df["az_pct"]).sum()
    w_obb = (df[wcol] * df["obb_pct"]).sum()
    if w_az + w_obb < 0.001:
        return None
    fig, ax = plt.subplots(figsize=(5.5, 5.5))
    _, _, autotexts = ax.pie(
        [w_az, w_obb], colors=["#1B4FBB", "#2D9D78"],
        autopct=lambda p: f"{p:.1f}%" if p >= 5 else "",
        pctdistance=0.70,
        wedgeprops=dict(width=0.58, edgecolor="white", linewidth=2.5),
        startangle=90,
    )
    for at in autotexts:
        at.set_fontsize(10); at.set_color("white"); at.set_fontweight("bold")
    ax.text(0, 0, "Asset\nAlloc.", ha="center", va="center",
            fontsize=11, fontweight="bold", color="#0D1B2A")
    fig.patch.set_facecolor("#FFFFFF")
    plt.tight_layout(pad=0.3)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig); buf.seek(0)
    return buf


def _mpl_annual_bar(annual_perf: dict, fund_name: str) -> io.BytesIO | None:
    if not annual_perf: return None
    years, vals = [], []
    for y,v in sorted(annual_perf.items()):
        try:
            num = float(v.replace("%","").replace(",",".").strip())
            years.append(y); vals.append(num)
        except: pass
    if not years: return None
    fig, ax = plt.subplots(figsize=(6,2.2))
    colors = ["#2D9D78" if v>=0 else "#E05252" for v in vals]
    bars = ax.bar(years, vals, color=colors, edgecolor="white", linewidth=0.8, width=0.6)
    ax.axhline(0, color="#94A3B8", linewidth=0.8, linestyle="-")
    for bar,val in zip(bars,vals):
        ax.text(bar.get_x()+bar.get_width()/2,
                bar.get_height()+(0.3 if val>=0 else -0.9),
                f"{val:+.1f}%", ha="center", va="bottom", fontsize=7, fontweight="bold",
                color="#2D9D78" if val>=0 else "#E05252")
    ax.set_ylabel("%", fontsize=8, color="#64748B")
    ax.tick_params(axis="both",labelsize=8,colors="#475569")
    ax.spines[["top","right","left"]].set_visible(False)
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.0f%%"))
    ax.grid(axis="y",alpha=0.3,linestyle="--")
    fig.patch.set_facecolor("#FFFFFF"); plt.tight_layout(pad=0.5)
    buf=io.BytesIO(); fig.savefig(buf,format="png",dpi=130,bbox_inches="tight",facecolor="white")
    plt.close(fig); buf.seek(0); return buf


# ════════════════════════════════════════════════════════════
# IBBOTSON CONE HELPERS
# ════════════════════════════════════════════════════════════

def _ibbotson_cone_png(mu: float, sigma: float, capitale: float,
                       orizzonte: int = 10, label: str = "Portafoglio") -> bytes:
    """PNG matplotlib del cono di Ibbotson (log-normal). Nessuna dipendenza da kaleido."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    t  = np.linspace(0, orizzonte, 300)
    mu_log = mu - sigma ** 2 / 2
    med   = capitale * np.exp(mu_log * t)
    up1   = capitale * np.exp((mu_log + sigma) * t)
    down1 = capitale * np.exp((mu_log - sigma) * t)
    up2   = capitale * np.exp((mu_log + 2 * sigma) * t)
    down2 = capitale * np.exp((mu_log - 2 * sigma) * t)

    fig, ax = plt.subplots(figsize=(11, 4), facecolor="white")
    ax.fill_between(t, down2, up2,   color="#BFDBFE", alpha=0.5, label="95% dei percorsi (±2σ)")
    ax.fill_between(t, down1, up1,   color="#3B82F6", alpha=0.35, label="68% dei percorsi (±1σ)")
    ax.plot(t, med,   color="#1B4FBB", lw=2,   label="Percorso centrale (mediana)")
    ax.axhline(capitale, color="#94A3B8", lw=1, ls="--", label=f"Capitale iniziale: € {capitale:,.0f}".replace(",", "."))
    y_min = min(capitale * 0.40, down2.min() * 0.92)
    y_max = up2.max() * 1.05
    ax.set_ylim(y_min, y_max)
    ax.set_xlim(0, orizzonte)
    ax.set_xlabel("Anni", fontsize=9)
    ax.set_ylabel("Valore portafoglio (€)", fontsize=9)
    ax.set_title(f"Cono di Ibbotson — {label}", fontsize=10, fontweight="bold", color="#0D1B2A")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"€ {v:,.0f}".replace(",", ".")))
    ax.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
    ax.legend(fontsize=7.5, loc="upper left")
    ax.grid(alpha=0.25, linestyle="--")
    fig.patch.set_facecolor("white")
    plt.tight_layout(pad=0.5)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=140, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _ibbotson_table_rows(mu: float, sigma: float, capitale: float,
                         years=(1, 3, 5, 10)) -> list[list]:
    """List of rows (one per year horizon) with log-normal scenario values."""
    import numpy as np
    mu_log = mu - sigma ** 2 / 2
    rows = []
    for y in years:
        med   = capitale * np.exp(mu_log * y)
        up1   = capitale * np.exp((mu_log + sigma) * y)
        down1 = capitale * np.exp((mu_log - sigma) * y)
        up2   = capitale * np.exp((mu_log + 2 * sigma) * y)
        down2 = capitale * np.exp((mu_log - 2 * sigma) * y)
        rows.append([y, down2, down1, med, up1, up2])
    return rows


# ── Prior forward-looking per sottocategoria ──────────────────────────────────
_AZ_SUBCAT_PRIOR: list[tuple[str, float]] = [
    ("emergent",        0.090), ("frontier",        0.095), ("asia",            0.085),
    ("pacifico",        0.085), ("cina",            0.090), ("china",           0.090),
    ("india",           0.095), ("latam",           0.090), ("latin",           0.090),
    ("africa",          0.095), ("giappone",        0.075), ("japan",           0.075),
    ("usa",             0.080), ("america",         0.080), ("nord america",    0.080),
    ("europa",          0.070), ("europe",          0.070), ("eurozona",        0.070),
    ("italia",          0.070), ("globali",         0.075), ("global",          0.075),
    ("worldwide",       0.075), ("internazional",   0.075),
    ("infrastructure",  0.065), ("infrastruttur",   0.065),
    ("healthcare",      0.075), ("technology",      0.080), ("tech",            0.080),
    ("thematic",        0.075), ("tematic",         0.075),
    ("high yield",      0.055), ("alto rendimento", 0.055), ("convertibil",     0.055),
    ("subordinati",     0.050), ("at1",             0.060), ("coco",            0.060),
    ("corporate",       0.035), ("societar",        0.035),
    ("inflation",       0.030), ("indicizzat",      0.030),
    ("governativ",      0.025), ("governo",         0.025), ("statali",         0.025),
    ("treasury",        0.025), ("aggregate",       0.030), ("income",          0.040),
    ("aggressiv",       0.060), ("dinamic",         0.058), ("moderato",        0.050),
    ("equilibrat",      0.050), ("conservativ",     0.035), ("prudente",        0.035),
    ("flessibil",       0.050), ("multi-asset",     0.050), ("multi asset",     0.050),
    ("allocation",      0.050), ("balanced",        0.050), ("fof",             0.050),
    ("absolute return", 0.040), ("ritorno assoluto",0.040), ("long short",      0.045),
    ("market neutral",  0.035), ("real asset",      0.050), ("commodit",        0.050),
    ("materie prime",   0.050), ("oro",             0.045), ("gold",            0.045),
    ("immobiliar",      0.055), ("reit",            0.055),
    ("monetar",         0.025), ("liquidit",        0.025), ("money market",    0.025),
]
_AZ_MU_MACRO: dict[str, float] = {
    "Azionari":            0.075,
    "Obbligazionari":      0.030,
    "Bilanciati/Flessibili": 0.050,
    "Alternativi":         0.040,
    "Altro":               0.050,
}
_AZ_VOL_FLOOR: dict[str, float] = {
    "Azionari": 0.110, "Bilanciati/Flessibili": 0.065,
    "Obbligazionari": 0.030, "Alternativi": 0.060, "Altro": 0.080,
}
_AZ_VOL_DEFAULT: dict[str, float] = {
    "Azionari": 0.150, "Bilanciati/Flessibili": 0.080,
    "Obbligazionari": 0.045, "Alternativi": 0.070, "Altro": 0.100,
}
_AZ_CAT_CORR: dict[tuple, float] = {
    ("Azionari",             "Azionari"):             0.78,
    ("Azionari",             "Bilanciati/Flessibili"): 0.55,
    ("Azionari",             "Obbligazionari"):       -0.10,
    ("Azionari",             "Alternativi"):           0.22,
    ("Azionari",             "Altro"):                 0.30,
    ("Bilanciati/Flessibili","Bilanciati/Flessibili"): 0.68,
    ("Bilanciati/Flessibili","Obbligazionari"):        0.42,
    ("Bilanciati/Flessibili","Alternativi"):           0.28,
    ("Bilanciati/Flessibili","Altro"):                 0.35,
    ("Obbligazionari",       "Obbligazionari"):        0.62,
    ("Obbligazionari",       "Alternativi"):           0.18,
    ("Obbligazionari",       "Altro"):                 0.20,
    ("Alternativi",          "Alternativi"):           0.32,
    ("Alternativi",          "Altro"):                 0.25,
    ("Altro",                "Altro"):                 0.50,
}


def _az_subcat_prior(nome: str, macro: str) -> float:
    n_low = nome.lower()
    for kw, prior in _AZ_SUBCAT_PRIOR:
        if kw in n_low:
            return prior
    return _AZ_MU_MACRO.get(macro, 0.060)


def _az_cat_corr(m1: str, m2: str) -> float:
    c = _AZ_CAT_CORR.get((m1, m2)) or _AZ_CAT_CORR.get((m2, m1))
    return c if c is not None else 0.25


def _az_portfolio_mu_sigma(
    d_sorted: "pd.DataFrame",
    wcol: str,
    fund_data: dict,
    factbook_data: dict,
    get_fb_val,
) -> tuple[float, float, int, int]:
    """
    μ e σ forward-looking del portafoglio con prior categoriali.
    Stesso approccio di portafogli_efficienti (optimizer.py).
    """
    import numpy as np

    rows_active = [(row["nome"], row.get("macro_cat","Altro"), float(row[wcol]))
                   for _, row in d_sorted.iterrows() if float(row[wcol]) > 0.001]
    if not rows_active:
        return 0.06, 0.10, 0, 0

    nomi  = [r[0] for r in rows_active]
    macro = [r[1] for r in rows_active]
    w     = np.array([r[2] for r in rows_active])
    w    /= w.sum()
    n     = len(nomi)

    def _get_vol(nome: str, mc: str) -> tuple[float, bool]:
        fd  = (fund_data or {}).get(nome, {})
        ana = fd.get("analysis", {})
        floor   = _AZ_VOL_FLOOR.get(mc, 0.080)
        default = _AZ_VOL_DEFAULT.get(mc, 0.100)
        for key in ("vol_3y", "vol_1y"):
            raw = ana.get(key, "") or get_fb_val(nome, key) or ""
            if raw and raw != "-":
                try:
                    v = float(str(raw).replace("%","").replace(",",".").strip())
                    if v > 1:
                        v /= 100.0
                    if v > 0:
                        return max(v, floor), True
                except Exception:
                    pass
        return default, False

    def _vol_missing_reason(nome: str) -> str:
        fd = (fund_data or {}).get(nome)
        if not fd:
            return "non censito su FondiDoc"
        ana = fd.get("analysis", {})
        v1 = str(ana.get("vol_1y", "") or "").strip()
        if not v1 or v1 in ("-", "n.d.", "N/D", "nd"):
            return "fondo con storico inferiore a 1 anno"
        return "fondo con storico inferiore a 1 anno"

    mu_vec  = np.array([_az_subcat_prior(nm, mc) for nm, mc in zip(nomi, macro)])
    vol_vec = []
    n_with_vol = 0
    missing_vol: list[tuple[str, str]] = []
    for nm, mc in zip(nomi, macro):
        v, ok = _get_vol(nm, mc)
        vol_vec.append(v)
        if ok:
            n_with_vol += 1
        else:
            missing_vol.append((nm, _vol_missing_reason(nm)))
    vol_vec = np.array(vol_vec)

    cov = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            if i == j:
                cov[i, j] = vol_vec[i] ** 2
            else:
                rho = _az_cat_corr(macro[i], macro[j])
                cov[i, j] = vol_vec[i] * vol_vec[j] * rho

    mu_ptf    = float(w @ mu_vec)
    var_ptf   = float(w @ cov @ w)
    sigma_ptf = float(np.sqrt(max(var_ptf, 0.0)))
    return mu_ptf, sigma_ptf, n_with_vol, n, missing_vol


# ════════════════════════════════════════════════════════════
# PDF GENERATION
# ════════════════════════════════════════════════════════════

def generate_pdf(df: pd.DataFrame, wcol: str, profile: str,
                 ptf_name: str, fund_data: dict = None,
                 fida_df: pd.DataFrame = None,
                 factbook_data: dict = None,
                 cache_date: str = "",
                 print_unp: bool = False,
                 qtl_charts: bool = False,
                 _progress_cb=None) -> bytes:
    """_progress_cb(fraction: float, text: str) — chiamata durante la cattura
    Quantalys per aggiornare una barra di avanzamento nell'UI."""

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.0*cm, rightMargin=1.0*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    ss = getSampleStyleSheet()
    def S(name,**kw): return ParagraphStyle(name,parent=ss["Normal"],**kw)

    T  = S("T",  fontName="Helvetica-Bold",  fontSize=22, textColor=rl_colors.HexColor("#0D1B2A"), spaceAfter=4,leading=28)
    EY = S("EY", fontName="Helvetica",       fontSize=8,  textColor=rl_colors.HexColor("#94A3B8"), spaceAfter=4,letterSpacing=1.5)
    SU = S("SU", fontName="Helvetica",       fontSize=10, textColor=rl_colors.HexColor("#64748B"), spaceAfter=4)
    SC = S("SC", fontName="Helvetica-Bold",  fontSize=11, textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=14,spaceAfter=8)
    BD = S("BD", fontName="Helvetica",       fontSize=8.5,textColor=rl_colors.HexColor("#1E293B"), leading=13, alignment=1)
    SM  = S("SM",  fontName="Helvetica",      fontSize=7.5,textColor=rl_colors.HexColor("#1E293B"), leading=11, alignment=1)
    SMC = S("SMC", fontName="Helvetica",      fontSize=7.5,textColor=rl_colors.HexColor("#1E293B"), leading=11, alignment=1)
    SML = S("SML", fontName="Helvetica",      fontSize=7.5,textColor=rl_colors.HexColor("#1E293B"), leading=11, alignment=0)
    FT = S("FT", fontName="Helvetica-Oblique",fontSize=7, textColor=rl_colors.HexColor("#94A3B8"), leading=10)
    FS = S("FS", fontName="Helvetica-Bold",  fontSize=13, textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=4,spaceAfter=2)
    FK = S("FK", fontName="Helvetica",       fontSize=7.5,textColor=rl_colors.HexColor("#64748B"), spaceAfter=2)
    LK = S("LK", fontName="Helvetica",       fontSize=7.5,textColor=rl_colors.HexColor("#1B4FBB"), spaceAfter=2)
    # HDR: sempre bianco+grassetto — per celle intestazione su sfondo scuro
    # (TEXTCOLOR di TableStyle NON sovrascrive il colore dei Paragraph — serve lo stile dedicato)
    HDR = S("HDR", fontName="Helvetica-Bold", fontSize=7.5,textColor=rl_colors.white, leading=11, alignment=1)
    HDRC= S("HDRC",fontName="Helvetica-Bold", fontSize=7.5,textColor=rl_colors.white, leading=11, alignment=1)

    story = []
    d_act = df[df[wcol]>0.001].copy()
    n_fondi = len(d_act)
    PW = 19 * cm   # printable width (A4 21cm - 2×1.0cm margins)

    # ISIN da foglio FIDA (fallback per fondi senza URL FondiDoc)
    isin_map = {}
    if fida_df is not None and not fida_df.empty and "isin" in fida_df.columns:
        isin_map = {r["nome"]: str(r["isin"]).strip() for _, r in fida_df.iterrows()
                    if r.get("isin") and str(r.get("isin","")).strip()}
    isin_map.update(MANUAL_ISIN_OVERRIDES)   # fondi senza ISIN nell'Excel
    w_az  = (d_act[wcol]*d_act["az_pct"]).sum()*100
    w_obb = (d_act[wcol]*d_act["obb_pct"]).sum()*100

    # ── ACCENT BAR ──────────────────────────────────────────
    story.append(Table([[""]], colWidths=[PW], rowHeights=[10],
        style=TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),rl_colors.HexColor("#0D1B2A")),
            ("LINEBELOW",(0,0),(-1,-1),3,rl_colors.HexColor("#C9A84C")),
        ])))
    story.append(Spacer(1,14))

    # ── TITLE BLOCK ─────────────────────────────────────────
    story.append(Paragraph("DEMO ANALISI", EY))
    story.append(Spacer(1,4))
    if "SUGGERITO" in ptf_name:
        _pdf_title = ptf_name.replace("SUGGERITO",
            'SUGGERITO<font size="11" color="#94A3B8"> da Global Persp.</font>', 1)
    elif ptf_name in ("PTF FULL", "PTF SHORT"):
        _pdf_title = f'{ptf_name}<font size="11" color="#94A3B8"> ispirato da Global Persp.</font>'
    else:
        _pdf_title = ptf_name
    story.append(Paragraph(f"Portafoglio {_pdf_title}", T))
    story.append(Paragraph(
        f"{PROFILE_ICONS.get(profile,'●')} Profilo {profile.title()}  ·  "
        f"Dati al {datetime.date.today().strftime('%d %B %Y')}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=14))

    # ── KPI ─────────────────────────────────────────────────
    KC = S("KC", fontName="Helvetica", fontSize=8.5,
           textColor=rl_colors.HexColor("#1E293B"), leading=13, alignment=1)
    def kpi_cell(v,l):
        return Paragraph(f'<font size="18"><b>{v}</b></font><br/>'
                         f'<font size="8" color="#64748B">{l}</font>', KC)
    kpi = Table(
        [[kpi_cell(str(n_fondi),"Fondi"),kpi_cell(f"{w_az:.1f}%","Quota Azionaria"),
          kpi_cell(f"{w_obb:.1f}%","Quota Obbligazionaria"),
          kpi_cell(datetime.date.today().strftime("%m/%Y"),"Data Report")]],
        colWidths=[PW/4]*4,
        rowHeights=[1.9*cm],
    )
    kpi.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.8,rl_colors.HexColor("#E2E8F0")),
        ("INNERGRID",(0,0),(-1,-1),0.8,rl_colors.HexColor("#E2E8F0")),
        ("BACKGROUND",(0,0),(-1,-1),rl_colors.HexColor("#F8FAFC")),
        ("PADDING",(0,0),(-1,-1),12),("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(kpi)

    # ── PIE CHARTS con legende ReportLab ────────────────────
    # Titolo sezione con meno spazio sopra per avvicinare il grafico al KPI
    SC_PIE = S("SCPIE", fontName="Helvetica-Bold", fontSize=11,
               textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=6, spaceAfter=5)
    story.append(Paragraph("Allocazione del Portafoglio", SC_PIE))

    PIE_W  = 6.5 * cm          # torta fondi
    LEG_W  = PW - PIE_W        # 11.5 cm per la legenda
    DOT_W  = 0.32 * cm
    # Ogni colonna di legenda (2 colonne affiancate)
    GAP_W  = 0.4 * cm          # gap tra le due colonne
    LC_W   = (LEG_W - DOT_W * 2 - GAP_W) / 2  # larghezza label per colonna

    LG = S("LG", fontName="Helvetica", fontSize=8,
           textColor=rl_colors.HexColor("#1E293B"), leading=11)

    def _dot(hex_color):
        t = Table([[""]], colWidths=[DOT_W], rowHeights=[DOT_W])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), rl_colors.HexColor(hex_color)),
        ]))
        return t

    # — Grafico 1: fondi con hyperlink (torta + legenda a 2 colonne affiancate) —
    pie_buf = _mpl_portfolio_pie(d_act, wcol, profile)
    pie_img = RLImage(pie_buf, width=PIE_W, height=PIE_W)
    d_leg   = d_act[d_act[wcol] > 0.005].sort_values(wcol, ascending=False)

    # Quantalys cache per fallback URL (ISIN → pagina fondo)
    _pdf_qtl = load_quantalys_cache()

    # Costruisci le celle della legenda
    leg_items = []
    for _, r in d_leg.iterrows():
        _rn      = r["nome"]                               # Excel key (for URL/data lookup)
        _rn_disp = r.get("nome_orig") or _rn               # GP name if available (for label)
        # 1) MANUAL direct  2) MANUAL fuzzy (vince su fida_urls)  3) FondiDoc cache
        _fd = fund_data or {}
        url = MANUAL_URL_OVERRIDES.get(_rn, "")
        if not url:
            _sk = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', _rn, flags=re.I).strip().lower()
            # 2) MANUAL fuzzy — PRIMA del cache direct/fuzzy
            if _sk:
                for _mk, _mu in MANUAL_URL_OVERRIDES.items():
                    _msk = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', _mk, flags=re.I).strip().lower()
                    if _msk and _msk in _sk and _mu:
                        url = _mu
                        break
            # 3) FondiDoc cache: direct poi fuzzy su _rn
            if not url:
                url = _fd.get(_rn, {}).get("url", "")
            if not url and _sk:
                for _fk, _fv in _fd.items():
                    if isinstance(_fv, dict) and _sk in _fk.lower() and _fv.get("url"):
                        url = _fv["url"]
                        break
            # 4) Fuzzy su nome_orig (GP name) se diverso dal nome risolto
            if not url and _rn_disp and _rn_disp != _rn:
                _sk2 = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', _rn_disp, flags=re.I).strip().lower()
                if _sk2:
                    for _fk, _fv in _fd.items():
                        if isinstance(_fv, dict) and _sk2 in _fk.lower() and _fv.get("url"):
                            url = _fv["url"]
                            break
            # 5) ISIN → pagina Quantalys (ultimo fallback)
            if not url:
                _isin_fb = isin_map.get(_rn, "") or isin_map.get(_rn_disp, "")
                if not _isin_fb and _sk:
                    for _ik, _iv in isin_map.items():
                        if _sk in _ik.lower():
                            _isin_fb = _iv; break
                if _isin_fb:
                    url = _pdf_qtl.get(_isin_fb, "")
        name_s = (_rn_disp[:24] + "…") if len(_rn_disp) > 24 else _rn_disp
        pct_s  = f"{r[wcol]*100:.1f}%"
        if url:
            lbl = Paragraph(
                f'<link href="{url}"><font color="#1B4FBB"><u>{name_s}</u></font></link>'
                f' <b>{pct_s}</b>', LG)
        else:
            lbl = Paragraph(f'{name_s} <b>{pct_s}</b>', LG)
        leg_items.append((_dot(r["color"]), lbl))

    # Disponi su 2 colonne: pari a sinistra, dispari a destra
    leg_rows_2c = []
    for i in range(0, len(leg_items), 2):
        d1, l1 = leg_items[i]
        if i + 1 < len(leg_items):
            d2, l2 = leg_items[i + 1]
        else:
            d2, l2 = Spacer(DOT_W, DOT_W), Paragraph("", LG)
        leg_rows_2c.append([d1, l1, d2, l2])

    leg_tbl = Table(leg_rows_2c,
                    colWidths=[DOT_W, LC_W, DOT_W + GAP_W, LC_W])
    leg_tbl.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ("LEFTPADDING",   (1,0), (1,-1),  4),
        ("LEFTPADDING",   (3,0), (3,-1),  4),
        ("LEFTPADDING",   (0,0), (0,-1),  0),
        ("LEFTPADDING",   (2,0), (2,-1),  GAP_W),
        ("RIGHTPADDING",  (0,0), (-1,-1), 2),
    ]))
    combo1 = Table([[pie_img, leg_tbl]], colWidths=[PIE_W, LEG_W])
    combo1.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE")]))

    # — Grafico 2: asset allocation — piccola, affiancata alla legenda (sotto la torta) —
    PIE_W2 = 5.0 * cm
    macro_buf = _mpl_macro_pie(d_act, wcol)
    macro_block = []
    if macro_buf:
        macro_img = RLImage(macro_buf, width=PIE_W2, height=PIE_W2)
        w_az_v  = (d_act[wcol] * d_act["az_pct"]).sum()
        w_obb_v = (d_act[wcol] * d_act["obb_pct"]).sum()
        macro_leg_rows = [
            [_dot("#1B4FBB"), Paragraph(f'Azionario  <b>{w_az_v*100:.1f}%</b>', LG)],
            [_dot("#2D9D78"), Paragraph(f'Obbligazionario  <b>{w_obb_v*100:.1f}%</b>', LG)],
        ]
        macro_leg_inner = Table(macro_leg_rows, colWidths=[DOT_W, 5*cm])
        macro_leg_inner.setStyle(TableStyle([
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (1,0), (1,-1),  5),
            ("LEFTPADDING",   (0,0), (0,-1),  0),
        ]))
        # Torta macro + legenda affiancate, allineate sinistra
        macro_row = Table([[macro_img, macro_leg_inner]],
                          colWidths=[PIE_W2, PW - PIE_W2])
        macro_row.setStyle(TableStyle([
            ("VALIGN",  (0,0), (-1,-1), "MIDDLE"),
            ("PADDING", (0,0), (-1,-1), 0),
        ]))
        macro_block = [Spacer(1, 6), macro_row]

    # Tutto il blocco grafici in KeepTogether → rimane sulla stessa pagina
    story.append(KeepTogether([combo1] + macro_block))
    story.append(PageBreak())

    # ════════════════════════════════════════════════════════
    # PAGE 2: RENDIMENTI 1-3-5 ANNI
    # ════════════════════════════════════════════════════════
    story.append(Paragraph("DEMO ANALISI", EY))
    story.append(Spacer(1,4))
    story.append(Paragraph("Tavola dei Rendimenti", T))
    _fb_loaded = bool(factbook_data)
    _fb_ref    = (factbook_data or {}).get("_ref_date", "")      # data dal frontespizio
    _fd_ref    = cache_date or datetime.date.today().strftime("%d/%m/%Y")  # data FondiDoc
    if _fb_loaded:
        _rend_src = (f"Fonte: FondiDoc aggiornata al {_fd_ref}"
                     + f"  ·  Fallback: Factbook AZ Investments"
                     + (f" al {_fb_ref}" if _fb_ref else ""))
    else:
        _rend_src = f"Fonte: FondiDoc aggiornata al {_fd_ref}"
    story.append(Paragraph(
        f"Performance per fondo  ·  Profilo {profile.title()}  ·  {_rend_src}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=12))

    # ── Helper: look up a performance metric from the factbook ────────────────
    _fb = factbook_data or {}
    _PERF_KEYS = {"ytd", "perf_1y", "perf_3y", "perf_5y"}

    def get_fb(nome: str, key: str) -> str:
        """Return factbook value for fund `nome` and metric `key`, or ''."""
        if not _fb:
            return ""
        if key not in _PERF_KEYS and key not in ("vol_1y", "vol_3y", "vol_5y"):
            return ""
        norm = _normalize_for_unp(nome)
        norm = _FUND_ALIASES.get(norm, norm)
        entry = _fb.get(norm)
        if not entry:
            # Substring fallback (same logic as lookup_unp)
            best, best_len = None, 0
            for fb_key, fb_val in _fb.items():
                if (fb_key in norm or norm in fb_key) and len(fb_key) > best_len:
                    best, best_len = fb_val, len(fb_key)
            entry = best
        if not entry:
            return ""
        return entry.get(key) or ""

    # ── Helper: weighted average of a metric across all funds ─────────────────
    def ptf_wavg(keys_list):
        """Weighted average per metric. Priorità FondiDoc; Factbook come fallback."""
        totals = {k: 0.0 for k in keys_list}
        cov_w  = {k: 0.0 for k in keys_list}
        for _, row in d_sorted.iterrows():
            fd  = (fund_data or {}).get(row["nome"], {})
            ana = fd.get("analysis", {})
            w   = row[wcol]
            for k in keys_list:
                raw = ana.get(k, "") or get_fb(row["nome"], k)
                try:
                    num = float(raw.replace("%","").replace(",",".").strip())
                    totals[k] += num * w
                    cov_w[k]  += w
                except Exception:
                    pass
        out = {}
        for k in keys_list:
            out[k] = f"{totals[k]/cov_w[k]:+.2f}%" if cov_w[k] > 0.01 else "-"
        return out

    # Paragraph style for portfolio summary row
    WH  = S("WH",  fontName="Helvetica-Bold", fontSize=8, textColor=rl_colors.white, leading=11, alignment=1)
    WHC = S("WHC", fontName="Helvetica-Bold", fontSize=8, textColor=rl_colors.white, leading=11, alignment=1)

    def pstyle_w(val):
        """White bold text coloured green/red for portfolio row."""
        try:
            v = float(val.replace("%","").replace(",",".").strip())
            c = "#7EFFC0" if v > 0 else ("#FFB3B3" if v < 0 else "#E2E8F0")
        except Exception:
            c = "#E2E8F0"
        return Paragraph(f'<font color="{c}"><b>{val}</b></font>', WH)

    def pstyle(val):
        try:
            v = float(val.replace("%","").replace(",","."))
            color = "#1A7A4A" if v>0 else ("#C0392B" if v<0 else "#475569")
            return f'<font color="{color}"><b>{val}</b></font>'
        except: return val

    # ── d_sorted deve essere definito PRIMA di chiamare ptf_wavg ──
    d_sorted = d_act.sort_values(wcol, ascending=False)

    # ── PERFORMANCE TABLE ────────────────────────────────────
    perf_keys = ["ytd","perf_1y","perf_3y","perf_5y","vol_1y","sharpe_1y"]
    ptf_p = ptf_wavg(perf_keys)

    perf_hdr = [Paragraph(f"<b>{t}</b>", HDR) for t in
                ["Fondo","Peso","YTD","1 Anno","3 Anni","5 Anni","Vol. 1A","Sharpe 1A"]]

    # Portfolio summary row (row index 1 — gold background)
    ptf_perf_row = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph(f"<b>100%</b>", WH),
        pstyle_w(ptf_p.get("ytd","-")),
        pstyle_w(ptf_p.get("perf_1y","-")),
        pstyle_w(ptf_p.get("perf_3y","-")),
        pstyle_w(ptf_p.get("perf_5y","-")),
        Paragraph(ptf_p.get("vol_1y","-"), WH),
        Paragraph(ptf_p.get("sharpe_1y","-"), WH),
    ]

    perf_rows = [perf_hdr, ptf_perf_row]

    for _, row in d_sorted.iterrows():
        fd  = (fund_data or {}).get(row["nome"], {})
        ana = fd.get("analysis", {})
        def gv(key, nome=row["nome"]):
            # Priorità FondiDoc; Factbook come fallback
            return ana.get(key, "") or get_fb(nome, key) or "-"
        perf_rows.append([
            Paragraph(row["nome"][:48], SM),
            Paragraph(f"<b>{row[wcol]*100:.1f}%</b>", SM),
            Paragraph(pstyle(gv("ytd")),     SM),
            Paragraph(pstyle(gv("perf_1y")), SM),
            Paragraph(pstyle(gv("perf_3y")), SM),
            Paragraph(pstyle(gv("perf_5y")), SM),
            Paragraph(gv("vol_1y"),           SM),
            Paragraph(gv("sharpe_1y"),        SM),
        ])

    perf_tbl = Table(perf_rows,
        colWidths=[6.2*cm,1.4*cm,1.4*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm],
        repeatRows=1)
    ts_perf = [
        ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),  # header
        ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
        # Portfolio summary row — deep forest green + gold accent
        ("BACKGROUND",(0,1),(-1,1), rl_colors.HexColor("#1B4332")),
        ("LINEBELOW",(0,1),(-1,1),  2, rl_colors.HexColor("#C9A84C")),
        # Funds
        ("FONTSIZE",(0,0),(-1,-1),  8),
        ("PADDING",(0,0),(-1,-1),   5),
        ("ROWBACKGROUNDS",(0,2),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",(0,0),(-1,-1),     "CENTER"),
        ("VALIGN",(0,0),(-1,-1),    "MIDDLE"),
    ]
    perf_tbl.setStyle(TableStyle(ts_perf))
    # KeepTogether: la tabella rendimenti non viene spezzata su due pagine
    NOTE_P = S("NTP", fontName="Helvetica-Oblique", fontSize=6.5,
               textColor=rl_colors.HexColor("#94A3B8"), leading=9)
    story.append(KeepTogether([
        perf_tbl,
        Spacer(1, 5),
        Paragraph(
            "◆ La riga <b>Portafoglio</b> riporta la media ponderata dei rendimenti dei singoli fondi, "
            "usando i pesi del profilo selezionato. Il calcolo include solo i fondi per cui il dato è "
            "disponibile su FondiDoc, rinormalizzando i pesi su di essi. "
            "I rendimenti a 3 e 5 anni sono tassi annualizzati: la media ponderata è un'approssimazione "
            "(il rendimento composito effettivo dipende dalla sequenza temporale e dalla correlazione tra fondi). "
            "Dati a titolo indicativo — non costituiscono consulenza di investimento.", NOTE_P),
    ]))
    story.append(Spacer(1,12))

    # ── RISK TABLE ───────────────────────────────────────────
    risk_keys = ["vol_1y","vol_3y","vol_5y","neg_vol_1y","sharpe_3y","sortino_1y"]
    ptf_r = ptf_wavg(risk_keys)

    risk_hdr = [Paragraph(f"<b>{t}</b>", HDR) for t in
                ["Fondo","Peso","Vol. 1A","Vol. 3A","Vol. 5A","Vol. Neg. 1A","Sharpe 3A","Sortino 1A"]]

    ptf_risk_row = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph("<b>100%</b>", WH),
        Paragraph(ptf_r.get("vol_1y","-"),     WH),
        Paragraph(ptf_r.get("vol_3y","-"),     WH),
        Paragraph(ptf_r.get("vol_5y","-"),     WH),
        Paragraph(ptf_r.get("neg_vol_1y","-"), WH),
        Paragraph(ptf_r.get("sharpe_3y","-"),  WH),
        Paragraph(ptf_r.get("sortino_1y","-"), WH),
    ]

    risk_rows = [risk_hdr, ptf_risk_row]
    for _, row in d_sorted.iterrows():
        fd  = (fund_data or {}).get(row["nome"], {})
        ana = fd.get("analysis", {})
        def gv_r(k): return ana.get(k,"-")
        risk_rows.append([
            Paragraph(row["nome"][:48], SM),
            Paragraph(f"{row[wcol]*100:.1f}%", SM),
            Paragraph(gv_r("vol_1y"),     SM), Paragraph(gv_r("vol_3y"),     SM), Paragraph(gv_r("vol_5y"),    SM),
            Paragraph(gv_r("neg_vol_1y"), SM), Paragraph(gv_r("sharpe_3y"),  SM), Paragraph(gv_r("sortino_1y"),SM),
        ])

    risk_tbl = Table(risk_rows,
        colWidths=[6.2*cm,1.4*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm,1.5*cm],
        repeatRows=1)
    ts_risk = [
        ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),
        ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
        ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
        ("BACKGROUND",(0,1),(-1,1), rl_colors.HexColor("#1B4332")),
        ("LINEBELOW",(0,1),(-1,1),  2, rl_colors.HexColor("#C9A84C")),
        ("FONTSIZE",(0,0),(-1,-1),  8),
        ("PADDING",(0,0),(-1,-1),   5),
        ("ROWBACKGROUNDS",(0,2),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",(0,0),(-1,-1),     "CENTER"),
        ("VALIGN",(0,0),(-1,-1),    "MIDDLE"),
    ]
    risk_tbl.setStyle(TableStyle(ts_risk))

    NOTE = S("NT", fontName="Helvetica-Oblique", fontSize=6.5,
             textColor=rl_colors.HexColor("#94A3B8"), leading=9)
    # KeepTogether: titolo + tabella rischio + nota nella stessa pagina
    story.append(KeepTogether([
        Paragraph("Metriche di Rischio", SC),
        risk_tbl,
        Spacer(1,6),
        Paragraph(
            "◆ I valori del Portafoglio sono medie ponderate per peso dei singoli fondi. "
            "La volatilità di portafoglio effettiva dipende anche dalle correlazioni tra i fondi. "
            "Dati forniti a titolo indicativo.", NOTE),
    ]))

    story.append(Spacer(1, 14))

    # ── ALLOCATION TABLE ─────────────────────────────────────
    def get_fi_metric(nome: str, key: str):
        """Return fixed-income metric (duration/credit_rating/ytm) from factbook or None."""
        if not _fb:
            return None
        norm = _normalize_for_unp(nome)
        norm = _FUND_ALIASES.get(norm, norm)
        entry = _fb.get(norm)
        if not entry or not isinstance(entry, dict):
            best, best_len = None, 0
            for fb_key, fb_val in _fb.items():
                if (isinstance(fb_val, dict)
                        and (fb_key in norm or norm in fb_key)
                        and len(fb_key) > best_len):
                    best, best_len = fb_val, len(fb_key)
            entry = best
        if not entry or not isinstance(entry, dict):
            return None
        return entry.get(key)

    # Per-fund helpers: prefer factbook composition over binary Excel values
    def _az_eff(row):
        v = get_fi_metric(row["nome"], "fb_az_pct")
        return v if v is not None else row["az_pct"]

    def _obb_eff(row):
        v = get_fi_metric(row["nome"], "fb_obb_pct")
        return v if v is not None else row["obb_pct"]

    # Portfolio-level composition (factbook-aware weighted averages)
    _ptf_az_wtd  = sum(_row[wcol] * _az_eff(_row)  for _, _row in d_sorted.iterrows())
    _ptf_obb_wtd = sum(_row[wcol] * _obb_eff(_row) for _, _row in d_sorted.iterrows())

    # Duration and rating weighted by effective obb% × w
    _ptf_dur_num = _ptf_dur_den = 0.0
    _ptf_rat_num = _ptf_rat_den = 0.0
    for _, _row in d_sorted.iterrows():
        _w   = _row[wcol]
        _obb = _obb_eff(_row)            # factbook obb% if available, else Excel
        _dur = get_fi_metric(_row["nome"], "duration")
        _rat = get_fi_metric(_row["nome"], "credit_rating")
        if isinstance(_dur, (int, float)) and _obb > 0:
            _ptf_dur_num += _w * _obb * _dur
            _ptf_dur_den += _w * _obb
        if isinstance(_rat, str) and _rat in RATING_SCALE and _obb > 0:
            _ptf_rat_num += _w * _obb * RATING_SCALE[_rat]
            _ptf_rat_den += _w * _obb

    _ptf_dur_str = (f"{_ptf_dur_num / _ptf_dur_den:.2f}"
                    if _ptf_dur_den > 0.001 else "-")
    if _ptf_rat_den > 0.001:
        _ri = max(1, min(22, round(_ptf_rat_num / _ptf_rat_den)))
        _ptf_rat_str = RATING_INVERSE.get(_ri, "-")
    else:
        _ptf_rat_str = "-"

    _alloc_hdr_items = [
        ("Fondo",        HDR), ("ISIN",         HDR), ("Peso",   HDR),
        ("% Azion.",    HDRC), ("% Obbl.",      HDRC),
        ("Duration",     HDR), ("Rating Medio",  HDR), ("Cat. FIDA", HDR),
        ("FIDA rating",  HDR), ("Morningstar",   HDR),
    ]
    alloc_hdr = [Paragraph(f"<b>{t}</b>", st) for t, st in _alloc_hdr_items]
    alloc_ptf = [
        Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
        Paragraph("",                                  WH),
        Paragraph("<b>100%</b>",                       WH),
        Paragraph(f"<b>{_ptf_az_wtd*100:.1f}%</b>",   WHC),
        Paragraph(f"<b>{_ptf_obb_wtd*100:.1f}%</b>",  WHC),
        Paragraph(f"<b>{_ptf_dur_str}</b>",            WH),
        Paragraph(f"<b>{_ptf_rat_str}</b>",            WH),
        Paragraph("",                                  WH),
        Paragraph("",                                  WH),
        Paragraph("",                                  WH),
    ]
    # FIDArating visual badges in PDF:
    # BACKGROUND per-cell works in ReportLab TableStyle and overrides
    # ROWBACKGROUNDS when placed after it.  Paragraph textColor must be
    # set via ParagraphStyle (not TEXTCOLOR in TableStyle — that is ignored
    # when the cell contains a Paragraph object).
    # 5=dark green bg / 4=medium green / 3=light green — all white text.
    # 1-2: no background, bold black (no red for print).
    _FIDA_BG_HEX = {"5": "#166534", "4": "#15803d", "3": "#16a34a"}

    def _fida_para(val: str):
        _v = str(val).strip()
        if _v in _FIDA_BG_HEX:
            return Paragraph(
                _v,
                S(f"SMF{_v}", fontName="Helvetica-Bold", fontSize=7.5,
                  textColor=rl_colors.white, leading=11, alignment=1))
        if _v in ("1", "2"):
            return Paragraph(
                _v,
                S(f"SMF{_v}", fontName="Helvetica-Bold", fontSize=7.5,
                  textColor=rl_colors.HexColor("#1E293B"), leading=11, alignment=1))
        return Paragraph(val, SM)   # "—" or unknown

    # Morningstar data for PDF (loaded from cache / session state)
    _ms_pdf = st.session_state.get("_ms_data") or load_ms_cache()

    # Morningstar amber/gold palette for PDF
    _MS_BG_HEX = {"5": "#78350F", "4": "#92400E", "3": "#B45309"}

    def _ms_para(val) -> Paragraph:
        """ReportLab Paragraph for a Morningstar rating value.
        Shows numeric value with star count in ASCII to stay within Helvetica charset.
        """
        try:
            v = int(val)
        except (TypeError, ValueError):
            return Paragraph("—", SM)
        label = f"{v} {'*'*v}"   # e.g. "4 ****" — ASCII-safe, no Unicode stars
        if str(v) in _MS_BG_HEX:
            return Paragraph(
                label,
                S(f"SMMSP{v}", fontName="Helvetica-Bold", fontSize=7,
                  textColor=rl_colors.white, leading=11, alignment=1))
        return Paragraph(
            label,
            S(f"SMMSd{v}", fontName="Helvetica", fontSize=7,
              textColor=rl_colors.HexColor("#475569"), leading=11, alignment=1))

    alloc_fund_rows = []
    _fida_vals = []   # keep to build BACKGROUND commands after the loop
    _ms_vals   = []
    for _, _row in d_sorted.iterrows():
        _dur2  = get_fi_metric(_row["nome"], "duration")
        _rat2  = get_fi_metric(_row["nome"], "credit_rating")
        _az_s  = _az_eff(_row)  * 100
        _obb_s = _obb_eff(_row) * 100
        _fd_ov2 = (fund_data or {}).get(_row["nome"], {}).get("overview", {})
        _cat2   = _fd_ov2.get("cat_assog") or "—"
        _fida2  = _fd_ov2.get("fida_rating") or "—"
        _ms2    = _ms_pdf.get(_row["nome"], {}).get("ms_rating")
        _fida_vals.append(str(_fida2).strip())
        _ms_vals.append(str(_ms2).strip() if _ms2 is not None else "—")
        _isin2 = isin_map.get(_row["nome"], "")
        if not _isin2:
            # Normalised fallback: look up isin_map by normalised name
            _n2 = _normalize_for_unp(_row["nome"])
            for _ik, _iv in isin_map.items():
                if _normalize_for_unp(_ik) == _n2:
                    _isin2 = _iv
                    break
        _disp_nome2 = _row.get("nome_orig") or _row["nome"]
        alloc_fund_rows.append([
            Paragraph(_disp_nome2[:48], SM),
            Paragraph(_isin2 or "—",                                     SM),
            Paragraph(f"{_row[wcol]*100:.1f}%",                          SM),
            Paragraph(f"{_az_s:.1f}%",                                   SMC),
            Paragraph(f"{_obb_s:.1f}%",                                  SMC),
            Paragraph(f"{_dur2:.2f}" if isinstance(_dur2, (int, float)) else "—", SM),
            Paragraph(_rat2 if isinstance(_rat2, str) else "—",           SM),
            Paragraph(_cat2,                                               SM),
            _fida_para(_fida2),
            _ms_para(_ms2),
        ])

    # Build per-row BACKGROUND commands for FIDArating (col 8) and Morningstar (col 9).
    # (ISIN column inserted at position 1 shifts FIDArating 7→8, Morningstar 8→9)
    _fida_bg_cmds = []
    for _fi, _fv in enumerate(_fida_vals):
        _bg_hex = _FIDA_BG_HEX.get(_fv)
        if _bg_hex:
            _tr = _fi + 2   # row 0=hdr, 1=ptf summary, 2+=fund rows
            _fida_bg_cmds.append(
                ("BACKGROUND", (8, _tr), (8, _tr),
                 rl_colors.HexColor(_bg_hex)))
    for _mi, _mv in enumerate(_ms_vals):
        _bg_hex_ms = _MS_BG_HEX.get(_mv)
        if _bg_hex_ms:
            _tr = _mi + 2
            _fida_bg_cmds.append(
                ("BACKGROUND", (9, _tr), (9, _tr),
                 rl_colors.HexColor(_bg_hex_ms)))

    # Fondo(3.4) ISIN(2.6) Peso(1.2) %Az(1.5) %Obb(1.3) Dur(1.6) Rat(1.6) Cat(2.1) FIDArtg(1.5) MS(2.2) = 19.0 cm
    alloc_tbl = Table(
        [alloc_hdr, alloc_ptf] + alloc_fund_rows,
        colWidths=[3.4*cm, 2.6*cm, 1.2*cm, 1.5*cm, 1.3*cm, 1.6*cm, 1.6*cm, 2.1*cm, 1.5*cm, 2.2*cm],
        repeatRows=1,
    )
    alloc_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0,0), (-1,0),  rl_colors.HexColor("#0D1B2A")),
        ("TEXTCOLOR",      (0,0), (-1,0),  rl_colors.white),
        ("FONTNAME",       (0,0), (-1,0),  "Helvetica-Bold"),
        ("BACKGROUND",     (0,1), (-1,1),  rl_colors.HexColor("#1B4332")),
        ("LINEBELOW",      (0,1), (-1,1),  2, rl_colors.HexColor("#C9A84C")),
        ("FONTSIZE",       (0,0), (-1,-1), 8),
        ("PADDING",        (0,0), (-1,-1), 5),
        ("ROWBACKGROUNDS", (0,2), (-1,-1),
         [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
        ("LINEBELOW",      (0,0), (-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
        ("ALIGN",          (0,0), (-1,-1), "CENTER"),
        ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
        *_fida_bg_cmds,   # coloured cell backgrounds — placed last to override
    ]))

    NOTE_A = S("NTA", fontName="Helvetica-Oblique", fontSize=6.5,
               textColor=rl_colors.HexColor("#94A3B8"), leading=9)
    story.append(KeepTogether([
        Paragraph("Scomposizione Azionario / Obbligazionario", SC),
        alloc_tbl,
        Spacer(1, 6),
        Paragraph(
            "◆ Le quote azionaria e obbligazionaria derivano dal Factbook AZ Investments"
            + (f" al {_fb_ref}" if _fb_ref else "")
            + " (pagine singoli fondi); in assenza del dato factbook si usa la "
              "classificazione binaria del foglio Excel. "
              "Duration e Rating Medio riguardano la sola componente obbligazionaria. "
              "Il Rating Medio di Portafoglio è la media ponderata degli score numerici "
              "(AAA=1 … D=22) × peso × quota obbligazionaria. "
              "Il simbolo — indica dato non disponibile.",
            NOTE_A),
    ]))

    story.append(Spacer(1, 14))

    if print_unp:
        # ── UNP / IUNP TABLE ─────────────────────────────────────
        # Pre-compute per-fund UNP/IUNP and portfolio weighted average
        _fund_unp: dict = {}
        _wtd_unp = _wtd_iunp = _cov_w = 0.0
        for _, _row in d_sorted.iterrows():
            _u, _iu = lookup_unp(_row["nome"])
            _fund_unp[_row["nome"]] = (_u, _iu)
            if _u is not None:
                _w = _row[wcol]
                _wtd_unp  += _u  * _w
                _wtd_iunp += _iu * _w
                _cov_w    += _w

        if _cov_w > 0.01:
            _ptf_unp_str  = f"{_wtd_unp  / _cov_w:.2f}%"
            _ptf_iunp_str = f"{_wtd_iunp / _cov_w:.2f}%"
        else:
            _ptf_unp_str = _ptf_iunp_str = "-"

        unp_hdr_row = [Paragraph(f"<b>{t}</b>", HDR) for t in
                       ["Fondo", "Peso", "%UNP", "%IUNP36", "FIDArating", "Morningstar"]]
        unp_ptf_row = [
            Paragraph(f"<b>◆ PORTAFOGLIO {ptf_name.upper()}</b>", WH),
            Paragraph("<b>100%</b>", WH),
            Paragraph(f"<b>{_ptf_unp_str}</b>",  WH),
            Paragraph(f"<b>{_ptf_iunp_str}</b>", WH),
            Paragraph("", WH),
            Paragraph("", WH),
        ]
        unp_fund_rows = []
        _unp_fida_vals: list = []
        _unp_ms_vals:   list = []
        for _, _row in d_sorted.iterrows():
            _u, _iu   = _fund_unp[_row["nome"]]
            _fd_ov_u  = (fund_data or {}).get(_row["nome"], {}).get("overview", {})
            _fida_u   = str(_fd_ov_u.get("fida_rating") or "—").strip()
            _ms_u     = _ms_pdf.get(_row["nome"], {}).get("ms_rating")
            _unp_fida_vals.append(_fida_u)
            _unp_ms_vals.append(str(_ms_u).strip() if _ms_u is not None else "—")
            unp_fund_rows.append([
                Paragraph(_row["nome"][:50], SM),
                Paragraph(f"{_row[wcol]*100:.1f}%", SM),
                Paragraph(f"{_u:.2f}%"  if _u  is not None else "—", SM),
                Paragraph(f"{_iu:.2f}%" if _iu is not None else "—", SM),
                _fida_para(_fida_u),
                _ms_para(_ms_u),
            ])

        # Per-cell background for FIDArating (col 4) and Morningstar (col 5)
        _unp_bg_cmds: list = []
        for _fi, _fv in enumerate(_unp_fida_vals):
            _bh = _FIDA_BG_HEX.get(_fv)
            if _bh:
                _unp_bg_cmds.append(("BACKGROUND", (4, _fi+2), (4, _fi+2), rl_colors.HexColor(_bh)))
        for _mi, _mv in enumerate(_unp_ms_vals):
            _bh = _MS_BG_HEX.get(_mv)
            if _bh:
                _unp_bg_cmds.append(("BACKGROUND", (5, _mi+2), (5, _mi+2), rl_colors.HexColor(_bh)))

        unp_tbl = Table(
            [unp_hdr_row, unp_ptf_row] + unp_fund_rows,
            colWidths=[6.0*cm, 1.5*cm, 2.0*cm, 2.0*cm, 2.0*cm, 4.5*cm],
            repeatRows=1,
        )
        unp_tbl.setStyle(TableStyle([
            ("BACKGROUND",     (0,0), (-1,0),  rl_colors.HexColor("#0D1B2A")),
            ("TEXTCOLOR",      (0,0), (-1,0),  rl_colors.white),
            ("FONTNAME",       (0,0), (-1,0),  "Helvetica-Bold"),
            ("BACKGROUND",     (0,1), (-1,1),  rl_colors.HexColor("#1B4332")),
            ("LINEBELOW",      (0,1), (-1,1),  2, rl_colors.HexColor("#C9A84C")),
            ("FONTSIZE",       (0,0), (-1,-1), 8),
            ("PADDING",        (0,0), (-1,-1), 5),
            ("ROWBACKGROUNDS", (0,2), (-1,-1),
             [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
            ("LINEBELOW",      (0,0), (-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
            ("ALIGN",          (0,0), (-1,-1), "CENTER"),
            ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
            *_unp_bg_cmds,
        ]))

        NOTE_U = S("NTU", fontName="Helvetica-Oblique", fontSize=6.5,
                   textColor=rl_colors.HexColor("#94A3B8"), leading=9)
        story.append(KeepTogether([
            Paragraph("UNP e IUNP dei Fondi in Portafoglio", SC),
            unp_tbl,
            Spacer(1, 6),
            Paragraph(
                "◆ UNP (Utile Netto di Portafoglio): commissione annua netta percepita dal consulente. "
                "IUNP36: indice UNP calcolato su orizzonte triennale. "
                "Fonte: Catalogo Prodotti & Servizi Azimut, settembre 2025. "
                "La riga Portafoglio è la media ponderata per peso dei fondi per cui il dato è disponibile. "
                "Il simbolo — indica che il fondo non è presente nel catalogo.",
                NOTE_U),
        ]))

    # ════════════════════════════════════════════════════════
    # CONO DI IBBOTSON
    # ════════════════════════════════════════════════════════
    try:
        _ib_mu, _ib_sig, _ib_n_ok, _ib_n_tot, _ib_missing = _az_portfolio_mu_sigma(
            d_sorted, wcol, fund_data, factbook_data, get_fb)
        _ib_cap = 100_000.0
        _ib_hor = 10

        CONE_H1 = S("CONE_H1", fontName="Helvetica-Bold", fontSize=14,
                    textColor=rl_colors.HexColor("#0D1B2A"), spaceBefore=6, spaceAfter=4)
        CONE_NT = S("CONE_NT", fontName="Helvetica-Oblique", fontSize=7,
                    textColor=rl_colors.HexColor("#64748B"), leading=10)
        CONE_SM = S("CONE_SM", fontName="Helvetica", fontSize=7.5,
                    textColor=rl_colors.HexColor("#1E293B"), leading=11, alignment=1)
        CONE_HD = S("CONE_HD", fontName="Helvetica-Bold", fontSize=7.5,
                    textColor=rl_colors.white, leading=11, alignment=1)

        _ib_png = _ibbotson_cone_png(_ib_mu, _ib_sig, _ib_cap, _ib_hor,
                                     label=f"Portafoglio {ptf_name}")
        _ib_img = RLImage(io.BytesIO(_ib_png), width=18*cm, height=6.5*cm)

        # Tabella: solo ±1σ (percentili 16°–84°)
        _ib_rows = _ibbotson_table_rows(_ib_mu, _ib_sig, _ib_cap, years=(1, 3, 5, 10))
        _ib_hdr_row = [
            Paragraph(f"<b>{h}</b>", CONE_HD)
            for h in ["Anni",
                       "Scenario sfavorevole\n5 su 6 finiscono\nsopra questa soglia",
                       "Caso centrale\n3 su 6 finiscono\nsopra questa soglia",
                       "Scenario favorevole\n1 su 6 finisce\nsopra questa soglia"]
        ]
        _ib_tbl_rows = [_ib_hdr_row]
        for _yr, _d2, _d1, _md, _u1, _u2 in _ib_rows:
            def _fmt(v): return f"€ {v:,.0f}".replace(",", ".")
            _ib_tbl_rows.append([
                Paragraph(str(_yr), CONE_SM),
                Paragraph(_fmt(_d1), CONE_SM),
                Paragraph(_fmt(_md), CONE_SM),
                Paragraph(_fmt(_u1), CONE_SM),
            ])
        _ib_tbl = Table(_ib_tbl_rows,
            colWidths=[1.4*cm, 5.8*cm, 5.8*cm, 5.8*cm])
        _ib_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), rl_colors.HexColor("#0D1B2A")),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1,-1), 8),
            ("PADDING",     (0, 0), (-1,-1), 6),
            ("ALIGN",       (0, 0), (-1,-1), "CENTER"),
            ("VALIGN",      (0, 0), (-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
            ("LINEBELOW",   (0, 0), (-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
            ("LINEBELOW",   (0, 0), (-1, 0), 1.5, rl_colors.HexColor("#C9A84C")),
        ]))

        _ib_rel = round(40 + 35 * (_ib_n_ok / max(_ib_n_tot, 1)))

        story.append(PageBreak())
        story.append(Paragraph("DEMO ANALISI", EY))
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            f"Cono di Ibbotson — Proiezione futura  |  Attendibilita' della stima: {_ib_rel}%",
            CONE_H1))
        story.append(Paragraph(
            f"Rendimento atteso: {_ib_mu*100:+.1f}% annuo  ·  "
            f"Volatilita' annua: {_ib_sig*100:.1f}%  ·  "
            f"Capitale di riferimento: € {_ib_cap:,.0f}".replace(",", "."),
            CONE_NT))
        story.append(Spacer(1, 6))
        story.append(_ib_img)
        story.append(Spacer(1, 8))
        story.append(_ib_tbl)
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "Come leggere il grafico: la linea centrale e' la mediana statistica dei percorsi simulati "
            f"(rendimento annuo composto {_ib_mu*100:+.1f}% meno il costo della varianza). "
            "La banda scura copre il 68% dei percorsi (±1 deviazione standard): "
            "in 2 anni su 3 il valore finale si collocherebbe in quell'intervallo. "
            "La banda chiara copre il 95% dei percorsi (±2 dev. std.): "
            "solo il 5% degli scenari cade al di fuori. "
            "La tabella mostra gli stessi tre scenari centrali (±1 dev. std.) per leggibilita'. "
            "Metodologia: modello log-normale di Ibbotson. "
            "Il rendimento atteso e' un prior forward-looking per categoria di fondo "
            f"(es. azionario globale ~7.5%, obbligazionario ~3%), non i rendimenti passati recenti. "
            "La volatilita' e' stimata da dati storici con soglia minima per categoria; "
            "la covarianza di portafoglio usa correlazioni categoriali (non empiriche). "
            "Non costituisce previsione garantita.",
                CONE_NT))
        if _ib_missing:
            _miss_str = "; ".join(f"{n} ({c})" for n, c in _ib_missing)
            story.append(Paragraph(
                f"&#9888; Fondi privi di dati di volatilita' storica ({len(_ib_missing)} su {_ib_n_tot}): "
                f"{_miss_str}. "
                "Per questi fondi la volatilita' e' sostituita dal valore di default categoriale.",
                CONE_NT))
    except Exception:
        pass

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════
    # PAGES 3+: SCHEDE SINGOLI FONDI
    # ════════════════════════════════════════════════════════
    story.append(Paragraph("DEMO ANALISI", EY))
    story.append(Spacer(1,4))
    story.append(Paragraph("Schede Analitiche dei Fondi", T))
    story.append(Paragraph(
        f"Profilo {profile.title()}  ·  Fonte: FIDA FondiDoc  ·  {datetime.date.today().strftime('%d %B %Y')}", SU))
    story.append(HRFlowable(width="100%",thickness=0.8,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=4))
    story.append(Paragraph(
        '🔍 <link href="https://www.morningstar.it/it/funds/SecuritySearchResults.aspx">'
        '<u>Motore di ricerca Morningstar</u></link>', LK))
    story.append(Spacer(1, 6))

    # ── Pre-conteggio per barra avanzamento Quantalys ───────────────────────
    _qtl_total = 0
    _qtl_done  = 0
    if qtl_charts and _progress_cb:
        for _, _pr in d_sorted.iterrows():
            _fd_pre   = (fund_data or {}).get(_pr["nome"], {})
            _isin_pre = _fd_pre.get("isin", "") or isin_map.get(_pr["nome"], "")
            if _isin_pre and _pdf_qtl.get(_isin_pre, ""):
                _qtl_total += 1
        if _qtl_total:
            _progress_cb(0.0, f"📊 Cattura grafici Quantalys (0/{_qtl_total})…")

    for idx, (_, row) in enumerate(d_sorted.iterrows()):
        fd  = (fund_data or {}).get(row["nome"], {})
        ov  = fd.get("overview",  {})
        ana = fd.get("analysis",  {})

        def gv(k,src=ana,fallback="-"): return src.get(k,fallback)

        srri_str = f"SRRI {gv('srri',ov,'—')}/7" if gv('srri',ov) != "-" else ""
        nav_str  = f"NAV {gv('nav')} € ({gv('last_update')})" if gv('nav') != "-" else ""
        rating_s = f"FIDArating {gv('fida_rating',ov)}" if gv('fida_rating',ov) not in ("-","—") else ""
        meta_extra = "  ·  ".join(x for x in [srri_str, rating_s, nav_str] if x)
        isin = fd.get("isin", "") or isin_map.get(row["nome"], "")
        isin_str = f"  ·  ISIN: <b>{isin}</b>" if isin else ""

        hdr_rows = [
            [Paragraph(f"<b>{row['nome']}</b>", FS)],
            [Paragraph(f"Peso: <b>{row[wcol]*100:.1f}%</b>  ·  {row['categoria']}{isin_str}", FK)],
            [Paragraph(meta_extra or "—", FK)],
        ]
        hdr_tbl = Table(hdr_rows, colWidths=[PW])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), rl_colors.HexColor("#F0F4F9")),
            ("LEFTPADDING",(0,0),(-1,-1), 8),
            ("RIGHTPADDING",(0,0),(-1,-1), 8),
            ("TOPPADDING",(0,0),(-1,0), 6),
            ("BOTTOMPADDING",(0,-1),(-1,-1), 6),
            ("TOPPADDING",(0,1),(-1,-1), 1),
            ("BOTTOMPADDING",(0,0),(-1,-2), 1),
            ("LINEBELOW",(0,-1),(-1,-1), 2, rl_colors.HexColor("#C9A84C")),
        ]))

        def pval(v):
            try:
                num = float(v.replace("%","").replace(",","."))
                c = "#1A7A4A" if num>0 else ("#C0392B" if num<0 else "#475569")
                return Paragraph(f'<font color="{c}"><b>{v}</b></font>', BD)
            except: return Paragraph(v, BD)

        perf_data = [
            [Paragraph("<b>Metrica</b>",HDR), Paragraph("<b>YTD</b>",HDR),
             Paragraph("<b>1 Anno</b>",HDR), Paragraph("<b>3 Anni</b>",HDR), Paragraph("<b>5 Anni</b>",HDR)],
            [Paragraph("Performance",SM),
             pval(gv("ytd")), pval(gv("perf_1y")), pval(gv("perf_3y")), pval(gv("perf_5y"))],
            [Paragraph("Volatilità",SM),
             Paragraph("—",SM), Paragraph(gv("vol_1y"),SM), Paragraph(gv("vol_3y"),SM), Paragraph(gv("vol_5y"),SM)],
            [Paragraph("Vol. Neg.",SM),
             Paragraph("—",SM), Paragraph(gv("neg_vol_1y"),SM), Paragraph(gv("neg_vol_3y"),SM), Paragraph(gv("neg_vol_5y"),SM)],
            [Paragraph("Sharpe",SM),
             Paragraph("—",SM), Paragraph("—",SM), Paragraph(gv("sharpe_3y"),SM), Paragraph(gv("sharpe_5y"),SM)],
            [Paragraph("Sortino",SM),
             Paragraph("—",SM), Paragraph(gv("sortino_1y"),SM), Paragraph("—",SM), Paragraph("—",SM)],
        ]
        # Larghezze colonne scheda: metriche + dettagli affiancati
        PERF_C = [2.5*cm, 1.5*cm, 1.8*cm, 1.8*cm, 1.8*cm]   # totale 9.4 cm
        DET_W  = PW - sum(PERF_C) - 0.6*cm                    # ~8.0 cm
        perf_tbl2 = Table(perf_data, colWidths=PERF_C)
        perf_tbl2.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), rl_colors.HexColor("#0D1B2A")),
            ("TEXTCOLOR",(0,0),(-1,0),  rl_colors.white),
            ("FONTNAME",(0,0),(-1,0),   "Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),  7),
            ("PADDING",(0,0),(-1,-1),   3),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.white,rl_colors.HexColor("#F8FAFC")]),
            ("LINEBELOW",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#E2E8F0")),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))

        det_data = [
            [Paragraph("<b>Dettagli Fondo</b>", BD)],
            [Paragraph(f"Data avvio: {gv('start_date',ov,'—')}", SML)],
            [Paragraph(f"Distribuzione: {gv('income',ov,'—')}", SML)],
            [Paragraph(f"Categoria: {gv('cat_assog',ov,'—')}", SML)],
            [Paragraph(f"Gestione: {gv('mgmt_fee',ov,'—')}  |  Perf.: {gv('perf_fee',ov,'—')}", SML)],
            [Paragraph(f"Sottoscrizione: {gv('sub_fee',ov,'—')}", SML)],
            [Paragraph(f"<b>FIDArating:</b> {gv('fida_rating',ov,'—')}  |  Score: {gv('fida_score',ov,'—')}", SML)],
        ]
        det_tbl = Table([[d[0]] for d in det_data], colWidths=[DET_W])
        det_tbl.setStyle(TableStyle([
            ("PADDING",(0,0),(-1,-1), 2),
            ("TOPPADDING",(0,0),(-1,0), 5),
            ("LINEBELOW",(0,0),(0,0), 0.8, rl_colors.HexColor("#C9A84C")),
            ("BACKGROUND",(0,0),(0,-1), rl_colors.HexColor("#F8FAFC")),
        ]))

        mid_row = Table([[perf_tbl2, det_tbl]],
                        colWidths=[sum(PERF_C), DET_W + 0.6*cm])
        mid_row.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1), "TOP"),
            ("PADDING",(0,0),(-1,-1), 0),
            ("LEFTPADDING",(1,0),(1,-1), 8),
        ]))

        annual  = ana.get("annual_perf")
        bar_buf = _mpl_annual_bar(annual, row["nome"]) if annual else None

        # KeepTogether: scheda compatta (≈ 2 per pagina)
        card = [Spacer(1,4), hdr_tbl, Spacer(1,4), mid_row]
        if bar_buf:
            card += [Spacer(1,3),
                     Paragraph("<b>Performance Annuale (%)</b>", SM),
                     RLImage(bar_buf, width=PW, height=2.4*cm)]

        # ── Grafici Quantalys (serie storica + performance comparata) ───────────
        if qtl_charts and isin:
            _qurl_chart = _pdf_qtl.get(isin, "")
            if _qurl_chart:
                _qtl_done += 1
                if _progress_cb and _qtl_total > 0:
                    _nm_short = row["nome"][:38] + ("…" if len(row["nome"]) > 38 else "")
                    _progress_cb(
                        (_qtl_done - 1) / _qtl_total,
                        f"📊 {_nm_short}  ({_qtl_done}/{_qtl_total})"
                    )
                _hist_url  = _qtl_to_historique_url(_qurl_chart)
                _qtl_png   = _capture_qtl_6charts(_hist_url)
                if _qtl_png:
                    _qtl_io  = io.BytesIO(_qtl_png)
                    # Calcola aspect ratio dall'immagine reale (include eventuale strip legenda)
                    try:
                        from PIL import Image as _PILImg
                        _qtl_pil = _PILImg.open(io.BytesIO(_qtl_png))
                        _qtl_ar  = _qtl_pil.height / max(_qtl_pil.width, 1)
                    except Exception:
                        _qtl_ar  = 739 / 1186   # fallback empirico
                    _qtl_w   = PW * 0.98
                    _qtl_h   = _qtl_w * _qtl_ar
                    card += [Spacer(1,4),
                             Paragraph(
                                 "<b>Analisi Quantalys</b>"
                                 " · Serie Storica &amp; Performance Comparata",
                                 SM),
                             RLImage(_qtl_io, width=_qtl_w, height=_qtl_h)]

        story.append(KeepTogether(card))

        # Separatore sottile tra schede
        if idx < len(d_sorted) - 1:
            story.append(HRFlowable(width="100%", thickness=0.5,
                                    color=rl_colors.HexColor("#CBD5E1"),
                                    spaceBefore=4, spaceAfter=4))

    # ── FOOTER ─────────────────────────────────────────────
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%",thickness=0.5,color=rl_colors.HexColor("#E2E8F0"),spaceAfter=6))
    story.append(Paragraph(
        "Documento generato automaticamente a scopo illustrativo. I dati di performance provengono da FIDA FondiDoc "
        "(fondidoc.it). I pesi indicati sono riferiti al portafoglio modello e non costituiscono offerta o consulenza "
        "di investimento. Rendimenti passati non garantiscono risultati futuri. © Azimut Group — uso interno.", FT))

    if _progress_cb:
        _progress_cb(0.97, "⚡ Assemblo PDF…")
    doc.build(story)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════
# FREE PORTFOLIO BUILDER
# ════════════════════════════════════════════════════════════

def free_portfolio_ui(data):
    fida = data.get("FIDA", pd.DataFrame())
    if fida.empty:
        st.error("❌ Foglio FIDA non trovato."); return None

    az_lookup = {}
    for sname in ["PTF FULL","PTF SHORT"]:
        if sname in data and not data[sname].empty:
            for _,r in data[sname].iterrows(): az_lookup[r["nome"]] = r["az_pct"]

    if "free_ptf" not in st.session_state: st.session_state.free_ptf = []
    st.markdown('<p class="sec-title">Costruttore Portafoglio Libero</p>',unsafe_allow_html=True)

    # ── FIDArating filter ─────────────────────────────────────────────────────
    _fd_live_free = st.session_state.get("_scomp_fd") or load_fund_cache()[0]
    # Build rating map: fund_name → "5"/"4"/"3"/"2"/"1"/"—"
    _fr_map = {
        r["nome"]: (str(_fd_live_free.get(r["nome"], {})
                        .get("overview", {}).get("fida_rating", "") or "").strip()
                    or "—")
        for _, r in fida.iterrows()
    }
    _has_ratings = any(v != "—" for v in _fr_map.values())

    # ── Morningstar filter ────────────────────────────────────────────────────
    _ms_live_free = st.session_state.get("_ms_data") or load_ms_cache()
    _ms_fr_map = {
        r["nome"]: (str(_ms_live_free.get(r["nome"], {}).get("ms_rating", "") or "").strip()
                    or "—")
        for _, r in fida.iterrows()
    }
    _has_ms_ratings = any(v != "—" for v in _ms_fr_map.values())

    _RATING_OPTS  = ["5", "4", "3", "2", "1", "—"]
    _FIDA_LABEL = {
        "5": "⭐⭐⭐⭐⭐  FIDArating 5",
        "4": "⭐⭐⭐⭐  FIDArating 4",
        "3": "⭐⭐⭐  FIDArating 3",
        "2": "⭐⭐  FIDArating 2",
        "1": "⭐  FIDArating 1",
        "—": "Nessun rating",
    }
    _MS_LABEL = {
        "5": "★★★★★  Morningstar 5",
        "4": "★★★★  Morningstar 4",
        "3": "★★★  Morningstar 3",
        "2": "★★  Morningstar 2",
        "1": "★  Morningstar 1",
        "—": "Nessun rating",
    }

    # Layout: two filter columns side by side
    _fcol1, _fcol2 = st.columns(2)
    with _fcol1:
        if _has_ratings:
            _sel_fida = st.multiselect(
                "🔵  Filtra per FIDArating",
                options=_RATING_OPTS,
                default=_RATING_OPTS,
                format_func=lambda x: _FIDA_LABEL[x],
                key="free_fida_filter",
            )
            _active_fida = set(_sel_fida) if _sel_fida else set(_RATING_OPTS)
        else:
            _active_fida = set(_RATING_OPTS)
            st.caption("ℹ️ FIDArating — scarica dati FondiDoc")
    with _fcol2:
        if _has_ms_ratings:
            _sel_ms = st.multiselect(
                "⭐  Filtra per Morningstar",
                options=_RATING_OPTS,
                default=_RATING_OPTS,
                format_func=lambda x: _MS_LABEL[x],
                key="free_ms_filter",
            )
            _active_ms = set(_sel_ms) if _sel_ms else set(_RATING_OPTS)
        else:
            _active_ms = set(_RATING_OPTS)
            st.caption("ℹ️ Morningstar — scarica rating FondiOnline")

    fida_filtered = fida[fida["nome"].apply(
        lambda n: (
            _fr_map.get(n, "—") in _active_fida
            and _ms_fr_map.get(n, "—") in _active_ms
        )
    )]
    if fida_filtered.empty:
        st.warning("⚠️ Nessun fondo corrisponde ai filtri selezionati.")
        fida_filtered = fida  # fallback: show all

    # Build option labels: include FIDArating, Morningstar badges and ISIN (for search)
    def _fund_option(r):
        fr   = _fr_map.get(r["nome"], "—")
        ms_r = _ms_fr_map.get(r["nome"], "—")
        ftag = f" · F{fr}"   if fr   != "—" else ""
        mtag = f" · M{ms_r}" if ms_r != "—" else ""
        isin = str(r.get("isin", "") or "").strip()
        isin_tag = f"  {isin}" if isin else ""
        if r["macro_cat"] != "Altro":
            return f"{r['nome']}{ftag}{mtag}  [{r['macro_cat']}]{isin_tag}"
        return f"{r['nome']}{ftag}{mtag}{isin_tag}"

    options = fida_filtered.apply(_fund_option, axis=1).tolist()

    # ── Aggiunge fondi pensione dalla cache FP ────────────────────────────────
    _fp_free = st.session_state.get("_fp_data") or load_fp_cache()
    _fp_names = [k for k in (_fp_free or {}) if k != "_ref_date"]

    def _fp_az(nome: str) -> float:
        """Stima az_pct dal nome del comparto pensionistico."""
        n = nome.lower()
        if "crescita"      in n: return 0.75
        if "accrescitivo"  in n: return 0.65
        if "equilibrato"   in n: return 0.50
        if "bilanciato"    in n: return 0.50
        if "conservativo"  in n: return 0.30
        if "obbligazionario" in n: return 0.05
        return 0.50

    _fp_options = [f"{n}  [🏦 Pensione]" for n in _fp_names]
    options = options + _fp_options

    # st.multiselect has native live search built into Streamlit (no Enter needed).
    # max_selections=1 limits it to a single fund, giving us a searchable picker.
    c1,c2,c3 = st.columns([3.5,1,0.8])
    with c1:
        _sel_list = st.multiselect(
            "🔍  Seleziona / cerca fondo:",
            options=options,
            max_selections=1,
            placeholder="Digita nome, ISIN o «pensione»…",
            key="sel_fund_ms",
        )
        sel = _sel_list[0] if _sel_list else (options[0] if options else "")
    with c2: w = st.number_input("Peso %",0.1,100.0,10.0,0.5,key="sel_w")
    with c3:
        st.markdown("<br>",unsafe_allow_html=True)
        if st.button("➕ Aggiungi",use_container_width=True):
            # Strip FIDArating "· FN", Morningstar "· MN", macro-cat "  [...]", ISIN "  LU..."
            fname = re.split(r'\s+·\s+[FM]\d|\s{2}\[|\s{2}[A-Z]{2}[A-Z0-9]{10}', sel)[0].strip()
            if any(f["nome"]==fname for f in st.session_state.free_ptf):
                st.toast("⚠️ Fondo già presente!",icon="⚠️")
            else:
                # Fondi pensione: non sono nel foglio FIDA
                _is_fp_fund = fname in _fp_names
                if _is_fp_fund:
                    _az = _fp_az(fname)
                    _mc = ("Obbligazionari" if _az < 0.15
                           else "Bilanciati/Flessibili" if _az < 0.70
                           else "Azionari")
                    st.session_state.free_ptf.append({
                        "nome": fname, "categoria": "Fondo Pensione",
                        "macro_cat": _mc, "az_pct": _az, "w_input": w})
                else:
                    fd = fida[fida["nome"]==fname].iloc[0] if not fida[fida["nome"]==fname].empty else None
                    mc = fd["macro_cat"] if fd is not None else "Altro"
                    az = az_lookup.get(fname, DEFAULT_AZ.get(mc, 0.5))
                    st.session_state.free_ptf.append({
                        "nome": fname,
                        "categoria": fd["categoria"] if fd is not None else "",
                        "macro_cat": mc, "az_pct": az, "w_input": w})
                st.rerun()

    # ── Carica portafoglio salvato ────────────────────────────────────────────
    _saved = load_saved_portfolios()
    if _saved:
        _saved_names = ["— Nuovo portafoglio —"] + sorted(_saved.keys())
        _sc1, _sc2 = st.columns([3, 1])
        with _sc1:
            _sel_saved = st.selectbox(
                "📂  Carica portafoglio salvato",
                options=_saved_names,
                key="sel_saved_ptf",
            )
        with _sc2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("📂 Carica", key="btn_load_saved",
                         use_container_width=True,
                         disabled=(_sel_saved == "— Nuovo portafoglio —")):
                _ptf_to_load = _saved.get(_sel_saved, {})
                _fondi_saved = _ptf_to_load.get("fondi", [])
                if _fondi_saved:
                    new_free = []
                    for _fs in _fondi_saved:
                        _fn = _fs.get("nome", "")
                        _fw = float(_fs.get("peso", 10.0))
                        _fd_row = fida[fida["nome"] == _fn]
                        if not _fd_row.empty:
                            _fd_r = _fd_row.iloc[0]
                            _mc = _fd_r["macro_cat"]
                            _az = az_lookup.get(_fn, DEFAULT_AZ.get(_mc, 0.5))
                            _cat = _fd_r["categoria"]
                        elif _fn in _fp_names:
                            # Fondo pensione salvato
                            _az  = _fp_az(_fn)
                            _mc  = ("Obbligazionari" if _az < 0.15
                                    else "Bilanciati/Flessibili" if _az < 0.70
                                    else "Azionari")
                            _cat = "Fondo Pensione"
                        else:
                            _mc, _az, _cat = "Altro", 0.5, ""
                        new_free.append({"nome": _fn, "categoria": _cat,
                                         "macro_cat": _mc, "az_pct": _az,
                                         "w_input": _fw})
                    st.session_state.free_ptf = new_free
                    st.toast(f"✅ Portafoglio '{_sel_saved}' caricato", icon="📂")
                    st.rerun()
        if _sel_saved != "— Nuovo portafoglio —":
            _ptf_info = _saved.get(_sel_saved, {})
            _ptf_date = _ptf_info.get("date", "")
            _n_fondi_s = len(_ptf_info.get("fondi", []))
            _dc1, _dc2 = st.columns([3, 1])
            with _dc1:
                st.caption(f"💾 Salvato il {_ptf_date} · {_n_fondi_s} fondi")
            with _dc2:
                if st.button("🗑️ Elimina", key="btn_del_saved",
                             use_container_width=True):
                    delete_portfolio(_sel_saved)
                    st.toast(f"🗑️ '{_sel_saved}' eliminato", icon="🗑️")
                    st.rerun()

    if not st.session_state.free_ptf: st.info("☝️ Aggiungi fondi."); return None
    st.markdown("**Fondi nel portafoglio:**")
    total_w = 0.0
    for i, fund in enumerate(st.session_state.free_ptf):
        r1,r2,r3 = st.columns([4,1.5,0.6])
        with r1: st.markdown(f"**{fund['nome']}** <span style='color:#64748b;font-size:.8rem;'>— {fund['macro_cat']}</span>",unsafe_allow_html=True)
        with r2:
            nw = st.number_input("Peso",0.0,100.0,float(fund["w_input"]),0.5,key=f"fw_{i}",label_visibility="collapsed")
            st.session_state.free_ptf[i]["w_input"] = nw
        with r3:
            if st.button("🗑️",key=f"del_{i}",use_container_width=True):
                st.session_state.free_ptf.pop(i); st.rerun()
        total_w += st.session_state.free_ptf[i]["w_input"]

    diff = abs(total_w-100.0)
    if diff<0.05: st.markdown(f'<div class="w-ok">✅ Somma pesi: <b>{total_w:.1f}%</b> — OK!</div>',unsafe_allow_html=True)
    else:         st.markdown(f'<div class="w-warn">⚠️ Somma pesi: <b>{total_w:.1f}%</b> (mancano {100-total_w:+.1f}%)</div>',unsafe_allow_html=True)
    if diff>0.5: return None

    # ── Salva portafoglio ─────────────────────────────────────────────────────
    st.markdown("---")
    _sv1, _sv2, _sv3 = st.columns([2.5, 1, 1])
    with _sv1:
        _ptf_save_name = st.text_input(
            "💾  Nome portafoglio",
            placeholder="es. Pippo Rossi — Conservativo",
            key="ptf_save_name",
            label_visibility="collapsed",
        )
    with _sv2:
        if st.button("💾 Salva", key="btn_save_ptf", use_container_width=True,
                     disabled=not _ptf_save_name.strip()):
            _fondi_to_save = [{"nome": f["nome"], "peso": f["w_input"]}
                              for f in st.session_state.free_ptf]
            save_portfolio(_ptf_save_name.strip(), _fondi_to_save)
            st.toast(f"✅ Portafoglio '{_ptf_save_name.strip()}' salvato", icon="💾")
            st.rerun()
    with _sv3:
        if st.button("🗑️ Svuota", key="btn_clear_ptf", use_container_width=True):
            st.session_state.free_ptf = []
            st.rerun()

    records = [{"nome":f["nome"],"categoria":f["categoria"],"gruppo":f["macro_cat"],"macro_cat":f["macro_cat"],"az_pct":f["az_pct"],"obb_pct":1-f["az_pct"],"r_weight":f["w_input"]/100,"w_cons":f["w_input"]/100,"w_equil":f["w_input"]/100,"w_accr":f["w_input"]/100} for f in st.session_state.free_ptf]
    df = pd.DataFrame(records)
    df["macro_cat"] = df["categoria"].apply(get_macro)
    df = assign_colors(df)
    for wc in ["w_cons","w_equil","w_accr"]:
        t = df[wc].sum(); df[wc] = df[wc]/t if t>0 else df[wc]
    return df


# ════════════════════════════════════════════════════════════
# GLOBAL PERSPECTIVES — PDF parsing & SUGGERITO portfolio
# ════════════════════════════════════════════════════════════

def _resolve_nome_for_fd(nome_pdf: str, fund_data: dict) -> str:
    """Map a PDF-format name (e.g. "AZ Allocation - Balanced Plus") to the
    actual key present in fund_data (usually an Excel-abbreviated form like
    "AZ F.1 All. Balanced Plus A Cap EUR").  Uses the same normalisation logic
    as UNP lookup.  Falls back to nome_pdf if no match is found."""
    if not fund_data or nome_pdf in fund_data:
        return nome_pdf
    norm = _normalize_for_unp(nome_pdf)
    norm = _FUND_ALIASES.get(norm, norm)
    best_key, best_len = None, 0
    for key in fund_data:
        k_norm = _normalize_for_unp(key)
        k_norm = _FUND_ALIASES.get(k_norm, k_norm)
        if k_norm == norm:
            return key                               # exact normalised match
        if (k_norm in norm or norm in k_norm) and len(k_norm) > best_len:
            best_key, best_len = key, len(k_norm)
    return best_key if best_key else nome_pdf


def parse_global_perspectives(pdf_bytes: bytes):
    """Parse a *Global Perspectives* quarterly PDF and return the three
    Azimut View scenario portfolios (Base, Bear, Bull), excluding private-
    market funds (ELTIF, RAIF, Demos, …).

    Returns
    -------
    dict | None
        ``{
            "Base": {
                "info": "Equity 32% · Bond 38% · Private Markets 30%",
                "funds": [
                    {"nome": "AZ Allocation - Balanced Plus",
                     "gruppo": "ALLOCATION",
                     "categoria": "Bilanciati/Flessibili",
                     "az_pct": 0.50, "obb_pct": 0.50, "weight": 0.045},
                    ...
                ],
                "subcat_weights": {"alloc_balanced": 25, ...},
            },
            "Bear": {...},
            "Bull": {...},
        }``
    or ``None`` if the PDF could not be recognised.
    """
    try:
        import pdfplumber
    except ImportError:
        return None
    import io as _io

    # ── 1. Extract text page by page ──────────────────────────────────────────
    # Estrae ENTRAMBI i formati in un solo passaggio:
    #  • extract_text()           → default, per section/fund parsing
    #  • extract_text(layout=True) → preserva colonne, per pesi torta
    try:
        pages        = []
        pages_layout = []
        with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
            for pg in pdf.pages:
                t = pg.extract_text() or ""
                if t:
                    pages.append(t)
                try:
                    tl = pg.extract_text(layout=True) or ""
                except Exception:
                    tl = ""
                pages_layout.append(tl)
    except Exception:
        return None
    if not pages:
        return None
    full        = "\n".join(pages)
    full_layout = "\n".join(pages_layout)

    # ── 2. Locate scenario section boundaries ─────────────────────────────────
    _SC_PATS: dict = {
        "Base": [r"AZIMUT\s+VIEW\s+SCENARIO\s+BASE", r"Scenario\s+Base"],
        "Bear": [r"AZIMUT\s+VIEW\s+SCENARIO\s+BEAR", r"Scenario\s+Bear"],
        "Bull": [r"AZIMUT\s+VIEW\s+SCENARIO\s+BULL", r"Scenario\s+Bull"],
    }
    positions: dict = {}
    for sc, pats in _SC_PATS.items():
        for pat in pats:
            m = re.search(pat, full, re.IGNORECASE)
            if m:
                positions[sc] = m.start()
                break
    if len(positions) < 3:
        return None

    sorted_sc = sorted(positions, key=lambda k: positions[k])
    sections: dict = {}
    for i, sc in enumerate(sorted_sc):
        start = positions[sc]
        end   = positions[sorted_sc[i + 1]] if i + 1 < len(sorted_sc) else len(full)
        sections[sc] = full[start:end]

    # ── 2b. Sezioni layout=True per estrazione colonne torta ──────────────────
    positions_l: dict = {}
    for sc, pats in _SC_PATS.items():
        for pat in pats:
            ml = re.search(pat, full_layout, re.IGNORECASE)
            if ml:
                positions_l[sc] = ml.start()
                break
    sorted_sl = sorted(positions_l, key=lambda k: positions_l[k])
    sections_l: dict = {}
    for i, sc in enumerate(sorted_sl):
        sl_s = positions_l[sc]
        sl_e = positions_l[sorted_sl[i + 1]] if i + 1 < len(sorted_sl) else len(full_layout)
        sections_l[sc] = full_layout[sl_s:sl_e]

    # ── 3. Sub-category meta ──────────────────────────────────────────────────
    _SUBCAT_LABELS: list = [
        ("alloc_balanced",  r"Allocation\s*[-–]\s*Balanced"),
        ("alloc_flexible",  r"Allocation\s*[-–]\s*Flex"),
        ("bond_aggregate",  r"Bond\s*[-–]\s*Aggregate"),
        ("bond_thematic",   r"Bond\s*[-–]\s*Thematic"),
        ("bond_em",         r"Bond\s*[-–]\s*Paesi\s+emergenti"),
        ("bond_target",     r"Bond\s*[-–]\s*Target"),
        ("equity_thematic", r"Equity\s*[-–]\s*Thematic"),
        ("equity_dev",      r"Equity\s*[-–]\s*Paesi\s+sviluppati"),
        ("equity_em",       r"Equity\s*[-–]\s*Paesi\s+emergenti"),
    ]
    _SUBCAT_GROUP = {
        "alloc_balanced": "ALLOCATION",   "alloc_flexible": "ALLOCATION",
        "bond_aggregate": "BOND",         "bond_thematic":  "BOND",
        "bond_em":        "BOND",         "bond_target":    "BOND",
        "equity_thematic":"AZIONARI (LONG)",
        "equity_dev":     "AZIONARI (LONG)", "equity_em": "AZIONARI (LONG)",
    }
    _SUBCAT_CAT = {
        "alloc_balanced":  "Bilanciati/Flessibili",
        "alloc_flexible":  "Bilanciati/Flessibili",
        "bond_aggregate":  "Obbligazionari", "bond_thematic": "Obbligazionari",
        "bond_em":         "Obbligazionari", "bond_target":   "Obbligazionari",
        "equity_thematic": "Azionari",
        "equity_dev":      "Azionari",       "equity_em":     "Azionari",
    }
    # Private-market keywords → skip these fund lines
    _PRIV_KW = frozenset([
        "raif", "eltif", "demos ", "yhox", "direct investments",
        "hybrid growth", "automobile", "infrastrutture", "real assets",
        "digitech fund", "young group", "alicrowd", "hipstr", " p103",
        "italia 500", "globALinvest", "borletti", "broadlight", "highpost",
        "roundshield", "pensinsula", "ophelia", "gp stakes", "kennedy lewis",
        "digital assets", "bcp asia", "valsabbina", "d-orbit",
        "escalator 1", "escalator 2",
    ])

    def _is_priv(name: str) -> bool:
        n = name.lower()
        return any(k in n for k in _PRIV_KW)

    # ── 4. Parse each scenario ────────────────────────────────────────────────
    result: dict = {}

    for sc_name, sect in sections.items():
        # ── 4a. Sub-category weights from pie-chart text ──────────────────────
        fc_m  = re.search(r"Fondi\s+consigliati", sect, re.IGNORECASE)
        pie   = sect[:fc_m.start()] if fc_m else sect
        sw: dict = {}

        # Strategia colonna: usa il testo layout=True che preserva le
        # posizioni x dei caratteri (come pdftotext -layout). Così la
        # torta a 2 colonne mantiene l'allineamento visivo e il numero
        # che appartiene a un'etichetta è quello alla stessa colonna
        # (stessa posizione orizzontale) sulle 2 righe precedenti.
        sect_l  = sections_l.get(sc_name, "")
        fc_ml   = re.search(r"Fondi\s+consigliati", sect_l, re.IGNORECASE) if sect_l else None
        pie_l   = (sect_l[:fc_ml.start()] if fc_ml else sect_l) if sect_l else ""
        pie_ll  = pie_l.split('\n') if pie_l else []

        for key, lbl_pat in _SUBCAT_LABELS:
            found = False

            # ── Prima scelta: colonna dal testo layout ────────────────────
            if pie_ll:
                for li, lline in enumerate(pie_ll):
                    lm = re.search(lbl_pat, lline, re.IGNORECASE)
                    if not lm:
                        continue
                    label_col = lm.start()
                    best_v, best_col_dist = None, 999
                    # Guarda la stessa riga (prima dell'etichetta) e le 2
                    # righe precedenti: cerca il numero con la posizione
                    # orizzontale più vicina a quella dell'etichetta.
                    for lj in range(max(0, li - 2), li + 1):
                        src = pie_ll[lj]
                        if lj == li:
                            src = src[:label_col]   # solo prima dell'etichetta
                        for mn in re.finditer(r'(\d{1,2})\s*%', src):
                            v = int(mn.group(1))
                            if 1 <= v <= 50:
                                col_dist = abs(mn.start() - label_col)
                                if col_dist < best_col_dist:
                                    best_v, best_col_dist = v, col_dist
                    if best_v is not None:
                        sw[key] = best_v
                        found = True
                        break

            # ── Fallback: distanza caratteri nel testo standard ───────────
            if not found:
                for m in re.finditer(lbl_pat, pie, re.IGNORECASE):
                    win_start = max(0, m.start() - 200)
                    win = pie[win_start: m.end() + 50]
                    best_v2, best_dist = None, 999999
                    for mn in re.finditer(r'(\d{1,2})\s*%', win):
                        v = int(mn.group(1))
                        if 1 <= v <= 50:
                            num_pos = win_start + mn.start()
                            dist = abs(num_pos - m.start())
                            if dist < best_dist:
                                best_v2, best_dist = v, dist
                    if best_v2 is not None:
                        sw[key] = best_v2
                        break

        # ── 4b. Parse "Fondi consigliati" section ─────────────────────────────
        if not fc_m:
            continue
        fc_txt = sect[fc_m.start():]

        cur_group:  str | None = None
        cur_subcat: str | None = None
        fund_subcat: dict = {}      # fund_name → subcat_key

        for line in fc_txt.split("\n"):
            l = line.strip()
            if not l:
                continue
            # — Group headers —
            if   re.match(r'^ALLOCATION$',    l, re.I): cur_group = "allocation";  cur_subcat = None
            elif re.match(r'^BOND$',          l, re.I): cur_group = "bond";         cur_subcat = None
            elif re.match(r'^EQUITY$',        l, re.I): cur_group = "equity";       cur_subcat = None
            elif re.match(r'^PRIVATE\s',      l, re.I): cur_group = "private";      cur_subcat = None
            # — Sub-category headers —
            elif re.match(r'^BALANCED$',      l, re.I): cur_subcat = "alloc_balanced"
            elif re.match(r'^FLEXIBLE$',      l, re.I): cur_subcat = "alloc_flexible"
            elif re.match(r'^AGGREGATE',      l, re.I): cur_subcat = "bond_aggregate"
            elif re.match(r'^THEMATIC$',      l, re.I):
                cur_subcat = "bond_thematic" if cur_group == "bond" else "equity_thematic"
            elif re.match(r'^TARGET',         l, re.I): cur_subcat = "bond_target"
            elif re.match(r'^PAESI\s+EMERGENTI$', l, re.I):
                cur_subcat = "bond_em" if cur_group == "bond" else "equity_em"
            elif re.match(r'^PAESI\s+SVILUPPATI$', l, re.I): cur_subcat = "equity_dev"
            elif re.match(r'^EMERGENTI$',     l, re.I):
                cur_subcat = "bond_em" if cur_group == "bond" else "equity_em"
            elif re.match(r'^SVILUPPATI$',    l, re.I): cur_subcat = "equity_dev"
            # — Fund lines (split on each "AZ Fund 1 -" to handle 2-column layout) —
            elif re.search(r'AZ\s+Fund\s+1\s*[-–]', l, re.I):
                for _part in re.split(r'(?=AZ\s+Fund\s+1\s*[-–])', l, flags=re.I):
                    _part = _part.strip()
                    _m = re.match(r'AZ\s+Fund\s+1\s*[-–]\s*(AZ\s+\S.+)', _part, re.I)
                    if _m and cur_subcat and cur_group != "private":
                        # Truncate at any additional "AZ Fund 1" still present
                        raw_nm = re.split(
                            r'\s+AZ\s+Fund\s+1\s*[-–]', _m.group(1), flags=re.I
                        )[0].strip().rstrip("*").strip()
                        if raw_nm and not _is_priv(raw_nm):
                            fund_subcat[raw_nm] = cur_subcat

        # ── 4c. Compute equal-weight per sub-category ─────────────────────────
        subcat_funds: dict = {}
        for fname, sc_key in fund_subcat.items():
            subcat_funds.setdefault(sc_key, []).append(fname)

        total_liq = sum(sw.get(k, 0) for k in subcat_funds)
        if total_liq == 0:
            total_liq = sum(sw.values()) or 70

        records: list = []
        for sc_key, funds in subcat_funds.items():
            w_sc = sw.get(sc_key, 0)
            if not funds:
                continue
            w_per = ((w_sc / len(funds)) / total_liq) if w_sc else (1.0 / max(len(fund_subcat), 1))
            grp = _SUBCAT_GROUP.get(sc_key, "ALLOCATION")
            cat = _SUBCAT_CAT.get(sc_key, "Altro")
            az  = DEFAULT_AZ.get(get_macro(cat), 0.5)
            for fname in funds:
                records.append({
                    "nome":     fname,
                    "gruppo":   grp,
                    "categoria": cat,
                    "az_pct":   az,
                    "obb_pct":  1.0 - az,
                    "weight":   w_per,
                    "subcat":   sc_key,
                })

        # ── 4d. Info string from summary paragraph ────────────────────────────
        info_m = re.search(
            r'azioni\s+(\d+)%.*?obbligazioni\s+(\d+)%.*?private\s+markets\s+(\d+)%',
            sect, re.IGNORECASE | re.DOTALL)
        info = (f"Equity {info_m.group(1)}% · Bond {info_m.group(2)}%"
                f" · Private Markets {info_m.group(3)}%") if info_m else ""

        result[sc_name] = {
            "info":           info,
            "funds":          records,
            "subcat_weights": sw,
        }

    # ── 5. Estrai edizione / trimestre dalla prima pagina ────────────────────────
    if result:
        _gp_edition = ""
        _cover = pages[0] if pages else ""
        # Cerca "Q1 2025" / "Q2 2026" ecc.
        _em = re.search(r'\bQ([1-4])\s*(20\d{2})\b', _cover)
        if _em:
            _gp_edition = f"Q{_em.group(1)} {_em.group(2)}"
        if not _gp_edition:
            # Cerca "1° Trimestre 2025" o "Trimestre 1 2025"
            _em = re.search(r'(\d)\s*[°o]?\s*[Tt]rimestre\s*(20\d{2})', _cover)
            if not _em:
                _em = re.search(r'[Tt]rimestre\s+(\d)\s*(20\d{2})', _cover)
            if _em:
                _qmap = {"1": "Q1", "2": "Q2", "3": "Q3", "4": "Q4"}
                _gp_edition = f"{_qmap.get(_em.group(1), 'Q?')} {_em.group(2)}"
        if not _gp_edition:
            # Cerca "N° 2" / "N. 2" / "Nr. 2" seguito da anno
            _em = re.search(r'N[°o\.r]*\s*(\d+)\D{0,10}(20\d{2})', _cover, re.IGNORECASE)
            if _em:
                _gp_edition = f"N° {_em.group(1)} - {_em.group(2)}"
        if not _gp_edition:
            # Cerca "N° 2" / "N. 2" senza anno esplicito vicino (anno su riga separata)
            _em_n = re.search(r'N[°o\.r]*\s*(\d+)', _cover, re.IGNORECASE)
            _em_y = re.search(r'\b(20\d{2})\b', _cover)
            if _em_n and _em_y:
                _gp_edition = f"N° {_em_n.group(1)} - {_em_y.group(1)}"
        if not _gp_edition:
            # Cerca "Edizione N" o "Edition N"
            _em = re.search(r'[Ee]dizione\s+(\d+)', _cover)
            if _em:
                _gp_edition = f"Ed. {_em.group(1)}"
        if not _gp_edition:
            # Ultimo tentativo: prime 400 chars del testo completo (copertina)
            _em = re.search(r'\bQ([1-4])\s*(20\d{2})\b', full[:400])
            if _em:
                _gp_edition = f"Q{_em.group(1)} {_em.group(2)}"
        if _gp_edition:
            result["_edition"] = _gp_edition

    return result if result else None


# ── Module-level badge helpers (used in suggerito_portfolio_ui) ──────────────
_FIDA_BG_GP  = {5:"#166534", 4:"#15803d", 3:"#16a34a", 2:"#64748B", 1:"#94A3B8"}
_MS_BG_GP    = {5:"#78350F", 4:"#92400E", 3:"#B45309"}
_MS_COL_GP   = {5:"#78350F", 4:"#92400E", 3:"#B45309", 2:"#475569", 1:"#94A3B8"}

def _fida_badge_gp(r) -> str:
    try:    v = int(r)
    except (TypeError, ValueError): return "<span style='color:#94A3B8;'>—</span>"
    bg = _FIDA_BG_GP.get(v)
    return (f"<span style='background:{bg};color:#fff;padding:2px 8px;"
            f"border-radius:4px;font-weight:700;font-size:.8rem;'>{v}</span>"
            if bg else f"<span style='color:#64748B;font-weight:700;font-size:.8rem;'>{v}</span>")

def _ms_badge_gp(ms_r) -> str:
    try:    v = int(ms_r)
    except (TypeError, ValueError): return "<span style='color:#94A3B8;'>—</span>"
    filled = "★" * v
    bg = _MS_BG_GP.get(v)
    if bg:
        return (f"<span style='background:{bg};color:#fff;padding:2px 8px;"
                f"border-radius:4px;font-weight:700;font-size:.8rem;'>{filled}</span>")
    _col = _MS_COL_GP.get(v, "#64748B")
    return (f"<span style='color:{_col};font-weight:700;"
            f"font-size:.8rem;'>{filled}</span>")

def _qtl_rating_cell(isin: str, ratings: dict) -> str:
    """Restituisce HTML del rating Quantalys (stelle + score) per un dato ISIN."""
    v = ratings.get(isin) if isin else None
    if not v:
        return "<span style='color:#CBD5E1;font-size:.75rem;'>—</span>"
    score  = v.get("score")
    globes = v.get("globes")
    if score is None:
        return "<span style='color:#CBD5E1;font-size:.75rem;'>—</span>"
    if globes:
        _gc = {1:"#EF4444",2:"#F97316",3:"#EAB308",4:"#22C55E",5:"#1B4FBB"}.get(globes,"#64748B")
        stars = (f"<span style='color:{_gc};font-size:.85rem;letter-spacing:1px;'>"
                 f"{'★'*globes}{'☆'*(5-globes)}</span>")
    else:
        stars = ""
    sc_col = ("#1B4FBB" if score >= 80 else "#22C55E" if score >= 60
              else "#EAB308" if score >= 40 else "#EF4444")
    sc_html = (f"<span style='font-size:.7rem;font-weight:700;color:{sc_col};"
               f"background:#F1F5F9;border-radius:3px;padding:1px 4px;'>{score}</span>")
    tip = f"Score {score}/100" + (f" · {globes} globi" if globes else "")
    return f"<span title='{tip}'>{stars}{'<br>' if stars else ''}{sc_html}</span>"


# Sub-category display names (Italian labels)
_SUBCAT_DISPLAY = {
    "alloc_balanced":  "Allocation – Balanced",
    "alloc_flexible":  "Allocation – Flexible",
    "bond_aggregate":  "Bond – Aggregate / Gov",
    "bond_thematic":   "Bond – Thematic",
    "bond_em":         "Bond – Paesi Emergenti",
    "bond_target":     "Bond – Target Maturity",
    "equity_thematic": "Equity – Thematic",
    "equity_dev":      "Equity – Paesi Sviluppati",
    "equity_em":       "Equity – Paesi Emergenti",
}


def suggerito_portfolio_ui(sc_name: str, gp_scenario: dict,
                           fund_data: dict, ms_data: dict,
                           extra_urls: dict | None = None):
    """Interactive portfolio builder for a SUGGERITO scenario.

    Shows macro-category headers with the scenario-suggested weight, then
    lists the recommended funds with FIDArating + Morningstar badges and a
    free peso-% input for each.  Returns a ready DataFrame when weights sum
    to 100 %, or None while the user is still editing.
    """
    funds = gp_scenario.get("funds", [])
    if not funds:
        return None

    sw = gp_scenario.get("subcat_weights", {})

    # Pre-build normalised lookup for extra_urls so GP names (e.g. "AZ Bond -
    # Paesi emergenti") match abbreviated Excel keys (e.g. "AZ F.1 Bd. Paesi
    # Emg A Cap EUR") even when the substring check would fail.
    _extra_urls_norm: dict = {}
    for _eu_k, _eu_v in (extra_urls or {}).items():
        if _eu_v:
            _extra_urls_norm[_normalize_for_unp(_eu_k)] = _eu_v

    # Group funds by subcategory, preserving parse order
    subcat_funds: dict = {}
    for f in funds:
        subcat_funds.setdefault(f["subcat"], []).append(f)

    # Per-scenario session-state key so weights reset when switching scenarios
    ss_key = f"_sg_w_{sc_name}"
    if ss_key not in st.session_state:
        # Distribuisce il peso suggerito di ogni sottocategoria equamente tra i
        # suoi fondi (es. alloc_balanced 25% / 6 fondi = 4,2% ciascuno).
        _init_w: dict = {}
        for _sc_k, _sc_fs in subcat_funds.items():
            _sc_pct = sw.get(_sc_k, 0.0)
            _n = len(_sc_fs)
            _per = round(_sc_pct / _n, 1) if _n else 0.0
            for _fnd in _sc_fs:
                _init_w[_fnd["nome"]] = _per
        st.session_state[ss_key] = _init_w
    ww: dict = st.session_state[ss_key]

    # ── Page header ───────────────────────────────────────────────────────────
    st.markdown('<p class="sec-title">Costruisci il Portafoglio Suggerito</p>',
                unsafe_allow_html=True)
    st.caption(
        "I pesi mostrati sono distribuiti equamente all'interno di ogni "
        "sottocategoria.  Modifica liberamente i valori e l'analisi si "
        "aggiorna automaticamente quando la somma raggiunge 100 %."
    )

    # ── Carica ratings Quantalys ──────────────────────────────────────────────
    _qtlr = load_quantalys_ratings()   # {ISIN: {"score": N, "globes": N}}

    # ── Column headers (only once, above all subcategories) ───────────────────
    _h1, _h2, _h3, _h4, _h5 = st.columns([4.5, 1.2, 1.2, 1.2, 1.4])
    _h1.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>Fondo</span>",
                 unsafe_allow_html=True)
    _h2.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>FIDArating</span>",
                 unsafe_allow_html=True)
    _h3.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>Morningstar</span>",
                 unsafe_allow_html=True)
    _h4.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>Quantalys</span>",
                 unsafe_allow_html=True)
    _h5.markdown("<span style='font-size:.7rem;color:#64748B;font-weight:600;"
                 "text-transform:uppercase;letter-spacing:.08em;'>Peso %</span>",
                 unsafe_allow_html=True)
    st.markdown("<hr style='margin:.15rem 0 .3rem 0;border-color:#e2e8f0;'>",
                unsafe_allow_html=True)

    # ── Per-subcategory sections ──────────────────────────────────────────────
    for sc_key, sc_funds in subcat_funds.items():
        w_sc   = sw.get(sc_key, 0)
        sc_lbl = _SUBCAT_DISPLAY.get(sc_key, sc_key)

        # — Subcategory header bar —
        st.markdown(
            f"<div style='background:linear-gradient(90deg,#0D1B2A,#162e52);"
            f"color:#fff;padding:.45rem 1rem;border-radius:6px;margin-top:.7rem;"
            f"display:flex;align-items:center;gap:.8rem;'>"
            f"<span style='font-weight:700;font-size:.88rem;flex:1;'>{sc_lbl}</span>"
            f"<span style='background:#C9A84C;color:#0D1B2A;padding:2px 9px;"
            f"border-radius:4px;font-size:.73rem;font-weight:700;white-space:nowrap;'>"
            f"Peso suggerito: {w_sc}%</span></div>",
            unsafe_allow_html=True)

        # — Fund rows —
        for f in sc_funds:
            fname    = f["nome"]
            resolved = _resolve_nome_for_fd(fname, fund_data)

            # Fuzzy fallback: se la chiave risolta non ha dati, cerca per nome
            # breve nel cache — aggiorna resolved così URL + rating usano la
            # stessa chiave arricchita di dati (fida_rating, ms_rating, url).
            _skey_f   = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', fname, flags=re.I).strip().lower()
            _fd_entry = (fund_data or {}).get(resolved) or {}
            if _skey_f and not _fd_entry.get("url") and not _fd_entry.get("overview"):
                for _k, _fv in (fund_data or {}).items():
                    if isinstance(_fv, dict) and _skey_f in _k.lower():
                        _fd_entry = _fv
                        resolved  = _k
                        break

            # Ratings from cache (usa resolved aggiornato)
            fd_ov  = _fd_entry.get("overview", {})
            fida_r = str(fd_ov.get("fida_rating") or "").strip() or "—"
            ms_r   = (ms_data or {}).get(resolved, {}).get("ms_rating")

            # Display name: strip "AZ [Family] - " prefix (sempre dall'fname originale)
            short = re.sub(r'^AZ\s+(?:Allocation|Bond|Equity)\s*[-–]\s*',
                           '', fname, flags=re.I).strip()

            # URL lookup: override manuale → cache → extra_urls → normalised → fuzzy
            url_sg = (
                MANUAL_URL_OVERRIDES.get(fname, "")
                or _fd_entry.get("url", "")
                or (extra_urls or {}).get(resolved, "")
                or (extra_urls or {}).get(fname, "")
                or _extra_urls_norm.get(_normalize_for_unp(fname), "")
            )
            if not url_sg and _skey_f:
                for _k, _eu in (extra_urls or {}).items():
                    if _skey_f in _k.lower() and _eu:
                        url_sg = _eu
                        break
            name_html = (
                f'<a href="{url_sg}" target="_blank" rel="noopener noreferrer" '
                f'style="color:#1B4FBB;text-decoration:underline;'
                f'text-underline-offset:2px;font-size:.84rem;font-weight:500;">'
                f'{short}</a>'
                if url_sg else
                f'<span style="font-size:.84rem;font-weight:500;color:#1e293b;">{short}</span>'
            )
            # Estrai ISIN dall'URL FondiDoc (formato: .../ISIN_nome-fondo)
            _sg_isin = ""
            _sg_url_chk = url_sg or _fd_entry.get("url", "")
            if _sg_url_chk:
                _sg_m = re.search(r'/([A-Z]{2}[A-Z0-9]{10})[_/]', _sg_url_chk)
                if _sg_m:
                    _sg_isin = _sg_m.group(1)

            c1, c2, c3, c4, c5 = st.columns([4.5, 1.2, 1.2, 1.2, 1.4])
            with c1:
                st.markdown(
                    f"<div style='padding:.55rem 0 .3rem 0;'>{name_html}</div>",
                    unsafe_allow_html=True)
            with c2:
                st.markdown(
                    f"<div style='padding:.5rem 0 .25rem 0;'>"
                    f"{_fida_badge_gp(fida_r)}</div>",
                    unsafe_allow_html=True)
            with c3:
                st.markdown(
                    f"<div style='padding:.5rem 0 .25rem 0;'>"
                    f"{_ms_badge_gp(ms_r)}</div>",
                    unsafe_allow_html=True)
            with c4:
                st.markdown(
                    f"<div style='padding:.5rem 0 .25rem 0;'>"
                    f"{_qtl_rating_cell(_sg_isin, _qtlr)}</div>",
                    unsafe_allow_html=True)
            with c5:
                _n_sc = len(subcat_funds.get(f["subcat"], [f]))
                _fb_w = round(sw.get(f["subcat"], 0.0) / _n_sc, 1)
                default_w = float(ww.get(fname, _fb_w))
                new_w = st.number_input(
                    "w", min_value=0.0, max_value=100.0,
                    value=default_w, step=0.5,
                    key=f"sg_{sc_name}_{fname[:35]}",
                    label_visibility="collapsed",
                )
                ww[fname] = new_w

        st.markdown("<hr style='margin:.25rem 0 0 0;border-color:#f1f5f9;'>",
                    unsafe_allow_html=True)

    # ── Sezione Private Markets (peso suggerito, nessun fondo) ────────────────
    _pm_pct = 0
    _gp_info = gp_scenario.get("info", "")
    _pm_m = re.search(r'Private\s+Markets\s+(\d+)%', _gp_info, re.IGNORECASE)
    if _pm_m:
        _pm_pct = int(_pm_m.group(1))
    else:
        _pm_pct = max(0, round(100 - sum(sw.values())))
    if _pm_pct > 0:
        st.markdown(
            f"<div style='background:linear-gradient(90deg,#3D2B1F,#5C3D2E);"
            f"color:#fff;padding:.45rem 1rem;border-radius:6px;margin-top:.7rem;"
            f"display:flex;align-items:center;gap:.8rem;'>"
            f"<span style='font-weight:700;font-size:.88rem;flex:1;'>"
            f"Private Markets</span>"
            f"<span style='background:#C9A84C;color:#0D1B2A;padding:2px 9px;"
            f"border-radius:4px;font-size:.73rem;font-weight:700;white-space:nowrap;'>"
            f"Peso suggerito: {_pm_pct}%</span></div>",
            unsafe_allow_html=True)
        st.markdown(
            "<p style='font-size:.82rem;color:#475569;font-weight:500;"
            "padding:.5rem .2rem .1rem .4rem;margin:0;'>"
            "I fondi Private Markets (ELTIF, RAIF, Demos, …) non sono inclusi "
            "nel portafoglio liquido — peso da considerare separatamente.</p>",
            unsafe_allow_html=True)
        st.markdown("<hr style='margin:.4rem 0 0 0;border-color:#f1f5f9;'>",
                    unsafe_allow_html=True)

    # ── Total weight indicator ────────────────────────────────────────────────
    # Accetta sia somma=pesi_liquidi (es. 70%) che somma=100%
    # (il secondo caso: l'utente ha incluso il peso PM nei fondi liquidi)
    _liq_tgt  = float(sum(sw.get(k, 0) for k in subcat_funds)) or 100.0
    total_w   = sum(ww.get(f["nome"], 0.0) for f in funds)
    diff_liq  = abs(total_w - _liq_tgt)
    diff_full = abs(total_w - 100.0)
    diff      = min(diff_liq, diff_full)
    st.markdown("<br>", unsafe_allow_html=True)
    if diff < 0.15:
        st.markdown(
            f'<div class="w-ok">✅ Somma pesi: <b>{total_w:.1f}%</b>'
            f' — Portafoglio pronto!</div>', unsafe_allow_html=True)
    else:
        # Mostra la distanza dal target più vicino
        if diff_liq <= diff_full:
            left = _liq_tgt - total_w
            tgt_lbl = f"target liquido {_liq_tgt:.0f}%"
        else:
            left = 100.0 - total_w
            tgt_lbl = "target 100%"
        st.markdown(
            f'<div class="w-warn">⚠️ Somma pesi: <b>{total_w:.1f}%</b>'
            f' ({"mancano" if left>0 else "eccedono"} {abs(left):.1f}%'
            f' al {tgt_lbl})</div>',
            unsafe_allow_html=True)

    if diff > 1.0:
        return None   # analysis only when weights are balanced

    # ── Build DataFrame ───────────────────────────────────────────────────────
    records: list = []
    for f in funds:
        peso = ww.get(f["nome"], 0.0)
        if peso <= 0:
            continue
        nome = _resolve_nome_for_fd(f["nome"], fund_data)
        # Fuzzy fallback su fund_data: usa la chiave che ha effettivamente dati
        if not (fund_data or {}).get(nome, {}).get("url"):
            _sk = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', f["nome"], flags=re.I).strip().lower()
            if _sk:
                for _k2, _fv2 in (fund_data or {}).items():
                    if isinstance(_fv2, dict) and _sk in _k2.lower() and _fv2.get("url"):
                        nome = _k2
                        break
        # Fallback su extra_urls (fida_urls): se fund_data è vuoto o manca l'URL
        # risolve il nome GP al nome FIDA completo per permettere il fetch FondiDoc
        if extra_urls and nome == f["nome"]:
            _gp_norm = _normalize_for_unp(f["nome"])
            _best_k, _best_l = None, 0
            for _fk in extra_urls:
                _fk_norm = _normalize_for_unp(_fk)
                if _fk_norm == _gp_norm:
                    _best_k = _fk
                    break
                if (_fk_norm in _gp_norm or _gp_norm in _fk_norm) and len(_fk_norm) > _best_l:
                    _best_k, _best_l = _fk, len(_fk_norm)
            if _best_k:
                nome = _best_k
        _gp_az = {
            "BOND":           DEFAULT_AZ["Obbligazionari"],
            "AZIONARI (LONG)": DEFAULT_AZ["Azionari"],
            "ALLOCATION":     DEFAULT_AZ["Bilanciati/Flessibili"],
        }.get(f["gruppo"], f["az_pct"])
        records.append({
            "nome":      nome,
            "nome_orig": f["nome"],   # GP-format name for display in composition panel
            "categoria": f["categoria"],
            "gruppo":    f["gruppo"],
            "macro_cat": get_macro(f["categoria"]),
            "az_pct":    _gp_az,
            "obb_pct":   1.0 - _gp_az,
            "r_weight":  peso / 100.0,
            "w_cons":    peso / 100.0,
            "w_equil":   peso / 100.0,
            "w_accr":    peso / 100.0,
        })
    if not records:
        return None
    df = pd.DataFrame(records)
    for wc in ("w_cons", "w_equil", "w_accr"):
        t = df[wc].sum()
        if t > 0:
            df[wc] /= t
    df["macro_cat"] = df["categoria"].apply(get_macro)
    df = assign_colors(df)
    st.markdown("<br>", unsafe_allow_html=True)
    return df


# ════════════════════════════════════════════════════════════
# MAIN APP
# ════════════════════════════════════════════════════════════

_APP_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@600;700&family=DM+Sans:opsz,wght@9..40,300;9..40,400;9..40,500;9..40,600&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
h1,h2,h3{font-family:'Cormorant Garamond',serif !important;}
[data-testid="stSidebar"]{background:linear-gradient(170deg,#06101e 0%,#0d1f3c 55%,#0a1628 100%);border-right:1px solid #1a3050;}
[data-testid="stSidebar"] .stFileUploader label,[data-testid="stSidebar"] .stRadio > label,[data-testid="stSidebar"] .stSelectbox > label{color:#4a6582 !important;font-size:.68rem !important;letter-spacing:.12em !important;text-transform:uppercase !important;font-weight:600 !important;}
[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p{color:#c0cfe0 !important;font-size:.9rem !important;}
[data-testid="stSidebar"] .stSelectbox>div>div{background:#132035 !important;border:1px solid #243d5a !important;color:#dde6f0 !important;border-radius:6px !important;}
[data-testid="stSidebar"] .stSelectbox svg{fill:#C9A84C !important;width:22px !important;height:22px !important;opacity:1 !important;}
[data-testid="stSidebar"] .stFileUploader>div{background:#132035 !important;border:1px dashed #2a4a6a !important;border-radius:8px !important;padding:.35rem .6rem !important;}
[data-testid="stSidebar"] .stFileUploader section{padding:.2rem 0 !important;min-height:unset !important;}
[data-testid="stSidebar"] .stFileUploader [data-testid="stFileUploaderDropzone"]{padding:.3rem .5rem !important;min-height:unset !important;}
[data-testid="stSidebar"] .stFileUploader [data-testid="stFileUploaderDropzoneInstructions"]{display:none !important;}
[data-testid="stSidebar"] .stFileUploader p,[data-testid="stSidebar"] .stFileUploader span{color:#8aa5c0 !important;font-size:.75rem !important;line-height:1.3 !important;}
[data-testid="stSidebar"] .stFileUploader{margin-bottom:.2rem !important;}
[data-testid="stSidebar"] hr{margin:.3rem 0 !important;border-color:#1a3050 !important;}
[data-testid="stSidebarContent"]{scrollbar-width:thin;scrollbar-color:#C9A84C #0d1f35;}
[data-testid="stSidebarContent"]::-webkit-scrollbar{width:9px !important;}
[data-testid="stSidebarContent"]::-webkit-scrollbar-track{background:#0d1f35 !important;border-radius:4px;}
[data-testid="stSidebarContent"]::-webkit-scrollbar-thumb{background:#C9A84C !important;border-radius:4px;border:2px solid #0d1f35;}
[data-testid="stSidebarContent"]::-webkit-scrollbar-thumb:hover{background:#e8c96a !important;}
section[data-testid="stSidebar"]>div:first-child{scrollbar-width:thin;scrollbar-color:#C9A84C #0d1f35;}
section[data-testid="stSidebar"]>div:first-child::-webkit-scrollbar{width:9px !important;}
section[data-testid="stSidebar"]>div:first-child::-webkit-scrollbar-track{background:#0d1f35 !important;}
section[data-testid="stSidebar"]>div:first-child::-webkit-scrollbar-thumb{background:#C9A84C !important;border-radius:4px;}
section[data-testid="stSidebar"]>div:first-child::-webkit-scrollbar-thumb:hover{background:#e8c96a !important;}
.main{background:#f6f8fb !important;}.block-container{padding-top:1.8rem !important;max-width:1300px;}
.az-header{background:linear-gradient(130deg,#081420 0%,#0f2644 50%,#162e52 100%);border-radius:16px;padding:2rem 2.5rem;margin-bottom:1.8rem;position:relative;overflow:hidden;box-shadow:0 8px 32px rgba(0,0,0,.15);}
.az-header::after{content:'';position:absolute;bottom:-60px;right:-40px;width:220px;height:220px;border-radius:50%;background:radial-gradient(circle,rgba(201,168,76,.18) 0%,transparent 70%);}
.az-eyebrow{font-size:.65rem;letter-spacing:.2em;color:#4a7098;text-transform:uppercase;font-weight:600;}
.az-title{font-family:'Cormorant Garamond',serif;font-size:2.1rem;font-weight:700;color:#f0f6ff;margin:.2rem 0 .4rem;line-height:1.1;}
.az-rule{width:38px;height:3px;background:#C9A84C;border-radius:2px;margin:.6rem 0;}
.az-meta{font-size:.88rem;color:#6b8fb0;}
.kpi{background:#fff;border:1px solid #e4eaf3;border-radius:12px;padding:1.2rem 1.4rem;box-shadow:0 1px 4px rgba(0,0,0,.05);}
.kpi-label{font-size:.65rem;text-transform:uppercase;letter-spacing:.1em;color:#94a3b8;font-weight:500;margin-bottom:.3rem;}
.kpi-value{font-size:1.9rem;font-weight:700;color:#0d1b2a;font-family:'Cormorant Garamond',serif;line-height:1;}
.kpi-sub{font-size:.75rem;color:#64748b;margin-top:.3rem;}
.sec-title{font-family:'Cormorant Garamond',serif;font-size:1.25rem;font-weight:600;color:#0d1b2a;border-bottom:2px solid #c9a84c;display:inline-block;padding-bottom:.4rem;margin-bottom:.9rem;}
.fund-group-hdr{background:#f0f4f9;padding:.45rem 1rem;font-size:.65rem;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:#64748b;border-bottom:1px solid #e2e8f0;}
.fund-row{display:flex;align-items:center;gap:10px;padding:.65rem 1rem;border-bottom:1px solid #f1f5f9;}
.fund-row:last-child{border-bottom:none;}
.fund-dot{width:8px;height:34px;border-radius:3px;flex-shrink:0;}
.fund-name{font-size:.83rem;color:#1e293b;font-weight:500;flex:1;min-width:0;}
.fund-cat{font-size:.68rem;color:#64748b;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.fund-pct{font-size:1rem;font-weight:700;color:#0d1b2a;min-width:2.8rem;text-align:right;}
[data-testid="stDownloadButton"]>button{background:linear-gradient(135deg,#0f2d6b 0%,#1b4fbb 100%) !important;color:#fff !important;font-size:1.05rem !important;font-weight:600 !important;padding:.9rem 2rem !important;border-radius:10px !important;border:none !important;width:100% !important;letter-spacing:.02em !important;box-shadow:0 4px 18px rgba(27,79,187,.35) !important;}
[data-testid="stDownloadButton"]>button:hover{box-shadow:0 6px 24px rgba(27,79,187,.55) !important;transform:translateY(-2px) !important;}
.w-ok{background:#d1fae5;border:1px solid #6ee7b7;border-radius:8px;padding:.7rem 1rem;font-size:.84rem;color:#065f46;}
.w-warn{background:#fef3c7;border:1px solid #fcd34d;border-radius:8px;padding:.7rem 1rem;font-size:.84rem;color:#92400e;}
@keyframes _aggiorna_blink{0%,100%{opacity:1;box-shadow:0 0 8px 3px #ef444466}50%{opacity:.45;box-shadow:0 0 18px 6px #ef4444cc}}
@keyframes _card_amber{0%,100%{opacity:1;box-shadow:0 0 8px 2px #d9770688}50%{opacity:.6;box-shadow:0 0 18px 6px #d97706cc}}
</style>
"""


def main():
    st.markdown(_APP_CSS, unsafe_allow_html=True)
    # Prima apertura della sessione → invita ad aggiornare i dati
    if "_session_needs_update" not in st.session_state:
        st.session_state["_session_needs_update"] = True
    _ms_with_rating = 0   # default; updated inside sidebar block below
    with st.sidebar:
        st.markdown("""<div style='padding:1.2rem 0 .4rem 0;'><div style='font-size:.6rem;letter-spacing:.22em;color:#3a5a78;text-transform:uppercase;font-weight:700;'>Analisi Portafoglio</div><div style='font-family:"Cormorant Garamond",serif;font-size:1.3rem;color:#dde8f5;font-weight:700;font-style:italic;margin-top:4px;line-height:1.3;'>Demo Analisi</div><div style='width:32px;height:3px;background:#C9A84C;border-radius:2px;margin-top:8px;'></div><div style='font-size:.6rem;color:#2a4a6a;margin-top:5px;'>v2.3 — Excel + GP cache persistente</div></div>""", unsafe_allow_html=True)
        st.markdown("<hr style='margin:.4rem 0 .5rem 0;border-color:#1a3050;'>", unsafe_allow_html=True)

        # ── Uploader Excel ────────────────────────────────────────────────────
        _xl_cache_raw, _xl_cache_date = load_excel_cache()
        if _xl_cache_date:
            _xl_hint = (f"💾 Cache: {_xl_cache_date} · carica per aggiornare")
        else:
            _xl_hint = "Nessuna cache — carica il file mensile."
        uploaded = st.file_uploader(
            "FILE EXCEL (PTF FULL + PTF SHORT + FIDA)",
            type=["xlsx","xls"],
            help=_xl_hint,
            key="uploader_xl",
        )
        if uploaded is not None:
            # Salva bytes + nome in session_state immediatamente (sopravvivono
            # a qualsiasi rerun successivo, anche programmativo)
            _xl_bytes_snap = uploaded.getvalue()
            if _xl_bytes_snap:
                st.session_state["_xl_pending_bytes"] = _xl_bytes_snap
                st.session_state["_xl_loaded_name"]   = uploaded.name
        elif st.session_state.get("_xl_loaded_name"):
            # Il widget è vuoto (Streamlit lo svuota su ogni rerun naturale)
            # ma il file era stato caricato in questa sessione → mostra badge
            st.markdown(
                f"<div style='background:#0d2b1a;border:1px solid #166534;"
                f"border-radius:6px;padding:5px 10px;margin:-4px 0 4px 0;"
                f"font-size:.78rem;color:#86efac;'>"
                f"✅&nbsp;<b>{st.session_state['_xl_loaded_name']}</b>"
                f"&nbsp;·&nbsp;attivo in sessione</div>",
                unsafe_allow_html=True)
        elif _xl_cache_date:
            st.markdown(f"<div style='font-size:.76rem;color:#7DD3FC;margin:-4px 0 6px 0;"
                        f"padding:3px 8px;background:#0c2236;border-radius:4px;"
                        f"border-left:3px solid #3b82f6;'>📂 Excel in cache &nbsp;·&nbsp; <b>{_xl_cache_date}</b></div>",
                        unsafe_allow_html=True)

        # ── Uploader Factbook ─────────────────────────────────────────────────
        uploaded_fb = st.file_uploader(
            "FACTBOOK PDF (prima estrazione)",
            type=["pdf"],
            help="Carica il Factbook PDF per estrarre Duration, Rating e Asset "
                 "Allocation. Dopo la prima estrazione scarica il file Excel "
                 "e ricaricalo la prossima volta: è più veloce.",
            key="uploader_fb",
        )
        if uploaded_fb is not None:
            _fb_bytes_snap = uploaded_fb.getvalue()
            if _fb_bytes_snap:
                st.session_state["_fb_pending_bytes"] = _fb_bytes_snap
                st.session_state["_fb_loaded_name"]   = uploaded_fb.name
        elif st.session_state.get("_fb_loaded_name"):
            _fb_dt_str   = st.session_state.get("_fb_doc_date", "")
            _fb_dt_extra = (f"&nbsp;·&nbsp;📅&nbsp;{_fb_dt_str}"
                            if _fb_dt_str else "")
            st.markdown(
                f"<div style='background:#0d2b1a;border:1px solid #166534;"
                f"border-radius:6px;padding:5px 10px;margin:-4px 0 4px 0;"
                f"font-size:.78rem;color:#86efac;'>"
                f"✅&nbsp;<b>{st.session_state['_fb_loaded_name']}</b>"
                f"{_fb_dt_extra}"
                f"&nbsp;·&nbsp;attivo in sessione</div>",
                unsafe_allow_html=True)

        uploaded_fb_xl = None  # rimosso: GITHUB_TOKEN salva automaticamente

        # Caption con la data di riferimento del Factbook.
        # Al primo render di sessione legge direttamente il JSON su disco
        # (la sidebar gira prima del main content che normalmente lo salva).
        _fb_doc_date_sb = st.session_state.get("_fb_doc_date", "")
        if not _fb_doc_date_sb:
            _fb_auto_q = load_factbook_auto()
            _fb_doc_date_sb = (_fb_auto_q or {}).get("_ref_date", "")
            if _fb_doc_date_sb:
                st.session_state["_fb_doc_date"] = _fb_doc_date_sb
        if _fb_doc_date_sb:
            st.markdown(f"<div style='font-size:.76rem;color:#7DD3FC;margin:-4px 0 6px 0;"
                        f"padding:3px 8px;background:#0c2236;border-radius:4px;"
                        f"border-left:3px solid #3b82f6;'>📅 Factbook in cache &nbsp;·&nbsp; <b>{_fb_doc_date_sb}</b></div>",
                        unsafe_allow_html=True)

        # ── Uploader Factbook Fondi Pensione ─────────────────────────────────
        _fp_cache       = load_fp_cache()
        _fp_cache_date  = _fp_cache.get("_ref_date", "")
        _fp_hint = (f"💾 Cache: {_fp_cache_date} · carica per aggiornare"
                    if _fp_cache_date else "Carica il Factbook Fondi Pensione PDF.")
        uploaded_fp = st.file_uploader(
            "FACTBOOK FONDI PENSIONE PDF",
            type=["pdf"],
            help=_fp_hint,
            key="uploader_fp",
        )
        if uploaded_fp is not None:
            _fp_bytes = uploaded_fp.getvalue()
            if _fp_bytes:
                with st.spinner("Estraggo dati fondi pensione…"):
                    _fp_parsed = parse_fp_factbook(_fp_bytes)
                if _fp_parsed:
                    save_fp_cache(_fp_parsed)
                    st.session_state["_fp_data"] = _fp_parsed
                    _fp_cache_date = _fp_parsed.get("_ref_date", "")
                    st.success(f"✅ Estratti {len(_fp_parsed)-1} fondi pensione"
                               + (f" · {_fp_cache_date}" if _fp_cache_date else ""))
                else:
                    st.warning("⚠️ Nessun dato estratto — verifica il formato del PDF.")
        if not st.session_state.get("_fp_data") and _fp_cache:
            st.session_state["_fp_data"] = _fp_cache
        if _fp_cache_date or st.session_state.get("_fp_data", {}).get("_ref_date"):
            _fp_dt = (_fp_cache_date
                      or st.session_state.get("_fp_data", {}).get("_ref_date", ""))
            st.markdown(f"<div style='font-size:.76rem;color:#7DD3FC;margin:-4px 0 6px 0;"
                        f"padding:3px 8px;background:#0c2236;border-radius:4px;"
                        f"border-left:3px solid #3b82f6;'>🏦 FP in cache &nbsp;·&nbsp; <b>{_fp_dt}</b></div>",
                        unsafe_allow_html=True)

        # ── Uploader GP ───────────────────────────────────────────────────────
        _gp_cache_data, _gp_cache_fname, _gp_cache_date = load_gp_cache()
        if _gp_cache_date:
            _gp_hint = (f"💾 Cache: {_gp_cache_date} · carica per aggiornare")
        else:
            _gp_hint = "Nessuna cache — carica il PDF trimestrale."
        uploaded_gp = st.file_uploader(
            "GLOBAL PERSPECTIVES PDF",
            type=["pdf"],
            help=_gp_hint,
            key="uploader_gp",
        )
        if uploaded_gp is not None:
            st.session_state["_gp_loaded_name"] = uploaded_gp.name
        if uploaded_gp is None:
            # Prendi edizione da session_state (render successivi) oppure
            # direttamente dalla cache su disco (primo render di sessione)
            _gp_ed_str = (st.session_state.get("_gp_doc_edition")
                          or (_gp_cache_data or {}).get("_edition", ""))
            if _gp_cache_date:
                _gp_label = f"<b>{_gp_ed_str}</b>" if _gp_ed_str else f"<b>{_gp_cache_date}</b>"
                st.markdown(f"<div style='font-size:.76rem;color:#7DD3FC;margin:-4px 0 6px 0;"
                            f"padding:3px 8px;background:#0c2236;border-radius:4px;"
                            f"border-left:3px solid #3b82f6;'>📂 GP in cache &nbsp;·&nbsp; {_gp_label}</div>",
                            unsafe_allow_html=True)
            elif st.session_state.get("_gp_loaded_name"):
                _gp_ed_extra = (f"&nbsp;·&nbsp;📅&nbsp;{_gp_ed_str}"
                                if _gp_ed_str else "")
                st.markdown(
                    f"<div style='background:#0d2b1a;border:1px solid #166534;"
                    f"border-radius:6px;padding:5px 10px;margin:-4px 0 4px 0;"
                    f"font-size:.78rem;color:#86efac;'>"
                    f"✅&nbsp;<b>{st.session_state['_gp_loaded_name']}</b>"
                    f"{_gp_ed_extra}"
                    f"&nbsp;·&nbsp;attivo in sessione</div>",
                    unsafe_allow_html=True)

        # ── Parsing GP (solo quando cambia file) ─────────────────────────────
        if uploaded_gp is not None:
            if st.session_state.get("_gp_filename") != uploaded_gp.name:
                with st.spinner("📄 Parsing Global Perspectives…"):
                    _gp_parsed = parse_global_perspectives(uploaded_gp.read())
                if _gp_parsed:
                    # ── Edizione: prima tenta dal testo PDF, poi dal nome file ──
                    _ed = _gp_parsed.get("_edition", "")
                    if not _ed:
                        # Fallback: AzimutGlobalPerspectives_ITA_032026.01.pdf
                        # Pattern _MMAAAA → mese+anno → trimestre
                        _fn_m = re.search(r'_(\d{2})(\d{4})', uploaded_gp.name)
                        if _fn_m:
                            _mo_n = int(_fn_m.group(1))
                            _yr_n = _fn_m.group(2)
                            _q_n  = ("Q1" if _mo_n <= 3 else
                                     "Q2" if _mo_n <= 6 else
                                     "Q3" if _mo_n <= 9 else "Q4")
                            _ed = f"{_q_n} {_yr_n}"
                        else:
                            # Fallback 2: Q1_2026 o Q1-2026 nel nome
                            _fn_q = re.search(
                                r'Q([1-4])[\s_\-]*(20\d{2})', uploaded_gp.name,
                                re.IGNORECASE)
                            if _fn_q:
                                _ed = f"Q{_fn_q.group(1)} {_fn_q.group(2)}"
                    if _ed:
                        _gp_parsed["_edition"] = _ed
                        st.session_state["_gp_doc_edition"] = _ed
                    st.session_state["_gp_data"]    = _gp_parsed
                    st.session_state["_gp_filename"] = uploaded_gp.name
                    # Nuovo PDF → il fetch FondiDoc va rifatto
                    st.session_state.pop("_gp_fetch_done", None)
                    # Salva su disco per le sessioni future (include _edition)
                    save_gp_cache(_gp_parsed, uploaded_gp.name)
                else:
                    st.session_state.pop("_gp_data", None)
                    st.warning("⚠️ PDF non riconosciuto — verifica che sia un "
                               "Global Perspectives Azimut.")
        else:
            # Nessun file caricato: usa cache su disco se disponibile
            if not st.session_state.get("_gp_data") and _gp_cache_data:
                st.session_state["_gp_data"]    = _gp_cache_data
                st.session_state["_gp_filename"] = _gp_cache_fname
                # Ripristina edizione: dalla cache, o dal nome file in cache
                _ed_c = _gp_cache_data.get("_edition", "")
                if not _ed_c and _gp_cache_fname:
                    _fn_mc = re.search(r'_(\d{2})(\d{4})', _gp_cache_fname)
                    if _fn_mc:
                        _mo_c = int(_fn_mc.group(1))
                        _q_c  = ("Q1" if _mo_c <= 3 else
                                 "Q2" if _mo_c <= 6 else
                                 "Q3" if _mo_c <= 9 else "Q4")
                        _ed_c = f"{_q_c} {_fn_mc.group(2)}"
                if _ed_c and "_gp_doc_edition" not in st.session_state:
                    st.session_state["_gp_doc_edition"] = _ed_c
            elif st.session_state.get("_gp_filename") and not _gp_cache_data:
                # Cache rimossa manualmente → pulisci session state
                st.session_state.pop("_gp_data",        None)
                st.session_state.pop("_gp_filename",     None)
                st.session_state.pop("_gp_doc_edition",  None)

        # ── Card stato dati ───────────────────────────────────────────────────
        st.markdown("<hr style='margin:.25rem 0 .3rem 0;border:none;border-top:1px solid #1a3050;'>", unsafe_allow_html=True)
        _fd_now = st.session_state.get("_scomp_fd") or load_fund_cache()[0]
        _ms_now = st.session_state.get("_ms_data") or load_ms_cache()
        _ms_with_rating = sum(1 for v in _ms_now.values() if v.get("ms_rating"))
        _gp_loaded_now  = bool(st.session_state.get("_gp_data"))

        _fd_line = (f"✅ <b>FondiDoc</b> — {len(_fd_now)} fondi"
                    if _fd_now else "⚠️ <b>FondiDoc</b> — non scaricato")
        _ms_line = (f"⭐ <b>Morningstar</b> — {_ms_with_rating} rating"
                    if _ms_with_rating else "⚠️ <b>Morningstar</b> — non scaricato")
        _gp_status_lines = ""
        _gp_miss = 0
        _n_gp    = 0
        if _gp_loaded_now:
            _gp_ok  = st.session_state["_gp_data"]
            _n_gp   = sum(len(v["funds"]) for v in _gp_ok.values()
                         if isinstance(v, dict) and "funds" in v)
            _fd_chk = _fd_now
            _gp_miss = len(set(
                f["nome"]
                for sc in _gp_ok.values()
                if isinstance(sc, dict) and "funds" in sc
                for f in sc.get("funds", [])
                if not (
                    _fd_chk.get(_resolve_nome_for_fd(f["nome"], _fd_chk), {}).get("url", "")
                    or _fd_chk.get(f["nome"], {}).get("url", "")
                )
            ))
            _gp_status_lines = (
                f"<br>🌐 <b>Global Perspectives</b> — {_n_gp} fondi"
                + (f" · ⚠️ {_gp_miss} senza dati" if _gp_miss else " · ✅ tutti aggiornati")
            )

        _all_ok   = bool(_fd_now and _ms_with_rating
                        and (_gp_miss == 0 if _gp_loaded_now else True))
        _any_data = bool(_fd_now or _ms_with_rating)
        _needs_upd = st.session_state.get("_session_needs_update", False)

        # card colore: ambra lampeggiante all'apertura / verde / giallo / rosso
        if _needs_upd and _any_data:
            # Dati in cache ma sessione fresca → invita ad aggiornare
            _card_bg, _card_brd, _card_clr = "#1a1200", "#b45309", "#fde68a"
            _card_extra_style = (
                "border-width:2px;"
                "box-shadow:0 0 10px 2px #d9770688;"
                "animation:_card_amber 1.4s ease-in-out infinite;")
            _card_anim_css = (
                "<style>@keyframes _card_amber{"
                "0%,100%{opacity:1;box-shadow:0 0 8px 2px #d9770688}"
                "50%{opacity:.6;box-shadow:0 0 18px 6px #d97706cc}}"
                "</style>")
        elif _all_ok:
            _card_bg, _card_brd, _card_clr = "#0d2b1a", "#166534", "#86efac"
            _card_extra_style = ""
            _card_anim_css    = ""
        elif _any_data:
            _card_bg, _card_brd, _card_clr = "#1a1a08", "#854d0e", "#fde68a"
            _card_extra_style = ""
            _card_anim_css    = ""
        else:
            _card_bg, _card_brd, _card_clr = "#3b0000", "#ef4444", "#fca5a5"
            _card_extra_style = (
                "border-width:2px;"
                "box-shadow:0 0 10px 2px #ef444488;"
                "animation:_card_alert 1s ease-in-out infinite;")
            _card_anim_css = (
                "<style>@keyframes _card_alert{"
                "0%,100%{opacity:1;box-shadow:0 0 10px 2px #ef444488}"
                "50%{opacity:.55;box-shadow:0 0 18px 5px #ef4444cc}}"
                "</style>")
        st.markdown(
            f"{_card_anim_css}"
            f"<div style='background:{_card_bg};border:2px solid {_card_brd};"
            f"border-radius:8px;padding:.5rem .85rem;font-size:.73rem;"
            f"color:{_card_clr};margin-bottom:.4rem;line-height:1.8;"
            f"{_card_extra_style}'>"
            f"{_fd_line}<br>{_ms_line}{_gp_status_lines}</div>",
            unsafe_allow_html=True)

        # ── Unico tasto Aggiorna Dati ─────────────────────────────────────────
        # "can update" = Excel disponibile (upload fresco OPPURE cache su disco) + GP
        _has_excel   = bool(uploaded or _xl_cache_raw)
        _can_update  = bool(_has_excel or _gp_loaded_now)
        _is_fetching = any(st.session_state.get(k) for k in (
            "_fetch_fd_requested", "_fetch_ms_requested", "_fetch_gp_requested"))

        if _is_fetching:
            # Tasto "in corso" — arancio pulsante, non cliccabile
            st.markdown(
                "<div style='background:linear-gradient(135deg,#92400E,#B45309);"
                "color:#fff;padding:.6rem 1rem;border-radius:8px;font-size:.88rem;"
                "font-weight:600;text-align:center;letter-spacing:.02em;"
                "animation:_aggiorna_pulse 1.2s ease-in-out infinite;opacity:.92;'>"
                "⏳  Aggiornamento in corso…</div>"
                "<style>@keyframes _aggiorna_pulse{"
                "0%{opacity:.92}50%{opacity:.55}100%{opacity:.92}}</style>",
                unsafe_allow_html=True)
        elif _can_update:
            # Colore tasto: rosso lampeggiante (apertura) / verde / giallo / rosso
            if _needs_upd:
                # App appena aperta → rosso lampeggiante con bordo rosso
                _btn_bg      = "linear-gradient(135deg,#7f1d1d,#DC2626)"
                _btn_brd     = "border:2px solid #ef4444 !important;"
                _wrap_anim   = "animation:_aggiorna_blink 1s ease-in-out infinite;"
                _wrap_radius = "border-radius:8px;overflow:hidden;"
            elif _all_ok:
                _btn_bg      = "linear-gradient(135deg,#14532d,#16A34A)"
                _btn_brd     = ""
                _wrap_anim   = ""
                _wrap_radius = ""
            elif _any_data:
                _btn_bg      = "linear-gradient(135deg,#78350f,#D97706)"
                _btn_brd     = ""
                _wrap_anim   = ""
                _wrap_radius = ""
            else:
                _btn_bg      = "linear-gradient(135deg,#7f1d1d,#DC2626)"
                _btn_brd     = "border:2px solid #ef4444 !important;"
                _wrap_anim   = "animation:_aggiorna_blink 1s ease-in-out infinite;"
                _wrap_radius = "border-radius:8px;overflow:hidden;"
            # Selettori separati: wrapper (animation) e button (colori)
            # Il wrapper div[stButton] non ha stili emotion → animation funziona
            _wrap_sel = "section[data-testid='stSidebar'] div[data-testid='stButton']"
            _btn_sel  = f"{_wrap_sel} > button"
            st.markdown(
                f"<style>"
                f"{_wrap_sel}{{{_wrap_anim}{_wrap_radius}}}"
                f"{_btn_sel}{{"
                f"background:{_btn_bg} !important;"
                f"color:#fff !important;{_btn_brd}"
                f"font-weight:600 !important;}}"
                f"{_btn_sel}:hover{{filter:brightness(1.18) !important;}}"
                f"</style>",
                unsafe_allow_html=True)
            if st.button("📥  Aggiorna Dati",
                         use_container_width=True,
                         help="Scarica in sequenza: FondiDoc (FIDArating + rendimenti), "
                              "Morningstar e — se il GP è caricato — dati fondi GP."):
                st.session_state["_session_needs_update"] = False
                if _has_excel:
                    st.session_state["_fetch_fd_requested"] = True
                    st.session_state["_fetch_ms_requested"] = True
                if _gp_loaded_now:
                    st.session_state["_fetch_gp_requested"] = True
                # NON chiamare st.rerun() qui: il click del bottone causa già un
                # rerun naturale in cui i file uploader rimangono visibili.
                # Il rerun programmato pulisce i widget; lo faremo una sola volta
                # alla fine, dopo tutti i fetch.
        else:
            st.caption("⬆️ Carica il file Excel o il PDF Global Perspectives")

        st.markdown("<hr style='margin:.25rem 0 .3rem 0;border:none;border-top:1px solid #1a3050;'>", unsafe_allow_html=True)
        _gp_loaded    = bool(st.session_state.get("_gp_data"))
        _ptf_options  = ["📋  PTF FULL", "⚡  PTF SHORT", "🎨  LIBERO"]
        if _gp_loaded:
            _ptf_options.append("🌐  SUGGERITO")
        # _ptf_type è la selezione persistente: NON è legata al widget radio
        # quindi non viene mai svuotata dal rerun che avviene prima del render
        # del radio (es. Aggiorna Dati chiama st.rerun() prima della riga del radio).
        # _ptf_choice_radio (widget key) può essere persa in quel rerun;
        # _ptf_type resta e fornisce il fallback corretto per il parametro index.
        _ptf_saved = st.session_state.get("_ptf_type", _ptf_options[0])
        if _ptf_saved not in _ptf_options:
            _ptf_saved = _ptf_options[0]
        _ptf_idx = _ptf_options.index(_ptf_saved)
        ptf_choice = st.radio("TIPO PORTAFOGLIO", _ptf_options,
                              index=_ptf_idx,
                              key="_ptf_choice_radio")
        # Aggiorna la selezione persistente ad ogni render
        st.session_state["_ptf_type"] = ptf_choice

        if "LIBERO" not in ptf_choice and "free_ptf" in st.session_state:
            del st.session_state["free_ptf"]

        # ── Profilo di rischio — sotto LIBERO (e PTF FULL/SHORT) ─────────────
        if "SUGGERITO" not in ptf_choice:
            profile = st.selectbox("PROFILO DI RISCHIO", PROFILES, index=0,
                                   key="_profile_select")
        else:
            # SUGGERITO non usa profilo — mantieni l'ultimo valore selezionato
            profile = st.session_state.get("_profile_select", PROFILES[0])

        # ── Scenario — sotto SUGGERITO ────────────────────────────────────────
        if "SUGGERITO" in ptf_choice:
            _gp_keys = [k for k in st.session_state.get("_gp_data", {}).keys()
                        if not k.startswith("_")]
            _SC_LABELS = {
                "Base": "⚖️  Scenario Base",
                "Bear": "🐻  Scenario Bear",
                "Bull": "🐂  Scenario Bull",
            }
            _sc_opts = [_SC_LABELS.get(k, k) for k in _gp_keys]
            if _sc_opts:
                _sc_sel = st.radio("SCENARIO", _sc_opts, key="_gp_sc_radio")
                st.session_state["_gp_sc_key"] = next(
                    (k for k, v in _SC_LABELS.items() if v == _sc_sel),
                    _gp_keys[0])
                # Show scenario info
                _sc_info = st.session_state["_gp_data"].get(
                    st.session_state["_gp_sc_key"], {}).get("info", "")
                if _sc_info:
                    st.caption(f"📊 {_sc_info}")

    ptf_label = ptf_choice.split("  ",1)[1] if "  " in ptf_choice else ptf_choice
    _is_suggerito = "SUGGERITO" in ptf_choice
    if _is_suggerito:
        _sc_key_hdr = st.session_state.get("_gp_sc_key", "Base")
        ptf_label   = f"SUGGERITO <span style='font-size:0.42em;font-weight:normal;opacity:0.65;vertical-align:middle;'>da Global Persp.</span> — Scenario {_sc_key_hdr}"
    elif ptf_label in ("PTF FULL", "PTF SHORT"):
        ptf_label   = f"{ptf_label} <span style='font-size:0.42em;font-weight:normal;opacity:0.65;vertical-align:middle;'>ispirato da Global Persp.</span>"

    # ── Auto-fetch GP links quando si entra in SUGGERITO con fondi mancanti ──
    # Scatta solo la prima volta (o dopo un nuovo PDF). Dopo qualsiasi fetch
    # (auto o manuale) _gp_fetch_done=True impedisce rifetch su cambio scenario.
    _is_already_fetching = any(st.session_state.get(k) for k in (
        "_fetch_fd_requested", "_fetch_ms_requested", "_fetch_gp_requested"))
    if (_is_suggerito and _gp_loaded_now and _gp_miss > 0
            and not _is_already_fetching
            and not st.session_state.get("_gp_fetch_done")):
        st.session_state["_gp_fetch_done"] = True
        st.session_state["_fetch_gp_requested"] = True
        # NON rerun qui: il fetch GP verrà eseguito nel blocco principale
        # sotto, nello stesso render, senza svuotare i file uploader.

    # ── Invalidate cached PDF when portfolio type or profile changes ──────────
    _ptf_key = f"{ptf_choice}|{profile}"
    if st.session_state.get("_last_ptf_key") != _ptf_key:
        for _k in ("_pdf_bytes_ready", "_pdf_fname_ready", "_pdf_lbl"):
            st.session_state.pop(_k, None)
        st.session_state["_last_ptf_key"] = _ptf_key

    st.markdown(f"""<div class="az-header"><div class="az-eyebrow">AZIMUT INVESTMENTS · AAS EMILIA ROMAGNA MARCHE UMBRIA</div><div class="az-rule"></div><div class="az-title">{ptf_label}</div><div class="az-meta">{PROFILE_ICONS.get(profile,'●')} Profilo {profile.title()} &nbsp;·&nbsp; {datetime.date.today().strftime('%d %B %Y')}</div></div>""",unsafe_allow_html=True)

    # ── Carica dati Excel (file fresco → salva cache; altrimenti usa cache) ─────
    # _xl_pending_bytes: bytes salvati subito nella sidebar quando il file è stato
    # caricato. Usati come fallback se un rerun programmativo ha svuotato il widget.
    _xl_pending = st.session_state.pop("_xl_pending_bytes", None)
    _xl_fresh_bytes = (uploaded.getvalue() if uploaded is not None else _xl_pending)

    if _xl_fresh_bytes is not None:
        with st.spinner("⏳ Caricamento dati…"):
            raw = parse_excel(_xl_fresh_bytes)
        _xl_from_cache = False
    elif _xl_cache_raw is not None:
        raw = _xl_cache_raw
        _xl_from_cache = True
    else:
        raw = {}
        _xl_from_cache = False

    # Assicura che MANUAL_URL_OVERRIDES sovrascriva qualsiasi URL sbagliato in
    # fida_urls (es. hyperlink Excel o risultato vecchio di st.cache_data).
    # Fatto PRIMA di save_excel_cache così la cache su disco è già corretta.
    if raw:
        _fida_existing = raw.get("fida_urls") or {}
        raw["fida_urls"] = {**_fida_existing, **MANUAL_URL_OVERRIDES}

    # Salva su disco dopo il patch — la cache conterrà sempre gli URL aggiornati
    if _xl_fresh_bytes is not None and raw:
        save_excel_cache(raw)

    # Controlla se ci sono dati Excel disponibili (upload o cache)
    _has_raw = bool(raw.get("PTF FULL") is not None
                    and not raw.get("PTF FULL", pd.DataFrame()).empty)

    if not _has_raw and not _is_suggerito:
        st.info("⬅️ **Carica il file Excel** nella barra laterale per iniziare.")
        return

    # ── Fetch FondiDoc / Morningstar / GP — tutti in un unico render ────────────
    # Strategia: eseguiamo tutti i fetch richiesti in SEQUENZA nello stesso render,
    # poi chiamiamo st.rerun() UNA SOLA VOLTA alla fine. In questo modo i file
    # uploader rimangono visibili per tutta la durata del caricamento; il singolo
    # rerun finale aggiorna la sidebar (contatori, stato cache) senza azzerare
    # i widget più volte di quanto necessario.
    _fetch_ran = False

    if _has_raw:
        if st.session_state.pop("_fetch_fd_requested", False):
            _fida_urls_all = raw.get("fida_urls", {})
            _sheets = [raw[s] for s in ("PTF FULL", "PTF SHORT")
                       if s in raw and not raw[s].empty]
            _df_all = (pd.concat(_sheets, ignore_index=True)
                       .drop_duplicates(subset=["nome"]) if _sheets else pd.DataFrame())
            # Aggiunge fondi GP (tutti gli scenari) risolti a nomi FIDA tramite fida_urls
            _gp_for_fetch = st.session_state.get("_gp_data") or _gp_cache_data or {}
            if _gp_for_fetch and _fida_urls_all:
                _existing = set(_df_all["nome"].tolist()) if not _df_all.empty else set()
                _gp_extra = set()
                for _sc_val in _gp_for_fetch.values():
                    if not isinstance(_sc_val, dict): continue
                    for _gf in _sc_val.get("funds", []):
                        _gn = _gf.get("nome", "")
                        if not _gn: continue
                        _gn_norm = _normalize_for_unp(_gn)
                        _best, _bl = None, 0
                        for _fk in _fida_urls_all:
                            _fk_norm = _normalize_for_unp(_fk)
                            if _fk_norm == _gn_norm:
                                _best = _fk; break
                            if (_fk_norm in _gn_norm or _gn_norm in _fk_norm) and len(_fk_norm) > _bl:
                                _best, _bl = _fk, len(_fk_norm)
                        if _best and _best not in _existing:
                            _gp_extra.add(_best)
                if _gp_extra:
                    _df_gp = pd.DataFrame([{"nome": n} for n in _gp_extra])
                    _df_all = pd.concat([_df_all, _df_gp], ignore_index=True).drop_duplicates(subset=["nome"])
            if not _df_all.empty:
                _pb_fd = st.progress(0, text="Scarico dati FondiDoc…")
                def _upd_fd(v): _pb_fd.progress(v, text=f"FondiDoc: {int(v*100)}%…")
                _fd_new = fetch_all_fund_data(_df_all, _fida_urls_all, _upd_fd)
                _pb_fd.empty()
                _fd_base_existing = load_fund_cache()[0]
                _fd_merged_new = dict(_fd_base_existing)
                for _fk, _fv in _fd_new.items():
                    if _fv and (_fv.get("analysis") or not _fd_merged_new.get(_fk, {}).get("analysis")):
                        _fd_merged_new[_fk] = _fv
                    elif _fv and not _fv.get("analysis") and _fd_merged_new.get(_fk, {}).get("analysis"):
                        _upd2 = dict(_fd_merged_new[_fk])
                        if _fv.get("url"):  _upd2["url"]  = _fv["url"]
                        if _fv.get("isin"): _upd2["isin"] = _fv["isin"]
                        _fd_merged_new[_fk] = _upd2
                    elif not _fv:
                        pass
                save_fund_cache(_fd_merged_new)
                st.session_state["_scomp_fd"] = _fd_merged_new
                _fetch_ran = True
            else:
                st.warning("⚠️ Nessun fondo trovato — verifica il file Excel.")

        if st.session_state.pop("_fetch_ms_requested", False):
            _fida_df   = raw.get("FIDA", pd.DataFrame())
            _sheets_ms = [raw[s] for s in ("PTF FULL", "PTF SHORT")
                          if s in raw and not raw[s].empty]
            _df_ms = (pd.concat(_sheets_ms, ignore_index=True)
                      .drop_duplicates(subset=["nome"]) if _sheets_ms else pd.DataFrame())
            if not _df_ms.empty:
                with st.spinner("⭐ Scarico rating Morningstar…"):
                    _ms_new = fetch_all_ms_ratings(_df_ms, _fida_df)
                _n_found = sum(1 for v in _ms_new.values() if v.get("ms_rating"))
                if _n_found > 0:
                    save_ms_cache(_ms_new)
                    st.session_state["_ms_data"] = _ms_new
                    st.success(f"⭐ Morningstar: {_n_found}/{len(_ms_new)} rating trovati")
                else:
                    _ms_cached = load_ms_cache()
                    _n_cached  = sum(1 for v in _ms_cached.values() if v.get("ms_rating"))
                    st.session_state["_ms_data"] = _ms_cached or _ms_new
                    if _n_cached > 0:
                        st.warning(f"⭐ Morningstar non raggiungibile — uso cache ({_n_cached} rating)")
                    else:
                        st.warning("⭐ Morningstar non raggiungibile — nessun dato in cache")
                _fetch_ran = True
            else:
                st.warning("⚠️ Nessun fondo trovato — verifica il file Excel.")
    else:
        st.session_state.pop("_fetch_fd_requested", None)
        st.session_state.pop("_fetch_ms_requested", None)

    # ── GP fund FondiDoc lookup (runs with or without Excel) ─────────────────
    if st.session_state.pop("_fetch_gp_requested", False):
        _gp_src  = st.session_state.get("_gp_data", {})
        _fd_base = st.session_state.get("_scomp_fd") or load_fund_cache()[0]
        if _gp_src:
            _pb_gp = st.progress(0, text="Cerco fondi GP su FondiDoc…")
            def _upd_gp(v):
                _pb_gp.progress(v, text=f"Ricerca fondi GP: {int(v*100)}%…")
            _quick = raw.get("fida_urls") or dict(MANUAL_URL_OVERRIDES)
            _gp_new = fetch_gp_urls_missing(_gp_src, _fd_base, _upd_gp, quick_urls=_quick)
            _pb_gp.empty()
            if _gp_new:
                _fd_merged = dict(_fd_base)
                for _gk, _gv in _gp_new.items():
                    _existing = _fd_merged.get(_gk, {})
                    if _existing.get("analysis") and not (_gv or {}).get("analysis"):
                        _upd = dict(_existing)
                        if (_gv or {}).get("url"):  _upd["url"]  = _gv["url"]
                        if (_gv or {}).get("isin"): _upd["isin"] = _gv["isin"]
                        _fd_merged[_gk] = _upd
                    else:
                        _fd_merged[_gk] = _gv
                save_fund_cache(_fd_merged)
                st.session_state["_scomp_fd"] = _fd_merged
                try:
                    _ms_existing = st.session_state.get("_ms_data") or load_ms_cache()
                    _sess_gp = requests.Session()
                    _ms_gp_new = {}
                    for _rn, _fd_v in _gp_new.items():
                        _isin_v = _fd_v.get("isin", "") if isinstance(_fd_v, dict) else ""
                        if _isin_v and _rn not in _ms_existing:
                            _r = _ms_rating_for_isin(_isin_v, _sess_gp)
                            if _r is not None:
                                _ms_gp_new[_rn] = {"ms_rating": _r, "fo_url": None}
                    if _ms_gp_new:
                        _ms_merged = {**_ms_existing, **_ms_gp_new}
                        save_ms_cache(_ms_merged)
                        st.session_state["_ms_data"] = _ms_merged
                except Exception:
                    pass
                st.success(
                    f"✅ Trovati dati FondiDoc per "
                    f"{len(_gp_new)}/{len(_gp_new)} fondi GP")
            else:
                st.warning(
                    "⚠️ Nessun dato trovato su FondiDoc per i fondi GP. "
                    "Potrebbe essere un problema di rete o di nomi.")
            st.session_state["_gp_fetch_done"] = True
            _fetch_ran = True

    # ── Unico rerun finale dopo tutti i fetch ─────────────────────────────────
    # Aggiorna sidebar (contatori, stato cache) senza svuotare i file uploader
    # più volte. I file rimangono visibili durante l'intero caricamento.
    if _fetch_ran:
        st.rerun()

    # ── Factbook data ──────────────────────────────────────────────────────────
    # Priority:
    #   1. Excel uploaded manually (override / fix)
    #   2. PDF uploaded (first-time extraction → auto-save to repo)
    #   3. Auto-load from data/factbook_dati.json committed in the repo
    factbook_data: dict = load_factbook_auto()
    _fb_source = f"repository ({len(factbook_data)} fondi)" if factbook_data else ""

    _fb_pending  = st.session_state.pop("_fb_pending_bytes",    None)
    _fb_xl_pending = None  # uploader rimosso
    _fb_fresh_bytes = (uploaded_fb.getvalue() if uploaded_fb is not None
                       else _fb_pending)
    if _fb_fresh_bytes is not None:
        # First-time (or refresh): parse PDF, auto-save, offer Excel
        with st.spinner("📖 Estraggo dati dal Factbook PDF…"):
            _new = parse_factbook(_fb_fresh_bytes)
        if _new:
            factbook_data = _new
            _fb_source = f"PDF ({len(_new)} fondi)"
            st.success(f"✅ Factbook estratto — {len(_new)} fondi trovati")
            # Auto-save to GitHub repo (needs GITHUB_TOKEN secret)
            with st.spinner("💾 Salvo dati nel repository…"):
                _saved = save_factbook_to_repo(_new)
            if _saved:
                st.info(
                    "🔄 Dati salvati nel repository. "
                    "Al prossimo accesso non servirà ricaricare il PDF "
                    "(l'app si riavvierà automaticamente entro ~1 min).")
            else:
                # Fallback: offer Excel download so user can re-use data
                xl_bytes = factbook_to_excel_bytes(_new)
                st.download_button(
                    label="💾  Scarica dati Factbook (Excel)",
                    data=xl_bytes,
                    file_name="factbook_dati.xlsx",
                    mime="application/vnd.openxmlformats-officedocument"
                         ".spreadsheetml.sheet",
                    help="Il salvataggio automatico non è riuscito (token "
                         "mancante?). Scarica questo Excel e caricalo "
                         "nella casella 'DATI FACTBOOK (Excel)' oppure "
                         "aggiorna il secret GITHUB_TOKEN.",
                )
        else:
            st.warning("⚠️ Factbook PDF caricato ma nessun dato estratto — "
                       "verrà usato FondiDoc")

    _fb_xl_fresh_bytes = (uploaded_fb_xl.getvalue() if uploaded_fb_xl is not None
                          else _fb_xl_pending)
    if _fb_xl_fresh_bytes is not None:
        # Manual override: user uploaded a corrected Excel
        _xl = factbook_from_excel(_fb_xl_fresh_bytes)
        if _xl:
            factbook_data = _xl
            _fb_source = f"Excel ({len(_xl)} fondi)"
            st.success(f"✅ Dati Factbook caricati da Excel — {len(_xl)} fondi")
        else:
            st.warning("⚠️ Excel Factbook vuoto — uso dati precedenti")

    # Salva data di riferimento del Factbook in session_state
    # (viene mostrata nella sidebar accanto all'uploader)
    _fb_ref_date = (factbook_data or {}).get("_ref_date", "")
    if _fb_ref_date:
        st.session_state["_fb_doc_date"] = _fb_ref_date

    if _is_suggerito:
        _gp_data_main = st.session_state.get("_gp_data", {})
        _sc_key_main  = st.session_state.get("_gp_sc_key", "Base")
        _sc_data_main = _gp_data_main.get(_sc_key_main)
        if not _sc_data_main:
            st.warning("📄 Carica il PDF **Global Perspectives** nella barra "
                       "laterale per vedere i portafogli suggeriti.")
            return
        _fd_for_gp = st.session_state.get("_scomp_fd") or load_fund_cache()[0]
        _ms_for_gp = st.session_state.get("_ms_data") or load_ms_cache()
        _fida_urls_gp = raw.get("fida_urls") or dict(MANUAL_URL_OVERRIDES)
        df = suggerito_portfolio_ui(_sc_key_main, _sc_data_main,
                                    _fd_for_gp, _ms_for_gp,
                                    extra_urls=_fida_urls_gp)
        if df is None or df.empty:
            return  # weights not balanced yet — builder is shown, analysis waits
    elif "LIBERO" in ptf_choice:
        df = free_portfolio_ui(raw)
    else:
        key = "PTF FULL" if "FULL" in ptf_choice else "PTF SHORT"
        if key not in raw or raw[key].empty:
            st.error(f"❌ Foglio '{key}' non trovato o vuoto."); return
        df = raw[key]

    if df is None or df.empty: return

    wcol   = PROFILE_W_COL[profile]
    df_act = df[df[wcol]>0.001].copy()

    # KPI row
    n_fondi = len(df_act)
    w_az    = (df_act[wcol]*df_act["az_pct"]).sum()*100
    w_obb   = (df_act[wcol]*df_act["obb_pct"]).sum()*100
    w_other = 0.0
    _other_label_kpi = ""

    # SUGGERITO: il DEFAULT_AZ per categoria (92% az, 50% bilanciati, 6% bond)
    # dà stime grossolane (es. 73%). Il PDF GP riporta direttamente la quota
    # equity/bond dello scenario → usala se disponibile.
    # Vengono estratti TUTTI i componenti (Equity, Bond, Private Markets…)
    # così da mostrare anche la quota "Economia Reale / Altro".
    if _is_suggerito:
        _sc_info_kpi = (st.session_state.get("_gp_data") or {}).get(
            st.session_state.get("_gp_sc_key", "Base"), {}).get("info", "")
        # Formato: "Equity 32% · Bond 38% · Private Markets 30%"
        _info_parts = re.findall(r'([A-Za-z][A-Za-z\s&]+?)\s+(\d+)\s*%', _sc_info_kpi)
        _other_parts = []
        for _pname, _pval in _info_parts:
            _pname = _pname.strip()
            if re.search(r'equity', _pname, re.I):
                w_az  = float(_pval)
            elif re.search(r'bond', _pname, re.I):
                w_obb = float(_pval)
            else:
                w_other += float(_pval)
                _other_parts.append(_pname)
        if len(_other_parts) == 1:
            _other_label_kpi = _other_parts[0]
        elif len(_other_parts) > 1:
            _other_label_kpi = "Economia Reale / Altro"

    srri = max(1, min(7, round(w_az/100*6+1)))

    _SRRI_LABELS = {
        1: "Rischio Molto Basso",
        2: "Rischio Basso",
        3: "Rischio Medio-Basso",
        4: "Rischio Medio",
        5: "Rischio Medio-Alto",
        6: "Rischio Alto",
        7: "Rischio Molto Alto",
    }
    _srri_sub = (
        f"<span style='font-weight:600;color:#0d1b2a;'>"
        f"{_SRRI_LABELS.get(srri,'—')}</span><br>"
        f"<span style='font-size:.68rem;line-height:1.5;'>"
        f"Indicatore sintetico europeo di rischio/rendimento<br>"
        f"Scala 1 (min) → 7 (max) · stima da quota azionaria</span>"
    )

    # Costruisci la lista KPI dinamicamente (aggiunge "Economia Reale" se presente)
    _kpi_items = [
        (str(n_fondi),      "Fondi in Portafoglio",     f"{df_act['gruppo'].nunique()} gruppi"),
        (f"{w_az:.1f}%",    "Quota Azionaria",          "ponderata per peso"),
        (f"{w_obb:.1f}%",   "Quota Obbligazionaria",    "ponderata per peso"),
    ]
    if w_other > 0.5:
        _kpi_items.append((f"{w_other:.1f}%", _other_label_kpi or "Economia Reale", "ponderata per peso"))
    _kpi_items.append((f"{srri} / 7", "Risk Score (SRRI proxy)", _srri_sub))

    _kpi_cols = st.columns(len(_kpi_items))
    for col, (val, lbl, sub) in zip(_kpi_cols, _kpi_items):
        col.markdown(
            f'<div class="kpi"><div class="kpi-label">{lbl}</div>'
            f'<div class="kpi-value">{val}</div>'
            f'<div class="kpi-sub">{sub}</div></div>',
            unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)

    # Charts + fund list
    col_l,col_r = st.columns([1.15,0.85],gap="large")
    with col_l:
        st.markdown('<p class="sec-title">Allocazione per Fondo</p>',unsafe_allow_html=True)
        st.plotly_chart(make_fund_pie(df_act,wcol,profile),use_container_width=True,config={"displayModeBar":False})
        st.markdown('<p class="sec-title">Allocazione per Macro-Categoria</p>',unsafe_allow_html=True)
        st.plotly_chart(make_macro_bar(df_act,wcol),use_container_width=True,config={"displayModeBar":False})
    with col_r:
        st.markdown('<p class="sec-title">Composizione del Portafoglio</p>',unsafe_allow_html=True)
        _gruppi = list(df_act["gruppo"].unique())
        # Icon map for known group names
        _GRP_ICON = {
            "ALLOCATION":     "🔀",
            "AZIONARI (LONG)":"📈",
            "BOND":           "🏛️",
        }
        _grp_tabs = st.tabs([
            f"{_GRP_ICON.get(g, '📁')}  {g}" for g in _gruppi
        ])
        for _gtab, gruppo in zip(_grp_tabs, _gruppi):
            with _gtab:
                sub = df_act[df_act["gruppo"]==gruppo].sort_values(wcol,ascending=False)
                rows_html = "".join([
                    f'<div class="fund-row">'
                    f'<div class="fund-dot" style="background:{r["color"]};"></div>'
                    f'<div style="flex:1;min-width:0;">'
                    f'<div class="fund-name">{r.get("nome_orig") or r["nome"]}</div>'
                    f'<div class="fund-cat">'
                    f'{r["categoria"][:48]+"…" if r["categoria"] and len(r["categoria"])>48 else (r["categoria"] or "—")}'
                    f'</div></div>'
                    f'<div class="fund-pct">{r[wcol]*100:.1f}%</div>'
                    f'</div>'
                    for _, r in sub.iterrows()
                ])
                st.markdown(
                    f'<div style="background:#fff;border:1px solid #e2e8f0;'
                    f'border-radius:10px;overflow:hidden;">{rows_html}</div>',
                    unsafe_allow_html=True)

    # ── Load cached FondiDoc data (bundled in repo) ──────────────────────────
    cached_fd, cache_date = load_fund_cache()

    # ── TABBED ANALYTICS TABLES ──────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<p class="sec-title">Analisi del Portafoglio</p>',
                unsafe_allow_html=True)

    # ── Shared helpers for all four tabs ──────────────────────────────────────
    def _fb_metric(nome: str, key: str):
        """Return metric from factbook_data (duration / credit_rating / az_pct …)."""
        if not factbook_data:
            return None
        norm = _normalize_for_unp(nome)
        norm = _FUND_ALIASES.get(norm, norm)
        entry = factbook_data.get(norm)
        if not entry or not isinstance(entry, dict):
            best, best_len = None, 0
            for _fk, _fv in factbook_data.items():
                if (isinstance(_fv, dict)
                        and (_fk in norm or norm in _fk)
                        and len(_fk) > best_len):
                    best, best_len = _fv, len(_fk)
            entry = best
        if not entry or not isinstance(entry, dict):
            return None
        return entry.get(key)

    # Prefer live FondiDoc data fetched in this session over on-disk cache
    _fd_live = st.session_state.get("_scomp_fd") or cached_fd

    # Morningstar ratings — from this session or from disk cache
    _ms_live = st.session_state.get("_ms_data") or load_ms_cache()

    # Morningstar color scale (amber/gold palette)
    _MS_COL = {5: "#78350F", 4: "#92400E", 3: "#B45309", 2: "#475569", 1: "#94A3B8"}
    _MS_BG  = {5: "#78350F", 4: "#92400E", 3: "#B45309"}   # bg only for top-3

    def _ms_badge_html(ms_r) -> str:
        """Return HTML span for a Morningstar rating integer.

        Only filled stars are shown (no empty stars): ☆ on a coloured background
        is visually indistinguishable from ★ in white text, which caused ratings
        like 3★ to appear as 5★.  Showing only the earned stars is unambiguous.
        """
        try:
            v = int(ms_r)
        except (TypeError, ValueError):
            return "<span style='color:#94A3B8;'>—</span>"
        filled = "★" * v          # e.g. "★★★" for 3 — no empty stars
        bg = _MS_BG.get(v)
        if bg:
            return (f"<span style='background:{bg};color:#fff;padding:2px 8px;"
                    f"border-radius:4px;font-weight:700;font-size:.8rem;'>"
                    f"{filled}</span>")
        col = _MS_COL.get(v, "#64748B")
        return (f"<span style='color:{col};font-weight:700;font-size:.8rem;'>"
                f"{filled}</span>")

    # Shared HTML style tokens
    _TH  = ("background:#0D1B2A;color:#fff;font-size:.74rem;"
            "padding:8px 10px;white-space:nowrap;")
    _TC  = "font-size:.77rem;padding:6px 10px;border-bottom:1px solid #f1f5f9;"
    _TP  = ("background:#1B4332;color:#fff;font-size:.77rem;font-weight:700;"
            "padding:6px 10px;border-bottom:2px solid #C9A84C;")

    # FIDArating color scale: 5=best (dark green) … 1=worst (dark red)
    _FIDA_COL = {5: "#166534", 4: "#15803d", 3: "#22c55e",
                 2: "#f87171", 1: "#b91c1c"}
    _FIDA_BG  = {5: "#166534", 4: "#15803d", 3: "#22c55e"}

    # Source note labels (reused in all tab footers)
    _fb_ref_date = (factbook_data or {}).get("_ref_date", "")
    _note_fb  = (f"Factbook AZ Investments al {_fb_ref_date}" if _fb_ref_date
                 else "Factbook AZ Investments" if factbook_data
                 else "n.d. — carica il Factbook PDF nella barra laterale")
    if st.session_state.get("_scomp_fd"):
        _note_fd = "FondiDoc live (questa sessione)"
    elif cached_fd:
        _note_fd = f"FondiDoc (cache {cache_date})"
    else:
        _note_fd = "n.d. — clicca «Genera PDF» per popolare i dati FondiDoc"
    _note_style = "font-size:.71rem;color:#94A3B8;margin-top:5px;"

    # Pre-sort funds by weight (descending) — shared across all tabs
    _df_sorted = df_act.sort_values(wcol, ascending=False)

    def _perf_val_col(raw) -> str:
        """Wrap a performance string in green/red HTML span."""
        s = str(raw) if raw is not None else "-"
        try:
            v   = float(s.replace("%", "").replace(",", ".").strip())
            col = "#1A7A4A" if v > 0 else ("#C0392B" if v < 0 else "#475569")
            return f"<span style='color:{col};font-weight:700;'>{s}</span>"
        except Exception:
            return f"<span style='color:#94A3B8;'>{s}</span>"

    def _get_ana(nome: str) -> dict:
        """Restituisce il dict 'analysis' cercando prima il nome diretto,
        poi con fuzzy GP→FIDA (serve per i fondi del portafoglio suggerito
        il cui nome non è stato risolto in chiave FIDA).

        Quando _fd_live è la sessione live (Excel caricato) potrebbe non
        contenere fondi GP del portafoglio suggerito: in quel caso usa
        cached_fd (fund_cache.json bundled) come fallback aggiuntivo."""
        def _lookup(src: dict) -> dict | None:
            e = src.get(nome)
            if not e:
                res = _resolve_nome_for_fd(nome, src)
                e = src.get(res)
            if not e and nome:
                _sk = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', nome,
                             flags=re.I).strip().lower()
                if _sk:
                    for _fk, _fv in src.items():
                        if isinstance(_fv, dict) and _sk in _fk.lower():
                            e = _fv; break
            return e

        entry = _lookup(_fd_live)
        # Fallback su cached_fd se:
        #  - entry è None (fondo non trovato nel live)
        #  - OPPURE entry esiste ma senza 'analysis' (URL/overview presenti
        #    ma fetch analisi fallito — es. dopo "Aggiorna Dati" per fondi GP)
        if not (entry or {}).get("analysis") and _fd_live is not cached_fd:
            entry_c = _lookup(cached_fd)
            if (entry_c or {}).get("analysis"):
                entry = entry_c
        return (entry or {}).get("analysis", {})

    def _perf_wavg(keys: list) -> dict:
        """Weighted average of performance/risk metrics across active funds."""
        totals = {k: 0.0 for k in keys}
        cov_w  = {k: 0.0 for k in keys}
        for _, _row in df_act.iterrows():
            _w   = _row[wcol]
            _ana = _get_ana(_row["nome"])
            for k in keys:
                raw = _ana.get(k, "") or _fb_metric(_row["nome"], k)
                try:
                    num = float(str(raw).replace("%", "").replace(",", ".").strip())
                    totals[k] += num * _w
                    cov_w[k]  += _w
                except Exception:
                    pass
        return {k: (f"{totals[k]/cov_w[k]:+.2f}%" if cov_w[k] > 0.01 else "-")
                for k in keys}

    def _html_table(hdr_cols: list, ptf_row: list, fund_rows: list) -> str:
        """Render a styled HTML table with header, portfolio summary row, fund rows."""
        def _align(i):
            return "left" if i == 0 else "center"
        hdr_html = "".join(
            f"<th style='{_TH}text-align:{_align(i)};'>{h}</th>"
            for i, h in enumerate(hdr_cols)
        )
        ptf_html = "".join(
            f"<td style='{_TP}text-align:{_align(i)};'>{v}</td>"
            for i, v in enumerate(ptf_row)
        )
        body_html = f"<tr>{ptf_html}</tr>"
        for fr in fund_rows:
            row_html = "".join(
                f"<td style='{_TC}text-align:{_align(i)};'>{v}</td>"
                for i, v in enumerate(fr)
            )
            body_html += f"<tr>{row_html}</tr>"
        return (
            f"<div style='overflow-x:auto;border-radius:10px;"
            f"border:1px solid #e2e8f0;background:#fff;'>"
            f"<table style='width:100%;border-collapse:collapse;'>"
            f"<thead><tr>{hdr_html}</tr></thead>"
            f"<tbody>{body_html}</tbody>"
            f"</table></div>"
        )

    _ptf_row_label = f"◆ PORTAFOGLIO {ptf_label.upper()}"

    # URL lookup: Excel hyperlinks first, FondiDoc cache as enriched fallback
    _fida_urls_raw = raw.get("fida_urls") or dict(MANUAL_URL_OVERRIDES)

    def _fund_url(nome: str) -> str:
        """Return the FondiDoc URL for a fund, or '' if not available.

        Priorità rigorosa:
          1. MANUAL_URL_OVERRIDES direct
          2. MANUAL_URL_OVERRIDES fuzzy   ← vince su TUTTO, anche su fida_urls
          3. FondiDoc cache direct
          4. fida_urls direct  (può contenere URL sbagliati dall'Excel)
          5. FondiDoc cache fuzzy
          6. fida_urls fuzzy
        """
        # 1. Match diretto
        if nome in MANUAL_URL_OVERRIDES:
            return MANUAL_URL_OVERRIDES[nome]

        # Calcola short key una sola volta
        _sk = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', nome, flags=re.I).strip().lower()

        # 2. MANUAL fuzzy — PRIMA di qualsiasi altro lookup (incluso fida_urls)
        if _sk:
            for _mk, _mu in MANUAL_URL_OVERRIDES.items():
                _msk = re.sub(r'^AZ\s+\S+\s*[-–]\s*', '', _mk, flags=re.I).strip().lower()
                if _msk and _msk in _sk and _mu:
                    return _mu

        # 3–4. Cache e fida_urls direct
        url = (_fd_live.get(nome, {}).get("url", "")
               or _fida_urls_raw.get(nome, ""))
        if url:
            return url

        # 5–6. Fuzzy su cache e fida_urls
        if _sk:
            for _fk, _fv in _fd_live.items():
                if isinstance(_fv, dict) and _sk in _fk.lower() and _fv.get("url"):
                    return _fv["url"]
            for _fk, _eu in _fida_urls_raw.items():
                if _sk in _fk.lower() and _eu:
                    return _eu
        return ""

    def _fund_link(nome: str) -> str:
        """Return fund name as HTML — hyperlinked if URL is available."""
        url = _fund_url(nome)
        if url:
            return (f'<a href="{url}" target="_blank" rel="noopener noreferrer" '
                    f'style="color:#1B4FBB;text-decoration:underline;'
                    f'text-underline-offset:2px;">{nome}</a>')
        return nome

    # ── Quantalys cache + ratings + ISIN map ────────────────────────────────
    _qtl_cache   = load_quantalys_cache()    # {ISIN: "https://www.quantalys.it/Fonds/..."}
    _qtl_ratings = load_quantalys_ratings()  # {ISIN: {"score": 87, "globes": 5}}
    _ms_cache    = load_morningstar_cache()  # {ISIN: "https://www.morningstar.it/..."}

    def _qtl_rating_html(isin: str) -> str:
        """Restituisce HTML con score e globi Quantalys per una riga della tabella."""
        v = _qtl_ratings.get(isin)
        if not v:
            return "<span style='color:#CBD5E1;font-size:.75rem;'>—</span>"
        score  = v.get("score")
        globes = v.get("globes")
        if score is None:
            return "<span style='color:#CBD5E1;font-size:.75rem;'>—</span>"
        if globes:
            _glob_col = {1:"#EF4444",2:"#F97316",3:"#EAB308",4:"#22C55E",5:"#1B4FBB"}.get(globes,"#64748B")
            stars = (f"<span style='color:{_glob_col};font-size:.9rem;letter-spacing:1px;'>"
                     f"{'★'*globes}{'☆'*(5-globes)}</span>")
        else:
            stars = ""
        score_col = ("#1B4FBB" if score >= 80 else "#22C55E" if score >= 60
                     else "#EAB308" if score >= 40 else "#EF4444")
        score_html = (f"<span style='font-size:.75rem;font-weight:700;color:{score_col};"
                      f"background:#F1F5F9;border-radius:3px;padding:1px 4px;'>{score}</span>")
        tooltip = f"Score {score}/100" + (f" · {globes} globi" if globes else "")
        return (f"<span title='{tooltip}'>{stars}"
                f"{'<br>' if stars else ''}{score_html}</span>")
    _fida_raw_ui = raw.get("FIDA", pd.DataFrame())
    _isin_map_ui: dict[str, str] = {}
    if isinstance(_fida_raw_ui, pd.DataFrame) and not _fida_raw_ui.empty and "isin" in _fida_raw_ui.columns:
        for _, _fr in _fida_raw_ui.iterrows():
            _fi = str(_fr.get("isin", "") or "").strip()
            _fn = str(_fr.get("nome", "") or "").strip()
            if _fi and _fn:
                _isin_map_ui[_fn] = _fi
    # Aggiungi ISIN mancanti dall'Excel (override manuale)
    _isin_map_ui.update(MANUAL_ISIN_OVERRIDES)

    def _qtl_concept_key(name: str) -> str:
        """Estrae il "nome-concetto" del fondo rimuovendo il prefisso AZ e il suffisso
        di classe (A/B Cap/Dis EUR ecc.) per poter abbinare classi diverse dello stesso fondo."""
        n = name.strip()
        # Rimuove prefisso tipo "AZ F.1 All. " / "AZ F.1 Eq. " / "AZ Allocation - " ecc.
        n = re.sub(r'^AZ\s+(?:F\.\d+\s+\w+[\. ]+|Fund\s+\d+\s*[-–]\s*|\w+\s*[-–]\s*)', '', n, flags=re.I).strip()
        # Rimuove suffisso di classe: "A Cap EUR", "B Dis EUR(i)", "A-HU Cap EUR Hdg", ecc.
        n = re.sub(r'\s+[A-Z](?:-[A-Z0-9]+)?\s+(?:Cap|Dis|Acc|Inc)\b.*$', '', n, flags=re.I).strip()
        return n.lower()

    # Mappa concetto → url (fallback quando l'ISIN specifico non è in cache)
    # Usa tutte le classi trovate, così se la classe B è in cache ma la A no, il fondo si linka comunque
    _qtl_concept_map: dict[str, str] = {}
    for _ck_nome, _ck_isin in _isin_map_ui.items():
        _ck_url = _qtl_cache.get(_ck_isin, "")
        if _ck_url:
            _ck = _qtl_concept_key(_ck_nome)
            if _ck and _ck not in _qtl_concept_map:
                _qtl_concept_map[_ck] = _ck_url

    # Estendi con ISIN da fida_urls (es. classe A non presente in FIDA):
    # URL fondidoc.it contengono l'ISIN nel percorso  →  estraiamolo per cercare in _qtl_cache
    _fidu_isin_re = re.compile(r'/([A-Z]{2}[A-Z0-9]{10})_')
    for _cu_nome, _cu_fdurl in (raw.get("fida_urls") or {}).items():
        _cu_m = _fidu_isin_re.search(str(_cu_fdurl))
        if not _cu_m:
            continue
        _cu_isin = _cu_m.group(1)
        _cu_url  = _qtl_cache.get(_cu_isin, "")
        if _cu_url:
            _cu_ck = _qtl_concept_key(_cu_nome)
            if _cu_ck and _cu_ck not in _qtl_concept_map:
                _qtl_concept_map[_cu_ck] = _cu_url

    tab1, tab_q, tab2, tab3, tab4, tab_fp = st.tabs([
        "📊  Scomposizione Az/Obb",
        "🔗  Quantalys · Morningstar",
        "📈  Rendimenti",
        "⚠️  Rischio",
        "💰  UNP / IUNP",
        "🏦  Fondi Pensione",
    ])

    # ── TAB 1 — SCOMPOSIZIONE ────────────────────────────────────────────────
    with tab1:
        _scomp_hdr = (
            f"<tr>"
            f"<th style='{_TH}text-align:left;'>Fondo</th>"
            f"<th style='{_TH}text-align:center;'>Peso</th>"
            f"<th style='{_TH}text-align:center;'>% Az.</th>"
            f"<th style='{_TH}text-align:center;'>% Obb.</th>"
            f"<th style='{_TH}text-align:center;'>Duration</th>"
            f"<th style='{_TH}text-align:center;'>Rating Medio</th>"
            f"<th style='{_TH}text-align:left;'>Cat. FIDA</th>"
            f"<th style='{_TH}text-align:center;'>FIDArating</th>"
            f"<th style='{_TH}text-align:center;'>Morningstar</th>"
            f"<th style='{_TH}text-align:center;'>Quantalys</th>"
            f"</tr>"
        )
        _tbl_body = ""
        for _, _tr in _df_sorted.iterrows():
            _dur   = _fb_metric(_tr["nome"], "duration")
            _rat   = _fb_metric(_tr["nome"], "credit_rating")
            _azfb  = _fb_metric(_tr["nome"], "fb_az_pct")
            _obfb  = _fb_metric(_tr["nome"], "fb_obb_pct")
            _az_d  = (_azfb if _azfb is not None else _tr["az_pct"]) * 100
            _ob_d  = (_obfb if _obfb is not None else _tr["obb_pct"]) * 100
            _fd_ov = _fd_live.get(_tr["nome"], {}).get("overview", {})
            _cat   = _fd_ov.get("cat_assog") or "—"
            _fida  = _fd_ov.get("fida_rating") or "—"
            _dur_s = f"{_dur:.2f} y" if isinstance(_dur, (int, float)) else "—"
            _rat_s = _rat if isinstance(_rat, str) else "—"
            _rat_w = "600" if _rat_s != "—" else "400"
            try:
                _fr_int = int(_fida)
                _fr_col = _FIDA_COL.get(_fr_int, "#64748B")
                _fr_bg  = _FIDA_BG.get(_fr_int)
            except (ValueError, TypeError):
                _fr_col, _fr_bg = "#64748B", None
            _fida_cell = (
                f"<span style='background:{_fr_bg};color:#fff;padding:2px 8px;"
                f"border-radius:4px;font-weight:700;'>{_fida}</span>"
                if _fr_bg else
                f"<span style='color:{_fr_col};font-weight:700;'>{_fida}</span>"
            )
            _ms_r = _ms_live.get(_tr["nome"], {}).get("ms_rating")
            _ms_cell = _ms_badge_html(_ms_r)
            # Quantalys rating — stesso lookup ISIN del tab_q
            _tr_isin = _isin_map_ui.get(_tr["nome"], "")
            if not _tr_isin:
                _tr_n = re.sub(r'[^a-z0-9]', '', _tr["nome"].lower())
                for _ik, _iv in _isin_map_ui.items():
                    if re.sub(r'[^a-z0-9]', '', _ik.lower()) == _tr_n:
                        _tr_isin = _iv; break
            _qtl_r_cell = _qtl_rating_html(_tr_isin)
            _tbl_body += (
                f"<tr>"
                f"<td style='{_TC}font-weight:500;'>{_fund_link(_tr['nome'])}</td>"
                f"<td style='{_TC}text-align:center;color:#1B4FBB;font-weight:600;'>"
                f"{_tr[wcol]*100:.1f}%</td>"
                f"<td style='{_TC}text-align:center;'>{_az_d:.1f}%</td>"
                f"<td style='{_TC}text-align:center;'>{_ob_d:.1f}%</td>"
                f"<td style='{_TC}text-align:center;'>{_dur_s}</td>"
                f"<td style='{_TC}text-align:center;font-weight:{_rat_w};'>{_rat_s}</td>"
                f"<td style='{_TC}color:#64748B;'>{_cat}</td>"
                f"<td style='{_TC}text-align:center;'>{_fida_cell}</td>"
                f"<td style='{_TC}text-align:center;'>{_ms_cell}</td>"
                f"<td style='{_TC}text-align:center;'>{_qtl_r_cell}</td>"
                f"</tr>"
            )
        if _tbl_body:
            st.markdown(
                f"<div style='overflow-x:auto;border-radius:10px;"
                f"border:1px solid #e2e8f0;background:#fff;'>"
                f"<table style='width:100%;border-collapse:collapse;'>"
                f"<thead>{_scomp_hdr}</thead><tbody>{_tbl_body}</tbody>"
                f"</table></div>",
                unsafe_allow_html=True)
            _ms_note = (f" &nbsp;·&nbsp; Morningstar: FondiOnline"
                        if _ms_with_rating else
                        " &nbsp;·&nbsp; Morningstar: clicca «Scarica Rating Morningstar»")
            st.markdown(
                f"<p style='{_note_style}'>"
                f"Duration &amp; Rating Medio: {_note_fb}"
                f" &nbsp;·&nbsp; Cat. FIDA &amp; FIDArating: {_note_fd}"
                f"{_ms_note}</p>",
                unsafe_allow_html=True)

    # ── TAB Q — QUANTALYS ────────────────────────────────────────────────────
    with tab_q:
        _TH_Q = ("background:#1B4FBB;color:#fff;padding:8px 12px;"
                 "font-size:.78rem;font-weight:700;white-space:nowrap;")
        _TC_Q = "padding:7px 12px;border-bottom:1px solid #f1f5f9;font-size:.82rem;"

        _q_hdr = (
            f"<tr>"
            f"<th style='{_TH_Q}text-align:left;'>Fondo</th>"
            f"<th style='{_TH_Q}text-align:center;'>Peso</th>"
            f"<th style='{_TH_Q}text-align:left;'>ISIN</th>"
            f"<th style='{_TH_Q}text-align:center;'>Quantalys</th>"
            f"<th style='{_TH_Q}text-align:center;background:#c2410c;'>Morningstar</th>"
            f"</tr>"
        )
        _q_body = ""
        _q_found = 0
        _ms_found = 0
        for _, _qr in _df_sorted.iterrows():
            _qnome = _qr["nome"]
            _qdisp = (_qr["nome_orig"]
                      if "nome_orig" in _df_sorted.columns and _qr.get("nome_orig")
                      else _qnome)
            _qpeso = f"{_qr[wcol]*100:.1f}%"
            # ISIN lookup (direct + normalized fallback)
            _qisin = _isin_map_ui.get(_qnome, "")
            if not _qisin:
                _qn_norm = _normalize_for_unp(_qnome)
                for _ik, _iv in _isin_map_ui.items():
                    if _normalize_for_unp(_ik) == _qn_norm:
                        _qisin = _iv
                        break
            # Fund name cell — look up URL by internal nome, display with nome_orig if available
            _q_furl = _fund_url(_qnome)
            _q_name_cell = (
                f"<a href='{_q_furl}' target='_blank' rel='noopener noreferrer' "
                f"style='color:#1B4FBB;text-decoration:underline;text-underline-offset:2px;'>"
                f"{_qdisp}</a>"
                if _q_furl else _qdisp
            )
            # Quantalys URL — prima ISIN diretto, poi fallback per nome-concetto
            _qurl = _qtl_cache.get(_qisin, "") if _qisin else ""
            if not _qurl:
                _qck = _qtl_concept_key(_qnome)
                _qurl = _qtl_concept_map.get(_qck, "")
                if not _qurl and _qdisp != _qnome:
                    _qck2 = _qtl_concept_key(_qdisp)
                    _qurl = _qtl_concept_map.get(_qck2, "")
            if _qurl:
                _q_found += 1
                _qtl_cell = (
                    f"<a href='{_qurl}' target='_blank' rel='noopener noreferrer' "
                    f"style='display:inline-block;padding:3px 12px;background:#1B4FBB;"
                    f"color:#fff;border-radius:5px;font-size:.77rem;font-weight:600;"
                    f"text-decoration:none;'>Apri &#x2197;</a>"
                )
            elif _qisin:
                _qtl_cell = "<span style='color:#94A3B8;font-size:.77rem;'>non trovato</span>"
            else:
                _qtl_cell = "<span style='color:#CBD5E1;font-size:.77rem;'>nessun ISIN</span>"
            # Morningstar URL — lookup diretto per ISIN
            _msurl = _ms_cache.get(_qisin, "") if _qisin else ""
            if _msurl:
                _ms_found += 1
                _ms_cell = (
                    f"<a href='{_msurl}' target='_blank' rel='noopener noreferrer' "
                    f"style='display:inline-block;padding:3px 12px;background:#c2410c;"
                    f"color:#fff;border-radius:5px;font-size:.77rem;font-weight:600;"
                    f"text-decoration:none;'>Apri &#x2197;</a>"
                )
            elif _qisin:
                _ms_cell = "<span style='color:#94A3B8;font-size:.77rem;'>non trovato</span>"
            else:
                _ms_cell = "<span style='color:#CBD5E1;font-size:.77rem;'>nessun ISIN</span>"
            _qisin_cell = (
                f"<span style='font-family:monospace;font-size:.78rem;color:#475569;'>{_qisin}</span>"
                if _qisin else
                "<span style='color:#CBD5E1;'>—</span>"
            )
            _q_body += (
                f"<tr>"
                f"<td style='{_TC_Q}font-weight:500;'>{_q_name_cell}</td>"
                f"<td style='{_TC_Q}text-align:center;color:#1B4FBB;font-weight:600;'>{_qpeso}</td>"
                f"<td style='{_TC_Q}'>{_qisin_cell}</td>"
                f"<td style='{_TC_Q}text-align:center;'>{_qtl_cell}</td>"
                f"<td style='{_TC_Q}text-align:center;'>{_ms_cell}</td>"
                f"</tr>"
            )
        if _q_body:
            st.markdown(
                f"<div style='overflow-x:auto;border-radius:10px;"
                f"border:1px solid #e2e8f0;background:#fff;'>"
                f"<table style='width:100%;border-collapse:collapse;'>"
                f"<thead>{_q_hdr}</thead><tbody>{_q_body}</tbody>"
                f"</table></div>",
                unsafe_allow_html=True)
            _note_parts = []
            if _qtl_cache:
                _q_pct = int(_q_found / len(_df_sorted) * 100) if len(_df_sorted) else 0
                _note_parts.append(
                    f"Quantalys: {_q_found}/{len(_df_sorted)} fondi ({_q_pct}%) · "
                    f"<a href='https://www.quantalys.it' target='_blank' "
                    f"style='color:#94A3B8;'>quantalys.it</a>"
                )
            if _ms_cache:
                _ms_pct = int(_ms_found / len(_df_sorted) * 100) if len(_df_sorted) else 0
                _note_parts.append(
                    f"Morningstar: {_ms_found}/{len(_df_sorted)} fondi ({_ms_pct}%) · "
                    f"<a href='https://www.morningstar.it' target='_blank' "
                    f"style='color:#94A3B8;'>morningstar.it</a>"
                )
            if _note_parts:
                st.markdown(
                    f"<p style='{_note_style}'>{' &nbsp;|&nbsp; '.join(_note_parts)}</p>",
                    unsafe_allow_html=True)
            if not _qtl_cache:
                st.info(
                    "Cache Quantalys non ancora disponibile. "
                    "Esegui `python build_quantalys_cache.py` per generarla.",
                    icon="ℹ️")

    # ── TAB 2 — RENDIMENTI ───────────────────────────────────────────────────
    with tab2:
        _pk      = ["ytd", "perf_1y", "perf_3y", "perf_5y", "vol_1y", "sharpe_1y"]
        _ptf_p   = _perf_wavg(_pk)
        _p_hdr   = ["Fondo", "Peso", "YTD", "1 Anno", "3 Anni", "5 Anni",
                    "Vol. 1A", "Sharpe 1A"]
        _p_ptf   = [_ptf_row_label, "100%",
                    _ptf_p.get("ytd",      "-"),
                    _ptf_p.get("perf_1y",  "-"),
                    _ptf_p.get("perf_3y",  "-"),
                    _ptf_p.get("perf_5y",  "-"),
                    _ptf_p.get("vol_1y",   "-"),
                    _ptf_p.get("sharpe_1y","-")]
        _p_funds = []
        for _, _pr in _df_sorted.iterrows():
            _np  = _pr["nome"]
            _ana = _get_ana(_np)
            def _gp(k, _n=_np, _a=_ana):
                v = _a.get(k, "") or _fb_metric(_n, k) or ""
                return str(v) if v else "-"
            _p_funds.append([
                _fund_link(_np),
                f"{_pr[wcol]*100:.1f}%",
                _perf_val_col(_gp("ytd")),
                _perf_val_col(_gp("perf_1y")),
                _perf_val_col(_gp("perf_3y")),
                _perf_val_col(_gp("perf_5y")),
                _gp("vol_1y"),
                _gp("sharpe_1y"),
            ])
        st.markdown(_html_table(_p_hdr, _p_ptf, _p_funds), unsafe_allow_html=True)
        st.markdown(
            f"<p style='{_note_style}'>"
            f"Rendimenti &amp; Rischio: {_note_fd} &nbsp;·&nbsp; Fallback: {_note_fb}</p>",
            unsafe_allow_html=True)

        # ── Cono di Ibbotson ─────────────────────────────────────────────────
        with st.expander("📐 Cono di Ibbotson — Proiezione futura", expanded=False):
            _ib_mu, _ib_sig, _ib_n_ok, _ib_n_tot, _ib_missing = _az_portfolio_mu_sigma(
                _df_sorted, wcol, _fd_live, factbook_data,
                lambda nome, key: _fb_metric(nome, key) or "")

            _c1, _c2 = st.columns([1, 2])
            with _c1:
                _ib_cap = st.number_input("Capitale iniziale (€)", min_value=1_000,
                                          max_value=10_000_000, value=100_000, step=5_000,
                                          key="_ib_cap_az")
                _ib_hor = st.slider("Orizzonte (anni)", 1, 30, 10, key="_ib_hor_az")
            with _c2:
                _ib_rel   = round(40 + 35 * (_ib_n_ok / max(_ib_n_tot, 1)))
                _ib_badge = "🟢" if _ib_rel >= 65 else ("🟡" if _ib_rel >= 50 else "🟠")
                st.metric("Attendibilità stima", f"{_ib_badge} {_ib_rel}%",
                          help=(
                              f"σ storica disponibile per {_ib_n_ok}/{_ib_n_tot} fondi. "
                              "μ sempre stimato da prior forward-looking per categoria (non rendimenti storici). "
                              "Il massimo raggiungibile è 75%, anche con tutti i dati di volatilità disponibili, "
                              "perché il rendimento atteso è sempre una stima qualitativa."
                          ))
                st.caption(
                    f"Rendimento atteso portafoglio: **{_ib_mu*100:+.2f}%** &nbsp;·&nbsp; "
                    f"Volatilità: **{_ib_sig*100:.2f}%** &nbsp;·&nbsp; "
                    f"_Prior forward-looking per categoria_")

                import numpy as np
                import plotly.graph_objects as go

                _t  = np.linspace(0, _ib_hor, 300)
                _mu_log = _ib_mu - _ib_sig ** 2 / 2
                _med    = _ib_cap * np.exp(_mu_log * _t)
                _up1    = _ib_cap * np.exp((_mu_log + _ib_sig) * _t)
                _dn1    = _ib_cap * np.exp((_mu_log - _ib_sig) * _t)
                _up2    = _ib_cap * np.exp((_mu_log + 2 * _ib_sig) * _t)
                _dn2    = _ib_cap * np.exp((_mu_log - 2 * _ib_sig) * _t)

                _fig_ib = go.Figure()
                _fig_ib.add_trace(go.Scatter(
                    x=np.concatenate([_t, _t[::-1]]),
                    y=np.concatenate([_up2, _dn2[::-1]]),
                    fill="toself", fillcolor="rgba(191,219,254,0.45)",
                    line=dict(width=0), name="95% dei percorsi (±2σ)"))
                _fig_ib.add_trace(go.Scatter(
                    x=np.concatenate([_t, _t[::-1]]),
                    y=np.concatenate([_up1, _dn1[::-1]]),
                    fill="toself", fillcolor="rgba(59,130,246,0.30)",
                    line=dict(width=0), name="68% dei percorsi (±1σ)"))
                _fig_ib.add_trace(go.Scatter(
                    x=_t, y=_med, mode="lines",
                    line=dict(color="#1B4FBB", width=2),
                    name="Percorso centrale (mediana)"))
                _fig_ib.add_hline(y=_ib_cap, line_dash="dot", line_color="#94A3B8",
                                  annotation_text=f"Capitale: € {_ib_cap:,.0f}".replace(",","."))
                _y_min = min(_ib_cap * 0.40, float(_dn2.min()) * 0.92)
                _y_max = float(_up2.max()) * 1.05
                _fig_ib.update_layout(
                    height=380,
                    margin=dict(l=10, r=10, t=30, b=10),
                    legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0),
                    xaxis=dict(title="Anni"),
                    yaxis=dict(title="Valore (€)", range=[_y_min, _y_max],
                               tickformat=",.0f"),
                    plot_bgcolor="white",
                    paper_bgcolor="white",
                )
                st.plotly_chart(_fig_ib, use_container_width=True)

                # Tabella ±1σ — HTML per header su due righe
                _ib_scen_rows = _ibbotson_table_rows(_ib_mu, _ib_sig, float(_ib_cap), years=(1,3,5,10))
                _TH_IB = ("background:#0D1B2A;color:#fff;font-size:.75rem;font-weight:600;"
                          "padding:8px 10px;text-align:center;line-height:1.35;")
                _TD_IB = ("font-size:.82rem;padding:7px 10px;text-align:center;"
                          "border-bottom:1px solid #E2E8F0;")
                _ib_html = (
                    "<table style='width:100%;border-collapse:collapse;margin-top:6px;'>"
                    "<thead><tr>"
                    f"<th style='{_TH_IB}width:8%;'>Anni</th>"
                    f"<th style='{_TH_IB}'>Scenario sfavorevole<br><span style='font-weight:400;font-size:.7rem;'>5 su 6 finiscono sopra questa soglia</span></th>"
                    f"<th style='{_TH_IB}'>Caso centrale<br><span style='font-weight:400;font-size:.7rem;'>3 su 6 finiscono sopra questa soglia</span></th>"
                    f"<th style='{_TH_IB}'>Scenario favorevole<br><span style='font-weight:400;font-size:.7rem;'>1 su 6 finisce sopra questa soglia</span></th>"
                    "</tr></thead><tbody>"
                )
                for _i, (_yr, _d2, _d1, _md, _u1, _u2) in enumerate(_ib_scen_rows):
                    _bg = "#F8FAFC" if _i % 2 else "#ffffff"
                    def _fv(v): return f"€ {v:,.0f}".replace(",",".")
                    _ib_html += (
                        f"<tr style='background:{_bg};'>"
                        f"<td style='{_TD_IB}font-weight:600;'>{_yr}</td>"
                        f"<td style='{_TD_IB}'>{_fv(_d1)}</td>"
                        f"<td style='{_TD_IB}'>{_fv(_md)}</td>"
                        f"<td style='{_TD_IB}'>{_fv(_u1)}</td>"
                        f"</tr>"
                    )
                _ib_html += "</tbody></table>"
                st.markdown(_ib_html, unsafe_allow_html=True)
                st.caption(
                    f"**Come leggere la tabella** — Caso centrale: rendimento annuo composto "
                    f"{_ib_mu*100:+.1f}% (prior categoriale) meno il costo della varianza "
                    f"(drag = σ²/2 = {_ib_sig**2/2*100:.2f}%). "
                    f"Sfavorevole/Favorevole = percorsi a ±1σ: in 2 anni su 3 il valore reale "
                    f"si collocherebbe tra questi due valori. "
                    f"Il grafico mostra anche la banda esterna ±2σ (95% dei percorsi). "
                    f"La dispersione cresce con il tempo per effetto del compounding: "
                    f"è matematicamente corretta ma non significa che i casi estremi siano probabili. "
                    f"Rendimento atteso: prior forward-looking per categoria (non dati storici recenti). "
                    f"Non costituisce previsione garantita.")
                if _ib_missing:
                    _miss_html = (
                        "<div style='margin-top:10px;padding:10px 14px;"
                        "background:#FFF7ED;border-left:3px solid #F59E0B;"
                        "border-radius:4px;font-size:.78rem;color:#78350F;'>"
                        f"<b>⚠️ {len(_ib_missing)} fondo/i privo/i di dati di volatilità storica</b> "
                        "— la volatilità è sostituita dal default di categoria "
                        f"(abbassa l'attendibilità a {_ib_rel}% invece del massimo 75%):<br>"
                        "<ul style='margin:4px 0 0 16px;'>"
                        + "".join(f"<li><b>{n}</b> — {causa}</li>" for n, causa in _ib_missing)
                        + "</ul></div>"
                    )
                    st.markdown(_miss_html, unsafe_allow_html=True)

    # ── TAB 3 — RISCHIO ──────────────────────────────────────────────────────
    with tab3:
        _rk      = ["vol_1y", "vol_3y", "vol_5y", "neg_vol_1y", "sharpe_3y", "sortino_1y"]
        _ptf_r   = _perf_wavg(_rk)
        _r_hdr   = ["Fondo", "Peso", "Vol. 1A", "Vol. 3A", "Vol. 5A",
                    "Vol. Neg. 1A", "Sharpe 3A", "Sortino 1A"]
        _r_ptf   = [_ptf_row_label, "100%",
                    _ptf_r.get("vol_1y",     "-"),
                    _ptf_r.get("vol_3y",     "-"),
                    _ptf_r.get("vol_5y",     "-"),
                    _ptf_r.get("neg_vol_1y", "-"),
                    _ptf_r.get("sharpe_3y",  "-"),
                    _ptf_r.get("sortino_1y", "-")]
        _r_funds = []
        for _, _rr in _df_sorted.iterrows():
            _nr  = _rr["nome"]
            _ana = _get_ana(_nr)
            def _gr(k, _a=_ana):
                return str(_a.get(k, "") or "") or "-"
            _r_funds.append([
                _fund_link(_nr),
                f"{_rr[wcol]*100:.1f}%",
                _gr("vol_1y"),
                _gr("vol_3y"),
                _gr("vol_5y"),
                _gr("neg_vol_1y"),
                _gr("sharpe_3y"),
                _gr("sortino_1y"),
            ])
        st.markdown(_html_table(_r_hdr, _r_ptf, _r_funds), unsafe_allow_html=True)
        st.markdown(
            f"<p style='{_note_style}'>Metriche di rischio: {_note_fd}</p>",
            unsafe_allow_html=True)

    # ── TAB 4 — UNP / IUNP ───────────────────────────────────────────────────
    with tab4:
        _u_wtd = _iu_wtd = _u_covw = 0.0
        _u_funds = []
        for _, _ur in _df_sorted.iterrows():
            _nu       = _ur["nome"]
            _uu, _iuu = lookup_unp(_nu)
            _wu       = _ur[wcol]
            if _uu is not None:
                _u_wtd  += _uu  * _wu
                _iu_wtd += _iuu * _wu
                _u_covw += _wu
            # FIDArating badge
            _fd_ov_u  = _fd_live.get(_nu, {}).get("overview", {})
            _fida_u   = _fd_ov_u.get("fida_rating") or "—"
            try:
                _fri_u  = int(_fida_u)
                _fcol_u = _FIDA_COL.get(_fri_u, "#64748B")
                _fbg_u  = _FIDA_BG.get(_fri_u)
            except (ValueError, TypeError):
                _fcol_u, _fbg_u = "#64748B", None
            _fida_cell_u = (
                f"<span style='background:{_fbg_u};color:#fff;padding:2px 8px;"
                f"border-radius:4px;font-weight:700;'>{_fida_u}</span>"
                if _fbg_u else
                f"<span style='color:{_fcol_u};font-weight:700;'>{_fida_u}</span>"
            )
            # Morningstar badge
            _ms_r_u    = _ms_live.get(_nu, {}).get("ms_rating")
            _ms_cell_u = _ms_badge_html(_ms_r_u)
            _u_funds.append([
                _fund_link(_nu),
                f"{_wu*100:.1f}%",
                f"{_uu:.2f}%"  if _uu  is not None else "—",
                f"{_iuu:.2f}%" if _iuu is not None else "—",
                _fida_cell_u,
                _ms_cell_u,
            ])
        _ptf_unp  = f"{_u_wtd/_u_covw:.2f}%"  if _u_covw > 0.01 else "-"
        _ptf_iunp = f"{_iu_wtd/_u_covw:.2f}%"  if _u_covw > 0.01 else "-"
        st.markdown(
            _html_table(
                ["Fondo", "Peso", "%UNP", "%IUNP36", "FIDArating", "Morningstar"],
                [_ptf_row_label, "100%", _ptf_unp, _ptf_iunp, "", ""],
                _u_funds,
            ),
            unsafe_allow_html=True)
        st.markdown(
            f"<p style='{_note_style}'>"
            f"UNP = Utile Netto di Portafoglio · IUNP36 = indice su orizzonte triennale. "
            f"Fonte: Catalogo Prodotti &amp; Servizi Azimut, settembre 2025.</p>",
            unsafe_allow_html=True)

    # ── TAB FP — FONDI PENSIONE ──────────────────────────────────────────────
    with tab_fp:
        _fp_data = st.session_state.get("_fp_data") or load_fp_cache()
        _fp_ref  = (_fp_data or {}).get("_ref_date", "")
        _fp_all  = {k: v for k, v in (_fp_data or {}).items()
                    if k != "_ref_date" and isinstance(v, dict)}
        # Mostra solo i FP presenti nel portafoglio attivo
        _ptf_nomi = set(df_act["nome"].tolist())
        _fp_funds = {k: v for k, v in _fp_all.items() if k in _ptf_nomi}
        if not _fp_funds:
            if not _fp_all:
                st.info(
                    "Nessun dato fondi pensione disponibile. "
                    "Carica il Factbook Fondi Pensione PDF dalla barra laterale.",
                    icon="🏦")
            else:
                st.info(
                    "Nessun fondo pensione nel portafoglio attivo. "
                    "Aggiungili dal costruttore LIBERO.",
                    icon="🏦")
        else:
            _fp_note = f"Fonte: Factbook Fondi Pensione · {_fp_ref}" if _fp_ref else "Fonte: Factbook Fondi Pensione"
            st.markdown(
                f"<p style='{_note_style}'><b>🏦 Fondi Pensione</b> &nbsp;·&nbsp; {_fp_note}</p>",
                unsafe_allow_html=True)
            _fp_hdr = ["Fondo", "Peso", "YTD", "1 Anno", "3 Anni", "5 Anni"]
            _fp_ptf = [f"<b>{len(_fp_funds)} fondi</b>", "", "", "", "", ""]

            def _pv(val):
                if not val or val == "-":
                    return "<span style='color:#94A3B8;'>—</span>"
                try:
                    v = float(str(val).replace("%","").replace("+","").replace(",","."))
                    c = "#1A7A4A" if v > 0 else ("#C0392B" if v < 0 else "#475569")
                    return f"<span style='color:{c};font-weight:700;'>{val}</span>"
                except Exception:
                    return val

            # Peso dal portafoglio attivo (ordine per peso decrescente)
            _fp_peso = {r["nome"]: r[wcol] for _, r in _df_sorted.iterrows()
                        if r["nome"] in _fp_funds}
            _fp_rows = []
            for _fn, _fv in sorted(_fp_funds.items(),
                                   key=lambda x: _fp_peso.get(x[0], 0), reverse=True):
                _wp = _fp_peso.get(_fn, 0)
                _fp_rows.append([
                    f"<span style='font-weight:500;'>{_fn}</span>",
                    f"{_wp*100:.1f}%",
                    _pv(_fv.get("ytd",     "")),
                    _pv(_fv.get("perf_1y", "")),
                    _pv(_fv.get("perf_3y", "")),
                    _pv(_fv.get("perf_5y", "")),
                ])
            st.markdown(
                _html_table(_fp_hdr, _fp_ptf, _fp_rows),
                unsafe_allow_html=True)
            st.markdown(
                f"<p style='{_note_style}'>{_fp_note} &nbsp;·&nbsp; "
                f"Rendimenti netti di imposta.</p>",
                unsafe_allow_html=True)

    # ── DOWNLOAD SECTION ─────────────────────────────────────
    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown('<p class="sec-title">Esporta Report PDF Completo</p>',unsafe_allow_html=True)

    fida_urls = raw.get("fida_urls", {})
    n_urls = sum(1 for nome in df_act["nome"].unique() if nome in fida_urls)

    col_btn,col_inf = st.columns([1,2])
    with col_inf:
        if factbook_data:
            _src_note = f"📖 Factbook: rendimenti per <b>{len(factbook_data)}</b> fondi"
        elif cached_fd:
            _src_note = f"💾 Cache FondiDoc aggiornata al: <b>{cache_date}</b>"
        else:
            _src_note = f"🌐 Dati live da FondiDoc per {n_urls}/{n_fondi} fondi"
        st.markdown(
            f"<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;"
            f"padding:1rem 1.25rem;'>"
            f"<div style='font-size:.8rem;color:#1d4ed8;font-weight:600;margin-bottom:.4rem;'>"
            f"Il report PDF (3 sezioni) contiene:</div>"
            f"<div style='font-size:.82rem;color:#1e40af;line-height:1.9;'>"
            f"✓ <b>Pag. 1</b> — Grafico a torta + KPI di portafoglio<br>"
            f"✓ <b>Pag. 2</b> — Tavola rendimenti YTD / 1A / 3A / 5A + Rischio<br>"
            f"✓ <b>Pag. 3+</b> — Schede analitiche per {n_fondi} fondi<br>"
            f"<span style='color:#3b82f6;'>{_src_note}</span></div></div>",
            unsafe_allow_html=True)

    fida_df  = raw.get("FIDA", pd.DataFrame())
    _is_free = "LIBERO" in ptf_choice

    # ── Opzioni PDF ───────────────────────────────────────────────────────────
    _print_unp = st.checkbox(
        "Includi tabella UNP/IUNP nel PDF",
        value=False,
        key="_pdf_print_unp",
        help="Aggiunge una tabella con i dati UNP e IUNP per ogni fondo (commissioni nette consulente).",
    )
    _qtl_charts_pdf = st.checkbox(
        "Includi grafici Quantalys nel PDF",
        value=False,
        key="_pdf_qtl_charts",
        help=(
            "Aggiunge sotto ogni scheda fondo i 6 grafici Quantalys "
            "(serie storiche 1/3/5 anni + fondo vs categoria). "
            "Richiede Playwright — la prima generazione può richiedere ~1 min."
        ),
    )
    if _qtl_charts_pdf:
        # Test Playwright visibile — clicca per verificare
        if st.button("🔍 Verifica Playwright", key="_btn_pw_test", help="Testa se Chromium è disponibile sul server"):
            with st.spinner("Test Playwright in corso…"):
                _pw_ok = False
                _pw_msg = ""
                try:
                    from playwright.sync_api import sync_playwright  # type: ignore
                    with sync_playwright() as _p:
                        _b = _p.chromium.launch(headless=True)
                        _pg = _b.new_page()
                        _pg.goto("https://www.quantalys.it/Fonds/Historique/825616",
                                 wait_until="domcontentloaded", timeout=30_000)
                        _pg.wait_for_selector(".qtjs-chart-histo svg", timeout=15_000)
                        _svg_n = _pg.evaluate("() => document.querySelectorAll('svg').length")
                        _b.close()
                    _pw_ok  = True
                    _pw_msg = f"✅ Playwright OK — SVG trovati: {_svg_n}"
                except ImportError:
                    _pw_msg = "❌ Playwright non installato (ImportError)"
                except Exception as _ex:
                    _pw_msg = f"❌ Errore: {_ex}"
            if _pw_ok:
                st.success(_pw_msg)
            else:
                st.error(_pw_msg)
    if _qtl_charts_pdf:
        # Verifica rapida: quanti fondi hanno URL Quantalys?
        _qtl_chk  = load_quantalys_cache()
        _fida_tmp = raw.get("FIDA", pd.DataFrame())
        _imap_tmp = ({r["nome"]: str(r["isin"]).strip()
                      for _, r in _fida_tmp.iterrows()
                      if str(r.get("isin","")).strip()}
                     if not _fida_tmp.empty and "isin" in _fida_tmp.columns else {})
        _imap_tmp.update(MANUAL_ISIN_OVERRIDES)

        _qtl_ok, _qtl_miss = [], []
        for _, _rr in df_act.iterrows():
            _rn   = _rr["nome"]
            _risin = _imap_tmp.get(_rn, "")
            if _qtl_chk.get(_risin, ""):
                _qtl_ok.append(_rn)
            else:
                _qtl_miss.append((_rn, _risin or "ISIN mancante"))

        if not _qtl_ok:
            st.warning(
                "⚠️ Nessun fondo del portafoglio ha un URL Quantalys in cache. "
                "I grafici non verranno inclusi.",
                icon="⚠️"
            )
        else:
            st.info(
                f"📊 Grafici Quantalys per **{len(_qtl_ok)}/{len(df_act)}** fondi. "
                "Prima cattura ~10s/fondo, poi in cache su disco.",
                icon="ℹ️"
            )

        if _qtl_miss:
            with st.expander(f"▸ {len(_qtl_miss)} fondo/i senza URL Quantalys — clicca per vedere", expanded=True):
                st.markdown(
                    "Incollami il link Quantalys (es. `https://www.quantalys.it/Fonds/12345`) "
                    "per ciascuno e lo aggiungo alla cache:\n"
                )
                for _mn, _mi in _qtl_miss:
                    st.markdown(f"- **{_mn}** · ISIN: `{_mi}`")

    # ── Cache key: invalidate when portfolio/profile/fund-data changes ──────
    _pdf_cache_key = (f"{_ptf_key}|{len(df_act)}|{len(_fd_live)}"
                      f"|unp{int(_print_unp)}|qtl{int(_qtl_charts_pdf)}"
                      + (f"|{hash(tuple(sorted(df_act['nome'].tolist())))}"
                         if _is_free else ""))

    # ── Auto-generate PDF for stable portfolios (FULL/SHORT/SUGGERITO) ──────
    # For LIBERO the weights change on every interaction, so we keep the button.
    if not _is_free and st.session_state.get("_pdf_cache_key") != _pdf_cache_key:
        _fname_auto = (f"Azimut_{ptf_label.replace(' ','_').replace('—','')}"
                       f"_{profile}_{datetime.date.today().strftime('%Y%m%d')}.pdf")
        _pb_auto = st.progress(0, text="⚡ Genero PDF…")
        def _upd_auto(v, txt="⚡ Genero PDF…"):
            _pb_auto.progress(min(float(v), 1.0), text=txt)
        try:
            _pdf_auto = generate_pdf(
                df_act, wcol, profile, ptf_label, _fd_live,
                fida_df=fida_df, factbook_data=factbook_data,
                cache_date=cache_date, print_unp=_print_unp,
                qtl_charts=_qtl_charts_pdf,
                _progress_cb=_upd_auto)
            _upd_auto(1.0, "✅ PDF pronto")
            st.session_state["_pdf_bytes_ready"]  = _pdf_auto
            st.session_state["_pdf_fname_ready"]  = _fname_auto
            st.session_state["_pdf_lbl"]          = (
                f"{len(_fd_live)} schede da FondiDoc" if _fd_live
                else "dati base (lancia Aggiorna Dati per arricchire)")
            st.session_state["_pdf_cache_key"]    = _pdf_cache_key
        except Exception as _pe:
            st.error(f"Errore generazione PDF: {_pe}")
        _pb_auto.empty()

    # ── Export AdvisorElite CSV
    with col_btn:
        try:
            _ae_lines = ['ISIN,Amount,']
            for _, _r in df_act.iterrows():
                _ae_isin = (_ae_fm := {**MANUAL_ISIN_OVERRIDES, **{r['nome']:str(r['isin']).strip() for _,r in raw.get('FIDA',pd.DataFrame()).iterrows() if r.get('isin')}}).get(_r.get('nome',''),'')
                _ae_peso = round(float(_r.get(wcol, 0)) * 100)
                if _ae_isin and _ae_peso > 0:
                    _ae_lines.append(str(_ae_isin) + ',' + str(_ae_peso))
            _ae_csv = '\n'.join(_ae_lines).encode('utf-8')
            st.download_button(
                'File per AdvisorElite',
                data=_ae_csv,
                file_name='file per advisorelite.csv',
                mime='text/csv',
                use_container_width=True,
            )
        except Exception:
            pass

    with col_btn:
        if st.session_state.get("_pdf_bytes_ready") and not _is_free:
            # One-click download for stable portfolios
            st.download_button(
                "📥  Scarica Report PDF",
                data=st.session_state["_pdf_bytes_ready"],
                file_name=st.session_state.get("_pdf_fname_ready", "report.pdf"),
                mime="application/pdf",
                use_container_width=True,
                type="primary",
            )
            st.caption(f"✅ {st.session_state.get('_pdf_lbl','PDF pronto')}")
        else:
            # LIBERO: manual generate (portfolio changes at every interaction)
            if st.session_state.get("_pdf_bytes_ready") and _is_free:
                st.download_button(
                    "📥  Scarica Report PDF",
                    data=st.session_state["_pdf_bytes_ready"],
                    file_name=st.session_state.get("_pdf_fname_ready", "report.pdf"),
                    mime="application/pdf",
                    use_container_width=True,
                )
                st.caption("Clicca 'Genera' per aggiornare con i pesi attuali.")
            if st.button("⚡  Genera PDF", use_container_width=True, type="primary"):
                for _k in ("_pdf_bytes_ready", "_pdf_fname_ready", "_pdf_lbl"):
                    st.session_state.pop(_k, None)
                pb = st.progress(0, text="Scarico dati FondiDoc…")
                def upd(v): pb.progress(v, text=f"FondiDoc: {int(v*100)}%…")
                fund_data = fetch_all_fund_data(df_act, fida_urls, upd)
                save_fund_cache(fund_data)
                st.session_state["_scomp_fd"] = fund_data
                pb.progress(0.0, text="⚡ Genero PDF…")
                def _upd_pdf(v, txt="⚡ Genero PDF…"):
                    pb.progress(min(float(v), 1.0), text=txt)
                try:
                    pdf_bytes = generate_pdf(
                        df_act, wcol, profile, ptf_label, fund_data,
                        fida_df=fida_df, factbook_data=factbook_data,
                        cache_date=cache_date, print_unp=_print_unp,
                        qtl_charts=_qtl_charts_pdf,
                        _progress_cb=_upd_pdf)
                    fname = (f"Azimut_{ptf_label.replace(' ','_')}_{profile}_"
                             f"{datetime.date.today().strftime('%Y%m%d')}.pdf")
                    st.session_state["_pdf_bytes_ready"]  = pdf_bytes
                    st.session_state["_pdf_fname_ready"]  = fname
                    st.session_state["_pdf_lbl"]          = f"{len(fund_data)} schede"
                    st.session_state["_pdf_cache_key"]    = _pdf_cache_key
                except Exception as _pe:
                    st.error(f"Errore PDF: {_pe}")
                pb.empty()
                st.rerun()

    st.markdown("<br><br>",unsafe_allow_html=True)


# ── Entry point ─────────────────────────────────────────────────────────────
# Call main() directly (not behind __name__ guard) so it always runs on
# Streamlit Cloud regardless of how the script is executed.
import traceback as _tb
try:
    main()
except BaseException as _e:
    # Re-raise Streamlit-internal control-flow exceptions so the framework
    # can handle them (RerunException → st.rerun(), StopException → st.stop(),
    # SystemExit, KeyboardInterrupt, etc.).
    _ename = type(_e).__name__
    if _ename in ("RerunException", "StopException") or \
       isinstance(_e, (SystemExit, KeyboardInterrupt)):
        raise
    # For everything else, show a readable error in the UI.
    try:
        st.error(f"**Errore imprevisto:** {_ename}: {_e}")
        with st.expander("🔍 Dettaglio tecnico (per il debug)", expanded=True):
            st.code(_tb.format_exc())
    except Exception:
        pass
